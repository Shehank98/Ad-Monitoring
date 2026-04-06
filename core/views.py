import io
import json
import logging
import os
import uuid
import pandas as pd
from datetime import date as date_cls, datetime

from django.conf import settings as django_settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Max, Min, Count
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

logger = logging.getLogger(__name__)

from accounts.decorators import role_required
from accounts.models import User
from accounts.views import create_user, edit_user, user_list

from .forms import AccountForm, ChannelForm, MonitoringUploadForm, ScheduleUploadForm
from .models import (
    Account, AuditLog, BrandMapping, Channel, ChannelOfficer,
    LMRBRow, MatchResult, MonitoringData, Schedule, ScheduleRow,
    SponsorshipLmrbAssignment, SummaryReportMeta, SystemSetting,
    TCRow, TransmissionReport,
    _ensure_defaults, get_setting, get_setting_list,
)


# ── JSON serialisation helper ─────────────────────────────────────────────────

class _JsonEnc(json.JSONEncoder):
    """Handles date, datetime, and Decimal so view data serialises cleanly."""
    def default(self, o):
        from decimal import Decimal
        if isinstance(o, (date_cls, datetime)):
            return o.isoformat()
        if isinstance(o, Decimal):
            return float(o)
        return super().default(o)


def _to_js(data):
    """Serialize Python data to a JS-safe JSON string (replaces </script> escapes)."""
    return json.dumps(data, cls=_JsonEnc).replace('</', '<\\/')


def _audit(request, action: str, detail: str = ''):
    """Record a sensitive admin action to AuditLog."""
    ip = (request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
          or request.META.get('REMOTE_ADDR'))
    AuditLog.objects.create(user=request.user, action=action, detail=detail, ip=ip or None)


def _theme_adtype(theme: str, brand_map: dict, spon_kw: list) -> str:
    """Classify a theme as 'commercial', 'sponsorship', or 'other'.

    Priority: BrandMapping lookup first, then keyword substring match.
    spon_kw is a list of lowercase keyword strings from SystemSetting.
    """
    key = (theme or '').lower().strip()
    tp = brand_map.get(key)
    if tp:
        return tp
    for kw in spon_kw:
        if kw and kw.lower() in key:
            return 'sponsorship'
    return 'other'


# ── Branding helpers ──────────────────────────────────────────────────────────

def _branding_url(asset_type: str) -> str:
    """Return the media URL for a branding asset (logo/tartan), or empty string if not uploaded."""
    branding_dir = os.path.join(django_settings.MEDIA_ROOT, 'branding')
    for ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg'):
        path = os.path.join(branding_dir, f'{asset_type}{ext}')
        if os.path.exists(path):
            return django_settings.MEDIA_URL + f'branding/{asset_type}{ext}'
    return ''


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_admin(user):
    return user.role in ('super_admin', 'admin')


def _account_access(user, account_id):
    if _is_admin(user):
        return True
    return user.accounts.filter(id=account_id).exists()


def _account_qs(user):
    if _is_admin(user):
        return Account.objects.all()
    return user.accounts.all()


def _detect_schedule_meta(df):
    """Auto-detect month, start_date, end_date from a schedule DataFrame."""
    df = df.copy()
    df.columns = df.columns.str.strip()
    month = ''
    start_date = end_date = None
    if 'Date' in df.columns:
        dates = pd.to_datetime(df['Date'], errors='coerce').dropna()
        if not dates.empty:
            start_date = dates.min().date()
            end_date   = dates.max().date()
            month      = dates.min().strftime('%B %Y')
    return month, start_date, end_date


def _detect_monitoring_meta(df, data_type):
    """
    Auto-detect channels and per-channel date ranges from a monitoring DataFrame.

    Returns list of dicts: [{channel, start_date, end_date, row_count}, ...]
    """
    df = df.copy()
    df.columns = df.columns.str.strip()

    if data_type == 'mediawatch':
        if {'Dd', 'Mn', 'Yr'}.issubset(df.columns):
            df['_date'] = pd.to_datetime(
                df['Yr'].astype(str) + '-' +
                df['Mn'].astype(str).str.zfill(2) + '-' +
                df['Dd'].astype(str).str.zfill(2),
                errors='coerce',
            )
        elif 'Date' in df.columns:
            df['_date'] = pd.to_datetime(df['Date'], errors='coerce')
        else:
            df['_date'] = pd.NaT
    else:  # maponline
        date_col = 'Prg Date' if 'Prg Date' in df.columns else 'Date'
        df['_date'] = pd.to_datetime(df.get(date_col, pd.Series(dtype='object')), errors='coerce')

    ch_col = 'Channel' if 'Channel' in df.columns else None
    channels = [str(c).strip() for c in df[ch_col].dropna().unique() if str(c).strip()] \
               if ch_col else ['Unknown']

    result = []
    for ch in channels:
        ch_df  = df[df[ch_col].astype(str).str.strip() == ch] if ch_col else df
        dates  = ch_df['_date'].dropna()
        result.append({
            'channel':    ch,
            'start_date': dates.min().date() if not dates.empty else None,
            'end_date':   dates.max().date() if not dates.empty else None,
            'row_count':  len(ch_df),
        })
    return result


def _safe_int(val):
    try:
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            return int(float(val))
    except (ValueError, TypeError):
        pass
    return None


def _safe_str(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    return str(val).strip()


def _safe_date(val):
    try:
        ts = pd.to_datetime(val, errors='coerce')
        return ts.date() if not pd.isna(ts) else None
    except Exception:
        return None


def _safe_decimal(val):
    from decimal import Decimal, InvalidOperation
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        return Decimal(str(val))
    except (InvalidOperation, ValueError, TypeError):
        return None


# ── Dashboard ─────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    from django.db.models import Count as _Count
    from django.utils import timezone as _tz
    user = request.user
    role = user.role

    if role == 'channel_officer':
        return redirect('/dashboard/channel-officer/')

    ctx  = {'user': user, 'today': _tz.localdate()}

    if role in ('super_admin', 'admin'):
        ctx['total_users']      = User.objects.count()
        ctx['active_users']     = User.objects.filter(is_active=True).count()
        ctx['total_schedules']  = Schedule.objects.count()
        ctx['total_mon']        = MonitoringData.objects.count()
        ctx['total_accounts']   = Account.objects.count()
        ctx['total_tc']         = TransmissionReport.objects.count()
        ctx['recent_schedules'] = (
            Schedule.objects.select_related('account', 'uploaded_by')
            .order_by('-uploaded_at')[:5]
        )
        ctx['recent_mon'] = (
            MonitoringData.objects.select_related('uploaded_by', 'account')
            .order_by('-uploaded_at')[:4]
        )
        ctx['recent_tc'] = (
            TransmissionReport.objects.select_related('account', 'uploaded_by')
            .order_by('-uploaded_at')[:4]
        )
        ctx['accounts_overview'] = list(
            Account.objects.annotate(
                sch_count=_Count('schedules', distinct=True),
                mon_count=_Count('monitoring_data', distinct=True),
                tc_count=_Count('tc_reports', distinct=True),
            ).order_by('name')
        )

        # ── A1: System health ─────────────────────────────────────────────────
        from django.db.models import Q as _Q, Max as _Max
        from datetime import timedelta as _td

        _total_results   = MatchResult.objects.count()
        _matched_results = MatchResult.objects.filter(status='matched').count()
        ctx['health_match_rate']    = (
            round(_matched_results / _total_results * 100) if _total_results else None
        )
        ctx['health_total_results'] = _total_results

        ctx['missing_lmrb'] = [a for a in ctx['accounts_overview']
                               if a.sch_count > 0 and a.mon_count == 0]
        ctx['missing_sch']  = [a for a in ctx['accounts_overview']
                               if a.mon_count > 0 and a.sch_count == 0]

        _stale_cutoff = _tz.now() - _td(days=7)
        _last_sch = dict(
            Schedule.objects.values('account_id')
            .annotate(last=_Max('uploaded_at'))
            .values_list('account_id', 'last')
        )
        _last_mon = dict(
            MonitoringData.objects.values('account_id')
            .annotate(last=_Max('uploaded_at'))
            .values_list('account_id', 'last')
        )
        _stale = []
        for _a in ctx['accounts_overview']:
            _dates = [d for d in [_last_sch.get(_a.id), _last_mon.get(_a.id)] if d]
            if _dates and max(_dates) < _stale_cutoff:
                _stale.append({'name': _a.name, 'days': (_tz.now() - max(_dates)).days})
        ctx['stale_accounts'] = _stale

        # ── A2: Per-account compliance % ─────────────────────────────────────
        _total_by_acct = dict(
            MatchResult.objects.values('schedule_row__account_id')
            .annotate(n=_Count('id'))
            .values_list('schedule_row__account_id', 'n')
        )
        _matched_by_acct = dict(
            MatchResult.objects.filter(status='matched')
            .values('schedule_row__account_id')
            .annotate(n=_Count('id'))
            .values_list('schedule_row__account_id', 'n')
        )
        for _a in ctx['accounts_overview']:
            _tot = _total_by_acct.get(_a.id, 0)
            _a.match_rate = (
                round(_matched_by_acct.get(_a.id, 0) / _tot * 100) if _tot else None
            )

        # ── A3: Recent engine runs ────────────────────────────────────────────
        ctx['recent_runs'] = list(
            MatchResult.objects.values(
                'schedule_row__account__name',
                'schedule_row__channel',
                'schedule_row__month',
            )
            .annotate(
                last_run=_Max('run_at'),
                total=_Count('id'),
                matched=_Count('id', filter=_Q(status='matched')),
            )
            .order_by('-last_run')[:5]
        )

    elif role == 'team_head':
        my_accounts = user.accounts.all()
        sch = Schedule.objects.filter(account__in=my_accounts).select_related('account')
        ctx['my_accounts']    = my_accounts.annotate(
            sch_count=_Count('schedules', distinct=True),
            mon_count=_Count('monitoring_data', distinct=True),
        )
        ctx['schedules']      = sch.order_by('-uploaded_at')[:5]
        ctx['schedule_count'] = sch.count()
        ctx['mon_count']      = MonitoringData.objects.filter(account__in=my_accounts).count()
        ctx['recent_mon']     = (
            MonitoringData.objects.filter(account__in=my_accounts)
            .select_related('account').order_by('-uploaded_at')[:4]
        )

    elif role == 'planner':
        my_accounts = user.accounts.all()
        sch = Schedule.objects.filter(account__in=my_accounts).select_related('account')
        ctx['my_accounts']    = my_accounts.annotate(
            sch_count=_Count('schedules', distinct=True),
        )
        ctx['my_schedules']   = sch.order_by('-uploaded_at')[:5]
        ctx['schedule_count'] = sch.count()

    elif role == 'operations':
        my_accounts = user.accounts.all()
        mon = MonitoringData.objects.filter(account__in=my_accounts)
        ctx['my_accounts']    = my_accounts.annotate(
            mon_count=_Count('monitoring_data', distinct=True),
        )
        ctx['my_uploads']   = mon.select_related('account').order_by('-uploaded_at')[:5]
        ctx['upload_count'] = mon.filter(uploaded_by=user).count()
        ctx['total_mon']    = mon.count()

    return render(request, 'dashboard/home.html', ctx)


# ── Account management ────────────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin'])
def account_list(request):
    accounts = Account.objects.all()
    form     = AccountForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            form = AccountForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, f'Account "{form.cleaned_data["name"]}" added.')
                return redirect('/dashboard/accounts/')
        elif action == 'delete':
            acc_id = request.POST.get('account_id')
            acc    = get_object_or_404(Account, id=acc_id)
            acc.delete()
            messages.success(request, f'Account "{acc.name}" deleted.')
            return redirect('/dashboard/accounts/')

    return render(request, 'admin_panel/accounts.html', {'accounts': accounts, 'form': form})


# ── Channel management ────────────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin'])
def channel_list(request):
    channels = Channel.objects.all()
    form     = ChannelForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            form = ChannelForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, f'Channel "{form.cleaned_data["name"]}" added.')
                return redirect('/dashboard/channels/')
        elif action == 'delete':
            ch_id = request.POST.get('channel_id')
            ch    = get_object_or_404(Channel, id=ch_id)
            ch.delete()
            messages.success(request, f'Channel "{ch.name}" deleted.')
            return redirect('/dashboard/channels/')

    return render(request, 'admin_panel/channels.html', {'channels': channels, 'form': form})


# ── Schedules ─────────────────────────────────────────────────────────────────

@login_required
def schedule_list(request):
    user = request.user
    qs   = Schedule.objects.select_related('account', 'uploaded_by')
    if not _is_admin(user):
        qs = qs.filter(account__in=user.accounts.all())

    account_id = request.GET.get('account')
    channel    = request.GET.get('channel', '').strip()
    month      = request.GET.get('month', '').strip()
    media_type = request.GET.get('media_type', '').strip()   # 'tv' | 'radio' | 'press' | ''
    q_search   = request.GET.get('q', '').strip()

    # Apply account + media_type first so months_list reflects the current scope
    if account_id:
        qs = qs.filter(account_id=account_id)
    if media_type:
        qs = qs.filter(media_type=media_type)

    # Build months list (before channel/month/q filters) for the month picker
    raw_months = (
        qs.values('month')
        .annotate(count=Count('id'), channel_count=Count('channel', distinct=True))
        .order_by()
    )
    def _month_key(m):
        try:
            return datetime.strptime(m['month'], '%B %Y')
        except Exception:
            return datetime.min
    sorted_months = sorted(raw_months, key=_month_key, reverse=True)  # newest first

    # Group months by year for the card grid
    from collections import defaultdict as _dd
    months_by_year = {}
    for m in sorted_months:
        try:
            year = datetime.strptime(m['month'], '%B %Y').year
        except Exception:
            year = 0
        months_by_year.setdefault(year, []).append(m)
    months_by_year_list = sorted(months_by_year.items(), reverse=True)  # newest year first

    # Apply remaining filters
    if channel:
        qs = qs.filter(channel__icontains=channel)
    if month:
        qs = qs.filter(month__icontains=month)
    if q_search:
        qs = qs.filter(original_filename__icontains=q_search)

    # Group filtered schedules by channel for the accordion view
    channel_groups = {}
    for s in qs.order_by('channel', 'schedule_number', '-version'):
        channel_groups.setdefault(s.channel, []).append(s)
    channel_groups_list = sorted(channel_groups.items())  # alphabetical by channel name

    today    = date_cls.today()
    accounts = _account_qs(user)
    channels = Channel.objects.all()
    return render(request, 'schedules/list.html', {
        'schedules':           qs,
        'channel_groups_list': channel_groups_list,
        'months_by_year_list': months_by_year_list,
        'accounts':            accounts,
        'channels':            channels,
        'filters':             {'account': account_id, 'channel': channel, 'month': month,
                                'media_type': media_type, 'q': q_search},
        'today':               today,
    })


@login_required
@role_required(['planner', 'super_admin', 'admin', 'team_head', 'operations'])
def schedule_upload(request):
    user = request.user
    form = ScheduleUploadForm()
    if not _is_admin(user):
        form.fields['account'].queryset = user.accounts.all()

    if request.method == 'POST':
        form = ScheduleUploadForm(request.POST, request.FILES)
        if not _is_admin(user):
            form.fields['account'].queryset = user.accounts.all()

        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                from verification.schedule_converter import convert_schedule_excel
                df = convert_schedule_excel(excel_file)
                if df is None or df.empty:
                    # Not pivot format - fall back to plain flat read
                    excel_file.seek(0)
                    df = pd.read_excel(excel_file)
                    df.columns = df.columns.str.strip()
            except Exception as e:
                messages.error(request, f'Cannot read Excel file: {e}')
                return render(request, 'schedules/upload.html', {'form': form})

            row_count   = len(df)
            account     = form.cleaned_data['account']
            channel_obj = form.cleaned_data['channel']

            month, start_date, end_date = _detect_schedule_meta(df)
            if not month:
                month = 'Unknown'

            sched_number   = form.cleaned_data['schedule_number'].strip()
            channel_name   = channel_obj.name
            from core.models import parse_channel_media_type as _pmt
            _sch_media_type, _ = _pmt(channel_name)
            dup_action     = request.POST.get('duplicate_action', '')

            # Check for duplicate ref number within same (account, channel, month)
            existing_same_ref = Schedule.objects.filter(
                account=account, channel=channel_name,
                month=month, schedule_number=sched_number,
                is_superseded=False,
            ).first()

            if existing_same_ref and dup_action not in ('replace', 'keep_both'):
                # Re-render the form with a confirmation prompt
                return render(request, 'schedules/upload.html', {
                    'form':          form,
                    'schedule_col_guide': [
                        ('PROGRAMME',   'Programme / show name',              True),
                        ('DAY',         'Day abbreviation (MON, TUE…)',       True),
                        ('TIME',        'Ad start time (HH:MM:SS)',           True),
                        ('End Time',    'Ad end time - column after TIME',    False),
                        ('DUR',         'Duration in seconds',                True),
                        ('BRAND',       'Brand name (optional inline col)',   False),
                        ('[date cols]', 'Date values → spot count per day',   True),
                    ],
                    'dup_ref_existing': existing_same_ref,
                    'dup_ref_form_data': request.POST,
                    'dup_ref_file_key':  excel_file.name,
                })

            # Handle replacement: mark existing as superseded + release matched LMRB
            if existing_same_ref and dup_action == 'replace':
                from core.models import LMRBRow as _LR, ScheduleRow as _SR
                # Release matched LMRB rows linked to superseded schedule rows
                sr_ids = list(existing_same_ref.rows.values_list('id', flat=True))
                _LR.objects.filter(schedule_matches__in=sr_ids).update(
                    is_matched=False, matched_at=None
                )
                existing_same_ref.rows.update(
                    is_matched=False, matched_lmrb=None, matched_at=None,
                    is_manual_matched=False,
                )
                existing_same_ref.is_superseded = True
                existing_same_ref.save(update_fields=['is_superseded'])
                messages.warning(request,
                    f'Schedule #{sched_number} (v{existing_same_ref.version}) '
                    f'was superseded. Previously matched LMRB rows have been released for re-matching.')

            # Version = next version per (account, channel, month)
            last_ver = (
                Schedule.objects
                .filter(account=account, channel=channel_name, month=month)
                .aggregate(Max('version'))['version__max'] or 0
            )
            version = last_ver + 1

            # ── Optional makeup fields ─────────────────────────────────────────
            makeup_end_date_raw = request.POST.get('makeup_end_date', '').strip()
            makeup_end_date = None
            if makeup_end_date_raw:
                from datetime import date as _date_type
                try:
                    from datetime import datetime as _dt
                    makeup_end_date = _dt.strptime(makeup_end_date_raw, '%Y-%m-%d').date()
                except ValueError:
                    pass

            parent_schedule_id_raw = request.POST.get('parent_schedule', '').strip()
            parent_schedule_obj = None
            if parent_schedule_id_raw:
                try:
                    parent_schedule_obj = Schedule.objects.get(
                        pk=int(parent_schedule_id_raw), account=account,
                    )
                except (Schedule.DoesNotExist, ValueError):
                    pass

            excel_file.seek(0)
            schedule = Schedule(
                account           = account,
                channel           = channel_name,
                media_type        = _sch_media_type,
                month             = month,
                schedule_number   = sched_number,
                original_filename = excel_file.name,
                row_count         = row_count,
                start_date        = start_date,
                end_date          = end_date,
                version           = version,
                uploaded_by       = user,
                makeup_end_date   = makeup_end_date,
                parent_schedule   = parent_schedule_obj,
            )
            schedule.file.save(excel_file.name, excel_file)
            schedule.save()

            # ── Parse Schedule rows into DB ────────────────────────────────────
            # Collect brand→sponsorship_type assignments submitted via the form.
            # Each brand can have MULTIPLE types selected (checkboxes), stored
            # pipe-separated, e.g. "-Intro|-Extro" for Opening/Closing.
            brand_type_map = {}
            for key in request.POST.keys():
                if key.startswith('spon_type_'):
                    brand_name = key[len('spon_type_'):]
                    vals = [v.strip() for v in request.POST.getlist(key) if v.strip()]
                    if vals:
                        brand_type_map[brand_name] = '|'.join(vals)
            _parse_schedule_rows(df, schedule, account, channel_name, month,
                                 brand_type_map=brand_type_map)

            messages.success(request,
                f'Schedule #{schedule.schedule_number} v{version} for {account} '
                f'({month}, {start_date} → {end_date}) uploaded - {row_count:,} rows.')

            # Auto-run verification for all available scopes
            try:
                from verification.engine import auto_run_all_for_account
                auto_run_all_for_account(account.id)
            except Exception:
                pass

            # Notify channel officers of any missed spots discovered
            try:
                _notify_missed_spots_whatsapp(account)
            except Exception:
                pass

            return redirect(f'/dashboard/schedules/upload/?uploaded={schedule.id}')

    col_guide = [
        ('PROGRAMME',   'Programme / show name',              True),
        ('DAY',         'Day abbreviation (MON, TUE…)',       True),
        ('TIME',        'Ad start time (HH:MM:SS)',           True),
        ('End Time',    'Ad end time - column after TIME',    False),
        ('DUR',         'Duration in seconds',                True),
        ('BRAND',       'Brand name (optional inline col)',   False),
        ('[date cols]', 'Date values → spot count per day',   True),
    ]
    # Provide available sponsorship type options from system settings
    spon_type_options = get_setting_list('lmrb_sponsorship_keywords')

    # Upload summary: shown after a successful upload redirect
    upload_summary = None
    uploaded_id = request.GET.get('uploaded')
    if uploaded_id:
        try:
            _sch = Schedule.objects.get(pk=uploaded_id)
            from django.db.models import Count as _Cnt
            _rows = (
                ScheduleRow.objects
                .filter(schedule=_sch)
                .values('ad_type', 'brand', 'duration')
                .annotate(spots=_Cnt('id'))
                .order_by('ad_type', 'brand', 'duration')
            )
            _commercial = [r for r in _rows if r['ad_type'] == 'COMMERCIAL BENEFITS']
            _sponsorship = [r for r in _rows if r['ad_type'] == 'SPONSORSHIP']
            upload_summary = {
                'schedule':    _sch,
                'commercial':  _commercial,
                'sponsorship': _sponsorship,
                'total_commercial': sum(r['spots'] for r in _commercial),
                'total_sponsorship': sum(r['spots'] for r in _sponsorship),
            }
        except Schedule.DoesNotExist:
            pass

    # Provide schedules that can be selected as parent (same account, all months)
    # for the makeup/reschedule parent_schedule dropdown.
    parent_schedule_choices = list(
        Schedule.objects
        .filter(account__in=(_is_admin(user) and Account.objects.all() or user.accounts.all()))
        .order_by('account__name', '-uploaded_at')
        .values('id', 'account__name', 'channel', 'month', 'schedule_number', 'version')
    )

    return render(request, 'schedules/upload.html', {
        'form': form,
        'schedule_col_guide': col_guide,
        'spon_type_options': spon_type_options,
        'upload_summary': upload_summary,
        'parent_schedule_choices': parent_schedule_choices,
    })


def _parse_schedule_rows(df, schedule, account, channel, month, brand_type_map=None):
    """Parse a schedule DataFrame and bulk-create ScheduleRow records."""
    brand_type_map = brand_type_map or {}
    rows = []
    for _, r in df.iterrows():
        ad_type = _safe_str(r.get('Advertisement_Type', '')).upper().strip()
        # Normalise: 'SPONSORSHIP BENEFITS' → 'SPONSORSHIP'
        if ad_type == 'SPONSORSHIP BENEFITS':
            ad_type = 'SPONSORSHIP'
        if ad_type not in ('COMMERCIAL BENEFITS', 'SPONSORSHIP'):
            continue
        brand = _safe_str(r.get('Brand', ''))
        spon_type = ''
        if ad_type == 'SPONSORSHIP':
            spon_type = brand_type_map.get(brand, '')
        rows.append(ScheduleRow(
            schedule         = schedule,
            account          = account,
            channel          = channel,
            month            = month,
            brand            = brand,
            product          = _safe_str(r.get('Product', '')),
            programme        = _safe_str(r.get('Programme', '')),
            date             = _safe_date(r.get('Date')),
            start_time       = _safe_str(r.get('Start_Time', '')),
            end_time         = _safe_str(r.get('End_Time', '')),
            duration         = _safe_int(r.get('Duration')),
            ad_type          = ad_type,
            sponsorship_type = spon_type,
        ))
    if rows:
        ScheduleRow.objects.bulk_create(rows, batch_size=500)


@login_required
@require_POST
def schedule_detect(request):
    """AJAX: parse an uploaded schedule file and return detected metadata."""
    excel_file = request.FILES.get('file')
    if not excel_file:
        return JsonResponse({'ok': False, 'error': 'No file provided'})
    try:
        from verification.schedule_converter import convert_schedule_excel
        df = convert_schedule_excel(excel_file)
        pivot = df is not None and not df.empty
        if not pivot:
            excel_file.seek(0)
            df = pd.read_excel(excel_file)
        month, start_date, end_date = _detect_schedule_meta(df)

        # Extract unique sponsorship brands if Advertisement_Type column exists
        spon_brands = []
        if 'Advertisement_Type' in df.columns and 'Brand' in df.columns:
            spon_mask = df['Advertisement_Type'].astype(str).str.strip().str.upper().isin(
                ['SPONSORSHIP BENEFITS', 'SPONSORSHIP']
            )
            spon_brands = sorted(
                b for b in df.loc[spon_mask, 'Brand'].dropna().unique()
                if str(b).strip()
            )

        return JsonResponse({
            'ok':          True,
            'month':       month,
            'start_date':  str(start_date) if start_date else '',
            'end_date':    str(end_date)   if end_date   else '',
            'row_count':   len(df),
            'is_pivot':    pivot,
            'spon_brands': spon_brands,
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required
def schedule_download(request, pk):
    """Serve the original schedule Excel file as a download."""
    schedule = get_object_or_404(Schedule, pk=pk)
    if not _is_admin(request.user) and schedule.account not in _account_qs(request.user):
        return HttpResponse('Access denied', status=403)
    try:
        file_path = schedule.file.path
        if not os.path.exists(file_path):
            return HttpResponse('File not found on server.', status=404)
        return FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename=schedule.original_filename or os.path.basename(file_path),
        )
    except Exception as e:
        return HttpResponse(f'Download failed: {e}', status=500)


@login_required
def schedule_delete(request, pk):
    schedule = get_object_or_404(Schedule, pk=pk)
    user     = request.user
    today    = date_cls.today()

    if _is_admin(user) or (schedule.uploaded_by == user and schedule.uploaded_at.date() == today):
        # Before deleting: clean up all associated data to maintain integrity
        sch_row_ids = list(schedule.rows.values_list('id', flat=True))

        if sch_row_ids:
            # Unlock any LMRBRows matched to schedule rows in this schedule
            LMRBRow.objects.filter(matched_schedule_id__in=sch_row_ids).update(
                is_matched=False, matched_schedule=None, matched_at=None,
            )
            # Reset TCRow schedule-match state for these schedule rows
            TCRow.objects.filter(matched_schedule_id__in=sch_row_ids).update(
                is_schedule_matched=False, matched_schedule=None,
            )
            # Delete MatchResult records for this schedule's rows
            MatchResult.objects.filter(schedule_row_id__in=sch_row_ids).delete()

        # Also delete MatchResult records scoped to this channel/month (engine-level)
        MatchResult.objects.filter(
            account=schedule.account, channel=schedule.channel, month=schedule.month,
        ).exclude(status='manual_match').delete()

        schedule.file.delete(save=False)
        schedule.delete()
        messages.success(request, 'Schedule deleted.')
    else:
        messages.error(request, 'You can only delete schedules you uploaded today.')
    return redirect('/dashboard/schedules/')


@login_required
@role_required(['operations', 'super_admin', 'admin', 'team_head'])
def schedule_lock(request):
    """AJAX POST — toggle the locked state of all Schedule records for one
    (account_id, channel, month) campaign scope.

    Body params:
        account_id  int
        channel     str
        month       str
        locked      '1' to lock, '0' to unlock
    Returns JSON {ok, locked, updated}.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)

    account_id = request.POST.get('account_id', '').strip()
    channel    = request.POST.get('channel', '').strip()
    month      = request.POST.get('month', '').strip()
    locked     = request.POST.get('locked', '1') == '1'

    if not (account_id and channel and month):
        return JsonResponse({'ok': False, 'error': 'Missing required params'}, status=400)

    updated = Schedule.objects.filter(
        account_id=account_id, channel=channel, month=month
    ).update(is_locked=locked)

    action = 'locked' if locked else 'unlocked'
    print(f"[schedule_lock] {action} {updated} schedule(s) for "
          f"account_id={account_id} channel={channel!r} month={month!r} by {request.user}")
    return JsonResponse({'ok': True, 'locked': locked, 'updated': updated})


# ── Monitoring data ───────────────────────────────────────────────────────────

def _cleanup_duplicate_monitoring_batches():
    """
    Auto-remove duplicate MonitoringData header records.

    When the same file (same account + original_filename) has been uploaded more
    than once, only the most-recently-uploaded batch (file_group_id) is kept;
    older batches are deleted.  LMRBRow data is NOT touched — the SHA-256
    dedup_key already prevents actual row duplication, so the underlying data
    is correct regardless of how many header records exist.

    Returns the number of MonitoringData records deleted.
    """
    from django.db.models import Count, Max

    # Find (account_id, original_filename) combos that have more than one distinct batch
    dupes = (
        MonitoringData.objects
        .values('account_id', 'original_filename')
        .annotate(batch_count=Count('file_group_id', distinct=True))
        .filter(batch_count__gt=1)
    )

    total_deleted = 0
    for d in dupes:
        # Identify the most-recent batch for this (account, filename)
        latest = (
            MonitoringData.objects
            .filter(account_id=d['account_id'], original_filename=d['original_filename'])
            .aggregate(latest_upload=Max('uploaded_at'))
        )['latest_upload']

        # Fetch the file_group_id of the winning batch
        keep_fgid = (
            MonitoringData.objects
            .filter(
                account_id=d['account_id'],
                original_filename=d['original_filename'],
                uploaded_at=latest,
            )
            .values_list('file_group_id', flat=True)
            .first()
        )

        deleted, _ = MonitoringData.objects.filter(
            account_id=d['account_id'],
            original_filename=d['original_filename'],
        ).exclude(file_group_id=keep_fgid).delete()

        if deleted:
            print(f"[cleanup_monitoring] removed {deleted} duplicate MonitoringData record(s) "
                  f"for account_id={d['account_id']} filename={d['original_filename']!r} "
                  f"(kept batch {keep_fgid})")
        total_deleted += deleted

    return total_deleted


@login_required
def monitoring_list(request):
    # Auto-clean duplicate MonitoringData batches (same file uploaded twice)
    _cleanup_duplicate_monitoring_batches()

    user = request.user
    qs   = MonitoringData.objects.select_related('account', 'uploaded_by')
    if not _is_admin(user):
        qs = qs.filter(account__in=user.accounts.all())

    dtype      = request.GET.get('type', '')
    channel    = request.GET.get('channel', '').strip()
    account_id = request.GET.get('account', '')
    if dtype:
        qs = qs.filter(data_type=dtype)
    if channel:
        qs = qs.filter(channel__icontains=channel)
    if account_id:
        qs = qs.filter(account_id=account_id)

    # Group records by file_group_id (newest first) so multi-channel uploads
    # appear together and can be deleted as one unit.
    sorted_data = list(qs.order_by('-uploaded_at'))
    groups_dict  = {}
    groups_order = []
    for d in sorted_data:
        fgid = d.file_group_id
        if fgid not in groups_dict:
            groups_dict[fgid]  = []
            groups_order.append(fgid)
        groups_dict[fgid].append(d)
    data_groups = [groups_dict[fgid] for fgid in groups_order]

    print(f"[monitoring_list] total_records={len(sorted_data)}  groups={len(data_groups)}")
    for i, grp in enumerate(data_groups):
        chs = [d.channel for d in grp]
        print(f"  group {i+1}: file_group_id={grp[0].file_group_id}  channels={chs}  uploaded_at={grp[0].uploaded_at}")

    cov_base = MonitoringData.objects.select_related('account')
    if not _is_admin(user):
        cov_base = cov_base.filter(account__in=user.accounts.all())
    coverage = (
        cov_base
        .values('account__name', 'channel', 'data_type')
        .annotate(from_date=Min('start_date'), to_date=Max('end_date'), uploads=Count('id'))
        .order_by('data_type', 'account__name', 'channel')
    )

    # ── Build batch summaries for display (one row per file_group_id) ─────────
    batch_summaries = []
    for fgid in groups_order:
        grp = groups_dict[fgid]
        first = grp[0]
        total_rows = sum(d.row_count or 0 for d in grp)
        all_dates = [(d.start_date, d.end_date) for d in grp if d.start_date and d.end_date]
        date_from = min(s for s, e in all_dates) if all_dates else None
        date_to   = max(e for s, e in all_dates) if all_dates else None
        ch_list   = sorted(set(d.channel for d in grp if d.channel))
        batch_summaries.append({
            'file_group_id': fgid,
            'first':         first,
            'data_type':     first.data_type,
            'account':       first.account,
            'channels':      ch_list,
            'date_from':     date_from,
            'date_to':       date_to,
            'total_rows':    total_rows,
            'uploaded_by':   first.uploaded_by,
            'uploaded_at':   first.uploaded_at,
            'group':         grp,         # still available for per-channel download
        })

    today    = date_cls.today()
    accounts = _account_qs(user)
    channels = Channel.objects.all()
    return render(request, 'monitoring/list.html', {
        'data_groups':     data_groups,
        'batch_summaries': batch_summaries,
        'coverage':        coverage,
        'filters':         {'type': dtype, 'channel': channel, 'account': account_id},
        'accounts':        accounts,
        'channels':        channels,
        'today':           today,
    })


@login_required
@role_required(['operations', 'super_admin', 'admin', 'team_head', 'planner'])
def monitoring_upload(request):
    """
    Upload a MapOnline or LMRB file.

    Channels and date ranges are AUTO-DETECTED from the file.
    One MonitoringData record is created per detected channel.
    Individual rows are parsed into LMRBRow (master table) with deduplication.
    """
    user       = request.user
    account_qs = _account_qs(user)
    form       = MonitoringUploadForm(account_queryset=account_qs)

    if request.method == 'POST':
        form = MonitoringUploadForm(request.POST, request.FILES, account_queryset=account_qs)
        if form.is_valid():
            excel_file = request.FILES['file']
            data_type  = form.cleaned_data['data_type']
            account    = form.cleaned_data['account']
            print(f"[monitoring_upload] file={excel_file.name}  data_type={data_type}  account={account}")
            try:
                df = pd.read_excel(excel_file)
                df.columns = df.columns.str.strip()
                print(f"[monitoring_upload] Excel loaded: {len(df)} rows  columns={list(df.columns)}")
            except Exception as e:
                print(f"[monitoring_upload] ERROR reading Excel: {e}")
                messages.error(request, f'Cannot read Excel file: {e}')
                return render(request, 'monitoring/upload.html', {'form': form})

            channel_metas = _detect_monitoring_meta(df, data_type)
            print(f"[monitoring_upload] detected {len(channel_metas)} channel(s): {channel_metas}")
            if not channel_metas:
                print("[monitoring_upload] ERROR: no channels detected")
                messages.error(request, 'No channels detected in the file.')
                return render(request, 'monitoring/upload.html', {'form': form})

            for meta in channel_metas:
                ch = meta['channel']
                if ch and ch != 'Unknown':
                    # Case-insensitive lookup first to avoid creating duplicates
                    # when LMRB file uses "SIRASA TV" but Channel model has "Sirasa TV".
                    if not Channel.objects.filter(name__iexact=ch).exists():
                        Channel.objects.create(name=ch)

            # ── Remove any existing MonitoringData records for the same file ──
            # LMRBRow dedup (SHA-256 key) prevents actual row duplication, but
            # without this check the header records accumulate and the list page
            # shows the same upload multiple times.
            existing_qs = MonitoringData.objects.filter(
                account=account,
                original_filename=excel_file.name,
            )
            existing_count = existing_qs.count()
            if existing_count:
                print(f"[monitoring_upload] deleting {existing_count} existing MonitoringData record(s) "
                      f"for account={account} filename={excel_file.name!r}")
                existing_qs.delete()

            group_id   = str(uuid.uuid4())
            saved_path = None
            excel_file.seek(0)
            print(f"[monitoring_upload] file_group_id={group_id}")

            for i, meta in enumerate(channel_metas):
                mon = MonitoringData(
                    account           = account,
                    data_type         = data_type,
                    channel           = meta['channel'],
                    start_date        = meta['start_date'],
                    end_date          = meta['end_date'],
                    original_filename = excel_file.name,
                    row_count         = meta['row_count'],
                    file_group_id     = group_id,
                    uploaded_by       = user,
                )
                if saved_path is None:
                    excel_file.seek(0)
                    mon.file.save(excel_file.name, excel_file, save=False)
                    saved_path = mon.file.name
                    print(f"[monitoring_upload] file saved → {saved_path}")
                else:
                    mon.file = saved_path
                mon.save()
                print(f"[monitoring_upload]   saved MonitoringData pk={mon.pk}  channel={meta['channel']}  rows={meta['row_count']}")

            # ── Parse LMRB rows into master DB table (append mode) ────────────
            # Read advertiser filter selections submitted by the user.
            # These are the advertiser names (from the Advertiser column in the
            # LMRB file) that belong to this account.  Only rows matching these
            # advertisers will be stored; all others are ignored.
            selected_advertisers = [
                v.strip() for v in request.POST.getlist('selected_advertisers')
                if v.strip()
            ]
            print(f"[monitoring_upload] parsing LMRB rows (append mode) …")
            print(f"[monitoring_upload] advertiser filter: {selected_advertisers or 'ALL'}")
            _parse_lmrb_rows(df, data_type, account, batch_id=uuid.UUID(group_id),
                             advertiser_filter=selected_advertisers or None)

            ch_names = ', '.join(m['channel'] for m in channel_metas)
            print(f"[monitoring_upload] DONE - {len(channel_metas)} channel(s): {ch_names}")
            messages.success(request,
                f'{"MapOnline" if data_type == "maponline" else "MediaWatch (LMRB)"} - '
                f'{account} - {len(channel_metas)} channel(s): {ch_names}. Uploaded successfully.')

            # Auto-run verification
            try:
                from verification.engine import auto_run_all_for_account
                print(f"[monitoring_upload] auto-running verification for account {account.id}")
                auto_run_all_for_account(account.id)
            except Exception as e:
                print(f"[monitoring_upload] auto-verification error (non-fatal): {e}")

            # WhatsApp: notify officers for any channel/month with missed spots
            try:
                _notify_missed_spots_whatsapp(account)
            except Exception as e:
                print(f"[monitoring_upload] whatsapp notify error (non-fatal): {e}")

            return redirect('/dashboard/monitoring/')

    return render(request, 'monitoring/upload.html', {'form': form})


def _parse_lmrb_rows(df, data_type, account, batch_id=None, advertiser_filter=None):
    """
    Parse monitoring DataFrame and append into the LMRBRow master table.

    APPEND MODE - rows are never deleted or replaced.
    Dedup key = SHA-256 of all meaningful columns.  If a row with the same key
    already exists it is silently skipped (bulk_create with ignore_conflicts=True).
    This means:
      • Uploading the same file twice → zero new rows (all skipped as duplicates)
      • Uploading a different date range → only truly new rows are inserted
      • Historical data is never erased by re-uploads

    batch_id: UUID (MonitoringData.file_group_id) so a batch can be bulk-deleted
    later via monitoring_delete_group.
    """
    df = df.copy()
    df.columns = df.columns.str.strip()  # ensure columns are stripped

    # ── Build a case-insensitive channel name lookup from the Channel model ─────
    _ch_canonical = {c.lower(): c for c in Channel.objects.values_list('name', flat=True)}

    def _canon_channel(raw: str) -> str:
        raw = str(raw).strip()
        canonical = _ch_canonical.get(raw.lower())
        if canonical:
            return canonical
        for prefix in ('tv - ', 'radio - '):
            if raw.lower().startswith(prefix):
                stripped = raw[len(prefix):]
                canonical = _ch_canonical.get(stripped.lower())
                if canonical:
                    return canonical
                return stripped
        return raw

    # ── Normalise to standard column names ─────────────────────────────────────
    if data_type == 'maponline':
        rename = {}
        _lmrb_ex_theme = get_setting_list('lmrb_extra_theme_aliases')
        _lmrb_ex_time  = get_setting_list('lmrb_extra_time_aliases')
        _lmrb_ex_dur   = get_setting_list('lmrb_extra_duration_aliases')
        _lmrb_ex_date  = get_setting_list('lmrb_extra_date_aliases')
        th_col = _find_col(df, 'Advt_Theme', 'Theme', *_lmrb_ex_theme)
        if th_col and th_col != 'Advt_Theme': rename[th_col] = 'Advt_Theme'
        dt_col = _find_col(df, 'Date', 'Prg Date', *_lmrb_ex_date)
        if dt_col and dt_col != 'Date': rename[dt_col] = 'Date'
        du_col = _find_col(df, 'Dur', 'Ad Dur', *_lmrb_ex_dur)
        if du_col and du_col != 'Dur': rename[du_col] = 'Dur'
        tm_col = _find_col(df, 'Advt_time', 'Ad Start', 'Advt_Time', *_lmrb_ex_time)
        if tm_col and tm_col != 'Advt_time': rename[tm_col] = 'Advt_time'
        pg_col = _find_col(df, 'Programme', 'Prg Name', 'Program')
        if pg_col and pg_col != 'Programme': rename[pg_col] = 'Programme'
        if rename:
            df.rename(columns=rename, inplace=True)
        if _find_col(df, 'Date') is not None:
            df['Date'] = pd.to_datetime(df[_find_col(df, 'Date')], errors='coerce')
    else:  # mediawatch
        if {'Dd', 'Mn', 'Yr'}.issubset(df.columns) and _find_col(df, 'Date') is None:
            df['Date'] = pd.to_datetime(
                df['Yr'].astype(str) + '-' +
                df['Mn'].astype(str).str.zfill(2) + '-' +
                df['Dd'].astype(str).str.zfill(2),
                errors='coerce',
            )
        elif _find_col(df, 'Date') is not None:
            df['Date'] = pd.to_datetime(df[_find_col(df, 'Date')], errors='coerce')

    dur_col = _find_col(df, 'Dur', 'Ad Dur', 'Duration')
    if dur_col is not None:
        df[dur_col] = pd.to_numeric(df[dur_col], errors='coerce')
        if dur_col != 'Dur':
            df.rename(columns={dur_col: 'Dur'}, inplace=True)

    ch_col = _find_col(df, 'Channel', 'Station')

    # ── Advertiser filter: normalise to lowercase set for fast membership test ──
    adv_col_name = _find_col(df, 'Advertiser', 'advertiser', 'ADVERTISER', 'Adv', 'Client')
    _adv_filter_set = (
        {a.lower() for a in advertiser_filter} if advertiser_filter else None
    )

    # ── Build rows dict keyed by dedup_key ────────────────────────────────────
    rows_by_key = {}

    for _, r in df.iterrows():
        channel   = _canon_channel(_safe_str(r.get(ch_col, 'Unknown')) if ch_col else 'Unknown')
        theme     = _safe_str(r.get('Advt_Theme', ''))
        advt_time = _safe_str(r.get('Advt_time', ''))
        dur       = _safe_int(r.get('Dur'))
        date_val  = _safe_date(r.get('Date'))
        brk_no    = _safe_int(r.get('BrkNo'))
        pos_in_brk = _safe_int(r.get('PosinBrk'))
        advertiser = _safe_str(r.get(adv_col_name, '') if adv_col_name else '')

        # Skip rows whose advertiser is not in the selected filter
        if _adv_filter_set and advertiser.lower() not in _adv_filter_set:
            continue
        product    = _safe_str(r.get('Product', ''))

        if not (theme and advt_time and date_val and channel):
            continue  # skip rows with missing key fields

        dedup_key = LMRBRow.make_dedup_key(
            account.id, channel, date_val, advt_time, theme, dur,
            brk_no=brk_no, pos_in_brk=pos_in_brk,
            advertiser=advertiser, product=product,
        )

        program_val = _safe_str(r.get('Program', r.get('Programme', r.get('Prg Name', ''))))

        # If same key appears twice in this upload, last row wins
        rows_by_key[dedup_key] = LMRBRow(
            account       = account,
            channel       = channel,
            date          = date_val,
            advt_theme    = theme,
            advt_time     = advt_time,
            duration      = dur,
            source        = data_type,
            dedup_key     = dedup_key,
            batch_id      = batch_id,
            # Extended columns
            product_group = _safe_str(r.get('Product_Group', '')),
            advertiser    = advertiser,
            product       = product,
            ads           = _safe_str(r.get('Ads', '')),
            program       = program_val,
            prog_time     = _safe_str(r.get('Prog_time', '')),
            ad_pos        = _safe_int(r.get('AdPos')),
            tot_ads       = _safe_int(r.get('TotAds')),
            brk_no        = brk_no,
            pos_in_brk    = pos_in_brk,
            ads_in_brk    = _safe_int(r.get('AdsinBrk')),
            lng           = _safe_str(r.get('Lng', '')),
            cost          = _safe_decimal(r.get('Cost')),
            day           = _safe_str(r.get('Day', '')),
        )

    print(f"[_parse_lmrb_rows] df rows={len(df)}  unique dedup_keys={len(rows_by_key)}")
    if not rows_by_key:
        print("[_parse_lmrb_rows] WARNING: no valid rows to insert (check column names and key fields)")
        return 0

    # ── Append-only insert - skip any rows that already exist ─────────────────
    new_rows = list(rows_by_key.values())
    created = LMRBRow.objects.bulk_create(new_rows, batch_size=500, ignore_conflicts=True)
    inserted = len(created)
    skipped  = len(new_rows) - inserted
    print(f"[_parse_lmrb_rows] inserted={inserted}  skipped(duplicate)={skipped}")
    return inserted


@login_required
@require_POST
def monitoring_detect(request):
    """AJAX: parse an uploaded monitoring file and return detected channels/dates + advertisers."""
    excel_file = request.FILES.get('file')
    data_type  = request.POST.get('data_type', 'mediawatch')
    if not excel_file:
        return JsonResponse({'ok': False, 'error': 'No file provided'})
    try:
        df = pd.read_excel(excel_file)
        df.columns = df.columns.str.strip()
        metas = _detect_monitoring_meta(df, data_type)

        # Detect unique advertisers so the uploader can filter by account
        adv_col = _find_col(df, 'Advertiser', 'advertiser', 'ADVERTISER', 'Adv', 'Client')
        advertisers = []
        if adv_col:
            advertisers = sorted(
                v for v in df[adv_col].dropna().astype(str).str.strip().unique()
                if v and v.lower() not in ('nan', 'none', '')
            )

        return JsonResponse({
            'ok': True,
            'channels': metas,
            'total_rows': len(df),
            'advertisers': advertisers,
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required
def monitoring_download(request, pk):
    """Serve the original monitoring Excel file as a download."""
    mon  = get_object_or_404(MonitoringData, pk=pk)
    user = request.user
    if not _is_admin(user) and mon.account not in _account_qs(user):
        return HttpResponse('Access denied', status=403)
    try:
        file_path = mon.file.path
        if not os.path.exists(file_path):
            return HttpResponse('File not found on server.', status=404)
        return FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename=mon.original_filename or os.path.basename(file_path),
        )
    except Exception as e:
        return HttpResponse(f'Download failed: {e}', status=500)


@login_required
def monitoring_delete(request, pk):
    mon   = get_object_or_404(MonitoringData, pk=pk)
    user  = request.user
    today = date_cls.today()
    print(f"[monitoring_delete] pk={pk}  channel={mon.channel}  file_group_id={mon.file_group_id}  user={user}")

    if not _is_admin(user) and mon.account and mon.account not in _account_qs(user):
        print(f"[monitoring_delete] ACCESS DENIED - user has no access to account {mon.account}")
        messages.error(request, 'You do not have access to this data.')
        return redirect('/dashboard/monitoring/')

    if _is_admin(user) or (mon.uploaded_by == user and mon.uploaded_at.date() == today):
        siblings = MonitoringData.objects.filter(file_group_id=mon.file_group_id).exclude(pk=mon.pk)
        sibling_count = siblings.count()
        print(f"[monitoring_delete] siblings in same group={sibling_count}")
        if not siblings.exists():
            print(f"[monitoring_delete] no siblings - deleting shared file {mon.file.name}")
            mon.file.delete(save=False)
        mon.delete()
        print(f"[monitoring_delete] deleted MonitoringData pk={pk}")
        messages.success(request, 'Dataset deleted.')
    else:
        print(f"[monitoring_delete] DENIED - not admin and not uploaded today by this user")
        messages.error(request, 'You can only delete datasets you uploaded today.')
    return redirect('/dashboard/monitoring/')


@login_required
def monitoring_delete_group(request, group_id):
    """
    Delete ALL MonitoringData records that share the same file_group_id.
    This lets users remove an entire multi-channel upload in one click.
    """
    user  = request.user
    today = date_cls.today()
    print(f"[monitoring_delete_group] group_id={group_id}  user={user}")

    group_qs = MonitoringData.objects.filter(file_group_id=group_id).select_related('account', 'uploaded_by')
    if not group_qs.exists():
        print(f"[monitoring_delete_group] ERROR: no records found for group_id={group_id}")
        messages.error(request, 'No records found for this upload group.')
        return redirect('/dashboard/monitoring/')

    first    = group_qs.first()
    channels = list(group_qs.values_list('channel', flat=True))
    count    = group_qs.count()
    print(f"[monitoring_delete_group] found {count} record(s): channels={channels}")
    print(f"[monitoring_delete_group] first.account={first.account}  first.uploaded_by={first.uploaded_by}  first.uploaded_at={first.uploaded_at}")

    if not _is_admin(user) and first.account and first.account not in _account_qs(user):
        print(f"[monitoring_delete_group] ACCESS DENIED - user has no access to account {first.account}")
        messages.error(request, 'You do not have access to this data.')
        return redirect('/dashboard/monitoring/')

    if not _is_admin(user) and not (first.uploaded_by == user and first.uploaded_at.date() == today):
        print(f"[monitoring_delete_group] DENIED - not admin and not uploaded today by this user")
        messages.error(request, 'You can only delete datasets you uploaded today.')
        return redirect('/dashboard/monitoring/')

    # Delete LMRBRows that belong to this upload batch
    import uuid as uuid_mod
    from core.models import ManualMatch, SponsorshipLmrbAssignment
    try:
        batch_uuid = uuid_mod.UUID(str(group_id))
        lmrb_qs = LMRBRow.objects.filter(batch_id=batch_uuid)
        lmrb_ids = list(lmrb_qs.values_list('id', flat=True))

        # 1. Unlock matched ScheduleRows before deleting
        matched_sch_ids = list(
            lmrb_qs.filter(is_matched=True).values_list('matched_schedule_id', flat=True)
        )
        if matched_sch_ids:
            ScheduleRow.objects.filter(id__in=matched_sch_ids).update(
                is_matched=False, matched_lmrb=None, matched_at=None,
            )

        # 2. Unlock TCRows that were LMRB-confirmed via these rows
        if lmrb_ids:
            from core.models import TCRow
            TCRow.objects.filter(matched_lmrb_id__in=lmrb_ids).update(
                is_lmrb_confirmed=False, matched_lmrb=None,
            )

        # 3. Remove ManualMatch records using these LMRB rows + unlock their ScheduleRows
        if lmrb_ids:
            manual_matches = ManualMatch.objects.filter(lmrb_row_id__in=lmrb_ids)
            mm_sch_ids = list(manual_matches.exclude(schedule_row=None).values_list('schedule_row_id', flat=True))
            if mm_sch_ids:
                ScheduleRow.objects.filter(id__in=mm_sch_ids).update(is_manual_matched=False)
            manual_matches.delete()

        # 4. Remove SponsorshipLmrbAssignment records + unlock their ScheduleRows
        if lmrb_ids:
            spon_assignments = SponsorshipLmrbAssignment.objects.filter(lmrb_row_id__in=lmrb_ids)
            spon_sch_ids = list(spon_assignments.values_list('schedule_row_id', flat=True))
            spon_assignments.delete()
            # Unlock sponsorship ScheduleRows (no is_matched for sponsorship, handled via assignment)

        # 5. Delete MatchResult records linked to these LMRB rows
        if lmrb_ids:
            from core.models import MatchResult
            MatchResult.objects.filter(lmrb_row_id__in=lmrb_ids).delete()
            # Also delete MatchResult records for ScheduleRows that were matched to these LMRBs
            if matched_sch_ids:
                MatchResult.objects.filter(schedule_row_id__in=matched_sch_ids,
                                           status='matched').delete()

        lmrb_count = lmrb_qs.count()
        lmrb_qs.delete()
        print(f"[monitoring_delete_group] deleted {lmrb_count} LMRBRow(s) for batch {group_id}")
    except Exception as e:
        print(f"[monitoring_delete_group] WARNING: could not delete LMRBRows: {e}")
        lmrb_count = 0

    # Delete the shared file once (all records in the group share the same file)
    if first.file:
        print(f"[monitoring_delete_group] deleting shared file: {first.file.name}")
        first.file.delete(save=False)

    group_qs.delete()
    print(f"[monitoring_delete_group] deleted {count} MonitoringData record(s) successfully")
    messages.success(request,
        f'Deleted upload batch ({", ".join(channels)}) - '
        f'{lmrb_count:,} LMRB row(s) removed.')
    return redirect('/dashboard/monitoring/')


# ── Brand mappings ────────────────────────────────────────────────────────────

@login_required
def brand_mapping_list(request):
    user       = request.user
    account_qs = _account_qs(user)
    account_id = request.GET.get('account', '')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            acc_id   = request.POST.get('account_id', '').strip()
            brand    = request.POST.get('brand', '').strip()
            themes_raw = request.POST.get('theme', '').strip()
            tc_theme = request.POST.get('tc_theme', '').strip()
            product  = request.POST.get('product', '').strip()
            dur_raw  = request.POST.get('duration', '').strip()
            duration = int(dur_raw) if dur_raw.isdigit() else None

            # Support multiple LMRB themes (pipe-separated)
            theme_list = [t.strip() for t in themes_raw.split('|') if t.strip()]

            if not (acc_id and brand and theme_list):
                messages.error(request, 'Account, Brand, and at least one LMRB Theme are required.')
            else:
                account = get_object_or_404(Account, id=acc_id)
                if not _is_admin(user) and account not in account_qs:
                    messages.error(request, 'No access to that account.')
                else:
                    created = 0
                    skipped = 0
                    for theme in theme_list:
                        exists = BrandMapping.objects.filter(
                            account=account, brand=brand, theme=theme,
                            duration=duration, product=product,
                        ).exists()
                        if exists:
                            skipped += 1
                        else:
                            BrandMapping.objects.create(
                                account=account, brand=brand, theme=theme,
                                tc_theme=tc_theme, duration=duration, product=product)
                            created += 1
                    dur_str = f' ({duration}s)' if duration else ''
                    if created:
                        messages.success(request, f'{created} mapping{"s" if created > 1 else ""} added for {brand}{dur_str}.')
                    if skipped:
                        messages.warning(request, f'{skipped} mapping{"s" if skipped > 1 else ""} already existed (skipped).')
            _qs = f'/dashboard/brand-mappings/?account={acc_id}'
            if product:
                from urllib.parse import quote
                _qs += f'&prefill_product={quote(product)}'
            return redirect(_qs)

        elif action == 'edit_field':
            mapping_id = request.POST.get('mapping_id')
            field_name = request.POST.get('field_name', '').strip()
            new_value  = request.POST.get('new_value', '').strip()
            if field_name not in ('product', 'brand', 'theme'):
                messages.error(request, 'Invalid field.')
            else:
                mapping = get_object_or_404(BrandMapping, id=mapping_id)
                if not _is_admin(user) and mapping.account not in account_qs:
                    messages.error(request, 'No access to that mapping.')
                else:
                    setattr(mapping, field_name, new_value)
                    mapping.save(update_fields=[field_name])
                    messages.success(request, f'Updated {field_name} for mapping #{mapping_id}.')
            return redirect(f'/dashboard/brand-mappings/?account={account_id or mapping.account_id}')

        elif action == 'edit_tc_theme':
            mapping_id = request.POST.get('mapping_id')
            tc_theme   = request.POST.get('tc_theme', '').strip()
            mapping    = get_object_or_404(BrandMapping, id=mapping_id)
            if not _is_admin(user) and mapping.account not in account_qs:
                messages.error(request, 'No access to that mapping.')
            else:
                mapping.tc_theme = tc_theme
                mapping.save(update_fields=['tc_theme'])
                messages.success(request, f'TC Theme updated for {mapping.brand}.')
            return redirect(f'/dashboard/brand-mappings/?account={account_id or mapping.account_id}')

        elif action == 'delete':
            mapping_id = request.POST.get('mapping_id')
            mapping    = get_object_or_404(BrandMapping, id=mapping_id)
            saved_acc  = mapping.account_id
            if not _is_admin(user) and mapping.account not in account_qs:
                messages.error(request, 'No access to that mapping.')
            else:
                mapping.delete()
                messages.success(request, 'Mapping deleted.')
            return redirect(f'/dashboard/brand-mappings/?account={account_id or saved_acc}')

    mappings = BrandMapping.objects.filter(account__in=account_qs).select_related('account')
    channel_filter = request.GET.get('channel', '')
    month_filter   = request.GET.get('month', '')
    if account_id:
        mappings = mappings.filter(account_id=account_id)
    if channel_filter:
        mappings = mappings.filter(
            brand__in=ScheduleRow.objects.filter(
                account_id=account_id, schedule__channel=channel_filter,
            ).values_list('brand', flat=True)
        ) if account_id else mappings
    if month_filter:
        mappings = mappings.filter(
            brand__in=ScheduleRow.objects.filter(
                account_id=account_id, schedule__month=month_filter,
            ).values_list('brand', flat=True)
        ) if account_id else mappings

    # Channels available for the selected account (for table filter)
    channels_for_account = sorted(set(
        Schedule.objects.filter(account_id=account_id).values_list('channel', flat=True)
    )) if account_id else []
    months_for_channel = sorted(set(
        Schedule.objects.filter(
            account_id=account_id, channel=channel_filter,
        ).values_list('month', flat=True)
    )) if account_id and channel_filter else []

    # ── Coverage panel (only when account + channel + month all set) ──────────
    coverage_data = None
    unmapped_lmrb_themes = []
    missing_variants = []
    if account_id and channel_filter and month_filter:
        sch_brands = list(
            ScheduleRow.objects
            .filter(account_id=account_id, schedule__channel=channel_filter, schedule__month=month_filter)
            .exclude(brand='')
            .values_list('brand', flat=True)
            .distinct()
            .order_by('brand')
        )
        existing = {
            bm.brand.lower().strip(): bm
            for bm in BrandMapping.objects.filter(account_id=account_id)
        }
        covered, partial_list, unmapped_list, mapped_list = 0, [], [], []
        for brand in sch_brands:
            bm = existing.get(brand.lower().strip())
            if bm is None:
                unmapped_list.append({'brand': brand})
            elif not bm.tc_theme.strip():
                partial_list.append({'brand': brand, 'theme': bm.theme})
            else:
                mapped_list.append({'brand': brand, 'theme': bm.theme, 'tc_theme': bm.tc_theme})
                covered += 1
        total = len(sch_brands)
        coverage_data = {
            'total':    total,
            'covered':  covered,
            'pct':      round(covered / total * 100) if total else 0,
            'unmapped': unmapped_list,
            'partial':  partial_list,
            'mapped':   mapped_list,
        }

        # ── Unmapped LMRB themes ─────────────────────────────────────────────
        # Find LMRB themes in this scope that have NO BrandMapping pointing to them.
        sch_dates = Schedule.objects.filter(
            account_id=account_id, channel=channel_filter, month=month_filter,
        ).aggregate(s=Min('start_date'), e=Max('end_date'))
        lmrb_qs = LMRBRow.objects.filter(
            account_id=account_id, channel__iexact=channel_filter,
        ).exclude(advt_theme='')
        if sch_dates['s']:
            lmrb_qs = lmrb_qs.filter(date__gte=sch_dates['s'])
        if sch_dates['e']:
            lmrb_qs = lmrb_qs.filter(date__lte=sch_dates['e'])
        all_lmrb_themes = sorted(set(lmrb_qs.values_list('advt_theme', flat=True)))

        # Build set of mapped LMRB themes (case-insensitive, supports wildcard *)
        mapped_themes_exact = set()
        mapped_themes_prefix = []
        for bm in BrandMapping.objects.filter(account_id=account_id):
            t = bm.theme.strip()
            if not t:
                continue
            if t.endswith('*'):
                mapped_themes_prefix.append(t[:-1].lower())
            else:
                mapped_themes_exact.add(t.lower())

        def _is_theme_mapped(theme):
            tl = theme.lower().strip()
            if tl in mapped_themes_exact:
                return True
            for prefix in mapped_themes_prefix:
                if tl.startswith(prefix):
                    return True
            return False

        unmapped_lmrb_themes = [t for t in all_lmrb_themes if not _is_theme_mapped(t)]

        # ── Missing language variants ────────────────────────────────────────
        # For each mapped theme with a (Sin)/(Tam)/(Eng) suffix, check if the
        # other language variants exist in LMRB data but lack a BrandMapping.
        import re
        _LANG_RE = re.compile(r'^(.*)\((Sin|Tam|Eng)\)$')
        _LANG_SUFFIXES = ['(Sin)', '(Tam)', '(Eng)']

        # All LMRB themes across ALL channels for this account (not just current)
        all_account_themes = set(
            LMRBRow.objects.filter(account_id=account_id)
            .exclude(advt_theme='')
            .values_list('advt_theme', flat=True)
        )
        all_account_lower = {t.lower(): t for t in all_account_themes}

        missing_variants = []  # list of {brand, mapped_theme, missing_theme, lang}
        for bm in BrandMapping.objects.filter(account_id=account_id):
            t = bm.theme.strip()
            m = _LANG_RE.match(t)
            if not m:
                continue
            base, mapped_lang = m.group(1), m.group(2)
            for sfx in _LANG_SUFFIXES:
                lang = sfx[1:-1]  # strip parens
                if lang == mapped_lang:
                    continue
                variant = base + sfx
                variant_lower = variant.lower()
                # Only show if this variant exists in LMRB but has no mapping
                if variant_lower in all_account_lower and not _is_theme_mapped(variant):
                    original_case = all_account_lower[variant_lower]
                    # Avoid duplicates
                    if not any(mv['missing_theme'].lower() == variant_lower and mv['brand'] == bm.brand
                               for mv in missing_variants):
                        missing_variants.append({
                            'brand': bm.brand,
                            'product': bm.product or '',
                            'mapped_theme': t,
                            'missing_theme': original_case,
                            'lang': lang,
                            'tc_theme': bm.tc_theme or '',
                            'duration': bm.duration,
                        })

    return render(request, 'admin_panel/brand_mappings.html', {
        'mappings':             mappings,
        'accounts':             account_qs,
        'filters':              {
            'account': account_id,
            'channel': channel_filter,
            'month':   month_filter,
        },
        'channels_for_account': channels_for_account,
        'months_for_channel':   months_for_channel,
        'coverage_data':        coverage_data,
        'unmapped_lmrb_themes': unmapped_lmrb_themes,
        'missing_variants':     missing_variants,
    })


@login_required
def brand_mapping_options(request):
    """
    AJAX: Return unique brands, LMRB themes, TC themes, and LMRB products for an account.
    Narrows results when channel / month / schedule_id / product are also provided.
    """
    account_id  = request.GET.get('account_id', '').strip()
    channel     = request.GET.get('channel', '').strip()
    month       = request.GET.get('month', '').strip()
    schedule_id = request.GET.get('schedule_id', '').strip()
    product     = request.GET.get('product', '').strip()
    if not account_id or not _account_access(request.user, account_id):
        return JsonResponse({'brands': [], 'themes': [], 'tc_themes': [], 'products': []})

    # ── Brands (from Schedule) ──────────────────────────────────────────────
    brands_qs = ScheduleRow.objects.filter(account_id=account_id).exclude(brand='')
    if schedule_id:
        brands_qs = brands_qs.filter(schedule_id=schedule_id)
    elif channel and month:
        brands_qs = brands_qs.filter(schedule__channel=channel, schedule__month=month)
    elif channel:
        brands_qs = brands_qs.filter(schedule__channel=channel)
    brands = sorted(set(brands_qs.values_list('brand', flat=True)))

    # ── Schedule date range for LMRB / TC filtering ─────────────────────────
    sch_filter = {'account_id': account_id}
    if schedule_id:
        sch_filter['id'] = schedule_id
    elif channel and month:
        sch_filter['channel'] = channel
        sch_filter['month']   = month
    elif channel:
        sch_filter['channel'] = channel
    sch_dates = Schedule.objects.filter(**sch_filter).aggregate(s=Min('start_date'), e=Max('end_date'))

    # ── LMRB base queryset (channel + date scoped) ─────────────────────────
    lmrb_base = LMRBRow.objects.filter(account_id=account_id)
    if channel:
        lmrb_base = lmrb_base.filter(channel__iexact=channel)
    if sch_dates['s']:
        lmrb_base = lmrb_base.filter(date__gte=sch_dates['s'])
    if sch_dates['e']:
        lmrb_base = lmrb_base.filter(date__lte=sch_dates['e'])

    # ── LMRB products ──────────────────────────────────────────────────────
    products = sorted(set(
        lmrb_base.exclude(product='').exclude(product__isnull=True)
        .values_list('product', flat=True)
    ))

    # ── LMRB themes (filtered by product if provided) ─────────────────────
    themes_qs = lmrb_base.exclude(advt_theme='')
    if product:
        themes_qs = themes_qs.filter(product__iexact=product)
    themes = sorted(set(themes_qs.values_list('advt_theme', flat=True)))

    # ── Theme → Product map (most common product per theme) ──────────────
    theme_product_map = {}
    tp_rows = (
        lmrb_base.exclude(advt_theme='')
        .exclude(product='').exclude(product__isnull=True)
        .values('advt_theme', 'product')
        .annotate(cnt=Count('id'))
        .order_by('advt_theme', '-cnt')
    )
    for row in tp_rows:
        t = row['advt_theme']
        if t not in theme_product_map:  # first = highest count
            theme_product_map[t] = row['product']

    # ── Check which LMRB themes are already mapped ────────────────────────
    mapped_exact = set()
    mapped_prefix = []
    for bm in BrandMapping.objects.filter(account_id=account_id):
        t = bm.theme.strip()
        if not t:
            continue
        if t.endswith('*'):
            mapped_prefix.append(t[:-1].lower())
        else:
            mapped_exact.add(t.lower())
    mapped_themes = []
    for t in themes:
        tl = t.lower().strip()
        is_mapped = tl in mapped_exact or any(tl.startswith(p) for p in mapped_prefix)
        if is_mapped:
            mapped_themes.append(t)

    # ── TC themes ─────────────────────────────────────────────────────────
    tc_qs = TCRow.objects.filter(account_id=account_id).exclude(tc_theme='')
    if channel:
        tc_qs = tc_qs.filter(channel__iexact=channel)
    if sch_dates['s']:
        tc_qs = tc_qs.filter(date__gte=sch_dates['s'])
    if sch_dates['e']:
        tc_qs = tc_qs.filter(date__lte=sch_dates['e'])
    tc_themes = sorted(set(tc_qs.values_list('tc_theme', flat=True)))

    return JsonResponse({
        'brands': brands, 'themes': themes, 'tc_themes': tc_themes,
        'products': products, 'mapped_themes': mapped_themes,
        'theme_product_map': theme_product_map,
    })


@login_required
def brand_mapping_channels(request):
    """AJAX: Return distinct channels from Schedules for an account."""
    account_id = request.GET.get('account_id', '').strip()
    if not account_id or not _account_access(request.user, account_id):
        return JsonResponse({'channels': []})
    channels = sorted(set(
        Schedule.objects.filter(account_id=account_id).values_list('channel', flat=True)
    ))
    return JsonResponse({'channels': channels})


@login_required
def brand_mapping_months(request):
    """AJAX: Return distinct months from Schedules for an account + channel."""
    account_id = request.GET.get('account_id', '').strip()
    channel    = request.GET.get('channel', '').strip()
    if not account_id or not _account_access(request.user, account_id):
        return JsonResponse({'months': []})
    months = sorted(set(
        Schedule.objects.filter(account_id=account_id, channel=channel)
        .values_list('month', flat=True)
    ))
    return JsonResponse({'months': months})


@login_required
def brand_mapping_schedules(request):
    """AJAX: Return schedules for an account + channel + month."""
    account_id = request.GET.get('account_id', '').strip()
    channel    = request.GET.get('channel', '').strip()
    month      = request.GET.get('month', '').strip()
    if not account_id or not _account_access(request.user, account_id):
        return JsonResponse({'schedules': []})
    schedules = list(
        Schedule.objects.filter(account_id=account_id, channel=channel, month=month)
        .order_by('version')
        .values('id', 'schedule_number', 'version', 'row_count')
    )
    return JsonResponse({'schedules': schedules})


# ── Monitoring Dashboard (Items 3 + 4) ───────────────────────────────────────

@login_required
def monitoring_dashboard(request):
    """
    Analytics dashboard: per-scope summary, 7 report tabs, export + PDF links.
    Accessible by operations, team_head, planner, admin, super_admin.
    """
    # MatchResult and ScheduleRow already imported at top level

    user       = request.user
    account_qs = _account_qs(user)

    account_id  = request.GET.get('account_id', '')
    channel     = request.GET.get('channel', '')
    month       = request.GET.get('month', '')
    schedule_id = request.GET.get('schedule_id', '').strip()

    selected_account   = None
    channels           = []
    months             = []
    schedules_in_scope = []
    stats    = {}
    tab_data = {}
    ch_summary     = []
    brand_summary  = []
    sponsorship_rows = []
    sponsorship_brand_summary = []
    lmrb_chart     = []
    commercial_chart_data  = []
    sponsorship_chart_data = []
    commercial_chart_json  = '[]'
    sponsorship_chart_json = '[]'
    lmrb_matched_rows   = []
    lmrb_unmatched_rows = []
    extra_aired_rows    = []

    if account_id:
        try:
            selected_account = account_qs.get(pk=account_id)
        except Account.DoesNotExist:
            pass

    if selected_account:
        channels = sorted(set(
            MatchResult.objects.filter(account_id=account_id)
            .values_list('channel', flat=True)
        ))
        if channel:
            months = sorted(set(
                MatchResult.objects.filter(account_id=account_id, channel=channel)
                .values_list('month', flat=True)
            ))

    if selected_account and channel and month:
        schedules_in_scope = list(
            Schedule.objects.filter(account_id=account_id, channel=channel, month=month)
            .order_by('schedule_number')
        )
        if not schedule_id and len(schedules_in_scope) == 1:
            schedule_id = str(schedules_in_scope[0].id)

        qs = MatchResult.objects.filter(account_id=account_id, channel=channel, month=month)
        if schedule_id:
            qs = qs.filter(schedule_row__schedule_id=schedule_id)
        total = qs.count()

        n_matched    = qs.filter(status='matched').count()
        n_prog_mis   = qs.filter(status='programme_mismatch').count()
        n_late       = qs.filter(status='late_telecast').count()
        n_not_aired  = qs.filter(status__in=['not_aired', 'no_mapping']).count()
        n_no_map     = qs.filter(status='no_mapping').count()
        n_aired      = n_matched + n_prog_mis + n_late

        # Use active schedules only (highest version per schedule_number) so
        # duplicate uploads don't inflate commercial / sponsorship counts.
        from verification.engine import active_schedule_ids as _active_ids
        _active_sch_ids = _active_ids(account_id, channel, month)
        _sr_base = ScheduleRow.objects.filter(schedule_id__in=_active_sch_ids)
        if schedule_id and int(schedule_id) in _active_sch_ids:
            _sr_base = _sr_base.filter(schedule_id=schedule_id)
        n_sponsorship = _sr_base.filter(ad_type='SPONSORSHIP').count()
        n_commercial  = _sr_base.filter(ad_type='COMMERCIAL BENEFITS').count()

        # Compliance is against the planned commercial rows, not just the verified
        # subset - rows past the LMRB date cap are counted as unverified (not ignored).
        # Programme Mismatch and Late Telecast both count as "aired" for compliance.
        compliance = round((n_matched + n_prog_mis + n_late) / n_commercial * 100, 1) if n_commercial else 0

        stats = {
            'total':        n_commercial,                  # planned commercial rows = real denominator
            'total_all':    n_commercial + n_sponsorship,  # combined for split-bar denominator
            'verified':     total,                         # MatchResult count
            'unverified':   max(0, n_commercial - total),  # rows past LMRB date cap
            'matched':      n_matched,
            'prog_mis':     n_prog_mis,
            'late':         n_late,
            'not_aired':    n_not_aired,
            'no_map':       n_no_map,
            'aired':        n_aired,
            'compliance':   compliance,
            'sponsorship':  n_sponsorship,
        }

        # Full Report: ALL active commercial schedule rows, not just verified ones.
        # Rows without a MatchResult (pending LMRB data) appear as synthetic dicts
        # with status='pending' so the table is always complete.
        _mr_by_sr_id = {
            mr.schedule_row_id: mr
            for mr in qs.select_related('lmrb_row').all()
        }
        _full_rows = []
        for _sr in _sr_base.filter(ad_type='COMMERCIAL BENEFITS').order_by('date', 'brand'):
            if _sr.id in _mr_by_sr_id:
                _full_rows.append(_mr_by_sr_id[_sr.id])
            else:
                _full_rows.append({
                    'brand':          _sr.brand,
                    'programme':      _sr.programme or '',
                    'scheduled_date': _sr.date,
                    'planned_start':  _sr.start_time or '',
                    'planned_end':    _sr.end_time   or '',
                    'duration':       _sr.duration,
                    'aired_date':     None,
                    'air_time':       None,
                    'status':         'pending',
                    'lmrb_row':       None,
                })

        tab_data = {
            'full':     _full_rows,
            # Programme mismatch = valid match (same day, brand, theme - time offset only)
            'matched':  list(qs.filter(status__in=['matched', 'programme_mismatch']).select_related('lmrb_row').order_by('scheduled_date', 'brand')),
            'not_aired': list(qs.filter(status__in=['not_aired', 'no_mapping']).order_by('scheduled_date', 'brand')),
            'prog_mis': list(qs.filter(status='programme_mismatch').select_related('lmrb_row').order_by('scheduled_date', 'brand')),
            'late':     list(qs.filter(status='late_telecast').select_related('lmrb_row').order_by('scheduled_date', 'brand')),
        }

        # Channel summary (only one channel in this scope)
        ch_summary = [{'channel': channel, **stats}]

        # Brand summary
        for br in qs.values_list('brand', flat=True).distinct().order_by('brand'):
            bq       = qs.filter(brand=br)
            bt       = bq.count()
            bm       = bq.filter(status='matched').count()
            brand_summary.append({
                'brand':      br,
                'total':      bt,
                'matched':    bm,
                'prog_mis':   bq.filter(status='programme_mismatch').count(),
                'late':       bq.filter(status='late_telecast').count(),
                'not_aired':  bq.filter(status__in=['not_aired', 'no_mapping']).count(),
                'compliance': round(bm / bt * 100, 1) if bt else 0,
            })

        # Sponsorship rows for the separate sponsorship tab (active schedules only)
        sponsorship_rows = list(
            _sr_base.filter(ad_type='SPONSORSHIP').order_by('date', 'start_time')
        )

        # Sponsorship brand breakdown (active schedules only)
        sponsorship_brand_summary = []
        for br in (_sr_base.filter(ad_type='SPONSORSHIP')
                   .values_list('brand', flat=True).distinct().order_by('brand')):
            sq = _sr_base.filter(ad_type='SPONSORSHIP', brand=br)
            st = sq.count()
            sm = sq.filter(is_matched=True).count()
            sponsorship_brand_summary.append({
                'brand': br,
                'total': st,
                'matched': sm,
                'not_matched': st - sm,
            })

        # commercial count already computed above; expose it in stats
        stats['commercial'] = n_commercial

        # ── LMRB queryset (scoped to schedule date range) ────────────────────
        def _time_bucket(t):
            try:
                h = int(str(t).split(':')[0])
                if 19 <= h <= 23:
                    return 'prime'
                elif 6 <= h <= 18:
                    return 'non_prime'
                else:
                    return 'other'
            except Exception:
                return 'other'

        sch_dates = ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month
        ).aggregate(d_min=Min('date'), d_max=Max('date'))
        lmrb_qs = LMRBRow.objects.filter(account_id=account_id, channel__iexact=channel)
        if sch_dates['d_min']:
            lmrb_qs = lmrb_qs.filter(date__gte=sch_dates['d_min'])
        if sch_dates['d_max']:
            lmrb_qs = lmrb_qs.filter(date__lte=sch_dates['d_max'])

        # Raw LMRB pool excludes sponsorship-matched rows so a row locked for
        # a SponsorshipLmrbAssignment never inflates commercial raw counts.
        lmrb_qs_raw = lmrb_qs.filter(is_sponsorship_matched=False)

        # ── Commercial chart data: Product → Brand → Theme (Planned vs Aired) ─
        # Only show chart data when verification results exist (MatchResult records).
        # Aired = MatchResult status in matched / programme_mismatch / late_telecast.
        aired_qs = qs.filter(
            status__in=['matched', 'programme_mismatch', 'late_telecast']
        ).select_related('lmrb_row', 'schedule_row')

        # theme_lower -> {prog: {count, prime, non_prime}}
        _theme_prog_details = {}
        for mr in aired_qs:
            t = (mr.theme or '').lower().strip()
            prog = ''
            time_val = ''
            if mr.lmrb_row:
                prog = mr.lmrb_row.program or ''
                time_val = mr.lmrb_row.advt_time or ''
            if not prog:
                prog = mr.programme or 'Unknown'
            bucket = _time_bucket(time_val) if time_val else 'other'
            entry = _theme_prog_details.setdefault(t, {}).setdefault(
                prog, {'count': 0, 'prime': 0, 'non_prime': 0}
            )
            entry['count'] += 1
            if bucket == 'prime':
                entry['prime'] += 1
            else:
                entry['non_prime'] += 1

        # theme_lower -> {date_str: count}  (for per-date aired bar chart)
        _theme_date_details = {}
        for mr in aired_qs:
            t = (mr.theme or '').lower().strip()
            date_str = str(mr.aired_date) if mr.aired_date else (
                str(mr.scheduled_date) if mr.scheduled_date else ''
            )
            if date_str:
                _theme_date_details.setdefault(t, {})
                _theme_date_details[t][date_str] = _theme_date_details[t].get(date_str, 0) + 1

        # theme_lower -> product name (from LMRBRow.product field)
        _theme_product_map = {}
        for lr in (lmrb_qs_raw.exclude(product='').exclude(product__isnull=True)
                   .values('advt_theme', 'product').distinct()):
            t = (lr['advt_theme'] or '').lower().strip()
            if t not in _theme_product_map:
                _theme_product_map[t] = lr['product']

        # (theme_lower, duration, product_lower) -> {count, progs, dates}
        # Keyed by LMRBRow.product so themes shared across products (e.g. "BB"
        # in both Mobitel Broadband and SLT Broadband) are kept separate.
        # Uses lmrb_qs_raw (excludes is_sponsorship_matched rows) so locked rows
        # don't inflate commercial raw counts.
        _raw_lmrb_by_tdp = {}
        for lr in lmrb_qs_raw.values(
            'advt_theme', 'duration', 'advt_time', 'date', 'program', 'product'
        ):
            t    = (lr['advt_theme'] or '').lower().strip()
            dur  = lr['duration']
            prod = (lr['product']   or '').lower().strip()
            key  = (t, dur, prod)
            prog     = lr['program'] or 'Unknown'
            time_val = lr['advt_time'] or ''
            date_str = str(lr['date']) if lr['date'] else ''
            bucket   = _time_bucket(time_val) if time_val else 'other'
            entry = _raw_lmrb_by_tdp.setdefault(
                key, {'count': 0, 'progs': {}, 'dates': {}}
            )
            entry['count'] += 1
            pe = entry['progs'].setdefault(
                prog, {'count': 0, 'prime': 0, 'non_prime': 0}
            )
            pe['count'] += 1
            if bucket == 'prime':
                pe['prime'] += 1
            else:
                pe['non_prime'] += 1
            if date_str:
                entry['dates'][date_str] = entry['dates'].get(date_str, 0) + 1

        def _agg_raw_lmrb(t_lower, duration, bm_product):
            """Aggregate raw LMRB count + programme details + date counts.

            bm_product - BrandMapping.product value.  When set, only LMRBRows
            whose product field matches are included (so "Mobitel Broadband BB"
            is never mixed with "SLT Broadband BB").  When the product field on
            the LMRB rows is empty or doesn't match (common for sponsorship/BB
            themes where LMRB data lacks a product column), the function falls
            back to matching by theme + duration only, so the chart never shows 0
            simply because of missing product metadata.
            duration=None aggregates across all durations.
            """
            def _scan(use_prod_filter):
                tot, progs, dates = 0, {}, {}
                pf = bm_product.lower().strip() if (use_prod_filter and bm_product) else None
                for (t, d, p), entry in _raw_lmrb_by_tdp.items():
                    if t != t_lower:
                        continue
                    if duration is not None and d != duration:
                        continue
                    if pf is not None and p != pf:
                        continue
                    tot += entry['count']
                    for prog, v in entry['progs'].items():
                        e = progs.setdefault(prog, {'count': 0, 'prime': 0, 'non_prime': 0})
                        e['count'] += v['count']
                        e['prime'] += v['prime']
                        e['non_prime'] += v['non_prime']
                    for dd, c in entry['dates'].items():
                        dates[dd] = dates.get(dd, 0) + c
                return tot, progs, dates

            total, merged_progs, merged_dates = _scan(use_prod_filter=True)
            # Fallback: if product filter wiped out all rows (LMRB rows have no
            # matching product metadata, e.g. BB / break-bumper themes), retry
            # without the product restriction so the count is still useful.
            if total == 0 and bm_product:
                total, merged_progs, merged_dates = _scan(use_prod_filter=False)
            return total, merged_progs, merged_dates

        def _build_chart_rows(ad_type_val):
            """Return flat list of dicts for one ad_type ('COMMERCIAL BENEFITS' or 'SPONSORSHIP').

            Product priority: BrandMapping.product → ScheduleRow.product →
                              LMRBRow.product (via _theme_product_map) → brand name.
            """
            rows = []
            brands = sorted(set(
                _sr_base.filter(ad_type=ad_type_val).values_list('brand', flat=True)
            ))
            for brand in brands:
                mappings = list(BrandMapping.objects.filter(
                    account_id=account_id, brand=brand
                ).order_by('theme'))
                if not mappings:
                    planned = _sr_base.filter(ad_type=ad_type_val, brand=brand).count()
                    # Use ScheduleRow.product if available for unmapped brands
                    sr_product = (
                        _sr_base.filter(ad_type=ad_type_val, brand=brand)
                        .exclude(product='').values_list('product', flat=True).first() or ''
                    )
                    # Sponsorship without BrandMapping: use sponsorship_type keyword
                    # to match raw LMRB observations within the schedule period.
                    kw_lmrb_count = 0
                    kw_progs: list = []
                    kw_pt = kw_non_pt = 0
                    kw_dates: list = []
                    if ad_type_val == 'SPONSORSHIP':
                        spon_type_val = (
                            _sr_base.filter(ad_type='SPONSORSHIP', brand=brand)
                            .exclude(sponsorship_type='')
                            .values_list('sponsorship_type', flat=True)
                            .first() or ''
                        )
                        if spon_type_val:
                            # sponsorship_type may be pipe-separated (e.g. "-Intro|-Extro");
                            # build an OR query so all selected keywords are included.
                            from django.db.models import Q as _Q
                            _spon_parts = [t.strip() for t in spon_type_val.split('|') if t.strip()]
                            _kw_q = _Q()
                            for _part in _spon_parts:
                                _kw_q |= _Q(advt_theme__icontains=_part)
                            _kw_all = lmrb_qs.filter(_kw_q)
                            # Product-aware filter: avoid mixing Mobitel and SLT
                            # rows that share the same sponsorship keyword.
                            _prod_hint = (
                                sr_product or brand.split()[0]
                            ).lower().strip()
                            kw_qs2 = _kw_all.filter(product__icontains=_prod_hint)
                            if not kw_qs2.exists():
                                kw_qs2 = _kw_all  # fallback when no product metadata
                            kw_lmrb_count = kw_qs2.count()
                            _kw_progs_raw: dict = {}
                            _kw_dates_raw: dict = {}
                            for _lr in kw_qs2.values('program', 'advt_time', 'date'):
                                _prog = _lr['program'] or 'Unknown'
                                _bucket = _time_bucket(_lr['advt_time'] or '') if _lr['advt_time'] else 'other'
                                _e = _kw_progs_raw.setdefault(_prog, {'count': 0, 'prime': 0, 'non_prime': 0})
                                _e['count'] += 1
                                if _bucket == 'prime':
                                    _e['prime'] += 1
                                    kw_pt += 1
                                else:
                                    _e['non_prime'] += 1
                                    kw_non_pt += 1
                                _d = str(_lr['date']) if _lr['date'] else ''
                                if _d:
                                    _kw_dates_raw[_d] = _kw_dates_raw.get(_d, 0) + 1
                            kw_progs = sorted(
                                [{'name': k, 'count': v['count'],
                                  'prime': v['prime'], 'non_prime': v['non_prime']}
                                 for k, v in _kw_progs_raw.items()],
                                key=lambda x: -x['count'],
                            )
                            kw_dates = [{'date': d, 'count': c}
                                        for d, c in sorted(_kw_dates_raw.items())]
                    rows.append({
                        'product': sr_product.strip() or brand,
                        'brand': brand,
                        'theme': brand,
                        'duration': None,
                        'planned': planned,
                        'aired': 0,
                        'lmrb_count': kw_lmrb_count,
                        'raw_lmrb_count': kw_lmrb_count,
                        'is_mapped': False,
                        'programmes': kw_progs,
                        'programmes_json': _to_js(kw_progs),
                        'pt_count': kw_pt,
                        'non_pt_count': kw_non_pt,
                        'dates_json': _to_js(kw_dates),
                    })
                    continue
                for bm in mappings:
                    t_lower = (bm.theme or '').lower().strip()
                    sr_q = _sr_base.filter(ad_type=ad_type_val, brand=brand)
                    if bm.duration is not None:
                        sr_q = sr_q.filter(duration=bm.duration)
                    planned = sr_q.count()

                    # Commercial aired = MatchResult (filter by brand to avoid
                    # double-counting when two brands share the same LMRB theme)
                    if ad_type_val == 'COMMERCIAL BENEFITS':
                        mr_q = qs.filter(
                            status__in=['matched', 'programme_mismatch', 'late_telecast'],
                            theme__iexact=bm.theme,
                            schedule_row__brand=brand,
                        )
                        if bm.duration is not None:
                            mr_q = mr_q.filter(duration=bm.duration)
                        aired = mr_q.count()
                        lmrb_count = aired
                        progs_raw = _theme_prog_details.get(t_lower, {})
                        progs = sorted(
                            [{'name': k, 'count': v['count'],
                              'prime': v['prime'], 'non_prime': v['non_prime']}
                             for k, v in progs_raw.items()],
                            key=lambda x: -x['count'],
                        )
                        dates_raw = _theme_date_details.get(t_lower, {})
                        dates_data = [{'date': d, 'count': c}
                                      for d, c in sorted(dates_raw.items())]
                    else:
                        # Sponsorship: use raw LMRB observations directly.
                        # No formal reconciliation is run for sponsorship rows -
                        # raw LMRB count within the schedule period is the best
                        # available proxy for "aired".
                        aired = 0
                        raw_lmrb_count, _raw_progs_map, _raw_dates_map = _agg_raw_lmrb(
                            t_lower, bm.duration, bm.product
                        )
                        lmrb_count = raw_lmrb_count
                        progs = sorted(
                            [{'name': k, 'count': v['count'],
                              'prime': v['prime'], 'non_prime': v['non_prime']}
                             for k, v in _raw_progs_map.items()],
                            key=lambda x: -x['count'],
                        )
                        dates_data = [{'date': d, 'count': c}
                                      for d, c in sorted(_raw_dates_map.items())]

                    pt_count    = sum(p['prime'] for p in progs)
                    non_pt_count = sum(p['non_prime'] for p in progs)

                    # For commercial: compute raw LMRB fallback if not already done above
                    if ad_type_val == 'COMMERCIAL BENEFITS':
                        raw_lmrb_count, _raw_progs_map, _raw_dates_map = _agg_raw_lmrb(
                            t_lower, bm.duration, bm.product
                        )
                        # Populate progs/dates from raw LMRB when no matched details exist
                        if not progs and raw_lmrb_count > 0:
                            progs = sorted(
                                [{'name': k, 'count': v['count'],
                                  'prime': v['prime'], 'non_prime': v['non_prime']}
                                 for k, v in _raw_progs_map.items()],
                                key=lambda x: -x['count'],
                            )
                            dates_data = [{'date': d, 'count': c}
                                          for d, c in sorted(_raw_dates_map.items())]
                            pt_count     = sum(p['prime']     for p in progs)
                            non_pt_count = sum(p['non_prime'] for p in progs)

                    # Product resolution: BrandMapping.product → ScheduleRow.product →
                    # LMRBRow product → brand
                    if bm.product.strip():
                        product = bm.product.strip()
                    else:
                        sr_product = (
                            sr_q.exclude(product='')
                            .values_list('product', flat=True).first() or ''
                        )
                        product = (sr_product.strip()
                                   or _theme_product_map.get(t_lower)
                                   or brand)

                    if planned == 0 and aired == 0 and lmrb_count == 0:
                        continue
                    # When a brand has multiple theme mappings (e.g. Sin/Tam
                    # language variants), hide mappings with zero LMRB evidence
                    # on this channel — they belong to a different channel.
                    if len(mappings) > 1 and aired == 0 and raw_lmrb_count == 0:
                        continue
                    rows.append({
                        'product': product,
                        'brand': brand,
                        'theme': bm.theme,
                        'duration': bm.duration,
                        'planned': planned,
                        'aired': aired,
                        'lmrb_count': lmrb_count,
                        'raw_lmrb_count': raw_lmrb_count,
                        'is_mapped': True,
                        'programmes': progs,
                        'programmes_json': _to_js(progs),
                        'pt_count': pt_count,
                        'non_pt_count': non_pt_count,
                        'dates_json': _to_js(dates_data),
                    })
            return rows

        def _group_by_product(rows):
            groups = {}
            for r in rows:
                grp = groups.setdefault(r['product'], [])
                grp.append(r)
            return [
                {
                    'product': p,
                    'themes': themes,
                    'total_planned': sum(t['planned'] for t in themes),
                    'total_aired': sum(
                        t['aired'] if t['aired'] > 0 else t['raw_lmrb_count']
                        for t in themes
                    ),
                    'total_lmrb': sum(
                        t['lmrb_count'] if t['lmrb_count'] > 0 else t['raw_lmrb_count']
                        for t in themes
                    ),
                }
                for p, themes in sorted(groups.items())
            ]

        commercial_chart_data = _group_by_product(_build_chart_rows('COMMERCIAL BENEFITS'))
        sponsorship_chart_data = _group_by_product(_build_chart_rows('SPONSORSHIP'))
        commercial_chart_json = _to_js(commercial_chart_data)
        sponsorship_chart_json = _to_js(sponsorship_chart_data)

        # Keep legacy lmrb_chart for the LMRB theme detail drawer
        lmrb_chart = list(
            lmrb_qs.values('advt_theme', 'duration')
            .annotate(count=Count('id'))
            .order_by('advt_theme', 'duration')
        )

        # ── LMRB theme detail (per theme, for detail drawer) ─────────────────
        lmrb_theme_detail = []
        for td in lmrb_chart:
            t_theme = td['advt_theme']
            t_dur   = td['duration']
            tqs     = lmrb_qs.filter(advt_theme=t_theme, duration=t_dur)
            progs   = list(
                tqs.exclude(program='').exclude(program__isnull=True)
                .values('program').annotate(cnt=Count('id')).order_by('-cnt')
            )
            dates     = tqs.aggregate(first=Min('date'), last=Max('date'))
            date_span = (
                (dates['last'] - dates['first']).days + 1
                if (dates['first'] and dates['last']) else 1
            )
            avg_per_day = round(td['count'] / date_span, 1) if date_span else td['count']
            bm      = BrandMapping.objects.filter(
                account_id=account_id, theme__iexact=t_theme
            ).first()
            sch_start = None
            if bm:
                sch_start = ScheduleRow.objects.filter(
                    account_id=account_id, channel=channel, month=month,
                    brand=bm.brand,
                ).aggregate(d=Min('date'))['d']
            # PT split for this theme+duration
            t_prime = t_non = 0
            for r in tqs.values_list('advt_time', flat=True):
                b = _time_bucket(r)
                if b == 'prime':
                    t_prime += 1
                else:
                    t_non += 1
            # Daily airings count
            daily = list(
                tqs.values('date').annotate(cnt=Count('id')).order_by('date')
            )
            lmrb_theme_detail.append({
                'theme':       t_theme,
                'duration':    t_dur,
                'count':       td['count'],
                'programmes':  progs,
                'first_aired': dates['first'].isoformat() if dates['first'] else None,
                'last_aired':  dates['last'].isoformat()  if dates['last']  else None,
                'sch_start':   sch_start.isoformat()      if sch_start      else None,
                'avg_per_day': avg_per_day,
                'prime':       t_prime,
                'non_prime':   t_non,
                'prime_pct':   round(t_prime / td['count'] * 100) if td['count'] else 0,
                'daily':       [{'date': str(r['date']), 'cnt': r['cnt']} for r in daily],
            })

        # ── Prime-time distribution (planned vs LMRB) ────────────────────────
        # Build theme → ad_type lookup via BrandMapping
        commercial_brands   = set(ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month,
            ad_type='COMMERCIAL BENEFITS',
        ).values_list('brand', flat=True))
        sponsorship_brands  = set(ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month,
            ad_type='SPONSORSHIP',
        ).values_list('brand', flat=True))
        theme_to_adtype = {}
        for bm in BrandMapping.objects.filter(account_id=account_id):
            key = (bm.theme or '').lower().strip()
            if bm.brand in commercial_brands:
                theme_to_adtype.setdefault(key, 'commercial')
            if bm.brand in sponsorship_brands:
                theme_to_adtype[key] = 'sponsorship'
        # Load sponsorship keyword list from settings (e.g. -BB, Tag, Com Break …)
        from .models import get_setting_list as _gsl
        spon_kw = _gsl('lmrb_sponsorship_keywords')

        def _pcts(d):
            total = d['prime'] + d['non_prime'] + d['other']
            return {
                **d,
                'total':          total,
                'prime_pct':      round(d['prime']     / total * 100) if total else 0,
                'non_prime_pct':  round(d['non_prime'] / total * 100) if total else 0,
                'other_pct':      round(d['other']     / total * 100) if total else 0,
            }

        # LMRB prime time distribution
        _pt_all  = {'prime': 0, 'non_prime': 0, 'other': 0}
        _pt_comm = {'prime': 0, 'non_prime': 0, 'other': 0}
        _pt_spon = {'prime': 0, 'non_prime': 0, 'other': 0}
        for lr in lmrb_qs.values('advt_theme', 'advt_time'):
            bucket = _time_bucket(lr['advt_time'])
            _pt_all[bucket] += 1
            tp = _theme_adtype(lr['advt_theme'], theme_to_adtype, spon_kw)
            if tp == 'commercial':
                _pt_comm[bucket] += 1
            elif tp == 'sponsorship':
                _pt_spon[bucket] += 1
        pt_all  = _pcts(_pt_all)
        pt_comm = _pcts(_pt_comm)
        pt_spon = _pcts(_pt_spon)

        # Schedule (planned) prime time distribution
        _sch_pt_comm = {'prime': 0, 'non_prime': 0, 'other': 0}
        _sch_pt_spon = {'prime': 0, 'non_prime': 0, 'other': 0}
        for sr in ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month,
        ).values('ad_type', 'start_time'):
            bucket = _time_bucket(sr['start_time'])
            if sr['ad_type'] == 'COMMERCIAL BENEFITS':
                _sch_pt_comm[bucket] += 1
            elif sr['ad_type'] == 'SPONSORSHIP':
                _sch_pt_spon[bucket] += 1
        sch_pt_comm = _pcts(_sch_pt_comm)
        sch_pt_spon = _pcts(_sch_pt_spon)

        # ── Matched LMRB rows (for LMRB Matched tab) ─────────────────────────
        lmrb_matched_rows = list(
            lmrb_qs.filter(is_matched=True).order_by('date', 'advt_time')
        )
        # ── Unmatched LMRB rows (for LMRB Unmatched tab) ─────────────────────
        lmrb_unmatched_rows = list(
            lmrb_qs.filter(is_matched=False).order_by('date', 'advt_time')
        )

        # ── Extra Aired rows ─────────────────────────────────────────────────
        # LMRB rows in the schedule period that are NOT matched, NOT
        # sponsorship-matched, NOT manual-matched, AND whose advt_theme
        # matches a BrandMapping for this account (i.e. they belong to our
        # brands — the advertiser got more airings than planned).
        _bm_themes_exact = set()
        _bm_themes_prefix = []
        for _bm in BrandMapping.objects.filter(account_id=account_id):
            _t = _bm.theme.strip()
            if not _t:
                continue
            if _t.endswith('*'):
                _bm_themes_prefix.append(_t[:-1].lower())
            else:
                _bm_themes_exact.add(_t.lower())

        def _theme_has_mapping(theme_val):
            tl = theme_val.lower().strip()
            if tl in _bm_themes_exact:
                return True
            for pfx in _bm_themes_prefix:
                if tl.startswith(pfx):
                    return True
            return False

        extra_aired_rows = []
        for lr in lmrb_unmatched_rows:
            if lr.is_sponsorship_matched or lr.is_manual_matched:
                continue
            if lr.advt_theme and _theme_has_mapping(lr.advt_theme):
                extra_aired_rows.append(lr)

        # ── Sponsorship Keyword-Type Breakdown ────────────────────────────────
        # For each keyword in lmrb_sponsorship_keywords, count LMRB rows that
        # contain that keyword (case-insensitive) grouped by PT / Non-PT, and
        # list the top programmes they aired in.
        # Also count planned sponsorship rows that have sponsorship_type matching
        # that keyword (set at schedule upload time).
        spon_kw_breakdown = []
        if spon_kw:
            for kw in spon_kw:
                if not kw:
                    continue
                kw_qs = lmrb_qs.filter(advt_theme__icontains=kw)
                kw_total_all = kw_qs.count()
                # Count planned rows whose sponsorship_type includes this keyword.
                # sponsorship_type may be pipe-separated (e.g. "-Intro|-Extro"),
                # so use icontains rather than iexact.
                kw_planned_all = _sr_base.filter(
                    ad_type='SPONSORSHIP', sponsorship_type__icontains=kw
                ).count()
                if kw_total_all == 0 and kw_planned_all == 0:
                    continue

                # ── Per-brand breakdown ──────────────────────────────────────────
                # Find all brands that have this sponsorship_type keyword so we can
                # separate e.g. "Mobitel Tags" from "SLT Tags" when both brands
                # share the same keyword.
                kw_brands_qs = (
                    _sr_base.filter(ad_type='SPONSORSHIP', sponsorship_type__icontains=kw)
                    .values_list('brand', flat=True).distinct()
                )
                kw_brand_list = sorted(set(kw_brands_qs))

                brand_entries = []
                for br in kw_brand_list:
                    br_planned = _sr_base.filter(
                        ad_type='SPONSORSHIP', brand=br
                    ).count()
                    # Derive product hint: BrandMapping.product → first word of brand
                    bm_prod = (
                        BrandMapping.objects.filter(account_id=account_id, brand=br)
                        .exclude(product='').values_list('product', flat=True).first() or ''
                    )
                    product_hint = (bm_prod or br.split()[0]).lower().strip()

                    # Product-aware LMRB filter with fallback
                    br_qs = kw_qs.filter(product__icontains=product_hint)
                    if not br_qs.exists():
                        br_qs = kw_qs  # fallback: no product metadata → use all

                    br_total = br_qs.count()
                    br_pt = br_non_pt = 0
                    for r in br_qs.values('advt_time'):
                        b = _time_bucket(r['advt_time'])
                        if b == 'prime':
                            br_pt += 1
                        elif b == 'non_prime':
                            br_non_pt += 1
                    br_progs = list(
                        br_qs.exclude(program='').exclude(program__isnull=True)
                        .values('program').annotate(cnt=Count('id')).order_by('-cnt')[:6]
                    )
                    brand_entries.append({
                        'brand':   br,
                        'planned': br_planned,
                        'total':   br_total,
                        'pt':      br_pt,
                        'non_pt':  br_non_pt,
                        'progs':   br_progs,
                    })

                # When only one brand uses this keyword (or no brands in schedule
                # but LMRB rows exist), keep the old aggregate behaviour as well.
                # Aggregate totals = sum of per-brand entries when brands exist,
                # else fall back to the raw keyword query total.
                if brand_entries:
                    kw_planned = sum(b['planned'] for b in brand_entries)
                    kw_total   = kw_total_all   # always show full LMRB keyword count
                    kw_pt      = sum(b['pt']     for b in brand_entries)
                    kw_non_pt  = sum(b['non_pt'] for b in brand_entries)
                    progs_agg: dict = {}
                    for be in brand_entries:
                        for p in be['progs']:
                            e = progs_agg.setdefault(p['program'], 0)
                            progs_agg[p['program']] = e + p['cnt']
                    progs = [{'program': p, 'cnt': c}
                             for p, c in sorted(progs_agg.items(), key=lambda x: -x[1])[:8]]
                else:
                    kw_planned = kw_planned_all
                    kw_total   = kw_total_all
                    kw_pt = kw_non_pt = 0
                    for r in kw_qs.values('advt_time'):
                        b = _time_bucket(r['advt_time'])
                        if b == 'prime':
                            kw_pt += 1
                        elif b == 'non_prime':
                            kw_non_pt += 1
                    progs = list(
                        kw_qs.exclude(program='').exclude(program__isnull=True)
                        .values('program').annotate(cnt=Count('id')).order_by('-cnt')[:8]
                    )

                spon_kw_breakdown.append({
                    'kw':      kw,
                    'planned': kw_planned,
                    'total':   kw_total,
                    'pt':      kw_pt,
                    'non_pt':  kw_non_pt,
                    'progs':   progs,
                    'brands':  brand_entries,
                })

        # Planned sponsorship PT/Non-PT for comparison
        spon_planned_pt     = _sch_pt_spon.get('prime', 0)
        spon_planned_non_pt = _sch_pt_spon.get('non_prime', 0)

        spon_kw_breakdown_json = _to_js(spon_kw_breakdown)

        # ── LMRB Sponsorship Pool ─────────────────────────────────────────────
        # All LMRB rows in scope that look like sponsorship (keyword match OR
        # BrandMapping-linked sponsorship brand), grouped by advt_theme.
        # Each row shows whether it has been assigned via SponsorshipLmrbAssignment.
        _assigned_lmrb_ids = set(
            SponsorshipLmrbAssignment.objects
            .filter(account_id=account_id)
            .values_list('lmrb_row_id', flat=True)
        )
        _spon_lmrb_theme_ids = set()
        for bm in BrandMapping.objects.filter(account_id=account_id):
            if bm.brand in sponsorship_brands:
                _spon_lmrb_theme_ids.add((bm.theme or '').lower().strip())

        lmrb_spon_pool = []
        for lr in lmrb_qs.order_by('advt_theme', 'date', 'advt_time'):
            t_lower = (lr.advt_theme or '').lower().strip()
            is_kw = any(kw and kw.lower() in t_lower for kw in spon_kw)
            is_mapped_spon = t_lower in _spon_lmrb_theme_ids
            if not (is_kw or is_mapped_spon):
                continue
            lmrb_spon_pool.append({
                'id':         lr.id,
                'date':       lr.date,
                'advt_time':  lr.advt_time,
                'advt_theme': lr.advt_theme,
                'duration':   lr.duration,
                'program':    lr.program or '-',
                'is_assigned': lr.id in _assigned_lmrb_ids,
                'is_mapped_spon': is_mapped_spon,
                'is_kw':      is_kw,
            })

        # ── Advt_Theme Analysis tab data ──────────────────────────────────────
        # For each unique Advt_Theme: daily airings, programme breakdown, PT split
        theme_analysis = []
        all_themes = sorted(set(lmrb_qs.values_list('advt_theme', flat=True)))
        for t in all_themes:
            t_qs = lmrb_qs.filter(advt_theme=t)
            t_total = t_qs.count()
            # Daily counts
            daily = list(
                t_qs.values('date').annotate(cnt=Count('id')).order_by('date')
            )
            # Programme breakdown
            progs = list(
                t_qs.exclude(program='').values('program').annotate(cnt=Count('id')).order_by('-cnt')[:10]
            )
            # PT split
            t_prime = t_non = 0
            for r in t_qs.values_list('advt_time', flat=True):
                bucket = _time_bucket(r)
                if bucket == 'prime':
                    t_prime += 1
                else:
                    t_non += 1
            theme_analysis.append({
                'theme':    t,
                'total':    t_total,
                'prime':    t_prime,
                'non_prime': t_non,
                'prime_pct': round(t_prime / t_total * 100) if t_total else 0,
                'daily':    daily,
                'programmes': progs,
            })
        theme_analysis_json = _to_js(theme_analysis)

    else:
        theme_analysis = []
        theme_analysis_json = '[]'
        spon_kw_breakdown = []
        spon_kw_breakdown_json = '[]'
        spon_planned_pt = 0
        spon_planned_non_pt = 0
        commercial_chart_data = []
        sponsorship_chart_data = []
        commercial_chart_json = '[]'
        sponsorship_chart_json = '[]'
        lmrb_spon_pool = []
        pt_all = pt_comm = pt_spon = sch_pt_comm = sch_pt_spon = {}

    return render(request, 'monitoring/dashboard.html', {
        'accounts':                  account_qs,
        'selected_account':          selected_account,
        'account_id':                account_id,
        'channels':                  channels,
        'months':                    months,
        'schedules_in_scope':        schedules_in_scope,
        'schedule_id':               schedule_id,
        'channel':                   channel,
        'month':                     month,
        'stats':                     stats,
        'tab_data':                  tab_data,
        'ch_summary':                ch_summary,
        'brand_summary':             brand_summary,
        'sponsorship_rows':          sponsorship_rows,
        'sponsorship_brand_summary': sponsorship_brand_summary,
        'lmrb_chart':                lmrb_chart,
        # Commercial / Sponsorship chart data (Product → Theme, Planned vs Aired)
        'commercial_chart_data':     commercial_chart_data,
        'sponsorship_chart_data':    sponsorship_chart_data,
        'commercial_chart_json':     commercial_chart_json,
        'sponsorship_chart_json':    sponsorship_chart_json,
        # Pre-serialised JSON - safe to emit with |safe in templates (no Python None/True/False)
        'lmrb_chart_json':           _to_js(list(lmrb_chart)),
        'lmrb_theme_detail':         lmrb_theme_detail if selected_account and channel and month else [],
        'lmrb_theme_detail_json':    _to_js(lmrb_theme_detail if selected_account and channel and month else []),
        'pt_all':                    pt_all if selected_account and channel and month else {},
        'pt_comm':                   pt_comm if selected_account and channel and month else {},
        'pt_spon':                   pt_spon if selected_account and channel and month else {},
        'sch_pt_comm':               sch_pt_comm if selected_account and channel and month else {},
        'sch_pt_spon':               sch_pt_spon if selected_account and channel and month else {},
        'lmrb_matched_rows':         lmrb_matched_rows,
        'lmrb_unmatched_rows':       lmrb_unmatched_rows,
        'extra_aired_rows':          extra_aired_rows,
        'theme_analysis':            theme_analysis,
        'theme_analysis_json':       theme_analysis_json,
        'spon_kw_breakdown':         spon_kw_breakdown if selected_account and channel and month else [],
        'spon_kw_breakdown_json':    spon_kw_breakdown_json if selected_account and channel and month else '[]',
        'spon_planned_pt':           spon_planned_pt if selected_account and channel and month else 0,
        'spon_planned_non_pt':       spon_planned_non_pt if selected_account and channel and month else 0,
        'lmrb_spon_pool':            lmrb_spon_pool,
    })


# ── Full Analytics Page ───────────────────────────────────────────────────────

@login_required
def analytics_full(request):
    """
    Comprehensive LMRB analytics page.
    Builds raw row JSON + pre-aggregated data for 25 charts across 6 sections.
    All chart computation + filtering is done client-side in JS using Chart.js.
    """
    from django.db.models import Avg, Sum

    user       = request.user
    account_qs = _account_qs(user)

    account_id = request.GET.get('account_id', '')
    channel    = request.GET.get('channel', '')
    month      = request.GET.get('month', '')
    # Prime-time window hours (configurable)
    prime_start_h = int(request.GET.get('prime_start', 18))
    prime_end_h   = int(request.GET.get('prime_end',   22))

    channels = []
    months   = []
    rows_json = '[]'
    kpis = {'total': 0, 'prime': 0, 'nonprime': 0, 'spon': 0, 'total_cost': None}
    filter_pg_list  = []
    filter_adv_list = []
    selected_account = None
    d_min = d_max = None

    if account_id and _account_access(user, account_id):
        try:
            selected_account = account_qs.get(id=account_id)
        except Exception:
            selected_account = None
        channels = list(
            ScheduleRow.objects.filter(account_id=account_id)
            .values_list('channel', flat=True).distinct().order_by('channel')
        )

    if account_id and channel and _account_access(user, account_id):
        months = list(
            ScheduleRow.objects.filter(account_id=account_id, channel=channel)
            .values_list('month', flat=True).distinct().order_by('month')
        )

    if account_id and channel and month and _account_access(user, account_id):
        # Date range from schedule
        dr = ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month
        ).aggregate(d_min=Min('date'), d_max=Max('date'))
        d_min, d_max = dr['d_min'], dr['d_max']

        # Build theme → ad_type mapping
        commercial_brands  = set(ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month,
            ad_type='COMMERCIAL BENEFITS',
        ).values_list('brand', flat=True))
        sponsorship_brands = set(ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month,
            ad_type='SPONSORSHIP',
        ).values_list('brand', flat=True))
        theme_to_adtype = {}
        for bm in BrandMapping.objects.filter(account_id=account_id):
            key = (bm.theme or '').lower().strip()
            if bm.brand in commercial_brands:
                theme_to_adtype.setdefault(key, 'commercial')
            if bm.brand in sponsorship_brands:
                theme_to_adtype[key] = 'sponsorship'
        from .models import get_setting_list as _gsl
        spon_kw = _gsl('lmrb_sponsorship_keywords')

        # Fetch all LMRB rows for the scope
        lmrb_qs = LMRBRow.objects.filter(
            account_id=account_id, channel__iexact=channel,
        )
        if d_min:
            lmrb_qs = lmrb_qs.filter(date__gte=d_min)
        if d_max:
            lmrb_qs = lmrb_qs.filter(date__lte=d_max)

        # Build compact row list for JS
        raw = []
        total_cost_sum = 0.0
        has_cost = False
        for row in lmrb_qs.values(
            'id', 'date', 'advt_theme', 'advt_time', 'duration',
            'program', 'product_group', 'advertiser',
            'ad_pos', 'tot_ads', 'brk_no', 'pos_in_brk', 'ads_in_brk',
            'cost', 'day',
        ):
            try:
                h = int(str(row['advt_time'] or '0').split(':')[0])
            except Exception:
                h = 0
            adtype = _theme_adtype(row['advt_theme'], theme_to_adtype, spon_kw)
            is_p   = prime_start_h <= h < prime_end_h
            cost_v = float(row['cost']) if row['cost'] is not None else None
            if cost_v is not None:
                has_cost = True
                total_cost_sum += cost_v
            # day-of-week: use 'day' field if present, else derive from date
            dow_name = (row['day'] or '').strip()
            if not dow_name and row['date']:
                dow_name = row['date'].strftime('%a')
            raw.append({
                'd':        str(row['date']),
                'dow':      row['date'].weekday() if row['date'] else 0,
                'downame':  dow_name,
                'h':        h,
                'theme':    row['advt_theme'] or '',
                'dur':      row['duration'],
                'prog':     row['program'] or '',
                'pg':       row['product_group'] or '',
                'adv':      row['advertiser'] or '',
                'adtype':   adtype,
                'prime':    is_p,
                'adpos':    row['ad_pos'],
                'totads':   row['tot_ads'],
                'brkno':    row['brk_no'],
                'posinbrk': row['pos_in_brk'],
                'adsinbrk': row['ads_in_brk'],
                'cost':     cost_v,
            })

        rows_json = _to_js(raw)

        # KPI summary
        kpis = {
            'total':      len(raw),
            'prime':      sum(1 for r in raw if r['prime']),
            'nonprime':   sum(1 for r in raw if not r['prime']),
            'spon':       sum(1 for r in raw if r['adtype'] == 'sponsorship'),
            'total_cost': round(total_cost_sum, 2) if has_cost else None,
        }
        if kpis['total']:
            kpis['prime_pct'] = round(kpis['prime'] / kpis['total'] * 100)
        else:
            kpis['prime_pct'] = 0

        # Filter option lists for the UI dropdowns
        filter_pg_list  = sorted({r['pg']  for r in raw if r['pg']})
        filter_adv_list = sorted({r['adv'] for r in raw if r['adv']})[:100]

    return render(request, 'monitoring/analytics.html', {
        'accounts':         account_qs,
        'selected_account': selected_account,
        'account_id':       account_id,
        'channels':         channels,
        'months':           months,
        'channel':          channel,
        'month':            month,
        'prime_start_h':    prime_start_h,
        'prime_end_h':      prime_end_h,
        'rows_json':        rows_json,
        'kpis':             kpis,
        'filter_pg_list':   filter_pg_list,
        'filter_adv_list':  filter_adv_list,
        'd_min':            d_min,
        'd_max':            d_max,
    })


# ── PDF Missed-Ad Report (Item 5) ─────────────────────────────────────────────

@login_required
def monitoring_pdf(request):
    """
    Generate a professional ReportLab PDF of Not-Aired / Programme-Mismatch
    rows for a given scope (account + channel + month).
    """
    # MatchResult already imported at top level

    account_id = request.GET.get('account_id', '')
    channel    = request.GET.get('channel', '')
    month      = request.GET.get('month', '')

    if not (account_id and channel and month):
        return HttpResponse('Missing parameters: account_id, channel, month', status=400)
    if not _account_access(request.user, account_id):
        return HttpResponse('Access denied', status=403)

    try:
        account = Account.objects.get(pk=account_id)
    except Account.DoesNotExist:
        return HttpResponse('Account not found', status=404)

    missed_qs = MatchResult.objects.filter(
        account_id=account_id, channel=channel, month=month,
        status__in=['not_aired', 'no_mapping', 'programme_mismatch', 'late_telecast'],
    ).order_by('scheduled_date', 'brand')

    sch_numbers = list(
        Schedule.objects.filter(account_id=account_id, channel=channel, month=month)
        .order_by('version')
        .values_list('schedule_number', flat=True)
    )
    schedule_number = ' / '.join(str(s) for s in sch_numbers if s)

    try:
        pdf_bytes = _build_missed_ad_pdf(
            account.name, channel, month, list(missed_qs),
            schedule_number=schedule_number,
        )
    except Exception as e:
        return HttpResponse(f'PDF generation failed: {e}', status=500)

    fname = f'missed_ads_{channel}_{month}.pdf'.replace(' ', '_')
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def _build_missed_ad_pdf(account_name, channel, month, rows, schedule_number=''):
    """Build a professional PDF bytes object using ReportLab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    NAVY   = colors.HexColor('#0b1726')
    BLUE   = colors.HexColor('#2563eb')
    LGRAY  = colors.HexColor('#f8fafc')
    MGRAY  = colors.HexColor('#e2e8f0')
    RED    = colors.HexColor('#dc2626')
    AMBER  = colors.HexColor('#d97706')
    VIOLET = colors.HexColor('#7c3aed')

    STATUS_COLOR = {
        'Not Aired':           RED,
        'No Brand Mapping':    colors.HexColor('#64748b'),
        'Programme Mismatch':  AMBER,
        'Late Telecast':       VIOLET,
    }

    h_title = ParagraphStyle('title', fontSize=16, textColor=NAVY,
                              fontName='Helvetica-Bold', spaceAfter=4)
    h_sub   = ParagraphStyle('sub',   fontSize=9,  textColor=colors.HexColor('#475569'),
                              fontName='Helvetica', spaceAfter=2)
    h_cell  = ParagraphStyle('cell',  fontSize=7.5, fontName='Helvetica')
    h_hdr   = ParagraphStyle('hdr',   fontSize=7.5, fontName='Helvetica-Bold',
                              textColor=colors.white)

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    story.append(Paragraph(f'Missed Ad Report - {channel}', h_title))
    meta_parts = [f'Account: {account_name}']
    if schedule_number:
        meta_parts.append(f'Schedule No: {schedule_number}')
    meta_parts.append(f'Month: {month}')
    story.append(Paragraph('  |  '.join(meta_parts), h_sub))
    story.append(Paragraph(
        f'Generated: {datetime.now().strftime("%d %B %Y %H:%M")}  |  '
        f'Total rows: {len(rows)}',
        h_sub,
    ))
    story.append(HRFlowable(width='100%', thickness=1, color=BLUE, spaceAfter=10))

    if not rows:
        story.append(Paragraph('No missed ads found for this scope.', styles['Normal']))
        doc.build(story)
        return buf.getvalue()

    # ── Table ─────────────────────────────────────────────────────────────────
    col_headers = [
        'Brand', 'Duration (s)', 'Programme',
        'Planned Date', 'Planned Start', 'Planned End',
        'Aired Date', 'Aired Time', 'Status',
    ]
    table_data = [[Paragraph(h, h_hdr) for h in col_headers]]

    for mr in rows:
        status_label = dict(mr.STATUS_CHOICES).get(mr.status, mr.status)
        table_data.append([
            Paragraph(mr.brand or '-', h_cell),
            Paragraph(str(mr.duration or '-'), h_cell),
            Paragraph(mr.programme or '-', h_cell),
            Paragraph(str(mr.scheduled_date or '-'), h_cell),
            Paragraph(mr.planned_start or '-', h_cell),
            Paragraph(mr.planned_end or '-', h_cell),
            Paragraph(str(mr.aired_date or '-'), h_cell),
            Paragraph(mr.air_time or '-', h_cell),
            Paragraph(status_label, ParagraphStyle(
                'st', fontSize=7.5, fontName='Helvetica-Bold',
                textColor=STATUS_COLOR.get(status_label, colors.black),
            )),
        ])

    # Column widths for landscape A4 (≈ 25 cm usable)
    col_widths = [4*cm, 1.8*cm, 4*cm, 2.2*cm, 2*cm, 2*cm, 2.2*cm, 2*cm, 3*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        # Header row
        ('BACKGROUND',  (0, 0), (-1, 0),  NAVY),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  7.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LGRAY, colors.white]),
        ('FONTSIZE',    (0, 1), (-1, -1), 7.5),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('GRID',        (0, 0), (-1, -1), 0.4, MGRAY),
        ('LINEBELOW',   (0, 0), (-1, 0),  1,   BLUE),
    ]))

    story.append(t)
    doc.build(story)
    return buf.getvalue()


@login_required
def full_ad_report_pdf(request):
    """
    Full Ad Report PDF - all MatchResult statuses for a scope:
    Matched, Programme Mismatch, Late Telecast, Not Aired, No Brand Mapping.

    Placed in the Verification section of the dashboard export bar and in
    the admin-export section.
    """
    account_id = request.GET.get('account_id', '')
    channel    = request.GET.get('channel', '')
    month      = request.GET.get('month', '')

    if not (account_id and channel and month):
        return HttpResponse('Missing parameters: account_id, channel, month', status=400)
    if not _account_access(request.user, account_id):
        return HttpResponse('Access denied', status=403)

    try:
        account = Account.objects.get(pk=account_id)
    except Account.DoesNotExist:
        return HttpResponse('Account not found', status=404)

    all_qs = MatchResult.objects.filter(
        account_id=account_id, channel=channel, month=month,
    ).order_by('scheduled_date', 'brand')

    sch_numbers = list(
        Schedule.objects.filter(account_id=account_id, channel=channel, month=month)
        .order_by('version')
        .values_list('schedule_number', flat=True)
    )
    schedule_number = ' / '.join(str(s) for s in sch_numbers if s)

    try:
        pdf_bytes = _build_full_ad_report_pdf(
            account.name, channel, month, list(all_qs),
            schedule_number=schedule_number,
        )
    except Exception as e:
        return HttpResponse(f'PDF generation failed: {e}', status=500)

    fname = f'full_ad_report_{channel}_{month}.pdf'.replace(' ', '_')
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def _build_full_ad_report_pdf(account_name, channel, month, rows, schedule_number=''):
    """Build Full Ad Report PDF - identical layout to missed-ad PDF but includes
    all statuses (matched, programme_mismatch, late_telecast, not_aired, no_mapping)
    with green colouring for Matched rows."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles    = getSampleStyleSheet()
    NAVY      = colors.HexColor('#0b1726')
    BLUE      = colors.HexColor('#2563eb')
    LGRAY     = colors.HexColor('#f8fafc')
    MGRAY     = colors.HexColor('#e2e8f0')
    RED       = colors.HexColor('#dc2626')
    AMBER     = colors.HexColor('#d97706')
    VIOLET    = colors.HexColor('#7c3aed')
    GREEN     = colors.HexColor('#16a34a')

    STATUS_COLOR = {
        'Matched':             GREEN,
        'Programme Mismatch':  AMBER,
        'Late Telecast':       VIOLET,
        'Not Aired':           RED,
        'No Brand Mapping':    colors.HexColor('#64748b'),
    }

    h_title = ParagraphStyle('title', fontSize=16, textColor=NAVY,
                              fontName='Helvetica-Bold', spaceAfter=4)
    h_sub   = ParagraphStyle('sub',   fontSize=9,  textColor=colors.HexColor('#475569'),
                              fontName='Helvetica', spaceAfter=2)
    h_cell  = ParagraphStyle('cell',  fontSize=7.5, fontName='Helvetica')
    h_hdr   = ParagraphStyle('hdr',   fontSize=7.5, fontName='Helvetica-Bold',
                              textColor=colors.white)

    story = []
    story.append(Paragraph(f'Full Ad Report - {channel}', h_title))
    meta_parts = [f'Account: {account_name}']
    if schedule_number:
        meta_parts.append(f'Schedule No: {schedule_number}')
    meta_parts.append(f'Month: {month}')
    story.append(Paragraph('  |  '.join(meta_parts), h_sub))
    story.append(Paragraph(
        f'Generated: {datetime.now().strftime("%d %B %Y %H:%M")}  |  '
        f'Total rows: {len(rows)}',
        h_sub,
    ))
    story.append(HRFlowable(width='100%', thickness=1, color=BLUE, spaceAfter=10))

    if not rows:
        story.append(Paragraph('No records found for this scope.', styles['Normal']))
        doc.build(story)
        return buf.getvalue()

    col_headers = [
        'Brand', 'Duration (s)', 'Programme',
        'Planned Date', 'Planned Start', 'Planned End',
        'Aired Date', 'Aired Time', 'Status',
    ]
    table_data = [[Paragraph(h, h_hdr) for h in col_headers]]

    for mr in rows:
        status_label = dict(mr.STATUS_CHOICES).get(mr.status, mr.status)
        table_data.append([
            Paragraph(mr.brand or '-', h_cell),
            Paragraph(str(mr.duration or '-'), h_cell),
            Paragraph(mr.programme or '-', h_cell),
            Paragraph(str(mr.scheduled_date or '-'), h_cell),
            Paragraph(mr.planned_start or '-', h_cell),
            Paragraph(mr.planned_end or '-', h_cell),
            Paragraph(str(mr.aired_date or '-'), h_cell),
            Paragraph(mr.air_time or '-', h_cell),
            Paragraph(status_label, ParagraphStyle(
                'st', fontSize=7.5, fontName='Helvetica-Bold',
                textColor=STATUS_COLOR.get(status_label, colors.black),
            )),
        ])

    col_widths = [4*cm, 1.8*cm, 4*cm, 2.2*cm, 2*cm, 2*cm, 2.2*cm, 2*cm, 3*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  NAVY),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  7.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LGRAY, colors.white]),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ('GRID',          (0, 0), (-1, -1), 0.4, MGRAY),
        ('LINEBELOW',     (0, 0), (-1, 0),  1,   BLUE),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# TC (Transmission Certificate) Upload & Management
# ═══════════════════════════════════════════════════════════════════════════════

def _find_col(df, *names):
    """Case-insensitive column finder. Returns the actual column name in df, or None."""
    lower_map = {c.lower().strip(): c for c in df.columns}
    for n in names:
        actual = lower_map.get(n.lower().strip())
        if actual is not None:
            return actual
    return None


def _detect_tc_meta(df):
    """
    Auto-detect channel, date range and row count from a TC DataFrame.
    Returns {'channel': str, 'start_date': date, 'end_date': date, 'row_count': int}
    """
    channel_col = _find_col(df, 'Channel', 'Station', 'CHANNEL')
    channel = _safe_str(df[channel_col].dropna().iloc[0]) if channel_col else ''

    date_col = _find_col(df, 'Date', 'Aired Date', 'Prg Date', 'AiredDate')

    dates = []
    if date_col:
        for v in df[date_col].dropna():
            d = _safe_date(v)
            if d:
                dates.append(d)

    start_date = min(dates) if dates else None
    end_date   = max(dates) if dates else None
    return {
        'channel':    channel,
        'start_date': start_date,
        'end_date':   end_date,
        'row_count':  len(df),
    }


def _parse_tc_rows(df, account, tc_report):
    """
    Parse TC DataFrame and upsert TCRow records.
    Handles various column naming conventions (case-insensitive).
    Returns count of rows inserted.
    """
    # ── Case-insensitive column normalisation ─────────────────────────────────
    # Build a rename map: actual_col → standard_name (only when needed)
    rename = {}

    def _ci_rename(standard, alts):
        """Map first found alt (case-insensitive) to standard name."""
        if _find_col(df, standard) is not None:
            return  # already present
        actual = _find_col(df, *alts)
        if actual is not None:
            rename[actual] = standard

    # Super-admin can add extra column aliases via Settings page
    _tc_ex_theme = get_setting_list('tc_extra_theme_aliases')
    _tc_ex_time  = get_setting_list('tc_extra_time_aliases')
    _tc_ex_date  = get_setting_list('tc_extra_date_aliases')
    _tc_ex_dur   = get_setting_list('tc_extra_duration_aliases')
    _tc_ex_prog  = get_setting_list('tc_extra_programme_aliases')

    _ci_rename('Channel',   ['Station', 'CHANNEL', 'channel'])
    _ci_rename('Date',      ['Aired Date', 'Prg Date', 'aired_date', 'AiredDate', 'Prg_Date',
                              *_tc_ex_date])
    _ci_rename('Programme', ['Program', 'Prg Name', 'PrgName', 'programme', *_tc_ex_prog])
    _ci_rename('TC_Theme',  ['Advt_Theme', 'Advt_theme', 'Theme', 'theme',
                              'Product', 'Description', 'Ad Name', 'AdName', 'Ad_Name',
                              *_tc_ex_theme])
    _ci_rename('Duration',  ['Dur', 'Seconds', 'Ad Dur', 'Duration_Sec', *_tc_ex_dur])
    _ci_rename('Aired_Time',['Advt_Time', 'Advt_time', 'advt_Time', 'Time',
                              'Aired Time', 'Ad Start', 'AdTime', 'AiredTime', *_tc_ex_time])

    if rename:
        df = df.rename(columns=rename)

    # ── Fallback defaults ─────────────────────────────────────────────────────
    if _find_col(df, 'TC_Theme') is None:
        df['TC_Theme'] = ''
    if _find_col(df, 'Aired_Time') is None:
        df['Aired_Time'] = ''
    if _find_col(df, 'Duration') is None:
        df['Duration'] = None
    if _find_col(df, 'Programme') is None:
        df['Programme'] = ''
    if _find_col(df, 'Channel') is None:
        df['Channel'] = tc_report.channel

    channel = tc_report.channel

    rows_by_key = {}
    for _, r in df.iterrows():
        theme      = _safe_str(r.get('TC_Theme', ''))
        aired_time = _safe_str(r.get('Aired_Time', ''))
        dur        = _safe_int(r.get('Duration'))
        date_val   = _safe_date(r.get('Date'))

        if not (theme and aired_time and date_val):
            continue

        dedup_key = TCRow.make_dedup_key(account.id, channel, date_val, aired_time, theme, dur)
        rows_by_key[dedup_key] = TCRow(
            account    = account,
            tc_report  = tc_report,
            channel    = channel,
            date       = date_val,
            programme  = _safe_str(r.get('Programme', '')),
            tc_theme   = theme,
            duration   = dur,
            aired_time = aired_time,
            dedup_key  = dedup_key,
        )

    if not rows_by_key:
        return 0

    # Delete any existing rows with same dedup keys (re-upload replaces)
    existing_keys = list(rows_by_key.keys())
    TCRow.objects.filter(dedup_key__in=existing_keys).delete()

    new_rows = list(rows_by_key.values())
    TCRow.objects.bulk_create(new_rows, batch_size=500)
    return len(new_rows)


@login_required
@require_POST
def tc_preview(request):
    """
    AJAX: Parse a TC file and return a data summary for user review.
    Does NOT save anything to the database.
    """
    account_id = request.POST.get('account_id', '').strip()
    channel    = request.POST.get('channel', '').strip()
    tc_file    = request.FILES.get('tc_file')

    if not tc_file:
        return JsonResponse({'ok': False, 'error': 'No file provided.'})

    if not _account_access(request.user, account_id):
        return JsonResponse({'ok': False, 'error': 'Access denied.'})

    # ── Parse the file (PDF or Excel) ─────────────────────────────────────────
    filename_lower = tc_file.name.lower()
    is_pdf = filename_lower.endswith('.pdf')

    if is_pdf:
        from verification.tc_converters.dispatch import get_converter
        converter = get_converter(channel)
        import tempfile, os
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                for chunk in tc_file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            df = converter.parse_pdf(tmp_path)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': f'Could not parse PDF: {e}'})
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
        if df.empty:
            return JsonResponse({'ok': False, 'error': 'No rows could be extracted from the PDF.'})
        df['Channel'] = channel
        parse_method = 'pdf'
    else:
        try:
            df = pd.read_excel(tc_file, header=0)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': f'Could not read file: {e}'})
        parse_method = 'excel'

    # ── Normalise column names (same logic as _parse_tc_rows) ─────────────────
    rename = {}

    def _ci_rename_local(standard, alts):
        if _find_col(df, standard) is not None:
            return
        actual = _find_col(df, *alts)
        if actual is not None:
            rename[actual] = standard

    _tc_ex_theme = get_setting_list('tc_extra_theme_aliases')
    _tc_ex_time  = get_setting_list('tc_extra_time_aliases')
    _tc_ex_date  = get_setting_list('tc_extra_date_aliases')
    _tc_ex_dur   = get_setting_list('tc_extra_duration_aliases')
    _tc_ex_prog  = get_setting_list('tc_extra_programme_aliases')

    _ci_rename_local('Channel',    ['Station', 'CHANNEL', 'channel'])
    _ci_rename_local('Date',       ['Aired Date', 'Prg Date', 'aired_date', 'AiredDate', 'Prg_Date', *_tc_ex_date])
    _ci_rename_local('Programme',  ['Program', 'Prg Name', 'PrgName', 'programme', *_tc_ex_prog])
    _ci_rename_local('TC_Theme',   ['Advt_Theme', 'Advt_theme', 'Theme', 'theme',
                                    'Product', 'Description', 'Ad Name', 'AdName', 'Ad_Name', *_tc_ex_theme])
    _ci_rename_local('Duration',   ['Dur', 'Seconds', 'Ad Dur', 'Duration_Sec', *_tc_ex_dur])
    _ci_rename_local('Aired_Time', ['Advt_Time', 'Advt_time', 'advt_Time', 'Time',
                                    'Aired Time', 'Ad Start', 'AdTime', 'AiredTime', *_tc_ex_time])
    if rename:
        df = df.rename(columns=rename)

    # Ensure required columns exist
    for col, default in [('TC_Theme', ''), ('Aired_Time', ''), ('Duration', None), ('Programme', '')]:
        if _find_col(df, col) is None:
            df[col] = default

    # ── Collect valid rows (same filters as _parse_tc_rows) ───────────────────
    rows = []
    skipped = 0
    for _, r in df.iterrows():
        theme      = _safe_str(r.get('TC_Theme', ''))
        aired_time = _safe_str(r.get('Aired_Time', ''))
        dur        = _safe_int(r.get('Duration'))
        date_val   = _safe_date(r.get('Date'))
        if not (theme and aired_time and date_val):
            skipped += 1
            continue
        rows.append({
            'date':      str(date_val),
            'programme': _safe_str(r.get('Programme', '')),
            'time':      aired_time,
            'theme':     theme,
            'duration':  dur,
        })

    if not rows:
        return JsonResponse({
            'ok': False,
            'error': f'No valid rows found (all {skipped} rows were missing Theme, Time, or Date). '
                     f'Check that column names are recognised.',
        })

    # ── Build summary ──────────────────────────────────────────────────────────
    from collections import Counter, defaultdict
    dates   = [r['date'] for r in rows]
    themes  = Counter()
    by_theme_dur = defaultdict(Counter)   # theme → {dur: count}

    for r in rows:
        key = (r['theme'], r['duration'])
        themes[key] += 1
        by_theme_dur[r['theme']][r['duration']] += 1

    theme_summary = []
    for (theme, dur), cnt in sorted(themes.items(), key=lambda x: -x[1]):
        theme_summary.append({'theme': theme, 'duration': dur, 'count': cnt})

    dur_breakdown = Counter(r['duration'] for r in rows)

    return JsonResponse({
        'ok':            True,
        'parse_method':  parse_method,
        'row_count':     len(rows),
        'skipped':       skipped,
        'date_range':    f"{min(dates)} → {max(dates)}",
        'themes':        theme_summary,
        'dur_breakdown': [{'duration': d, 'count': c}
                          for d, c in sorted(dur_breakdown.items(), key=lambda x: -(x[1]))],
        'sample_rows':   rows[:10],
    })


@login_required
@require_POST
def tc_upload_parsed(request):
    """
    Accept pre-parsed TC rows from the client-side PDF.js converter and save them.
    Rows arrive as a JSON array in POST['rows_json'].
    Each row: {date: "YYYY-MM-DD", programme, time, theme, duration}
    """
    account_id  = request.POST.get('account_id', '').strip()
    channel     = request.POST.get('channel', '').strip()
    month       = request.POST.get('month', '').strip()
    schedule_id = request.POST.get('schedule_id', '').strip()
    filename    = request.POST.get('filename', 'tc_upload.pdf')
    rows_json   = request.POST.get('rows_json', '[]')

    if not (account_id and channel and month):
        messages.error(request, 'Account, Channel and Month are required.')
        return redirect('/dashboard/tc/upload/')

    if not _account_access(request.user, account_id):
        messages.error(request, 'Access denied.')
        return redirect('/dashboard/tc/upload/')

    try:
        raw_rows = json.loads(rows_json)
    except Exception:
        messages.error(request, 'Invalid row data received.')
        return redirect('/dashboard/tc/upload/')

    if not raw_rows:
        messages.error(request, 'No rows to upload.')
        return redirect('/dashboard/tc/upload/')

    account = get_object_or_404(Account, id=account_id)

    schedule_obj = None
    if schedule_id:
        try:
            schedule_obj = Schedule.objects.get(id=schedule_id, account=account)
        except Schedule.DoesNotExist:
            pass

    # Build DataFrame from the pre-parsed JSON rows
    records = []
    for r in raw_rows:
        date_val = _safe_date(r.get('date'))
        if date_val is None:
            continue
        records.append({
            'Date':       date_val,
            'Channel':    channel,
            'Programme':  str(r.get('programme') or ''),
            'Aired_Time': str(r.get('time') or ''),
            'TC_Theme':   str(r.get('theme') or ''),
            'Duration':   _safe_int(r.get('duration')),
        })

    if not records:
        messages.error(request, 'No valid rows found - all rows were missing a valid date.')
        return redirect('/dashboard/tc/upload/')

    df = pd.DataFrame(records)
    meta = _detect_tc_meta(df)

    tc_report = TransmissionReport.objects.create(
        account           = account,
        channel           = channel,
        month             = month,
        schedule          = schedule_obj,
        original_filename = filename,
        start_date        = meta.get('start_date'),
        end_date          = meta.get('end_date'),
        uploaded_by       = request.user,
    )

    rows_count = _parse_tc_rows(df, account, tc_report)
    tc_report.row_count = rows_count
    tc_report.save()

    messages.success(
        request,
        f'TC Report uploaded successfully: {rows_count} rows from "{filename}".'
    )
    if request.user.role == 'channel_officer':
        return redirect('/dashboard/channel-officer/')
    return redirect('/dashboard/tc/')


@login_required
def tc_list(request):
    user       = request.user
    account_qs = _account_qs(user)
    account_id = request.GET.get('account_id', '')
    channel    = request.GET.get('channel', '')

    reports = TransmissionReport.objects.filter(
        account__in=account_qs
    ).select_related('account', 'uploaded_by').order_by('-uploaded_at')

    if account_id:
        reports = reports.filter(account_id=account_id)
    if channel:
        reports = reports.filter(channel=channel)

    channels = sorted(set(
        TransmissionReport.objects.filter(account__in=account_qs)
        .values_list('channel', flat=True)
    ))

    return render(request, 'tc/list.html', {
        'reports':    reports,
        'accounts':   account_qs,
        'channels':   channels,
        'account_id': account_id,
        'channel':    channel,
    })


@login_required
@role_required(['super_admin', 'admin', 'operations'])
def tc_pdf_convert(request):
    """
    GET  - render the PDF TC converter page.
    POST - receive converted rows as JSON, save to DB, return summary.
    """
    user       = request.user
    account_qs = _account_qs(user)

    if request.method == 'POST':
        import json as _json
        from datetime import datetime as _dt

        try:
            body       = _json.loads(request.body)
            account_id = body.get('account_id', '').strip()
            channel    = body.get('channel', '').strip()
            rows       = body.get('rows', [])

            if not (account_id and channel and rows):
                return JsonResponse({'ok': False, 'error': 'Missing account, channel, or rows.'})

            account = Account.objects.get(pk=account_id)
            if not _account_access(user, account_id):
                return JsonResponse({'ok': False, 'error': 'Access denied.'})

            # Parse dates and derive month / date range
            from datetime import date as _date_t
            valid_dates = []
            parsed_rows = []
            for r in rows:
                raw_date = r.get('DATE', '')
                try:
                    # Converter outputs D/M/YYYY
                    parts = raw_date.split('/')
                    d = _date_t(int(parts[2]), int(parts[1]), int(parts[0]))
                    valid_dates.append(d)
                    parsed_rows.append({**r, '_date': d})
                except Exception:
                    continue

            if not parsed_rows:
                return JsonResponse({'ok': False, 'error': 'No valid rows with parseable dates.'})

            start_date = min(valid_dates)
            end_date   = max(valid_dates)
            month      = start_date.strftime('%B %Y')

            tc_report = TransmissionReport.objects.create(
                account           = account,
                channel           = channel,
                month             = month,
                file              = None,
                original_filename = f'PDF_Converted_{channel}_{month}.xlsx',
                row_count         = 0,
                start_date        = start_date,
                end_date          = end_date,
                uploaded_by       = user,
            )

            new_rows = []
            for r in parsed_rows:
                theme      = (r.get('Advt_theme') or '').strip()
                aired_time = (r.get('Advt_time')  or '').strip()
                dur        = _safe_int(r.get('Duration'))
                date_val   = r['_date']
                if not (theme and aired_time):
                    continue
                dedup_key = TCRow.make_dedup_key(account.id, channel, date_val, aired_time, theme, dur)
                new_rows.append(TCRow(
                    account    = account,
                    tc_report  = tc_report,
                    channel    = channel,
                    date       = date_val,
                    programme  = (r.get('PROGRAMME') or '').strip(),
                    tc_theme   = theme,
                    duration   = dur,
                    aired_time = aired_time,
                    dedup_key  = dedup_key,
                ))

            # Dedup + save
            existing_keys = [rw.dedup_key for rw in new_rows]
            TCRow.objects.filter(dedup_key__in=existing_keys).delete()
            TCRow.objects.bulk_create(new_rows, batch_size=500)
            tc_report.row_count = len(new_rows)
            tc_report.save(update_fields=['row_count'])

            # Summary: group by theme × duration
            from collections import Counter
            summary = {}
            for rw in new_rows:
                key = (rw.tc_theme, rw.duration)
                summary[key] = summary.get(key, 0) + 1
            summary_list = [
                {'theme': k[0], 'duration': k[1], 'count': v}
                for k, v in sorted(summary.items(), key=lambda x: -x[1])
            ]

            return JsonResponse({
                'ok':         True,
                'report_id':  tc_report.id,
                'month':      month,
                'row_count':  len(new_rows),
                'summary':    summary_list,
            })

        except Account.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Account not found.'})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)})

    return render(request, 'tc/pdf_convert.html', {
        'accounts': account_qs,
    })


@login_required
@role_required(['super_admin', 'admin', 'team_head', 'operations', 'planner', 'channel_officer'])
def tc_upload(request):
    user       = request.user
    account_qs = _account_qs(user)

    # channel_officer: restrict to their assigned accounts (via User.accounts M2M)
    officer_profile = None
    if user.role == 'channel_officer':
        # Pick the profile matching the selected account+channel (or first)
        sel_ch = request.GET.get('channel', request.POST.get('channel', '')).strip()
        sel_ac = request.GET.get('account_id', request.POST.get('account_id', '')).strip()
        profiles_qs = ChannelOfficer.objects.filter(user=user).select_related('account')
        if sel_ch and sel_ac:
            officer_profile = profiles_qs.filter(account_id=sel_ac, channel=sel_ch).first()
        if officer_profile is None:
            officer_profile = profiles_qs.first()
        # Use the accounts assigned to this user (multi-account support)
        user_acct_qs = user.accounts.all()
        if user_acct_qs.exists():
            account_qs = user_acct_qs

    if request.method == 'POST':
        account_id = request.POST.get('account_id', '').strip()
        channel    = request.POST.get('channel', '').strip()
        month      = request.POST.get('month', '').strip()
        tc_file    = request.FILES.get('tc_file')
        schedule_id = request.POST.get('schedule_id', '').strip()

        if not (account_id and channel and month and tc_file):
            messages.error(request, 'Account, Channel, Month and TC file are required.')
            return redirect('/dashboard/tc/upload/')

        if not _account_access(user, account_id):
            messages.error(request, 'Access denied.')
            return redirect('/dashboard/tc/upload/')

        account = get_object_or_404(Account, id=account_id)

        # ── Detect file type and parse accordingly ────────────────────────────
        filename_lower = tc_file.name.lower()
        is_pdf = filename_lower.endswith('.pdf')

        if is_pdf:
            from verification.tc_converters.dispatch import get_converter
            converter = get_converter(channel)
            if converter is None:
                messages.error(
                    request,
                    f'No PDF converter is configured for channel "{channel}". '
                    f'Please upload an Excel file instead, or contact your administrator.'
                )
                return redirect('/dashboard/tc/upload/')

            import tempfile, os
            try:
                suffix = '.pdf'
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    for chunk in tc_file.chunks():
                        tmp.write(chunk)
                    tmp_path = tmp.name
                df = converter.parse_pdf(tmp_path)
            except Exception as e:
                messages.error(request, f'Could not parse PDF TC file: {e}')
                return redirect('/dashboard/tc/upload/')
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

            if df.empty:
                messages.error(request, 'No data rows could be extracted from the PDF.')
                return redirect('/dashboard/tc/upload/')

            # PDF parser already returns standard column names; ensure Channel present
            df['Channel'] = channel
        else:
            try:
                df = pd.read_excel(tc_file, header=0)
            except Exception as e:
                messages.error(request, f'Could not read TC file: {e}')
                return redirect('/dashboard/tc/upload/')

        schedule_obj = None
        if schedule_id:
            try:
                schedule_obj = Schedule.objects.get(id=schedule_id, account=account)
            except Schedule.DoesNotExist:
                pass

        meta = _detect_tc_meta(df)

        tc_report = TransmissionReport.objects.create(
            account           = account,
            channel           = channel,
            month             = month,
            schedule          = schedule_obj,
            file              = tc_file,
            original_filename = tc_file.name,
            row_count         = 0,
            start_date        = meta['start_date'],
            end_date          = meta['end_date'],
            uploaded_by       = user,
        )

        count = _parse_tc_rows(df, account, tc_report)
        tc_report.row_count = count
        tc_report.save(update_fields=['row_count'])

        messages.success(request, f'TC uploaded: {count} rows for {channel} / {month}.')
        if user.role == 'channel_officer':
            return redirect('/dashboard/channel-officer/')
        return redirect('/dashboard/tc/')

    # GET - show upload form (accept pre-fill params from WhatsApp links)
    schedules = Schedule.objects.filter(account__in=account_qs).select_related('account').order_by('-uploaded_at')
    prefill = {
        'account_id':  request.GET.get('account_id', ''),
        'channel':     request.GET.get('channel', ''),
        'month':       request.GET.get('month', ''),
        'schedule_pk': request.GET.get('schedule_pk', ''),
    }
    # channel_officer: auto-fill from their profile if not already set
    if officer_profile and not prefill['account_id']:
        prefill['account_id'] = str(officer_profile.account_id)
        prefill['channel']    = officer_profile.channel
    return render(request, 'tc/upload.html', {
        'accounts':       account_qs,
        'schedules':      schedules,
        'prefill':        prefill,
        'is_officer':     user.role == 'channel_officer',
    })


@login_required
def tc_detect(request):
    """AJAX: Detect channel + dates from an uploaded TC file."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'})

    tc_file = request.FILES.get('tc_file')
    if not tc_file:
        return JsonResponse({'ok': False, 'error': 'No file'})

    # PDF files: we can't detect metadata at AJAX time without a full parse;
    # return ok=True with empty fields so the form keeps what the user typed.
    if tc_file.name.lower().endswith('.pdf'):
        return JsonResponse({
            'ok':         True,
            'channel':    '',
            'start_date': '',
            'end_date':   '',
            'row_count':  0,
            'columns':    [],
            'is_pdf':     True,
        })

    try:
        df   = pd.read_excel(tc_file, header=0)
        meta = _detect_tc_meta(df)
        return JsonResponse({
            'ok':         True,
            'channel':    meta['channel'],
            'start_date': str(meta['start_date']) if meta['start_date'] else '',
            'end_date':   str(meta['end_date'])   if meta['end_date']   else '',
            'row_count':  meta['row_count'],
            'columns':    list(df.columns[:20]),
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required
@require_POST
def tc_delete(request, pk):
    report = get_object_or_404(TransmissionReport, pk=pk)
    user   = request.user

    # Channel officers can only delete TCs for their own account+channel combos
    if user.role == 'channel_officer':
        allowed_channels = list(
            ChannelOfficer.objects.filter(user=user).values_list('channel', flat=True)
        )
        if (report.account_id not in user.accounts.values_list('id', flat=True)
                or report.channel not in allowed_channels):
            messages.error(request, 'Access denied.')
            return redirect('/dashboard/channel-officer/')
    elif not _account_access(user, report.account_id):
        messages.error(request, 'Access denied.')
        return redirect('/dashboard/tc/')

    channel = report.channel
    month   = report.month
    report.delete()
    messages.success(request, f'TC report for {channel} / {month} deleted.')
    if user.role == 'channel_officer':
        return redirect('/dashboard/channel-officer/')
    return redirect('/dashboard/tc/')


@login_required
def tc_reconcile(request):
    """Run TC-Schedule + TC-LMRB reconciliation for a scope."""
    from verification.tc_engine import reconcile_tc

    account_id  = request.POST.get('account_id') or request.GET.get('account_id', '')
    channel     = request.POST.get('channel')    or request.GET.get('channel', '')
    month       = request.POST.get('month')      or request.GET.get('month', '')
    mode        = request.POST.get('mode', 'smart')
    schedule_id = (request.POST.get('schedule_id') or request.GET.get('schedule_id', '')).strip()

    if not (account_id and channel and month):
        messages.error(request, 'Account, channel and month are required.')
        return redirect('/dashboard/tc/')

    if not _account_access(request.user, account_id):
        messages.error(request, 'Access denied.')
        return redirect('/dashboard/tc/')

    sid = int(schedule_id) if schedule_id else None
    try:
        result = reconcile_tc(account_id, channel, month, mode=mode, schedule_id=sid)
        msg = (
            f'Reconciliation complete: {result["matched"]} matched, '
            f'{result["extra"]} extra, {result["lmrb_confirmed"]} LMRB-confirmed.'
        )
        messages.success(request, msg)

        # WhatsApp: notify operations officer that reconciliation is done
        try:
            _notify_reconciliation_done_whatsapp(account_id, channel, month, result)
        except Exception as e:
            print(f"[tc_reconcile] whatsapp notify error (non-fatal): {e}")

        # ── Diagnostic: if 0 TC-Schedule matches, show the unique TC themes so
        # the user knows what to enter in BrandMapping → tc_theme field.
        if result['matched'] == 0 and result['extra'] > 0:
            unique_themes = list(
                TCRow.objects.filter(
                    account_id=account_id, channel=channel, tc_report__month=month,
                ).values_list('tc_theme', flat=True).distinct().order_by('tc_theme')
            )
            # Also check if BrandMappings exist with tc_theme set
            mapped_themes = list(
                BrandMapping.objects.filter(account_id=account_id)
                .exclude(tc_theme='').values_list('tc_theme', flat=True).distinct()
            )
            if not mapped_themes:
                themes_str = ', '.join(f'"{t}"' for t in unique_themes[:10])
                messages.warning(
                    request,
                    f'No BrandMapping tc_theme values are configured. '
                    f'The TC file contains these themes: {themes_str}. '
                    f'Go to Brand Mappings and set the "TC Theme" field for each brand '
                    f'to match exactly what appears in the TC file.'
                )
    except Exception as e:
        messages.error(request, f'Reconciliation failed: {e}')

    redir = f'/dashboard/summary/?account_id={account_id}&channel={channel}&month={month}'
    if schedule_id:
        redir += f'&schedule_id={schedule_id}'
    return redirect(redir)


# ═══════════════════════════════════════════════════════════════════════════════
# TC Three-Way Comparison  (Plan vs LMRB vs TC)
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def tc_three_way(request):
    """
    Three-way evidence view for each planned ad in a scope:
      PLAN  - what was scheduled  (ScheduleRow)
      LMRB  - what 3rd-party monitoring observed  (matched LMRBRow)
      TC    - what the channel's own certificate says  (matched TCRow)

    Matching priority applied by the engines (already in DB):
      1. Theme  (via BrandMapping)
      2. Duration
      3. Date   (LMRB within schedule date range; TC late-aired rule: date >= schedule date)
      4. Advt_Time / Aired_Time (±5 sec tolerance for TC-LMRB cross-check)
    """
    user       = request.user
    account_qs = _account_qs(user)

    account_id  = request.GET.get('account_id', '')
    channel     = request.GET.get('channel', '')
    month       = request.GET.get('month', '')
    schedule_id = request.GET.get('schedule_id', '').strip()

    selected_account   = None
    channels           = []
    months             = []
    schedules_in_scope = []
    rows               = []

    if account_id:
        try:
            selected_account = account_qs.get(pk=account_id)
        except Account.DoesNotExist:
            pass

    if selected_account:
        channels = sorted(set(
            ScheduleRow.objects.filter(account_id=account_id)
            .values_list('channel', flat=True)
        ))
        if channel:
            months = sorted(set(
                ScheduleRow.objects.filter(account_id=account_id, channel=channel)
                .values_list('month', flat=True)
            ))

    if selected_account and channel and month:
        if not _account_access(user, account_id):
            messages.error(request, 'Access denied.')
            return redirect('/dashboard/tc/detail/')

        schedules_in_scope = list(
            Schedule.objects.filter(account_id=account_id, channel=channel, month=month)
            .order_by('schedule_number')
        )
        if not schedule_id and len(schedules_in_scope) == 1:
            schedule_id = str(schedules_in_scope[0].id)

        from core.models import ManualMatch
        from verification.tc_engine import _time_to_secs as _tts3w

        def _pt(t):
            """Parse HH:MM:SS → seconds, or None."""
            return _tts3w(str(t)) if t else None

        sch_filter = dict(account_id=account_id, channel=channel, month=month)
        if schedule_id:
            sch_filter['schedule_id'] = schedule_id
        sch_qs = (
            ScheduleRow.objects
            .filter(**sch_filter)
            .select_related('matched_lmrb')
            .prefetch_related('tc_matches', 'tc_matches__matched_lmrb')
            .order_by('date', 'start_time', 'brand')
        )

        # Build manual match lookup: {schedule_row_id: ManualMatch}
        mm_lookup = {
            mm.schedule_row_id: mm
            for mm in ManualMatch.objects.filter(
                account_id=account_id, channel=channel, month=month,
                schedule_row__isnull=False,
            ).select_related('tc_row', 'lmrb_row')
        }

        # Pre-load LMRB rows for nearest-match search on tc_only rows.
        # (is_lmrb_confirmed lives on TCRow, not LMRBRow - no such filter here.)
        scope_lmrb_pool = list(
            LMRBRow.objects.filter(
                account_id=account_id,
                channel=channel,
            ).values('advt_time', 'duration', 'date')
        )

        for sr in sch_qs:
            lmrb = sr.matched_lmrb
            tc   = sr.tc_matches.filter(is_schedule_matched=True).first()

            # Check for manual match (3way) which also sets is_schedule_matched
            mm = mm_lookup.get(sr.id)
            is_manual = mm is not None

            # If engine didn't find a TC match but there's a manual 3way link, use that
            if not tc and mm and mm.tc_row:
                tc = mm.tc_row

            # Determine row status for colour-coding
            if tc and tc.is_lmrb_confirmed:
                status = 'aired'          # confirmed by both TC and LMRB
            elif tc and not tc.is_lmrb_confirmed:
                status = 'tc_only'        # TC says aired but LMRB doesn't confirm
            elif lmrb or (mm and mm.lmrb_row):
                status = 'lmrb_only'      # LMRB found a match but no TC record
            elif is_manual:
                status = 'manual'         # manually matched, no TC/LMRB confirmation
            else:
                status = 'not_aired'      # neither source confirms it

            # Use manual LMRB if engine didn't lock one
            if not lmrb and mm and mm.lmrb_row:
                lmrb = mm.lmrb_row

            # ── Display LMRB: prefer the LMRB that TC was confirmed against ──────
            # When TC is confirmed, tc.matched_lmrb is the LMRB row within ±tolerance
            # seconds of the TC aired time - i.e. the SAME physical broadcast.
            # Using sr.matched_lmrb (Schedule↔LMRB engine result) can show a
            # different LMRB from a completely different time slot, which confuses
            # users into thinking the pair is wrong.  When TC is not confirmed we
            # fall back to sr.matched_lmrb so something useful is still shown.
            display_lmrb = lmrb
            if tc and tc.is_lmrb_confirmed and tc.matched_lmrb_id:
                display_lmrb = tc.matched_lmrb

            # ── Timing deltas ───────────────────────────────────────────
            plan_start_s  = _pt(sr.start_time)
            plan_end_s    = _pt(sr.end_time)
            lmrb_time_s   = _pt(display_lmrb.advt_time if display_lmrb else None)
            tc_time_s     = _pt(tc.aired_time if tc else None)

            # LMRB time delta relative to plan window (0 = within, positive = seconds outside)
            lmrb_delta = None
            if lmrb_time_s is not None and plan_start_s is not None and plan_end_s is not None:
                if plan_start_s <= lmrb_time_s <= plan_end_s:
                    lmrb_delta = 0
                elif lmrb_time_s > plan_end_s:
                    lmrb_delta = lmrb_time_s - plan_end_s
                else:
                    lmrb_delta = plan_start_s - lmrb_time_s

            # TC↔LMRB time gap (seconds)
            tc_lmrb_delta = None
            if tc_time_s is not None and lmrb_time_s is not None:
                tc_lmrb_delta = abs(tc_time_s - lmrb_time_s)

            # ── Nearest unconfirmed LMRB for tc_only rows ───────────────
            near_lmrb_time  = ''
            near_lmrb_delta = None
            tc_date_val     = tc.date if tc else None
            if status == 'tc_only' and tc_time_s is not None:
                best_d, best_t = None, ''
                for lr in scope_lmrb_pool:
                    if lr['duration'] != sr.duration:
                        continue
                    if lr['date'] != tc_date_val:
                        continue
                    lt_s = _pt(str(lr['advt_time']))
                    if lt_s is None:
                        continue
                    d = abs(tc_time_s - lt_s)
                    if best_d is None or d < best_d:
                        best_d, best_t = d, str(lr['advt_time'])
                near_lmrb_time  = best_t
                near_lmrb_delta = best_d

            # ── Confidence score (0-100) ─────────────────────────────────
            lmrb_date_val = display_lmrb.date if display_lmrb else None
            if status == 'aired':
                score = 100
                if lmrb_date_val != sr.date:
                    score -= 10
                if lmrb_delta is not None and lmrb_delta > 0:
                    score -= min(20, lmrb_delta // 30)
                if tc_lmrb_delta is not None and tc_lmrb_delta > 0:
                    score -= min(10, tc_lmrb_delta)
                score = max(60, score)
            elif status == 'tc_only':
                score = 55
                if tc_date_val == sr.date:
                    score += 10
                if near_lmrb_delta is not None:
                    if near_lmrb_delta <= 30:
                        score += 15
                    elif near_lmrb_delta <= 120:
                        score += 7
                score = min(79, score)
            elif status == 'lmrb_only':
                score = 50
                if lmrb_date_val == sr.date:
                    score += 15
                if lmrb_delta == 0:
                    score += 15
                score = min(74, score)
            elif status == 'manual':
                score = 70
            else:
                score = 0

            if score >= 85:
                conf_label = 'excellent'
            elif score >= 65:
                conf_label = 'good'
            elif score >= 40:
                conf_label = 'partial'
            elif score > 0:
                conf_label = 'low'
            else:
                conf_label = 'none'

            rows.append({
                'brand':          sr.brand,
                'ad_type':        sr.ad_type,
                'duration':       sr.duration,
                # Plan (Schedule)
                'plan_date':      sr.date,
                'plan_programme': sr.programme,
                'plan_start':     sr.start_time,
                'plan_end':       sr.end_time,
                # LMRB (3rd-party monitoring) - uses TC-confirmed LMRB when available
                'lmrb_date':      display_lmrb.date       if display_lmrb else None,
                'lmrb_programme': display_lmrb.program    if display_lmrb else '',
                'lmrb_time':      display_lmrb.advt_time  if display_lmrb else '',
                'lmrb_theme':     display_lmrb.advt_theme if display_lmrb else '',
                # TC (channel certificate)
                'tc_date':        tc_date_val,
                'tc_programme':   tc.programme     if tc else '',
                'tc_time':        tc.aired_time    if tc else '',
                'tc_theme':       tc.tc_theme      if tc else '',
                # Status flags
                'has_lmrb':            display_lmrb is not None,
                'has_tc':              tc is not None,
                'tc_lmrb_confirmed':   tc.is_lmrb_confirmed if tc else False,
                'is_manual':           is_manual,
                'status':              status,
                # Match quality / confidence
                'confidence_pct':   score,
                'confidence_label': conf_label,
                'lmrb_delta_secs':  lmrb_delta,
                'tc_lmrb_delta_secs': tc_lmrb_delta,
                'near_lmrb_time':   near_lmrb_time,
                'near_lmrb_delta':  near_lmrb_delta,
            })

    return render(request, 'tc/detail.html', {
        'accounts':           account_qs,
        'selected_account':   selected_account,
        'account_id':         account_id,
        'channels':           channels,
        'months':             months,
        'schedules_in_scope': schedules_in_scope,
        'schedule_id':        schedule_id,
        'channel':            channel,
        'month':              month,
        'rows':               rows,
        'total':              len(rows),
        'n_aired':            sum(1 for r in rows if r['status'] == 'aired'),
        'n_tc_only':          sum(1 for r in rows if r['status'] == 'tc_only'),
        'n_lmrb_only':        sum(1 for r in rows if r['status'] == 'lmrb_only'),
        'n_not_aired':        sum(1 for r in rows if r['status'] == 'not_aired'),
    })


# ═══════════════════════════════════════════════════════════════════════════════
# Summary Sheet Report
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def summary_report(request):
    """View + edit Summary Sheet metadata; preview the reconciliation results."""
    from verification.tc_engine import build_summary_data

    user       = request.user
    account_qs = _account_qs(user)

    account_id  = request.GET.get('account_id', '')
    channel     = request.GET.get('channel', '')
    month       = request.GET.get('month', '')
    schedule_id = request.GET.get('schedule_id', '').strip()

    # Save metadata if POST
    if request.method == 'POST':
        account_id  = request.POST.get('account_id', '').strip()
        channel     = request.POST.get('channel', '').strip()
        month       = request.POST.get('month', '').strip()
        schedule_id = request.POST.get('schedule_id', '').strip()

        if account_id and channel and month and _account_access(user, account_id):
            account = get_object_or_404(Account, id=account_id)
            meta, _ = SummaryReportMeta.objects.get_or_create(
                account=account, channel=channel, month=month
            )
            meta.supplier_invoice_no = request.POST.get('supplier_invoice_no', '').strip()
            meta.po_no               = request.POST.get('po_no', '').strip()
            meta.invoice_no          = request.POST.get('invoice_no', '').strip()
            meta.notes               = request.POST.get('notes', '').strip()
            meta.prepared_by         = request.POST.get('prepared_by', '').strip()
            meta.checked_by          = request.POST.get('checked_by', '').strip()
            meta.authorised_by       = request.POST.get('authorised_by', '').strip()
            meta.save()
            messages.success(request, 'Summary metadata saved.')
        qs_str = f'account_id={account_id}&channel={channel}&month={month}'
        if schedule_id:
            qs_str += f'&schedule_id={schedule_id}'
        return redirect(f'/dashboard/summary/?{qs_str}')

    # Dropdown data
    channels          = []
    months            = []
    schedules_in_scope = []
    selected_account  = None
    summary_data      = None
    meta              = None
    schedule_obj      = None
    tc_report         = None

    if account_id:
        try:
            selected_account = account_qs.get(pk=account_id)
        except Account.DoesNotExist:
            pass

    if selected_account:
        channels = sorted(set(
            ScheduleRow.objects.filter(account_id=account_id)
            .values_list('channel', flat=True)
        ))
        if channel:
            months = sorted(set(
                ScheduleRow.objects.filter(account_id=account_id, channel=channel)
                .values_list('month', flat=True)
            ))

    if selected_account and channel and month:
        if not _account_access(user, account_id):
            messages.error(request, 'Access denied.')
            return redirect('/dashboard/summary/')

        # Build schedule selector for this scope
        schedules_in_scope = list(
            Schedule.objects.filter(account_id=account_id, channel=channel, month=month)
            .order_by('schedule_number')
        )
        # Auto-select when only one schedule exists
        if not schedule_id and len(schedules_in_scope) == 1:
            schedule_id = str(schedules_in_scope[0].id)

        sid = int(schedule_id) if schedule_id else None
        try:
            summary_data = build_summary_data(account_id, channel, month, schedule_id=sid)
        except Exception as e:
            messages.warning(request, f'Could not build summary: {e}')

        meta = SummaryReportMeta.objects.filter(
            account_id=account_id, channel=channel, month=month
        ).first()

        if schedule_id:
            schedule_obj = Schedule.objects.filter(id=schedule_id).first()
        else:
            schedule_obj = schedules_in_scope[0] if schedules_in_scope else None

        tc_report = TransmissionReport.objects.filter(
            account_id=account_id, channel=channel, month=month
        ).order_by('-uploaded_at').first()

    return render(request, 'summary/report.html', {
        'accounts':           account_qs,
        'selected_account':   selected_account,
        'account_id':         account_id,
        'channels':           channels,
        'months':             months,
        'schedules_in_scope': schedules_in_scope,
        'schedule_id':        schedule_id,
        'channel':            channel,
        'month':              month,
        'summary_data':       summary_data,
        'meta':               meta,
        'schedule_obj':       schedule_obj,
        'tc_report':          tc_report,
    })


@login_required
def summary_excel(request):
    """Download Summary Sheet as a formatted Excel workbook."""
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    from verification.tc_engine import build_summary_data

    account_id = request.GET.get('account_id', '')
    channel    = request.GET.get('channel', '')
    month      = request.GET.get('month', '')

    if not (account_id and channel and month):
        messages.error(request, 'Incomplete parameters.')
        return redirect('/dashboard/summary/')

    if not _account_access(request.user, account_id):
        messages.error(request, 'Access denied.')
        return redirect('/dashboard/summary/')

    account  = get_object_or_404(Account, id=account_id)
    data     = build_summary_data(account_id, channel, month)
    meta     = SummaryReportMeta.objects.filter(
        account_id=account_id, channel=channel, month=month
    ).first()
    sched    = Schedule.objects.filter(
        account_id=account_id, channel=channel, month=month
    ).order_by('-uploaded_at').first()

    estimate_no = sched.schedule_number if sched else ''

    # ── Workbook setup ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Summary'

    # Styles
    NAVY  = PatternFill('solid', fgColor='0F2340')
    LBLUE = PatternFill('solid', fgColor='DBEAFE')
    LGRAY = PatternFill('solid', fgColor='F8FAFC')
    DGRAY = PatternFill('solid', fgColor='E2E8F0')
    GOLD  = PatternFill('solid', fgColor='FEF9C3')
    GRNFILL = PatternFill('solid', fgColor='DCFCE7')
    REDFILL  = PatternFill('solid', fgColor='FEE2E2')

    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    bold11    = Font(bold=True, size=11)
    bold10    = Font(bold=True, size=10)
    norm10    = Font(size=10)
    title_font = Font(bold=True, size=14, color='0F2340')

    thin = Side(style='thin', color='CBD5E1')
    thick_side = Side(style='medium', color='2563EB')
    def border(left=True, right=True, top=True, bottom=True):
        return Border(
            left=Side(style='thin', color='CBD5E1') if left else Side(),
            right=Side(style='thin', color='CBD5E1') if right else Side(),
            top=Side(style='thin', color='CBD5E1') if top else Side(),
            bottom=Side(style='thin', color='CBD5E1') if bottom else Side(),
        )

    centre = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')

    # Column widths (7 data columns - Avg 30s removed)
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12

    row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:G{row}')
    c = ws.cell(row, 1, 'SUMMARY SHEET REPORT')
    c.font = title_font; c.alignment = centre; c.fill = LGRAY
    ws.row_dimensions[row].height = 28
    row += 1

    # ── Company info rows ─────────────────────────────────────────────────────
    company_info = [
        ('Phoenix O & M (Pvt) Ltd', 'MONTH',                month),
        ('No 16, Barnes Place',     'CHANNEL',               channel),
        ('Colombo 7',               'ESTIMATE NO',           estimate_no),
        ('',                        'SUPPLIER INVOICE NO',   meta.supplier_invoice_no if meta else ''),
        ('',                        'PO NO',                 meta.po_no if meta else ''),
        ('',                        'INVOICE NO',            meta.invoice_no if meta else ''),
    ]
    for addr, label, value in company_info:
        ws.merge_cells(f'A{row}:C{row}')
        c = ws.cell(row, 1, addr); c.font = norm10; c.alignment = left
        ws.merge_cells(f'D{row}:E{row}')
        c = ws.cell(row, 4, label + '  :'); c.font = bold10; c.alignment = right
        ws.merge_cells(f'F{row}:G{row}')
        c = ws.cell(row, 6, value); c.font = norm10; c.alignment = left
        row += 1

    row += 1  # blank

    # ── Column headers ────────────────────────────────────────────────────────
    headers = ['PRODUCT', 'DUR', 'PLANNED', 'AIRED', 'MISSED', 'EXTRA', '3RD PARTY']
    for col_i, h in enumerate(headers, start=1):
        c = ws.cell(row, col_i, h)
        c.font = hdr_font; c.fill = NAVY; c.alignment = centre
        c.border = border()
    ws.row_dimensions[row].height = 20
    row += 1

    # ── Commercial Benefits section ───────────────────────────────────────────
    if data['commercial']:
        ws.merge_cells(f'A{row}:G{row}')
        c = ws.cell(row, 1, 'Commercial Benefits'); c.font = bold10; c.fill = LBLUE
        c.alignment = left
        row += 1

        for item in data['commercial']:
            vals = [item['product'], item['dur'], item['planned'], item['aired'],
                    item['missed'], item['extra'], item['third_party']]
            fill = REDFILL if item['missed'] > 0 else None
            for col_i, v in enumerate(vals, start=1):
                c = ws.cell(row, col_i, v)
                c.font = norm10
                c.alignment = centre if col_i > 1 else left
                c.border = border()
                if fill:
                    c.fill = fill
            row += 1

        # Grand Total
        t = data['commercial_total']
        total_vals = ['Grand Total', '', t['planned'], t['aired'], t['missed'],
                      t['extra'], t['third_party']]
        for col_i, v in enumerate(total_vals, start=1):
            c = ws.cell(row, col_i, v)
            c.font = bold10; c.fill = DGRAY; c.alignment = centre if col_i > 1 else left
            c.border = border()
        row += 1

    row += 1  # blank

    # ── Sponsorship Benefits section ──────────────────────────────────────────
    if data['sponsorship']:
        ws.merge_cells(f'A{row}:G{row}')
        c = ws.cell(row, 1, 'Sponsorship Benefits'); c.font = bold11; c.fill = GOLD
        c.alignment = left
        row += 1

        for section in data['sponsorship']:
            # Programme heading
            ws.merge_cells(f'A{row}:G{row}')
            c = ws.cell(row, 1, section['programme'].upper())
            c.font = bold10; c.fill = LBLUE; c.alignment = left
            row += 1

            for item in section['rows']:
                vals = [item['product'], item['dur'], item['planned'], item['aired'],
                        item['missed'], item['extra'], item['third_party']]
                for col_i, v in enumerate(vals, start=1):
                    c = ws.cell(row, col_i, v)
                    c.font = norm10
                    c.alignment = centre if col_i > 1 else left
                    c.border = border()
                row += 1
            # No per-programme subtotal - only grand total at end

        # Sponsorship Grand Total
        st = data['sponsorship_total']
        ws.merge_cells(f'A{row}:G{row}')
        c = ws.cell(row, 1, 'SPONSORSHIP GRAND TOTAL'); c.font = bold11; c.fill = GOLD
        c.alignment = left
        row += 1
        total_vals = ['Grand Total', '', st['planned'], st['aired'], st['missed'],
                      st['extra'], st['third_party']]
        for col_i, v in enumerate(total_vals, start=1):
            c = ws.cell(row, col_i, v)
            c.font = bold10; c.fill = DGRAY; c.alignment = centre if col_i > 1 else left
            c.border = border()
        row += 1

    row += 1  # blank

    # ── Notes ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row, 1, 'NOTE:').font = bold10
    row += 1
    notes_text = meta.notes if meta and meta.notes else 'Channel transmission attached\nSponsorship Benefits confirmation email attached'
    for line in notes_text.split('\n'):
        ws.merge_cells(f'A{row}:G{row}')
        c = ws.cell(row, 1, line.strip()); c.font = norm10
        row += 1

    row += 2  # blank

    # ── Signatures (3 across 7 columns: A-B, C-D-E, F-G) ─────────────────────
    sig_labels = ['PREPARED BY', 'CHECKED BY', 'AUTHORISED BY']
    sig_values = [
        meta.prepared_by if meta else '',
        meta.checked_by  if meta else '',
        meta.authorised_by if meta else '',
    ]
    sig_ranges = [(1, 2), (3, 5), (6, 7)]
    for (start_col, end_col), lbl in zip(sig_ranges, sig_labels):
        ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row,   end_column=end_col,
        )
        c = ws.cell(row, start_col, '………………………………………………')
        c.font = norm10; c.alignment = centre
        ws.merge_cells(
            start_row=row+1, start_column=start_col,
            end_row=row+1,   end_column=end_col,
        )
        c = ws.cell(row+1, start_col, lbl); c.font = bold10; c.alignment = centre

    # ── Matched LMRB sheet (second sheet) ────────────────────────────────────
    ws_lmrb = wb.create_sheet('Matched LMRB')
    _write_matched_lmrb_sheet(ws_lmrb, account_id, channel, month)

    # ── Build response ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'Summary_{account.name}_{channel}_{month}.xlsx'.replace(' ', '_')
    resp  = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def summary_pdf(request):
    """Premium Reconciliation PDF: full-bleed cover + summary tables + matched LMRB."""
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib.utils import ImageReader as _IR
    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageTemplate, NextPageTemplate,
        Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from verification.tc_engine import build_summary_data
    from django.db.models import Min, Max

    account_id = request.GET.get('account_id', '')
    channel    = request.GET.get('channel', '')
    month      = request.GET.get('month', '')

    if not (account_id and channel and month):
        messages.error(request, 'Incomplete parameters.')
        return redirect('/dashboard/summary/')
    if not _account_access(request.user, account_id):
        messages.error(request, 'Access denied.')
        return redirect('/dashboard/summary/')

    account = get_object_or_404(Account, id=account_id)
    data    = build_summary_data(account_id, channel, month)
    meta    = SummaryReportMeta.objects.filter(
        account_id=account_id, channel=channel, month=month
    ).first()
    sched = Schedule.objects.filter(
        account_id=account_id, channel=channel, month=month
    ).order_by('-uploaded_at').first()
    estimate_no = sched.schedule_number if sched else ''

    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY      = HexColor('#0D1F3C')   # primary deep navy
    NAVY_MID  = HexColor('#1A3353')   # slightly lighter navy (accents)
    GOLD      = HexColor('#C8A951')   # warm gold accent
    GOLD_LITE = HexColor('#F0DFA0')   # pale gold tint
    WHITE     = colors.white
    OFF_WHITE = HexColor('#F7F9FC')
    SLATE_100 = HexColor('#F1F5F9')
    SLATE_200 = HexColor('#E2E8F0')
    SLATE_400 = HexColor('#94A3B8')
    SLATE_600 = HexColor('#475569')
    SLATE_900 = HexColor('#0F172A')
    GREEN     = HexColor('#15803D')
    GREEN_BG  = HexColor('#DCFCE7')
    RED       = HexColor('#B91C1C')
    RED_BG    = HexColor('#FEE2E2')
    AMBER     = HexColor('#B45309')
    AMBER_BG  = HexColor('#FEF3C7')

    # ── Page geometry ────────────────────────────────────────────────────────
    PORT_W, PORT_H = A4                     # 595.28 × 841.89 pt
    LAND_W, LAND_H = rl_landscape(A4)
    MARGIN  = 1.5 * cm
    FOOT_H  = 0.85 * cm
    # Cover: navy band occupies top 40% of page
    BAND_H  = PORT_H * 0.40

    # ── Logo path ────────────────────────────────────────────────────────────
    logo_path = None
    logo_url  = _branding_url('logo')
    if logo_url:
        try:
            lp = os.path.join(
                django_settings.MEDIA_ROOT,
                logo_url.replace(django_settings.MEDIA_URL, '', 1),
            )
            if os.path.exists(lp):
                logo_path = lp
        except Exception:
            pass

    # ── Report meta ───────────────────────────────────────────────────────────
    report_ts   = datetime.now().strftime('%d %B %Y  %H:%M')
    report_date = datetime.now().strftime('%d %B %Y')
    invoice_ref = ''
    if meta:
        invoice_ref = meta.invoice_no or meta.supplier_invoice_no or meta.po_no or ''

    sch_dates = Schedule.objects.filter(
        account_id=account_id, channel=channel, month=month
    ).aggregate(d_min=Min('start_date'), d_max=Max('end_date'))
    date_min = sch_dates.get('d_min')
    date_max = sch_dates.get('d_max')
    if date_min and date_max:
        campaign_period = f"{date_min.strftime('%d %b %Y')} \u2013 {date_max.strftime('%d %b %Y')}"
    elif date_min:
        campaign_period = date_min.strftime('%d %b %Y')
    else:
        campaign_period = '\u2014'

    comm_total    = data.get('commercial_total', {})
    spon_total    = data.get('sponsorship_total', {})
    total_planned = (comm_total.get('planned', 0) or 0) + (spon_total.get('planned', 0) or 0)
    total_aired   = (comm_total.get('aired',   0) or 0) + (spon_total.get('aired',   0) or 0)
    total_missed  = (comm_total.get('missed',  0) or 0) + (spon_total.get('missed',  0) or 0)
    total_extra   = (comm_total.get('extra',   0) or 0) + (spon_total.get('extra',   0) or 0)

    # ── Canvas page callbacks ─────────────────────────────────────────────────

    def _footer(canvas, pw, page_num):
        """Draw hairline + footer text."""
        canvas.setStrokeColor(SLATE_200)
        canvas.setLineWidth(0.4)
        canvas.line(MARGIN, FOOT_H, pw - MARGIN, FOOT_H)
        canvas.setFont('Helvetica', 6)
        canvas.setFillColor(SLATE_400)
        left = f'CONFIDENTIAL  \u2022  {account.name}  \u2022  {channel}  \u2022  {month}'
        if invoice_ref:
            left += f'  \u2022  Ref: {invoice_ref}'
        canvas.drawString(MARGIN, FOOT_H * 0.35, left)
        canvas.drawRightString(pw - MARGIN, FOOT_H * 0.35, f'Page {page_num}')

    def _inner_page_header(canvas, pw, ph):
        """Thin navy bar + gold accent line at top of every non-cover page."""
        bar_h = 0.6 * cm
        # Navy bar
        canvas.setFillColor(NAVY)
        canvas.rect(0, ph - bar_h, pw, bar_h, stroke=0, fill=1)
        # Gold accent under bar
        canvas.setFillColor(GOLD)
        canvas.rect(0, ph - bar_h - 1.5, pw, 1.5, stroke=0, fill=1)
        # Context text in bar
        canvas.setFont('Helvetica-Bold', 7)
        canvas.setFillColor(SLATE_400)
        canvas.drawString(MARGIN, ph - bar_h + 0.15 * cm,
                          'ADVERTISEMENT MONITORING  \u2014  RECONCILIATION REPORT')
        canvas.setFillColor(WHITE)
        canvas.setFont('Helvetica', 6.5)
        canvas.drawRightString(pw - MARGIN, ph - bar_h + 0.15 * cm,
                               f'{account.name}  \u00b7  {channel}  \u00b7  {month}')
        # Logo on right inside bar
        if logo_path:
            try:
                _img = _IR(logo_path)
                canvas.drawImage(_img,
                                 pw - MARGIN - 3.0 * cm, ph - bar_h + 0.05 * cm,
                                 width=2.8 * cm, height=0.5 * cm,
                                 preserveAspectRatio=True, mask='auto')
            except Exception:
                pass

    def _cover_page(canvas, doc):
        canvas.saveState()
        W, H = PORT_W, PORT_H

        # ── 1. Left gold accent stripe (full height) ─────────────────────────
        canvas.setFillColor(GOLD)
        canvas.rect(0, 0, 3.5, H, stroke=0, fill=1)

        # ── 2. Full-bleed navy band (top 40%) ────────────────────────────────
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - BAND_H, W, BAND_H, stroke=0, fill=1)

        # ── 3. Subtle diagonal accent inside navy band (bottom-right) ────────
        canvas.setFillColor(NAVY_MID)
        canvas.setStrokeColor(NAVY_MID)
        path = canvas.beginPath()
        path.moveTo(W * 0.55, H - BAND_H)
        path.lineTo(W, H - BAND_H + BAND_H * 0.45)
        path.lineTo(W, H - BAND_H)
        path.close()
        canvas.drawPath(path, stroke=0, fill=1)

        # ── 4. Gold separator line beneath the navy band ─────────────────────
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - BAND_H - 3, W, 3, stroke=0, fill=1)

        # ── 5. Agency tag + report type label ────────────────────────────────
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(SLATE_400)
        canvas.drawString(1.2 * cm, H - 1.1 * cm,
                          'PHOENIX O & M (PVT) LTD  \u2022  ADVERTISEMENT MONITORING')

        # ── 6. Main title (two lines) ─────────────────────────────────────────
        canvas.setFillColor(WHITE)
        canvas.setFont('Helvetica-Bold', 34)
        canvas.drawString(1.2 * cm, H - BAND_H + 5.5 * cm, 'RECONCILIATION')
        canvas.setFont('Helvetica-Bold', 34)
        canvas.drawString(1.2 * cm, H - BAND_H + 3.8 * cm, 'SUMMARY REPORT')
        # Thin gold underline beneath title
        canvas.setStrokeColor(GOLD)
        canvas.setLineWidth(1.5)
        canvas.line(1.2 * cm, H - BAND_H + 3.5 * cm, 10 * cm, H - BAND_H + 3.5 * cm)

        # ── 7. Channel / Month pill in navy band ──────────────────────────────
        pill_x = 1.2 * cm
        pill_y = H - BAND_H + 2.3 * cm
        canvas.setFillColor(NAVY_MID)
        canvas.roundRect(pill_x, pill_y, 8 * cm, 0.65 * cm, 3, stroke=0, fill=1)
        canvas.setFont('Helvetica-Bold', 8)
        canvas.setFillColor(GOLD_LITE)
        canvas.drawString(pill_x + 0.3 * cm, pill_y + 0.18 * cm,
                          f'{channel.upper()}   \u2022   {month.upper()}')

        # ── 8. Logo in navy band (top-right) ─────────────────────────────────
        if logo_path:
            try:
                _img = _IR(logo_path)
                canvas.drawImage(_img,
                                 W - MARGIN - 4 * cm, H - MARGIN - 1.8 * cm,
                                 width=3.6 * cm, height=1.4 * cm,
                                 preserveAspectRatio=True, mask='auto')
            except Exception:
                pass

        # ── 9. KPI metrics strip (bottom of white area) ──────────────────────
        kpi_y   = FOOT_H + 0.8 * cm
        kpi_h   = 2.4 * cm
        kpi_w   = (W - 2 * MARGIN - 3 * 0.35 * cm) / 4
        kpi_gap = 0.35 * cm

        kpi_data = [
            ('TOTAL PLANNED', str(total_planned), NAVY,  WHITE,    NAVY_MID),
            ('TOTAL AIRED',   str(total_aired),   GREEN, GREEN_BG, GREEN),
            ('TOTAL MISSED',  str(total_missed),  RED,   RED_BG,   RED),
            ('EXTRA AIRED',   str(total_extra),   AMBER, AMBER_BG, AMBER),
        ]
        for i, (label, val, accent, bg, txt_color) in enumerate(kpi_data):
            x = MARGIN + i * (kpi_w + kpi_gap)
            # Box background
            canvas.setFillColor(bg)
            canvas.setStrokeColor(SLATE_200)
            canvas.setLineWidth(0.4)
            canvas.roundRect(x, kpi_y, kpi_w, kpi_h, 4, stroke=1, fill=1)
            # Top accent bar
            canvas.setFillColor(accent)
            canvas.roundRect(x, kpi_y + kpi_h - 0.22 * cm, kpi_w, 0.22 * cm, 2, stroke=0, fill=1)
            # Large value
            canvas.setFont('Helvetica-Bold', 26)
            canvas.setFillColor(txt_color)
            canvas.drawCentredString(x + kpi_w / 2, kpi_y + 0.9 * cm, val)
            # Label
            canvas.setFont('Helvetica-Bold', 6.5)
            canvas.setFillColor(SLATE_600)
            canvas.drawCentredString(x + kpi_w / 2, kpi_y + 0.3 * cm, label)

        _footer(canvas, W, doc.page)
        canvas.restoreState()

    def _portrait_page(canvas, doc):
        canvas.saveState()
        _inner_page_header(canvas, PORT_W, PORT_H)
        _footer(canvas, PORT_W, doc.page)
        canvas.restoreState()

    def _landscape_page(canvas, doc):
        canvas.saveState()
        _inner_page_header(canvas, LAND_W, LAND_H)
        _footer(canvas, LAND_W, doc.page)
        canvas.restoreState()

    # ── Frames ───────────────────────────────────────────────────────────────
    # Cover frame: the white area below the band, above KPIs
    _cover_content_top    = PORT_H - BAND_H - 4 * mm      # just below gold line
    _cover_content_bottom = FOOT_H + 3.4 * cm             # above KPI strip
    _cover_frame = Frame(
        MARGIN, _cover_content_bottom,
        PORT_W - 2 * MARGIN,
        _cover_content_top - _cover_content_bottom,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        id='cover',
    )
    _HDR_H = 0.6 * cm + 1.5 + 3 * mm          # bar + gold line + padding
    def _mk_frame(w, h, fid):
        return Frame(
            MARGIN, FOOT_H + 2 * mm,
            w - 2 * MARGIN,
            h - _HDR_H - FOOT_H - 4 * mm,
            leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
            id=fid,
        )

    # ── Document ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    doc = BaseDocTemplate(buf, pagesize=A4,
                          leftMargin=0, rightMargin=0,
                          topMargin=0, bottomMargin=0)
    doc.addPageTemplates([
        PageTemplate(id='cover',     frames=[_cover_frame],
                     onPage=_cover_page, pagesize=A4),
        PageTemplate(id='portrait',  frames=[_mk_frame(PORT_W, PORT_H, 'portrait')],
                     onPage=_portrait_page, pagesize=A4),
        PageTemplate(id='landscape', frames=[_mk_frame(LAND_W, LAND_H, 'landscape')],
                     onPage=_landscape_page, pagesize=rl_landscape(A4)),
    ])

    # ── Paragraph styles ─────────────────────────────────────────────────────
    def _ps(name, **kw):
        kw.setdefault('fontName', 'Helvetica')
        return ParagraphStyle(name, **kw)

    S = dict(
        section   = _ps('s_sect',   fontSize=9.5, textColor=NAVY, fontName='Helvetica-Bold',
                        spaceBefore=6, spaceAfter=4),
        sub       = _ps('s_sub',    fontSize=8,   textColor=SLATE_600, spaceAfter=2),
        cell      = _ps('s_cell',   fontSize=7.5, textColor=SLATE_900),
        cell_r    = _ps('s_cellr',  fontSize=7.5, textColor=SLATE_900, alignment=TA_RIGHT),
        cell_sm   = _ps('s_csm',    fontSize=5.5, textColor=SLATE_900),
        hdr_wht   = _ps('s_hwht',   fontSize=7.5, textColor=WHITE, fontName='Helvetica-Bold',
                        alignment=TA_CENTER),
        gt        = _ps('s_gt',     fontSize=7.5, textColor=SLATE_900, fontName='Helvetica-Bold'),
        gt_r      = _ps('s_gtr',    fontSize=7.5, textColor=SLATE_900, fontName='Helvetica-Bold',
                        alignment=TA_RIGHT),
        prog_lbl  = _ps('s_prog',   fontSize=8,   textColor=NAVY_MID, fontName='Helvetica-Bold',
                        spaceBefore=6, spaceAfter=2),
        # Cover-specific
        cov_lbl   = _ps('c_lbl',    fontSize=8.5, textColor=SLATE_600, fontName='Helvetica-Bold'),
        cov_val   = _ps('c_val',    fontSize=9.5, textColor=SLATE_900, fontName='Helvetica-Bold'),
        cov_conf  = _ps('c_conf',   fontSize=7.5, textColor=SLATE_400, fontName='Helvetica-Bold',
                        alignment=TA_CENTER),
        cov_sig_l = _ps('c_sigl',   fontSize=8,   textColor=SLATE_600, fontName='Helvetica-Bold'),
        cov_sig_v = _ps('c_sigv',   fontSize=9,   textColor=SLATE_900, fontName='Helvetica-Bold'),
    )

    story = []

    # ═══════════════════════════════════════════════════════════════════════════
    # PAGE 1 - COVER  (drawn by canvas; Platypus renders info card in white area)
    # ═══════════════════════════════════════════════════════════════════════════

    content_w = PORT_W - 2 * MARGIN

    # Info card ─ clean borderless rows with subtle dividers
    cover_meta = [
        ('Client / Account',   account.name),
        ('Agency',             'Phoenix O & M (Pvt) Ltd'),
        ('Channel',            channel),
        ('Campaign Period',    campaign_period),
        ('Estimate / RO No.',  estimate_no or '\u2014'),
        ('Report Date',        report_date),
    ]
    if meta:
        if meta.invoice_no:
            cover_meta.append(('Invoice No.',         meta.invoice_no))
        if meta.po_no:
            cover_meta.append(('Release Order Ref.',  meta.po_no))
        if meta.supplier_invoice_no:
            cover_meta.append(('Supplier Invoice No.', meta.supplier_invoice_no))

    info_tbl_data = [
        [Paragraph(k, S['cov_lbl']), Paragraph(str(v), S['cov_val'])]
        for k, v in cover_meta
    ]
    info_tbl = Table(info_tbl_data, colWidths=[5.2 * cm, content_w - 5.2 * cm], hAlign='LEFT')
    _info_row_count = len(info_tbl_data)
    _info_style = [
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (0, -1),  0),
        ('LEFTPADDING',   (1, 0), (1, -1),  10),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND',    (0, 0), (0, -1),  OFF_WHITE),
    ]
    # Subtle dividers between rows (skip last)
    for _r in range(_info_row_count - 1):
        _info_style.append(('LINEBELOW', (0, _r), (-1, _r), 0.3, SLATE_200))
    # Gold left border on label column
    _info_style.append(('LINEBEFORE', (0, 0), (0, -1), 2, GOLD))
    info_tbl.setStyle(TableStyle(_info_style))

    story.append(Spacer(1, 0.4 * cm))
    story.append(info_tbl)

    # Signature block
    if meta and any([meta.prepared_by, meta.checked_by, meta.authorised_by]):
        story.append(Spacer(1, 0.5 * cm))
        sig_data = [[
            Table([
                [Paragraph(role, S['cov_sig_l'])],
                [Paragraph(name or '\u2014', S['cov_sig_v'])],
            ], colWidths=[4.8 * cm],
               style=TableStyle([
                   ('TOPPADDING',    (0, 0), (-1, -1), 3),
                   ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                   ('LEFTPADDING',   (0, 0), (-1, -1), 0),
                   ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
               ]))
            for role, name in [
                ('Prepared by',    meta.prepared_by or ''),
                ('Checked by',     meta.checked_by or ''),
                ('Authorised by',  meta.authorised_by or ''),
            ]
        ]]
        sig_tbl = Table(sig_data, colWidths=[5.2 * cm, 5.2 * cm, 5.2 * cm], hAlign='LEFT')
        sig_tbl.setStyle(TableStyle([
            ('LINEABOVE',    (0, 0), (-1, 0), 0.4, SLATE_200),
            ('TOPPADDING',   (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 6),
            ('LEFTPADDING',  (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(sig_tbl)

    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(
        'STRICTLY CONFIDENTIAL \u2014 This document is intended solely for authorised recipients.',
        S['cov_conf'],
    ))

    # ═══════════════════════════════════════════════════════════════════════════
    # PAGE 2 - RECONCILIATION SUMMARY TABLES (Portrait)
    # ═══════════════════════════════════════════════════════════════════════════
    story.append(NextPageTemplate('portrait'))
    story.append(PageBreak())

    # Section heading helper - left gold bar via table padding trick
    def _section_heading(title, subtitle=''):
        rows = [[Paragraph(title, S['section'])]]
        if subtitle:
            rows.append([Paragraph(subtitle, S['sub'])])
        t = Table(rows, colWidths=[content_w])
        t.setStyle(TableStyle([
            ('LINEBEFORE',   (0, 0), (0, -1), 3, GOLD),
            ('LEFTPADDING',  (0, 0), (-1, -1), 8),
            ('TOPPADDING',   (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BACKGROUND',   (0, 0), (-1, -1), OFF_WHITE),
        ]))
        return t

    story.append(_section_heading(
        'RECONCILIATION SUMMARY',
        f'{account.name}  \u00b7  {channel}  \u00b7  {month}'
        + (f'  \u00b7  Estimate No: {estimate_no}' if estimate_no else ''),
    ))
    story.append(Spacer(1, 0.35 * cm))

    # Summary table builder
    SUM_HEADERS = ['Product / Brand', 'Dur', 'Planned', 'Aired', 'Missed', 'Extra', '3rd Party']
    SUM_WIDTHS  = [5.5 * cm, 1.5 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm]

    def _num_cell(val, zero_style=None):
        """Return a Paragraph for a numeric table cell, styled by value."""
        s = zero_style or S['cell_r']
        return Paragraph(str(val), s)

    def _missed_style(val):
        if val and int(val) > 0:
            return _ps(f'mc_{val}', fontSize=7.5, textColor=RED,
                       fontName='Helvetica-Bold', alignment=TA_RIGHT)
        return S['cell_r']

    def _extra_style(val):
        if val and int(val) > 0:
            return _ps(f'ec_{val}', fontSize=7.5, textColor=GREEN,
                       fontName='Helvetica-Bold', alignment=TA_RIGHT)
        return S['cell_r']

    def _build_sum_table(rows, total_row, has_total=True):
        hdr = [Paragraph(h, S['hdr_wht']) for h in SUM_HEADERS]
        tdata = [hdr]
        for r in rows:
            missed = r.get('missed', 0) or 0
            extra  = r.get('extra',  0) or 0
            tdata.append([
                Paragraph(str(r.get('product', '')), S['cell']),
                Paragraph(str(r.get('dur') or '\u2014'), S['cell_r']),
                _num_cell(r.get('planned',     0)),
                _num_cell(r.get('aired',       0)),
                Paragraph(str(missed), _missed_style(missed)),
                Paragraph(str(extra),  _extra_style(extra)),
                _num_cell(r.get('third_party', 0)),
            ])
        if has_total:
            t = total_row
            gt_missed = t.get('missed', 0) or 0
            gt_extra  = t.get('extra',  0) or 0
            tdata.append([
                Paragraph('Grand Total', S['gt']),
                Paragraph('', S['cell']),
                Paragraph(str(t.get('planned',     0)), S['gt_r']),
                Paragraph(str(t.get('aired',       0)), S['gt_r']),
                Paragraph(str(gt_missed), _ps('gm', fontSize=7.5, fontName='Helvetica-Bold',
                                             textColor=RED if gt_missed > 0 else SLATE_900,
                                             alignment=TA_RIGHT)),
                Paragraph(str(gt_extra),  _ps('ge', fontSize=7.5, fontName='Helvetica-Bold',
                                             textColor=GREEN if gt_extra > 0 else SLATE_900,
                                             alignment=TA_RIGHT)),
                Paragraph(str(t.get('third_party', 0)), S['gt_r']),
            ])
        n   = len(tdata)
        tbl = Table(tdata, colWidths=SUM_WIDTHS, repeatRows=1)
        style = [
            # Header row
            ('BACKGROUND',    (0, 0), (-1, 0),  NAVY),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  WHITE),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('ALIGN',         (1, 0), (-1, 0),  'CENTER'),
            ('LINEBELOW',     (0, 0), (-1, 0),  2, GOLD),
            # Data rows
            ('FONTSIZE',      (0, 0), (-1, -1), 7.5),
            ('ROWBACKGROUNDS',(0, 1), (-1, n - 2 if has_total else -1),
                              [SLATE_100, WHITE]),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
            ('LINEBELOW',     (0, 0), (-1, -1), 0.25, SLATE_200),
            # Totals row
        ]
        if has_total:
            style += [
                ('BACKGROUND',    (0, -1), (-1, -1), HexColor('#EEF2FF')),
                ('LINEABOVE',     (0, -1), (-1, -1), 1.2, NAVY),
                ('LINEBELOW',     (0, -1), (-1, -1), 1.2, NAVY),
            ]
        tbl.setStyle(TableStyle(style))
        return tbl

    if data.get('commercial'):
        story.append(Paragraph('Commercial Benefits', S['prog_lbl']))
        story.append(Spacer(1, 0.15 * cm))
        story.append(_build_sum_table(data['commercial'], data['commercial_total']))
        story.append(Spacer(1, 0.5 * cm))

    if data.get('sponsorship'):
        story.append(Paragraph('Sponsorship Benefits', S['prog_lbl']))
        story.append(Spacer(1, 0.1 * cm))
        for section in data['sponsorship']:
            prog_label = Table(
                [[Paragraph(section['programme'].upper(), S['sub'])]],
                colWidths=[content_w],
            )
            prog_label.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, -1), HexColor('#F0F4FF')),
                ('LEFTPADDING',   (0, 0), (-1, -1), 8),
                ('TOPPADDING',    (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LINEBEFORE',    (0, 0), (0, -1),  2, NAVY_MID),
            ]))
            story.append(prog_label)
            story.append(Spacer(1, 0.1 * cm))
            story.append(_build_sum_table(section['rows'], {}, has_total=False))
            story.append(Spacer(1, 0.25 * cm))
        story.append(Paragraph('SPONSORSHIP GRAND TOTAL', S['prog_lbl']))
        story.append(Spacer(1, 0.1 * cm))
        story.append(_build_sum_table([], data['sponsorship_total']))

    if meta and meta.notes:
        story.append(Spacer(1, 0.5 * cm))
        story.append(_section_heading('Notes'))
        story.append(Spacer(1, 0.2 * cm))
        for line in meta.notes.splitlines():
            if line.strip():
                story.append(Paragraph(line, S['sub']))

    # ═══════════════════════════════════════════════════════════════════════════
    # PAGE 3+ - MATCHED LMRB REPORT (Landscape)
    # ═══════════════════════════════════════════════════════════════════════════
    story.append(NextPageTemplate('landscape'))
    story.append(PageBreak())

    land_content_w = LAND_W - 2 * MARGIN
    story.append(_section_heading(
        'MATCHED LMRB REPORT',
        f'{account.name}  \u00b7  {channel}  \u00b7  {month}'
        '  \u00b7  Commercial (TC-confirmed) + Sponsorship (assigned)',
    ))
    story.append(Spacer(1, 0.35 * cm))

    # Fetch rows
    commercial_lmrb_qs = LMRBRow.objects.filter(account_id=account_id, channel__iexact=channel)
    if date_min and date_max:
        commercial_lmrb_qs = commercial_lmrb_qs.filter(date__range=(date_min, date_max))
    commercial_lmrb_qs = (
        commercial_lmrb_qs
        .filter(tc_confirmations__isnull=False)
        .distinct()
        .order_by('date', 'program', 'advt_time')
    )
    spon_lmrb_ids = set(
        SponsorshipLmrbAssignment.objects.filter(
            account_id=account_id,
            schedule_row__channel=channel,
            schedule_row__month=month,
        ).values_list('lmrb_row_id', flat=True)
    )
    commercial_ids       = set(commercial_lmrb_qs.values_list('id', flat=True))
    sponsorship_only_ids = spon_lmrb_ids - commercial_ids
    sponsorship_only_qs  = (
        LMRBRow.objects.filter(id__in=sponsorship_only_ids)
        .order_by('date', 'program', 'advt_time')
    )
    all_lmrb_rows = list(commercial_lmrb_qs) + list(sponsorship_only_qs)

    LMRB_HEADERS = [
        'Prod Group', 'Advertiser', 'Product', 'Advt Theme', 'Ads',
        'Channel', 'Program', 'Date', 'Day', 'Prog Time', 'Advt Time',
        'Ad Pos', 'Tot Ads', 'Brk No', 'Pos/Brk', 'Ads/Brk',
        'Lng', 'Dur', 'Cost',
    ]
    _lw = land_content_w
    LMRB_WIDTHS = [
        _lw * 0.08, _lw * 0.10, _lw * 0.08, _lw * 0.12, _lw * 0.03,
        _lw * 0.08, _lw * 0.10, _lw * 0.07, _lw * 0.04, _lw * 0.05, _lw * 0.05,
        _lw * 0.04, _lw * 0.04, _lw * 0.04, _lw * 0.04, _lw * 0.04,
        _lw * 0.03, _lw * 0.03, _lw * 0.05,
    ]

    if all_lmrb_rows:
        lmrb_data = [[Paragraph(h, S['hdr_wht']) for h in LMRB_HEADERS]]
        for lr in all_lmrb_rows:
            cost_str = f'{lr.cost:,.2f}' if lr.cost is not None else ''
            date_str = lr.date.strftime('%d/%m/%y') if lr.date else ''
            lmrb_data.append([
                Paragraph(str(lr.product_group or ''), S['cell_sm']),
                Paragraph(str(lr.advertiser or ''),    S['cell_sm']),
                Paragraph(str(lr.product or ''),       S['cell_sm']),
                Paragraph(str(lr.advt_theme or ''),    S['cell_sm']),
                Paragraph(str(lr.ads or ''),           S['cell_sm']),
                Paragraph(str(lr.channel or ''),       S['cell_sm']),
                Paragraph(str(lr.program or ''),       S['cell_sm']),
                Paragraph(date_str,                    S['cell_sm']),
                Paragraph(str(lr.day or ''),           S['cell_sm']),
                Paragraph(str(lr.prog_time or ''),     S['cell_sm']),
                Paragraph(str(lr.advt_time or ''),     S['cell_sm']),
                Paragraph(str(lr.ad_pos     if lr.ad_pos     is not None else ''), S['cell_sm']),
                Paragraph(str(lr.tot_ads    if lr.tot_ads    is not None else ''), S['cell_sm']),
                Paragraph(str(lr.brk_no     if lr.brk_no     is not None else ''), S['cell_sm']),
                Paragraph(str(lr.pos_in_brk if lr.pos_in_brk is not None else ''), S['cell_sm']),
                Paragraph(str(lr.ads_in_brk if lr.ads_in_brk is not None else ''), S['cell_sm']),
                Paragraph(str(lr.lng or ''), S['cell_sm']),
                Paragraph(str(lr.duration or ''), S['cell_sm']),
                Paragraph(cost_str,          S['cell_sm']),
            ])
        lmrb_tbl = Table(lmrb_data, colWidths=LMRB_WIDTHS, repeatRows=1)
        lmrb_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0),  NAVY),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  WHITE),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('LINEBELOW',     (0, 0), (-1, 0),  2, GOLD),
            ('FONTSIZE',      (0, 0), (-1, -1), 5.5),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [SLATE_100, WHITE]),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING',   (0, 0), (-1, -1), 2),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
            ('LINEBELOW',     (0, 1), (-1, -1), 0.2, SLATE_200),
        ]))
        story.append(lmrb_tbl)
    else:
        story.append(Paragraph('No matched LMRB rows found for this scope.', S['sub']))

    # ── Build & respond ───────────────────────────────────────────────────────
    doc.build(story)
    buf.seek(0)
    fname = f'Reconciliation_{account.name}_{channel}_{month}.pdf'.replace(' ', '_')
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def _write_matched_lmrb_sheet(ws, account_id, channel, month):
    """
    Shared helper: write all matched LMRB rows (commercial + sponsorship) into
    an existing openpyxl worksheet.  Returns the row count written.

    Commercial matched  = LMRBRow.is_matched=True (Schedule↔LMRB engine)
    Sponsorship matched = LMRBRow rows linked via SponsorshipLmrbAssignment
    Both sets are deduplicated so a row appearing in both groups is written once.
    """
    import openpyxl  # noqa - imported by callers too; ensure available
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.db.models import Min, Max
    from core.models import SponsorshipLmrbAssignment

    sch_dates = Schedule.objects.filter(
        account_id=account_id, channel=channel, month=month,
    ).aggregate(d_min=Min('start_date'), d_max=Max('end_date'))
    date_min = sch_dates.get('d_min')
    date_max = sch_dates.get('d_max')

    # Commercial matched (auto-matched by Schedule↔LMRB engine)
    comm_qs = LMRBRow.objects.filter(
        account_id=account_id, channel__iexact=channel, is_matched=True,
    )
    if date_min and date_max:
        comm_qs = comm_qs.filter(date__range=(date_min, date_max))

    # Sponsorship matched (via SponsorshipLmrbAssignment)
    spon_ids = set(
        SponsorshipLmrbAssignment.objects.filter(
            account_id=account_id,
            schedule_row__channel=channel,
            schedule_row__month=month,
        ).values_list('lmrb_row_id', flat=True)
    )

    # TC-confirmed sponsorship LMRB rows: TCRow.matched_lmrb where the TCRow is
    # linked to a SPONSORSHIP ScheduleRow.  These are confirmed by TC reconciliation
    # but may not have a SponsorshipLmrbAssignment record yet; they must still appear
    # in the final LMRB export so the complete aired picture is captured.
    tc_spon_lmrb_ids = set(
        TCRow.objects.filter(
            account_id=account_id, channel=channel,
            tc_report__month=month,
            is_schedule_matched=True,
            is_lmrb_confirmed=True,
            matched_schedule__ad_type='SPONSORSHIP',
            matched_lmrb__isnull=False,
        ).values_list('matched_lmrb_id', flat=True)
    )

    # Combine, deduplicate by id, sort by date+time
    all_ids = set(comm_qs.values_list('id', flat=True)) | spon_ids | tc_spon_lmrb_ids
    combined = list(
        LMRBRow.objects.filter(id__in=all_ids).order_by('date', 'advt_time')
    )

    HDR_FILL = PatternFill('solid', fgColor='0F2340')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    norm     = Font(size=10)
    centre   = Alignment(horizontal='center', vertical='center')

    headers = [
        'Product_Group', 'Advertiser', 'Product', 'Advt_Theme', 'Ads',
        'Channel', 'Program',
        'Dd', 'Mn', 'Yr', 'Day', 'Prog_time', 'Advt_time',
        'AdPos', 'TotAds', 'BrkNo', 'PosinBrk', 'AdsinBrk',
        'Lng', 'Dur', 'Cost',
    ]
    for col_i, h in enumerate(headers, start=1):
        c = ws.cell(1, col_i, h)
        c.font = hdr_font; c.fill = HDR_FILL; c.alignment = centre

    for row_i, lr in enumerate(combined, start=2):
        dd = lr.date.day   if lr.date else ''
        mn = lr.date.month if lr.date else ''
        yr = lr.date.year  if lr.date else ''
        vals = [
            lr.product_group or '',
            lr.advertiser or '',
            lr.product or '',
            lr.advt_theme or '',
            lr.ads or '',
            lr.channel or '',
            lr.program or '',
            dd, mn, yr,
            lr.day or '',
            lr.prog_time or '',
            lr.advt_time or '',
            lr.ad_pos     if lr.ad_pos     is not None else '',
            lr.tot_ads    if lr.tot_ads    is not None else '',
            lr.brk_no     if lr.brk_no     is not None else '',
            lr.pos_in_brk if lr.pos_in_brk is not None else '',
            lr.ads_in_brk if lr.ads_in_brk is not None else '',
            lr.lng or '',
            lr.duration   if lr.duration   is not None else '',
            float(lr.cost) if lr.cost is not None else '',
        ]
        for col_i, v in enumerate(vals, start=1):
            ws.cell(row_i, col_i, v).font = norm

    return len(combined)


@login_required
def matched_lmrb_excel(request):
    """
    Export all matched LMRB rows (commercial auto-matched + sponsorship manually
    matched) for the selected scope as a standalone Excel file.
    """
    import openpyxl

    account_id = request.GET.get('account_id', '')
    channel    = request.GET.get('channel', '')
    month      = request.GET.get('month', '')

    if not (account_id and channel and month):
        return HttpResponse('Missing parameters: account_id, channel, month', status=400)
    if not _account_access(request.user, account_id):
        return HttpResponse('Access denied', status=403)

    account = get_object_or_404(Account, id=account_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Matched LMRB'

    _write_matched_lmrb_sheet(ws, account_id, channel, month)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'Matched_LMRB_{account.name}_{channel}_{month}.xlsx'.replace(' ', '_')
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
@role_required(['super_admin', 'admin'])
def admin_export(request):
    """
    Admin-only full database export.

    GET:  Show the filter form.
    POST: Generate and download the Excel file.

    Exports to Excel with separate sheets:
      - Matched LMRB
      - Unmatched LMRB
      - Schedule Rows
      - TC Rows
      - Full Ad Report (MatchResult)
      - Brand Mappings
      - Manual Matches
      - Sponsorship Assignments

    Filters: account, channel, month, schedule_number, brand, date_from, date_to.
    Each sheet has a metadata row at the top with row count.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    user       = request.user
    account_qs = _account_qs(user)

    if request.method == 'GET':
        account_id = request.GET.get('account_id', '')
        sel_channel = request.GET.get('channel', '')
        channels, months, schedule_numbers = [], [], []
        if account_id and _account_access(user, account_id):
            channels = sorted(set(
                list(LMRBRow.objects.filter(account_id=account_id).values_list('channel', flat=True))
                + list(Schedule.objects.filter(account_id=account_id).values_list('channel', flat=True))
            ))
            if sel_channel:
                months = sorted(set(
                    Schedule.objects.filter(account_id=account_id, channel=sel_channel)
                    .values_list('month', flat=True)
                ))
        return render(request, 'admin_export.html', {
            'accounts': account_qs,
            'account_id': account_id,
            'channels': channels,
            'sel_channel': sel_channel,
            'months': months,
        })

    # ── POST - build export ──────────────────────────────────────────────────
    account_id      = request.POST.get('account_id', '').strip()
    channel         = request.POST.get('channel', '').strip()
    month           = request.POST.get('month', '').strip()
    schedule_number = request.POST.get('schedule_number', '').strip()
    brand           = request.POST.get('brand', '').strip()
    date_from       = request.POST.get('date_from', '').strip()
    date_to         = request.POST.get('date_to', '').strip()
    scope           = request.POST.get('scope', 'both')

    if not account_id or not _account_access(user, account_id):
        messages.error(request, 'Select a valid account.')
        return redirect('/dashboard/admin-export/')

    account = get_object_or_404(Account, id=account_id)

    # ── Shared date helpers ──────────────────────────────────────────────────
    from datetime import date as _date
    _from_d = None
    _to_d   = None
    try:
        if date_from:
            _from_d = _date.fromisoformat(date_from)
    except ValueError:
        pass
    try:
        if date_to:
            _to_d = _date.fromisoformat(date_to)
    except ValueError:
        pass

    # ── Styling helpers ──────────────────────────────────────────────────────
    wb        = openpyxl.Workbook()
    HDR_FILL  = PatternFill('solid', fgColor='0F2340')
    META_FILL = PatternFill('solid', fgColor='FEF9C3')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    meta_font = Font(bold=True, size=10)
    norm_font = Font(size=10)
    centre_al = Alignment(horizontal='center', vertical='center')

    filter_label = f'Account: {account.name}'
    if channel:         filter_label += f' | Channel: {channel}'
    if month:           filter_label += f' | Month: {month}'
    if schedule_number: filter_label += f' | Schedule No: {schedule_number}'

    def _meta_row(ws, ncols, label, n_rows):
        ws.merge_cells(f'A1:{chr(64 + min(ncols, 26))}1')
        c = ws.cell(1, 1, f'{label} | {filter_label} | Total rows: {n_rows}')
        c.font = meta_font; c.fill = META_FILL

    def _hdr_row(ws, headers):
        for col_i, h in enumerate(headers, start=1):
            c = ws.cell(2, col_i, h)
            c.font = hdr_font; c.fill = HDR_FILL; c.alignment = centre_al

    def _norm_rows(ws, n_rows, n_cols):
        for row_i in range(3, n_rows + 3):
            for col_i in range(1, n_cols + 1):
                ws.cell(row_i, col_i).font = norm_font

    # ── 1. LMRB sheets ──────────────────────────────────────────────────────
    lmrb_qs = LMRBRow.objects.filter(account_id=account_id)
    if channel:   lmrb_qs = lmrb_qs.filter(channel__iexact=channel)
    if month:     lmrb_qs = lmrb_qs.filter(date__range=_month_date_range(month))
    if brand:     lmrb_qs = lmrb_qs.filter(advt_theme__icontains=brand)
    if _from_d:   lmrb_qs = lmrb_qs.filter(date__gte=_from_d)
    if _to_d:     lmrb_qs = lmrb_qs.filter(date__lte=_to_d)

    matched_qs   = lmrb_qs.filter(is_matched=True).order_by('date', 'advt_time')
    unmatched_qs = lmrb_qs.filter(is_matched=False).order_by('date', 'advt_time')

    LMRB_HEADERS = [
        'Date', 'Channel', 'Advt_Theme', 'Advt_Time', 'Duration', 'Source',
        'Product_Group', 'Advertiser', 'Product', 'Ads', 'Program', 'Prog_Time',
        'Ad_Pos', 'Tot_Ads', 'Brk_No', 'Pos_In_Brk', 'Ads_In_Brk',
        'Lng', 'Cost', 'Day', 'Is_Matched', 'Is_Sponsorship_Matched', 'Is_Manual_Matched',
    ]

    def _write_lmrb_sheet(ws_sheet, qs, label):
        rows = list(qs)
        _meta_row(ws_sheet, len(LMRB_HEADERS), label, len(rows))
        _hdr_row(ws_sheet, LMRB_HEADERS)
        for row_i, lr in enumerate(rows, start=3):
            ws_sheet.cell(row_i,  1, str(lr.date) if lr.date else '')
            ws_sheet.cell(row_i,  2, lr.channel or '')
            ws_sheet.cell(row_i,  3, lr.advt_theme or '')
            ws_sheet.cell(row_i,  4, lr.advt_time or '')
            ws_sheet.cell(row_i,  5, lr.duration)
            ws_sheet.cell(row_i,  6, lr.source or '')
            ws_sheet.cell(row_i,  7, lr.product_group or '')
            ws_sheet.cell(row_i,  8, lr.advertiser or '')
            ws_sheet.cell(row_i,  9, lr.product or '')
            ws_sheet.cell(row_i, 10, lr.ads or '')
            ws_sheet.cell(row_i, 11, lr.program or '')
            ws_sheet.cell(row_i, 12, lr.prog_time or '')
            ws_sheet.cell(row_i, 13, lr.ad_pos)
            ws_sheet.cell(row_i, 14, lr.tot_ads)
            ws_sheet.cell(row_i, 15, lr.brk_no)
            ws_sheet.cell(row_i, 16, lr.pos_in_brk)
            ws_sheet.cell(row_i, 17, lr.ads_in_brk)
            ws_sheet.cell(row_i, 18, lr.lng or '')
            ws_sheet.cell(row_i, 19, float(lr.cost) if lr.cost is not None else '')
            ws_sheet.cell(row_i, 20, lr.day or '')
            ws_sheet.cell(row_i, 21, 'Yes' if lr.is_matched else 'No')
            ws_sheet.cell(row_i, 22, 'Yes' if lr.is_sponsorship_matched else 'No')
            ws_sheet.cell(row_i, 23, 'Yes' if lr.is_manual_matched else 'No')
        _norm_rows(ws_sheet, len(rows), len(LMRB_HEADERS))

    if scope == 'unmatched':
        ws_un = wb.active; ws_un.title = 'Unmatched LMRB'
        _write_lmrb_sheet(ws_un, unmatched_qs, 'Unmatched LMRB')
    else:
        ws_matched = wb.active; ws_matched.title = 'Matched LMRB'
        _write_lmrb_sheet(ws_matched, matched_qs, 'Matched LMRB')
        if scope == 'both':
            ws_un = wb.create_sheet('Unmatched LMRB')
            _write_lmrb_sheet(ws_un, unmatched_qs, 'Unmatched LMRB')

    # ── 2. Schedule Rows sheet ───────────────────────────────────────────────
    sr_qs = ScheduleRow.objects.select_related('schedule', 'matched_lmrb').filter(account_id=account_id)
    if channel:         sr_qs = sr_qs.filter(channel__iexact=channel)
    if month:           sr_qs = sr_qs.filter(month=month)
    if schedule_number: sr_qs = sr_qs.filter(schedule__schedule_number=schedule_number)
    if brand:           sr_qs = sr_qs.filter(brand__icontains=brand)
    if _from_d:         sr_qs = sr_qs.filter(date__gte=_from_d)
    if _to_d:           sr_qs = sr_qs.filter(date__lte=_to_d)
    sr_qs = sr_qs.order_by('month', 'date', 'start_time')
    sr_rows = list(sr_qs)

    ws_sr = wb.create_sheet('Schedule Rows')
    SR_HEADERS = [
        'Schedule No', 'Version', 'Channel', 'Month', 'Brand', 'Programme',
        'Date', 'Start Time', 'End Time', 'Duration (s)', 'Ad Type',
        'Is Matched', 'Is Manual Matched', 'Matched LMRB Theme', 'Matched At',
    ]
    _meta_row(ws_sr, len(SR_HEADERS), 'Schedule Rows', len(sr_rows))
    _hdr_row(ws_sr, SR_HEADERS)
    for row_i, sr in enumerate(sr_rows, start=3):
        ws_sr.cell(row_i,  1, sr.schedule.schedule_number if sr.schedule else '')
        ws_sr.cell(row_i,  2, sr.schedule.version if sr.schedule else '')
        ws_sr.cell(row_i,  3, sr.channel or '')
        ws_sr.cell(row_i,  4, sr.month or '')
        ws_sr.cell(row_i,  5, sr.brand or '')
        ws_sr.cell(row_i,  6, sr.programme or '')
        ws_sr.cell(row_i,  7, str(sr.date) if sr.date else '')
        ws_sr.cell(row_i,  8, sr.start_time or '')
        ws_sr.cell(row_i,  9, sr.end_time or '')
        ws_sr.cell(row_i, 10, sr.duration)
        ws_sr.cell(row_i, 11, sr.ad_type or '')
        ws_sr.cell(row_i, 12, 'Yes' if sr.is_matched else 'No')
        ws_sr.cell(row_i, 13, 'Yes' if sr.is_manual_matched else 'No')
        ws_sr.cell(row_i, 14, sr.matched_lmrb.advt_theme if sr.matched_lmrb_id else '')
        ws_sr.cell(row_i, 15, str(sr.matched_at) if sr.matched_at else '')
    _norm_rows(ws_sr, len(sr_rows), len(SR_HEADERS))

    # ── 3. TC Rows sheet ─────────────────────────────────────────────────────
    tc_qs = TCRow.objects.select_related('tc_report', 'matched_schedule').filter(account_id=account_id)
    if channel: tc_qs = tc_qs.filter(channel__iexact=channel)
    if month:   tc_qs = tc_qs.filter(tc_report__month=month)
    if _from_d: tc_qs = tc_qs.filter(date__gte=_from_d)
    if _to_d:   tc_qs = tc_qs.filter(date__lte=_to_d)
    tc_qs = tc_qs.order_by('date', 'aired_time')
    tc_rows = list(tc_qs)

    ws_tc = wb.create_sheet('TC Rows')
    TC_HEADERS = [
        'Channel', 'Month', 'Date', 'Programme', 'TC Theme', 'Duration (s)',
        'Aired Time', 'Is Schedule Matched', 'Matched Schedule Brand',
        'Matched Schedule Date', 'Is LMRB Confirmed', 'Is Extra',
    ]
    _meta_row(ws_tc, len(TC_HEADERS), 'TC Rows', len(tc_rows))
    _hdr_row(ws_tc, TC_HEADERS)
    for row_i, tc in enumerate(tc_rows, start=3):
        ws_tc.cell(row_i,  1, tc.channel or '')
        ws_tc.cell(row_i,  2, tc.tc_report.month if tc.tc_report else '')
        ws_tc.cell(row_i,  3, str(tc.date) if tc.date else '')
        ws_tc.cell(row_i,  4, tc.programme or '')
        ws_tc.cell(row_i,  5, tc.tc_theme or '')
        ws_tc.cell(row_i,  6, tc.duration)
        ws_tc.cell(row_i,  7, tc.aired_time or '')
        ws_tc.cell(row_i,  8, 'Yes' if tc.is_schedule_matched else 'No')
        ws_tc.cell(row_i,  9, tc.matched_schedule.brand if tc.matched_schedule_id else '')
        ws_tc.cell(row_i, 10, str(tc.matched_schedule.date) if tc.matched_schedule_id and tc.matched_schedule.date else '')
        ws_tc.cell(row_i, 11, 'Yes' if tc.is_lmrb_confirmed else 'No')
        ws_tc.cell(row_i, 12, 'Yes' if tc.is_extra else 'No')
    _norm_rows(ws_tc, len(tc_rows), len(TC_HEADERS))

    # ── 4. Full Ad Report sheet (MatchResult) ────────────────────────────────
    mr_qs = MatchResult.objects.filter(account_id=account_id)
    if channel: mr_qs = mr_qs.filter(channel__iexact=channel)
    if month:   mr_qs = mr_qs.filter(month=month)
    if _from_d: mr_qs = mr_qs.filter(scheduled_date__gte=_from_d)
    if _to_d:   mr_qs = mr_qs.filter(scheduled_date__lte=_to_d)
    if brand:   mr_qs = mr_qs.filter(brand__icontains=brand)
    mr_qs = mr_qs.order_by('month', 'scheduled_date', 'brand')
    mr_rows = list(mr_qs)

    ws_report = wb.create_sheet('Full Ad Report')
    AR_HEADERS = [
        'Month', 'Channel', 'Brand', 'Theme', 'Duration', 'Programme',
        'Planned Date', 'Planned Start', 'Planned End',
        'Aired Date', 'Aired Time', 'Source', 'Status',
    ]
    STATUS_LABEL = dict(MatchResult.STATUS_CHOICES)
    _meta_row(ws_report, len(AR_HEADERS), 'Full Ad Report', len(mr_rows))
    _hdr_row(ws_report, AR_HEADERS)
    for row_i, mr in enumerate(mr_rows, start=3):
        ws_report.cell(row_i,  1, mr.month or '')
        ws_report.cell(row_i,  2, getattr(mr, 'channel', '') or '')
        ws_report.cell(row_i,  3, mr.brand or '')
        ws_report.cell(row_i,  4, mr.theme or '')
        ws_report.cell(row_i,  5, mr.duration)
        ws_report.cell(row_i,  6, mr.programme or '')
        ws_report.cell(row_i,  7, str(mr.scheduled_date) if mr.scheduled_date else '')
        ws_report.cell(row_i,  8, mr.planned_start or '')
        ws_report.cell(row_i,  9, mr.planned_end or '')
        ws_report.cell(row_i, 10, str(mr.aired_date) if mr.aired_date else '')
        ws_report.cell(row_i, 11, mr.air_time or '')
        ws_report.cell(row_i, 12, mr.source or '')
        ws_report.cell(row_i, 13, STATUS_LABEL.get(mr.status, mr.status))
        for col_i in range(1, len(AR_HEADERS) + 1):
            ws_report.cell(row_i, col_i).font = norm_font

    # ── 5. Brand Mappings sheet ───────────────────────────────────────────────
    bm_qs = BrandMapping.objects.filter(account_id=account_id)
    if brand: bm_qs = bm_qs.filter(brand__icontains=brand)
    bm_qs = bm_qs.order_by('brand', 'theme')
    bm_rows = list(bm_qs)

    ws_bm = wb.create_sheet('Brand Mappings')
    BM_HEADERS = ['Brand', 'LMRB Theme', 'TC Theme', 'Duration Filter']
    _meta_row(ws_bm, len(BM_HEADERS), 'Brand Mappings', len(bm_rows))
    _hdr_row(ws_bm, BM_HEADERS)
    for row_i, bm in enumerate(bm_rows, start=3):
        ws_bm.cell(row_i, 1, bm.brand or '')
        ws_bm.cell(row_i, 2, bm.theme or '')
        ws_bm.cell(row_i, 3, bm.tc_theme or '')
        ws_bm.cell(row_i, 4, bm.duration if bm.duration is not None else '')
    _norm_rows(ws_bm, len(bm_rows), len(BM_HEADERS))

    # ── 6. Manual Matches sheet ───────────────────────────────────────────────
    mm_qs = ManualMatch.objects.select_related(
        'schedule_row', 'tc_row', 'lmrb_row', 'matched_by'
    ).filter(account_id=account_id)
    if channel: mm_qs = mm_qs.filter(channel__iexact=channel)
    if month:   mm_qs = mm_qs.filter(month=month)
    mm_qs = mm_qs.order_by('month', 'channel', 'matched_at')
    mm_rows = list(mm_qs)

    ws_mm = wb.create_sheet('Manual Matches')
    MM_HEADERS = [
        'Channel', 'Month', 'Match Mode',
        'Schedule Brand', 'Schedule Date', 'Schedule Start',
        'TC Theme', 'TC Date', 'TC Aired Time',
        'LMRB Theme', 'LMRB Date', 'LMRB Time',
        'Note', 'Matched By', 'Matched At',
    ]
    _meta_row(ws_mm, len(MM_HEADERS), 'Manual Matches', len(mm_rows))
    _hdr_row(ws_mm, MM_HEADERS)
    for row_i, mm in enumerate(mm_rows, start=3):
        ws_mm.cell(row_i,  1, mm.channel or '')
        ws_mm.cell(row_i,  2, mm.month or '')
        ws_mm.cell(row_i,  3, mm.match_mode or '')
        ws_mm.cell(row_i,  4, mm.schedule_row.brand if mm.schedule_row_id else '')
        ws_mm.cell(row_i,  5, str(mm.schedule_row.date) if mm.schedule_row_id and mm.schedule_row.date else '')
        ws_mm.cell(row_i,  6, mm.schedule_row.start_time if mm.schedule_row_id else '')
        ws_mm.cell(row_i,  7, mm.tc_row.tc_theme if mm.tc_row_id else '')
        ws_mm.cell(row_i,  8, str(mm.tc_row.date) if mm.tc_row_id and mm.tc_row.date else '')
        ws_mm.cell(row_i,  9, mm.tc_row.aired_time if mm.tc_row_id else '')
        ws_mm.cell(row_i, 10, mm.lmrb_row.advt_theme if mm.lmrb_row_id else '')
        ws_mm.cell(row_i, 11, str(mm.lmrb_row.date) if mm.lmrb_row_id and mm.lmrb_row.date else '')
        ws_mm.cell(row_i, 12, mm.lmrb_row.advt_time if mm.lmrb_row_id else '')
        ws_mm.cell(row_i, 13, mm.note or '')
        ws_mm.cell(row_i, 14, mm.matched_by.name if mm.matched_by_id else '')
        ws_mm.cell(row_i, 15, str(mm.matched_at) if mm.matched_at else '')
    _norm_rows(ws_mm, len(mm_rows), len(MM_HEADERS))

    # ── 7. Sponsorship Assignments sheet ─────────────────────────────────────
    sp_qs = SponsorshipLmrbAssignment.objects.select_related(
        'schedule_row', 'lmrb_row', 'matched_by'
    ).filter(account_id=account_id)
    if channel: sp_qs = sp_qs.filter(schedule_row__channel__iexact=channel)
    if month:   sp_qs = sp_qs.filter(schedule_row__month=month)
    sp_qs = sp_qs.order_by('schedule_row__month', 'schedule_row__date')
    sp_rows = list(sp_qs)

    ws_sp = wb.create_sheet('Sponsorship Assignments')
    SP_HEADERS = [
        'Channel', 'Month', 'Schedule Brand', 'Schedule Date',
        'Schedule Start', 'Schedule Duration (s)',
        'LMRB Theme', 'LMRB Date', 'LMRB Time', 'LMRB Duration (s)',
        'Match Type', 'Matched By', 'Matched At',
    ]
    _meta_row(ws_sp, len(SP_HEADERS), 'Sponsorship Assignments', len(sp_rows))
    _hdr_row(ws_sp, SP_HEADERS)
    for row_i, sp in enumerate(sp_rows, start=3):
        sr = sp.schedule_row
        lr = sp.lmrb_row
        ws_sp.cell(row_i,  1, sr.channel if sr else '')
        ws_sp.cell(row_i,  2, sr.month if sr else '')
        ws_sp.cell(row_i,  3, sr.brand if sr else '')
        ws_sp.cell(row_i,  4, str(sr.date) if sr and sr.date else '')
        ws_sp.cell(row_i,  5, sr.start_time if sr else '')
        ws_sp.cell(row_i,  6, sr.duration if sr else '')
        ws_sp.cell(row_i,  7, lr.advt_theme if lr else '')
        ws_sp.cell(row_i,  8, str(lr.date) if lr and lr.date else '')
        ws_sp.cell(row_i,  9, lr.advt_time if lr else '')
        ws_sp.cell(row_i, 10, lr.duration if lr else '')
        ws_sp.cell(row_i, 11, sp.match_type or '')
        ws_sp.cell(row_i, 12, sp.matched_by.name if sp.matched_by_id else 'Auto')
        ws_sp.cell(row_i, 13, str(sp.matched_at) if sp.matched_at else '')
    _norm_rows(ws_sp, len(sp_rows), len(SP_HEADERS))

    # ── Build filename & response ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    parts = [account.name]
    if channel:         parts.append(channel)
    if month:           parts.append(month)
    if schedule_number: parts.append(f'Sch{schedule_number}')
    fname = f'AdminExport_{"_".join(parts)}.xlsx'.replace(' ', '_')
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def _month_date_range(month_str):
    """Return (first_date, last_date) for a 'Month YYYY' string, or (None, None)."""
    import calendar
    from datetime import date as _date, datetime
    try:
        dt = datetime.strptime(month_str, '%B %Y').date()
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        return (_date(dt.year, dt.month, 1), _date(dt.year, dt.month, last_day))
    except Exception:
        return (None, None)


# ── System Settings (super_admin only) ────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin'])
def system_settings(request):
    """
    Site-wide configuration page - super_admin only.

    On first visit, any missing settings are auto-created from SETTING_DEFAULTS
    so the page always shows the full list even on a fresh installation.
    """
    # Ensure all default settings exist in DB
    _ensure_defaults()

    if request.method == 'POST':
        updated = 0
        for s in SystemSetting.objects.all():
            new_val = request.POST.get(s.key, '').strip()
            if s.value != new_val:
                s.value = new_val
                s.save()
                updated += 1
        if updated:
            messages.success(request, f'Settings saved - {updated} value(s) updated.')
            _audit(request, 'settings_change', f'{updated} setting(s) updated.')
        else:
            messages.success(request, 'Settings saved (no changes).')
        return redirect('system_settings')

    # Group settings by their category display label
    from collections import OrderedDict
    CATEGORY_ORDER = ['reconciliation', 'tc_parsing', 'lmrb_parsing', 'whatsapp']
    CATEGORY_LABELS = {
        'reconciliation': 'Reconciliation',
        'tc_parsing':     'TC File Parsing',
        'lmrb_parsing':   'LMRB / MapOnline File Parsing',
        'whatsapp':       'WhatsApp Notifications',
    }
    all_settings = list(SystemSetting.objects.all())
    categories = OrderedDict()
    for cat_key in CATEGORY_ORDER:
        cat_settings = [s for s in all_settings if s.category == cat_key]
        if cat_settings:
            categories[CATEGORY_LABELS.get(cat_key, cat_key)] = cat_settings

    return render(request, 'admin_panel/settings.html', {
        'categories': categories,
        'branding_logo_url':   _branding_url('logo'),
        'branding_tartan_url': _branding_url('tartan'),
    })


# ── Branding upload ─────────────────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin'])
def branding_upload(request):
    """Upload logo or tartan pattern image for the login page / sidebar."""
    if request.method == 'POST':
        asset_type = request.POST.get('asset_type')   # 'logo' or 'tartan'
        uploaded   = request.FILES.get('file')
        if asset_type not in ('logo', 'tartan') or not uploaded:
            messages.error(request, 'Invalid upload - specify logo or tartan and provide a file.')
            return redirect('system_settings')
        # Determine extension
        orig_name = uploaded.name.lower()
        ext = '.png'
        for e in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg'):
            if orig_name.endswith(e):
                ext = e
                break
        dest_dir = os.path.join(django_settings.MEDIA_ROOT, 'branding')
        os.makedirs(dest_dir, exist_ok=True)
        dest_path = os.path.join(dest_dir, f'{asset_type}{ext}')
        # Remove previous files with different extensions
        for e in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg'):
            old = os.path.join(dest_dir, f'{asset_type}{e}')
            if old != dest_path and os.path.exists(old):
                os.remove(old)
        with open(dest_path, 'wb') as fh:
            for chunk in uploaded.chunks():
                fh.write(chunk)
        label = 'Logo' if asset_type == 'logo' else 'Tartan pattern'
        messages.success(request, f'{label} uploaded successfully.')
    return redirect('system_settings')


# ── Sponsorship Reconciliation ─────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head', 'planner'])
@require_POST
def sponsorship_reconcile(request):
    """
    POST: Run auto Step 1 sponsorship reconciliation for a scope.
    Accepts mode='smart' (default) or mode='reset'.
    """
    from verification.sponsorship_engine import reconcile_sponsorship

    account_id  = request.POST.get('account_id', '').strip()
    channel     = request.POST.get('channel', '').strip()
    month       = request.POST.get('month', '').strip()
    mode        = request.POST.get('mode', 'smart')
    schedule_id = request.POST.get('schedule_id', '').strip()

    if not (account_id and channel and month):
        messages.error(request, 'Incomplete parameters.')
        return redirect('/dashboard/summary/')
    if not _account_access(request.user, account_id):
        messages.error(request, 'Access denied.')
        return redirect('/dashboard/summary/')

    sid = int(schedule_id) if schedule_id else None
    result = reconcile_sponsorship(int(account_id), channel, month, mode=mode, schedule_id=sid)
    messages.success(
        request,
        f'Sponsorship reconciliation complete: '
        f'{result["assigned"]} new assignments, '
        f'{result["already_assigned"]} already assigned, '
        f'{result["total_spon_rows"]} total sponsorship rows.'
    )
    redir = f'/dashboard/summary/?account_id={account_id}&channel={channel}&month={month}'
    if schedule_id:
        redir += f'&schedule_id={schedule_id}'
    return redirect(redir)


@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head'])
def sponsorship_candidates(request):
    """
    GET (AJAX): Return unmatched LMRBRow list available for manual sponsorship
    assignment.  Optionally filtered by brand and duration query params.
    """
    from verification.sponsorship_engine import lmrb_candidates

    account_id = request.GET.get('account_id', '').strip()
    channel    = request.GET.get('channel', '').strip()
    month      = request.GET.get('month', '').strip()
    brand      = request.GET.get('brand', '').strip()
    duration   = request.GET.get('duration', '').strip()

    if not (account_id and channel and month):
        return JsonResponse({'error': 'Incomplete parameters.'}, status=400)
    if not _account_access(request.user, account_id):
        return JsonResponse({'error': 'Access denied.'}, status=403)

    rows = lmrb_candidates(int(account_id), channel, month)

    # Optional server-side filter by brand mapping theme.
    # NOTE: `brand` here is actually the *product* name from the summary row
    # (openSponPicker passes row.product).  We resolve themes by looking up all
    # BrandMappings whose .product matches that value; if none match by product
    # we fall back to matching by .brand so direct brand names also work.
    if brand or duration:
        from core.models import BrandMapping as BM
        bm_all = list(BM.objects.filter(account_id=account_id))
        brand_lower = brand.lower().strip() if brand else ''
        themes = set()

        # Pass 1 - match BrandMapping by product (row.product from template)
        for bm in bm_all:
            if brand_lower and (bm.product or '').lower().strip() != brand_lower:
                continue
            if duration:
                try:
                    if bm.duration is not None and int(bm.duration) != int(duration):
                        continue
                except (ValueError, TypeError):
                    pass
            if bm.theme:
                themes.add(bm.theme.lower().strip())

        # Pass 2 - fallback: match by brand name directly
        if not themes and brand_lower:
            for bm in bm_all:
                if bm.brand.lower().strip() != brand_lower:
                    continue
                if duration:
                    try:
                        if bm.duration is not None and int(bm.duration) != int(duration):
                            continue
                    except (ValueError, TypeError):
                        pass
                if bm.theme:
                    themes.add(bm.theme.lower().strip())

        if themes:
            rows = [
                r for r in rows
                if str(r.get('advt_theme', '')).lower().strip() in themes
            ]
        if duration:
            try:
                dur_int = int(duration)
                rows = [r for r in rows if r.get('duration') == dur_int]
            except ValueError:
                pass

    for r in rows:
        if hasattr(r.get('date'), 'isoformat'):
            r['date'] = r['date'].isoformat()

    return JsonResponse({'candidates': rows})


@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head', 'planner'])
@require_POST
def sponsorship_assign(request):
    """
    POST (AJAX): Manually assign LMRB rows to sponsorship schedule rows.
    Body: JSON {"assignments": [[schedule_row_id, lmrb_row_id], ...],
                "account_id": ..., "channel": ..., "month": ...}
    """
    import json as _json
    from verification.sponsorship_engine import manual_assign

    try:
        body = _json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid JSON.'}, status=400)

    account_id = str(body.get('account_id', '')).strip()
    channel    = str(body.get('channel', '')).strip()
    month      = str(body.get('month', '')).strip()

    if not (account_id and channel and month):
        return JsonResponse({'error': 'Incomplete parameters.'}, status=400)
    if not _account_access(request.user, account_id):
        return JsonResponse({'error': 'Access denied.'}, status=403)

    try:
        assignments = [(int(sr), int(lr)) for sr, lr in body.get('assignments', [])]
    except (ValueError, TypeError, KeyError):
        return JsonResponse({'error': 'Invalid assignments payload.'}, status=400)

    result = manual_assign(int(account_id), channel, month, assignments, request.user)
    return JsonResponse(result)


@login_required
@role_required(['super_admin', 'admin', 'team_head', 'planner'])
@require_POST
def sponsorship_reset(request):
    """POST: Delete all SponsorshipLmrbAssignments for a scope and unlock LMRBRows."""
    from verification.sponsorship_engine import reset_sponsorship

    account_id = request.POST.get('account_id', '').strip()
    channel    = request.POST.get('channel', '').strip()
    month      = request.POST.get('month', '').strip()

    if not (account_id and channel and month):
        messages.error(request, 'Incomplete parameters.')
        return redirect('/dashboard/summary/')
    if not _account_access(request.user, account_id):
        messages.error(request, 'Access denied.')
        return redirect('/dashboard/summary/')

    result = reset_sponsorship(int(account_id), channel, month)
    messages.success(
        request,
        f'Sponsorship reconciliation reset: '
        f'{result["deleted"]} assignments removed, '
        f'{result["lmrb_unlocked"]} LMRB rows unlocked.'
    )
    return redirect(
        f'/dashboard/summary/?account_id={account_id}&channel={channel}&month={month}'
    )


@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head'])
def sponsorship_unmatched_rows(request):
    """
    GET (AJAX): Return unmatched SPONSORSHIP ScheduleRow ids for a brand/duration/month,
    so the manual picker can pair selected LMRB rows with pending schedule rows.
    """
    account_id = request.GET.get('account_id', '').strip()
    channel    = request.GET.get('channel', '').strip()
    month      = request.GET.get('month', '').strip()
    brand      = request.GET.get('brand', '').strip()
    duration   = request.GET.get('duration', '').strip()

    if not (account_id and channel and month):
        return JsonResponse({'error': 'Incomplete parameters.'}, status=400)
    if not _account_access(request.user, account_id):
        return JsonResponse({'error': 'Access denied.'}, status=403)

    qs = ScheduleRow.objects.filter(
        account_id=account_id, channel=channel, month=month,
        ad_type='SPONSORSHIP', brand=brand,
    ).exclude(sponsorship_assignment__isnull=False)

    if duration:
        try:
            qs = qs.filter(duration=int(duration))
        except ValueError:
            pass

    ids = list(qs.order_by('date', 'start_time').values_list('id', flat=True))
    return JsonResponse({'schedule_row_ids': ids})


# ── Manual Reconciliation ──────────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head', 'planner'])
def manual_reconciliation(request):
    """
    Three-panel view: unmatched schedule rows (left), unmatched TC spots (middle),
    unmatched LMRB rows (right). Supports Schedule+LMRB, 3-way, and TC+LMRB modes.
    Existing ManualMatch records for the scope are shown below with a De-match button.
    """
    from core.models import ManualMatch

    user       = request.user
    account_qs = _account_qs(user)

    account_id = request.GET.get('account_id', '')
    channel    = request.GET.get('channel', '')
    month      = request.GET.get('month', '')

    channels           = []
    months             = []
    unmatched_schedule = []
    unmatched_tc       = []
    unmatched_lmrb     = []
    manual_matches     = []
    sch_date_range     = (None, None)

    if account_id:
        if _account_access(user, account_id):
            channels = list(
                ScheduleRow.objects.filter(account_id=account_id)
                .values_list('channel', flat=True).distinct().order_by('channel')
            )

    if account_id and channel:
        if _account_access(user, account_id):
            months = list(
                ScheduleRow.objects.filter(account_id=account_id, channel=channel)
                .values_list('month', flat=True).distinct().order_by('month')
            )

    if account_id and channel and month and _account_access(user, account_id):
        # Sponsorship schedule rows that have already been matched via
        # SponsorshipLmrbAssignment must not appear in the unmatched panel.
        from core.models import SponsorshipLmrbAssignment as _SLA
        matched_spon_sr_ids = set(
            _SLA.objects.filter(
                account_id=account_id,
                schedule_row__channel=channel,
                schedule_row__month=month,
            ).values_list('schedule_row_id', flat=True)
        )

        # Unmatched schedule rows (both COMMERCIAL BENEFITS and SPONSORSHIP)
        unmatched_schedule = list(
            ScheduleRow.objects.filter(
                account_id=account_id, channel=channel, month=month,
                ad_type__in=['COMMERCIAL BENEFITS', 'SPONSORSHIP'],
                is_matched=False,
                is_manual_matched=False,
            ).exclude(id__in=matched_spon_sr_ids)
            .order_by('date', 'start_time', 'brand')
        )

        # Determine schedule date range for display purposes
        from django.db.models import Min as _Min, Max as _Max
        dr = ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month,
        ).aggregate(mn=_Min('date'), mx=_Max('date'))
        sch_date_range = (dr['mn'], dr['mx'])

        # Unmatched TC rows - all TC rows not yet linked to a schedule or a
        # ManualMatch.  This deliberately includes is_extra=True rows (spots the
        # reconcile engine couldn't match to any schedule row) so they can be
        # manually paired with an LMRB row via the tc_lmrb match mode.
        unmatched_tc = list(
            TCRow.objects.filter(
                account_id=account_id,
                channel=channel,
                tc_report__month=month,
                manual_match__isnull=True,
                is_schedule_matched=False,
            ).select_related('tc_report')
            .order_by('tc_theme', 'duration', 'date', 'aired_time')[:500]
        )

        # Unmatched LMRB rows for this channel - NOT date-filtered so out-of-range
        # rows (the whole purpose of manual matching) are visible.
        unmatched_lmrb = list(
            LMRBRow.objects.filter(
                account_id=account_id,
                channel__iexact=channel,
                is_matched=False,
                is_manual_matched=False,
                is_sponsorship_matched=False,
            ).order_by('advt_theme', 'duration', 'date', 'advt_time')[:500]
        )

        manual_matches = list(
            ManualMatch.objects.filter(
                account_id=account_id, channel=channel, month=month,
            ).select_related('schedule_row', 'tc_row', 'lmrb_row', 'matched_by')
        )

    return render(request, 'manual_reconciliation/index.html', {
        'accounts':           account_qs,
        'account_id':         account_id,
        'channels':           channels,
        'channel':            channel,
        'months':             months,
        'month':              month,
        'unmatched_schedule': unmatched_schedule,
        'unmatched_tc':       unmatched_tc,
        'unmatched_lmrb':     unmatched_lmrb,
        'manual_matches':     manual_matches,
        'sch_date_range':     sch_date_range,
    })


@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head', 'planner'])
@require_POST
def manual_match_create(request):
    """
    POST: Create a ManualMatch.  Supports three modes:
      schedule_lmrb - Schedule + LMRB  (original)
      3way          - Schedule + TC + LMRB
      tc_lmrb       - TC + LMRB only (no schedule row)
    """
    from core.models import ManualMatch, MatchResult

    account_id      = request.POST.get('account_id', '').strip()
    channel         = request.POST.get('channel', '').strip()
    month           = request.POST.get('month', '').strip()
    match_mode      = request.POST.get('match_mode', 'schedule_lmrb').strip()
    schedule_row_id = request.POST.get('schedule_row_id', '').strip()
    tc_row_id       = request.POST.get('tc_row_id', '').strip()
    lmrb_row_id     = request.POST.get('lmrb_row_id', '').strip()
    note            = request.POST.get('note', '').strip()

    redirect_url = (
        f'/dashboard/manual/?account_id={account_id}'
        f'&channel={channel}&month={month}'
    )

    if not all([account_id, channel, month, lmrb_row_id]):
        messages.error(request, 'Incomplete parameters.')
        return redirect(redirect_url)
    if match_mode in ('schedule_lmrb', '3way') and not schedule_row_id:
        messages.error(request, 'A schedule row is required for this match mode.')
        return redirect(redirect_url)
    if match_mode in ('3way', 'tc_lmrb') and not tc_row_id:
        messages.error(request, 'A TC row is required for this match mode.')
        return redirect(redirect_url)
    if not _account_access(request.user, account_id):
        messages.error(request, 'Access denied.')
        return redirect(redirect_url)

    # Fetch schedule row if needed
    sr = None
    if schedule_row_id:
        sr = get_object_or_404(
            ScheduleRow, id=schedule_row_id, account_id=account_id,
            channel=channel, month=month,
            is_matched=False, is_manual_matched=False,
        )

    # Fetch TC row if needed
    tr = None
    if tc_row_id:
        tr = get_object_or_404(
            TCRow, id=tc_row_id, account_id=account_id, channel=channel,
            manual_match__isnull=True,
        )

    # Fetch LMRB row
    lr = get_object_or_404(
        LMRBRow, id=lmrb_row_id, account_id=account_id,
        is_matched=False, is_manual_matched=False, is_sponsorship_matched=False,
    )

    # Race-condition guards
    if ManualMatch.objects.filter(lmrb_row=lr).exists():
        messages.error(request, 'LMRB row already manually matched.')
        return redirect(redirect_url)
    if sr and ManualMatch.objects.filter(schedule_row=sr).exists():
        messages.error(request, 'Schedule row already manually matched.')
        return redirect(redirect_url)
    if tr and ManualMatch.objects.filter(tc_row=tr).exists():
        messages.error(request, 'TC row already manually matched.')
        return redirect(redirect_url)

    # Create the manual match
    ManualMatch.objects.create(
        account_id   = account_id,
        channel      = channel,
        month        = month,
        match_mode   = match_mode,
        schedule_row = sr,
        tc_row       = tr,
        lmrb_row     = lr,
        note         = note,
        matched_by   = request.user,
    )

    # Lock rows
    if sr:
        ScheduleRow.objects.filter(id=sr.id).update(is_manual_matched=True)
    LMRBRow.objects.filter(id=lr.id).update(is_manual_matched=True)

    # For 3way and tc_lmrb modes: also lock the TCRow so it appears as matched
    # in TC reconciliation views (is_schedule_matched=True on TCRow).
    if tr and match_mode in ('3way', 'tc_lmrb'):
        update_fields = {'is_schedule_matched': True}
        if sr:
            update_fields['matched_schedule_id'] = sr.id
        TCRow.objects.filter(id=tr.id).update(**update_fields)

    # Remove stale engine MatchResult for this schedule row
    if sr:
        MatchResult.objects.filter(
            schedule_row=sr,
            status__in=['not_aired', 'late_telecast', 'programme_mismatch'],
        ).delete()

    if sr:
        desc = f'{sr.brand} ({sr.duration}s, {sr.date})'
    elif tr:
        desc = f'TC: {tr.tc_theme} ({tr.date})'
    else:
        desc = 'unknown'
    messages.success(
        request,
        f'Manually matched [{match_mode}]: {desc} ← LMRB: {lr.advt_theme} ({lr.date} {lr.advt_time})',
    )
    return redirect(redirect_url)


@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head', 'planner'])
@require_POST
def manual_dematch(request, pk):
    """
    POST: Remove a ManualMatch and unlock both rows.
    """
    from core.models import ManualMatch

    account_id = request.POST.get('account_id', '').strip()
    channel    = request.POST.get('channel', '').strip()
    month      = request.POST.get('month', '').strip()

    redirect_url = (
        f'/dashboard/manual/?account_id={account_id}'
        f'&channel={channel}&month={month}'
    )

    if not _account_access(request.user, account_id):
        messages.error(request, 'Access denied.')
        return redirect(redirect_url)

    mm = get_object_or_404(ManualMatch, pk=pk, account_id=account_id)

    sr_id = mm.schedule_row_id
    lr_id = mm.lmrb_row_id
    if mm.schedule_row_id:
        sr_desc = f'{mm.schedule_row.brand} ({mm.schedule_row.date})'
    elif mm.tc_row_id:
        sr_desc = f'TC: {mm.tc_row.tc_theme} ({mm.tc_row.date})'
    else:
        sr_desc = 'entry'

    tc_row_id_to_unlock = mm.tc_row_id
    tc_mode = mm.match_mode

    mm.delete()

    # Unlock rows
    if sr_id:
        ScheduleRow.objects.filter(id=sr_id).update(is_manual_matched=False)
    LMRBRow.objects.filter(id=lr_id).update(is_manual_matched=False)

    # Unlock TCRow if it was locked by this manual match
    if tc_row_id_to_unlock and tc_mode in ('3way', 'tc_lmrb'):
        TCRow.objects.filter(id=tc_row_id_to_unlock).update(
            is_schedule_matched=False, matched_schedule=None,
        )

    messages.success(request, f'De-matched: {sr_desc}. All rows are now available again.')
    return redirect(redirect_url)


# ── Commercial Tags Pool ───────────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head'])
def commercial_candidates(request):
    """
    GET (AJAX): Return unmatched LMRBRow list available for manual commercial
    Tags assignment.  Filtered by brand theme mapping and optionally by duration.
    Only returns rows that are not commercially matched, not sponsorship matched,
    and not manually matched (i.e., fully available).
    """
    from django.db.models import Q as _Q, Min as _Min, Max as _Max

    account_id = request.GET.get('account_id', '').strip()
    channel    = request.GET.get('channel', '').strip()
    month      = request.GET.get('month', '').strip()
    brand      = request.GET.get('brand', '').strip()
    duration   = request.GET.get('duration', '').strip()

    if not (account_id and channel and month):
        return JsonResponse({'error': 'Incomplete parameters.'}, status=400)
    if not _account_access(request.user, account_id):
        return JsonResponse({'error': 'Access denied.'}, status=403)

    # Get scope date range from schedule
    agg = Schedule.objects.filter(account_id=account_id, channel=channel, month=month
        ).aggregate(d_min=_Min('start_date'), d_max=_Max('end_date'))
    d_min, d_max = agg['d_min'], agg['d_max']
    if not d_min or not d_max:
        ragg = ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month,
        ).aggregate(d_min=_Min('date'), d_max=_Max('date'))
        d_min = d_min or ragg['d_min']
        d_max = d_max or ragg['d_max']

    # Base queryset: fully unmatched LMRB rows in scope
    qs = LMRBRow.objects.filter(
        account_id=account_id, channel__iexact=channel,
        is_matched=False, is_sponsorship_matched=False, is_manual_matched=False,
    )
    if d_min:
        qs = qs.filter(date__gte=d_min)
    if d_max:
        qs = qs.filter(date__lte=d_max)

    # Filter by theme matching the brand
    if brand:
        themes = set()
        for bm in BrandMapping.objects.filter(account_id=account_id):
            if bm.brand.lower().strip() != brand.lower().strip():
                continue
            if duration:
                try:
                    if bm.duration is not None and int(bm.duration) != int(duration):
                        continue
                except (ValueError, TypeError):
                    pass
            if bm.theme:
                themes.add(bm.theme.lower().strip())
        if themes:
            theme_q = _Q()
            for t in themes:
                theme_q |= _Q(advt_theme__iexact=t)
            qs = qs.filter(theme_q)

    if duration:
        try:
            qs = qs.filter(duration=int(duration))
        except ValueError:
            pass

    rows = list(qs.order_by('advt_theme', 'date', 'advt_time').values(
        'id', 'date', 'advt_time', 'advt_theme', 'duration', 'program', 'source',
    ))
    for r in rows:
        if hasattr(r.get('date'), 'isoformat'):
            r['date'] = r['date'].isoformat()

    return JsonResponse({'candidates': rows})


@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head'])
def commercial_unmatched_rows(request):
    """
    GET (AJAX): Return unmatched COMMERCIAL BENEFITS ScheduleRow IDs for a
    brand/duration/month so the picker can pair LMRB rows with schedule slots.
    """
    account_id = request.GET.get('account_id', '').strip()
    channel    = request.GET.get('channel', '').strip()
    month      = request.GET.get('month', '').strip()
    brand      = request.GET.get('brand', '').strip()
    duration   = request.GET.get('duration', '').strip()

    if not (account_id and channel and month):
        return JsonResponse({'error': 'Incomplete parameters.'}, status=400)
    if not _account_access(request.user, account_id):
        return JsonResponse({'error': 'Access denied.'}, status=403)

    qs = ScheduleRow.objects.filter(
        account_id=account_id, channel=channel, month=month,
        ad_type='COMMERCIAL BENEFITS', brand=brand,
        is_matched=False, is_manual_matched=False,
    )
    if duration:
        try:
            qs = qs.filter(duration=int(duration))
        except ValueError:
            pass

    ids = list(qs.order_by('date', 'start_time').values_list('id', flat=True))
    return JsonResponse({'schedule_row_ids': ids})


@login_required
@role_required(['super_admin', 'admin', 'operations', 'team_head', 'planner'])
@require_POST
def commercial_assign(request):
    """
    POST (AJAX): Manually assign LMRB rows to commercial BENEFITS schedule rows
    as 'Tags'.  Creates ManualMatch records (schedule_lmrb mode) which are then
    counted in build_summary_data's manual_aired column.

    Body: JSON {
      "account_id": ..., "channel": ..., "month": ...,
      "assignments": [[schedule_row_id, lmrb_row_id], ...]
    }
    """
    import json as _json
    from core.models import ManualMatch as _MM

    try:
        body = _json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid JSON.'}, status=400)

    account_id = str(body.get('account_id', '')).strip()
    channel    = str(body.get('channel', '')).strip()
    month      = str(body.get('month', '')).strip()

    if not (account_id and channel and month):
        return JsonResponse({'error': 'Incomplete parameters.'}, status=400)
    if not _account_access(request.user, account_id):
        return JsonResponse({'error': 'Access denied.'}, status=403)

    try:
        assignments = [(int(sr), int(lr)) for sr, lr in body.get('assignments', [])]
    except (ValueError, TypeError, KeyError):
        return JsonResponse({'error': 'Invalid assignments payload.'}, status=400)

    created = 0
    skipped = 0

    for sr_id, lr_id in assignments:
        try:
            sr = ScheduleRow.objects.get(
                id=sr_id, account_id=account_id, channel=channel,
                month=month, ad_type='COMMERCIAL BENEFITS',
                is_matched=False, is_manual_matched=False,
            )
        except ScheduleRow.DoesNotExist:
            skipped += 1
            continue

        try:
            lr = LMRBRow.objects.get(
                id=lr_id, account_id=account_id,
                is_matched=False, is_sponsorship_matched=False,
                is_manual_matched=False,
            )
        except LMRBRow.DoesNotExist:
            skipped += 1
            continue

        # Race-condition guard
        if _MM.objects.filter(lmrb_row=lr).exists() or _MM.objects.filter(schedule_row=sr).exists():
            skipped += 1
            continue

        _MM.objects.create(
            account_id=account_id, channel=channel, month=month,
            match_mode='schedule_lmrb',
            schedule_row=sr, lmrb_row=lr,
            note='Commercial Tags – manually assigned via summary picker',
            matched_by=request.user,
        )
        ScheduleRow.objects.filter(id=sr_id).update(is_manual_matched=True)
        LMRBRow.objects.filter(id=lr_id).update(is_manual_matched=True)
        created += 1

    return JsonResponse({'created': created, 'skipped': skipped})


# ── Audit Log (C1) ─────────────────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin'])
def audit_log(request):
    """Show the last 500 AuditLog entries - super_admin / admin only."""
    from django.db.models import Q as _Q
    action_filter = request.GET.get('action', '')
    user_filter   = request.GET.get('user', '')

    qs = AuditLog.objects.select_related('user').order_by('-timestamp')
    if action_filter:
        qs = qs.filter(action=action_filter)
    if user_filter:
        qs = qs.filter(user__id=user_filter)

    from accounts.models import User as _User
    return render(request, 'admin_panel/audit_log.html', {
        'logs':           qs[:500],
        'action_choices': AuditLog.ACTION_CHOICES,
        'action_filter':  action_filter,
        'user_filter':    user_filter,
        'staff_users':    _User.objects.filter(role__in=['super_admin', 'admin'])
                                       .order_by('name'),
    })


# ── Admin Analytics (B1 + B2) ──────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin'])
def admin_analytics(request):
    """
    Cross-account analytics page for super_admin / admin.
    B1: Compliance bar chart, pass breakdown donut, monthly trend.
    B2: Brand mapping coverage per account (unmapped LMRB themes).
    """
    from django.db.models import Q as _Q, Count as _Count

    # ── B1 Chart 1: Per-account compliance % ──────────────────────────────────
    _acct_totals = {
        r['account_id']: r['n']
        for r in MatchResult.objects.values('account_id').annotate(n=_Count('id'))
    }
    _acct_matched = {
        r['account_id']: r['n']
        for r in MatchResult.objects.filter(status='matched')
                 .values('account_id').annotate(n=_Count('id'))
    }
    _acct_names = {a.id: a.name for a in Account.objects.all()}

    acct_compliance = []
    for aid, total in sorted(_acct_totals.items(), key=lambda x: _acct_names.get(x[0], '')):
        matched = _acct_matched.get(aid, 0)
        acct_compliance.append({
            'name': _acct_names.get(aid, str(aid)),
            'pct':  round(matched / total * 100) if total else 0,
            'matched': matched,
            'total':   total,
        })

    # ── B1 Chart 2: Global status breakdown ───────────────────────────────────
    _global = MatchResult.objects.aggregate(
        matched   = _Count('id', filter=_Q(status='matched')),
        prog_mis  = _Count('id', filter=_Q(status='programme_mismatch')),
        late      = _Count('id', filter=_Q(status='late_telecast')),
        not_aired = _Count('id', filter=_Q(status='not_aired')),
        no_map    = _Count('id', filter=_Q(status='no_mapping')),
        manual    = _Count('id', filter=_Q(status='manual_match')),
    )
    _global['total'] = sum(_global.values())

    # ── B1 Chart 3: Monthly trend ──────────────────────────────────────────────
    monthly_raw = list(
        MatchResult.objects.values('month')
        .annotate(
            total   = _Count('id'),
            matched = _Count('id', filter=_Q(status='matched')),
        )
        .order_by('month')
    )
    # Sort by parsed date so "January 2025" < "February 2025" etc.
    import calendar as _cal
    def _month_key(m):
        parts = m['month'].split()
        if len(parts) == 2:
            try:
                mon_num = list(_cal.month_name).index(parts[0])
                return (int(parts[1]), mon_num)
            except (ValueError, IndexError):
                pass
        return (9999, 0)
    monthly_raw.sort(key=_month_key)

    # ── B2: Brand mapping coverage per account ────────────────────────────────
    mapping_coverage = []
    for account in Account.objects.order_by('name'):
        lmrb_themes = list(
            LMRBRow.objects.filter(account=account)
            .exclude(advt_theme='')
            .values_list('advt_theme', flat=True)
            .distinct()
        )
        mapped_raw = list(
            BrandMapping.objects.filter(account=account)
            .values_list('theme', flat=True)
        )
        # Build normalized prefix/exact lookup
        mapped_norm = set()
        prefixes = []
        for t in mapped_raw:
            t_strip = t.strip()
            if t_strip.endswith('*'):
                prefixes.append(t_strip[:-1].lower())
            else:
                mapped_norm.add(t_strip.lower())

        def _is_covered(theme):
            t_low = theme.strip().lower()
            if t_low in mapped_norm:
                return True
            return any(t_low.startswith(p) for p in prefixes)

        unmapped = sorted([t for t in lmrb_themes if not _is_covered(t)])
        total = len(lmrb_themes)
        covered = total - len(unmapped)
        if total > 0:
            mapping_coverage.append({
                'account':  account.name,
                'total':    total,
                'covered':  covered,
                'pct':      round(covered / total * 100),
                'unmapped': unmapped[:30],
            })

    ctx = {
        'acct_compliance':    acct_compliance,
        'acct_labels_js':     _to_js([a['name'] for a in acct_compliance]),
        'acct_rates_js':      _to_js([a['pct']  for a in acct_compliance]),
        'global_status':      _global,
        'status_labels_js':   _to_js(['Matched', 'Prog. Mismatch', 'Late Telecast',
                                       'Not Aired', 'No Mapping', 'Manual']),
        'status_values_js':   _to_js([_global['matched'], _global['prog_mis'],
                                       _global['late'], _global['not_aired'],
                                       _global['no_map'], _global['manual']]),
        'monthly_labels_js':  _to_js([m['month']   for m in monthly_raw]),
        'monthly_total_js':   _to_js([m['total']   for m in monthly_raw]),
        'monthly_matched_js': _to_js([m['matched'] for m in monthly_raw]),
        'mapping_coverage':   mapping_coverage,
    }
    return render(request, 'admin_panel/admin_analytics.html', ctx)


# ── Database Admin Tools ───────────────────────────────────────────────────────

@login_required
@role_required(['super_admin', 'admin'])
def db_tools(request):
    """
    Database management page - admin/super_admin only.
    Provides data reset, selective delete, and database backup tools.
    """
    from django.db import connection
    from django.db.models import Count as _Count

    from competitor.models import CompetitorAd as _CompetitorAd, CompetitorUploadBatch as _CompetitorBatch

    def _counts():
        return {
            'schedules':        Schedule.objects.count(),
            'schedule_rows':    ScheduleRow.objects.count(),
            'lmrb_rows':        LMRBRow.objects.count(),
            'monitoring':       MonitoringData.objects.count(),
            'match_results':    MatchResult.objects.count(),
            'tc_rows':          TCRow.objects.count(),
            'tc_reports':       TransmissionReport.objects.count(),
            'competitor_ads':   _CompetitorAd.objects.count(),
            'competitor_batches': _CompetitorBatch.objects.count(),
        }

    msg = None
    msg_type = 'success'

    if request.method == 'POST':
        action = request.POST.get('action')
        confirm = request.POST.get('confirm', '')

        if action == 'reset_match_results':
            n = MatchResult.objects.exclude(status='manual_match').delete()[0]
            # Also reset is_matched flags on ScheduleRow and LMRBRow
            ScheduleRow.objects.filter(is_matched=True).update(
                is_matched=False, matched_lmrb=None, matched_at=None
            )
            LMRBRow.objects.filter(is_matched=True).update(
                is_matched=False, matched_schedule=None, matched_at=None
            )
            msg = f'Reset complete: {n} MatchResult records deleted and row locks cleared. Re-run verification to rebuild results.'
            _audit(request, 'db_reset', f'{n} MatchResult records deleted.')

        elif action == 'delete_duplicate_schedules':
            # Keep only the highest-version Schedule per (account, channel, month, schedule_number)
            from django.db.models import Max as _Max
            keeper_ids = set()
            for combo in Schedule.objects.values('account_id', 'channel', 'month', 'schedule_number').distinct():
                best = Schedule.objects.filter(**combo).order_by('-version').values_list('id', flat=True).first()
                if best:
                    keeper_ids.add(best)
            dupes = Schedule.objects.exclude(id__in=keeper_ids)
            n_schedules = dupes.count()
            if n_schedules == 0:
                msg = 'No duplicate schedules found - database is already clean.'
                msg_type = 'info'
            else:
                # Delete ScheduleRows belonging to duplicate schedules, then the schedules
                n_rows = ScheduleRow.objects.filter(schedule__in=dupes).count()
                MatchResult.objects.filter(schedule_row__schedule__in=dupes).delete()
                ScheduleRow.objects.filter(schedule__in=dupes).delete()
                dupes.delete()
                msg = (f'Removed {n_schedules} duplicate Schedule record(s) and {n_rows} ScheduleRow(s). '
                       f'Only the latest version of each schedule number is kept.')
                _audit(request, 'db_dedup', msg)

        elif action == 'delete_schedules' and confirm == 'DELETE':
            MatchResult.objects.all().delete()
            ScheduleRow.objects.all().delete()
            Schedule.objects.all().delete()
            msg = 'All schedule data deleted.'
            _audit(request, 'db_delete', 'Deleted all Schedule, ScheduleRow, and MatchResult records.')

        elif action == 'delete_lmrb' and confirm == 'DELETE':
            LMRBRow.objects.all().delete()
            MonitoringData.objects.all().delete()
            msg = 'All LMRB / monitoring data deleted.'
            _audit(request, 'db_delete', 'Deleted all LMRBRow and MonitoringData records.')

        elif action == 'delete_tc' and confirm == 'DELETE':
            TCRow.objects.all().delete()
            TransmissionReport.objects.all().delete()
            msg = 'All TC data deleted.'
            _audit(request, 'db_delete', 'Deleted all TCRow and TransmissionReport records.')

        elif action == 'delete_all' and confirm == 'DELETE_ALL':
            MatchResult.objects.all().delete()
            ScheduleRow.objects.all().delete()
            Schedule.objects.all().delete()
            LMRBRow.objects.all().delete()
            MonitoringData.objects.all().delete()
            TCRow.objects.all().delete()
            TransmissionReport.objects.all().delete()
            # Reset row locks
            ScheduleRow.objects.update(is_matched=False, matched_lmrb=None, matched_at=None)
            LMRBRow.objects.update(is_matched=False, matched_schedule=None, matched_at=None)
            msg = 'ALL data deleted. The system is now empty.'
            _audit(request, 'db_delete', 'NUCLEAR DELETE - all data wiped.')

        elif action == 'delete_competitor_account' and confirm == 'DELETE':
            from competitor.models import CompetitorAd as _CA, CompetitorUploadBatch as _CB
            acct_id = request.POST.get('competitor_account_id', '')
            if acct_id:
                try:
                    from core.models import Account as _Acc
                    acct = _Acc.objects.get(id=int(acct_id))
                    n_ads = _CA.objects.filter(account=acct).count()
                    n_batches = _CB.objects.filter(account=acct).count()
                    _CB.objects.filter(account=acct).delete()  # cascades to CompetitorAd
                    msg = f'Deleted {n_ads:,} competitor ad rows and {n_batches} batches for "{acct.name}".'
                    _audit(request, 'db_delete', msg)
                except (Account.DoesNotExist, ValueError):
                    msg = 'Account not found.'
                    msg_type = 'error'
            else:
                msg = 'No account selected.'
                msg_type = 'error'

        elif action == 'delete_competitor_all' and confirm == 'DELETE':
            from competitor.models import CompetitorAd as _CA, CompetitorUploadBatch as _CB
            n_ads = _CA.objects.count()
            n_batches = _CB.objects.count()
            _CB.objects.all().delete()  # cascades to all CompetitorAd rows
            msg = f'Deleted all {n_ads:,} competitor ad rows and {n_batches} batches.'
            _audit(request, 'db_delete', msg)

        elif action in ('delete_competitor_account', 'delete_competitor_all'):
            msg = 'Delete cancelled - confirmation text did not match.'
            msg_type = 'error'

        elif action in ('delete_schedules', 'delete_lmrb', 'delete_tc', 'delete_all'):
            msg = 'Delete cancelled - confirmation text did not match.'
            msg_type = 'error'

        elif action == 'backup_db':
            # Serve the SQLite file directly if running SQLite
            db_path = django_settings.DATABASES['default'].get('NAME', '')
            if db_path and db_path.endswith('.sqlite3') and os.path.exists(db_path):
                import mimetypes
                return FileResponse(
                    open(db_path, 'rb'),
                    as_attachment=True,
                    filename='ad_monitor_backup.sqlite3',
                    content_type='application/octet-stream',
                )
            else:
                msg = ('This deployment uses PostgreSQL. Use Railway\'s built-in backup feature '
                       'or run `pg_dump` from the CLI to export your database.')
                msg_type = 'info'

    # ── B3: Data Quality Scorecard ────────────────────────────────────────────
    from django.db.models import Q as _Q, Max as _Max

    _lmrb_total = LMRBRow.objects.count()
    if _lmrb_total:
        _lmrb_agg = LMRBRow.objects.aggregate(
            has_duration  = _Count('id', filter=_Q(duration__isnull=False)),
            has_programme = _Count('id', filter=_Q(program__gt='')),
            has_cost      = _Count('id', filter=_Q(cost__isnull=False)),
            has_brk_no    = _Count('id', filter=_Q(brk_no__isnull=False)),
            has_pos_brk   = _Count('id', filter=_Q(pos_in_brk__isnull=False)),
        )
        def _pct(n): return round(n / _lmrb_total * 100)
        lmrb_quality = {
            'total':      _lmrb_total,
            'duration':   _pct(_lmrb_agg['has_duration']),
            'programme':  _pct(_lmrb_agg['has_programme']),
            'cost':       _pct(_lmrb_agg['has_cost']),
            'break_info': _pct(_lmrb_agg['has_brk_no']),
            'pos_in_brk': _pct(_lmrb_agg['has_pos_brk']),
        }
    else:
        lmrb_quality = None

    _tc_total = TCRow.objects.count()
    if _tc_total:
        _tc_agg = TCRow.objects.aggregate(
            has_programme = _Count('id', filter=_Q(programme__gt='')),
            has_duration  = _Count('id', filter=_Q(duration__isnull=False)),
        )
        def _tpct(n): return round(n / _tc_total * 100)
        tc_quality = {
            'total':     _tc_total,
            'programme': _tpct(_tc_agg['has_programme']),
            'duration':  _tpct(_tc_agg['has_duration']),
        }
    else:
        tc_quality = None

    _mr_total = MatchResult.objects.count()
    if _mr_total:
        _mr_agg = MatchResult.objects.aggregate(
            matched     = _Count('id', filter=_Q(status='matched')),
            prog_mis    = _Count('id', filter=_Q(status='programme_mismatch')),
            late        = _Count('id', filter=_Q(status='late_telecast')),
            not_aired   = _Count('id', filter=_Q(status='not_aired')),
            no_mapping  = _Count('id', filter=_Q(status='no_mapping')),
            manual      = _Count('id', filter=_Q(status='manual_match')),
        )
        match_quality = {**_mr_agg, 'total': _mr_total}
    else:
        match_quality = None

    freshness = {
        'schedule':   Schedule.objects.aggregate(last=_Max('uploaded_at'))['last'],
        'lmrb':       MonitoringData.objects.aggregate(last=_Max('uploaded_at'))['last'],
        'tc':         TransmissionReport.objects.aggregate(last=_Max('uploaded_at'))['last'],
    }

    from competitor.models import CompetitorUploadBatch as _CB2
    competitor_account_summary = list(
        _CB2.objects.values('account__id', 'account__name')
        .annotate(
            batches=_Count('id'),
            ads=_Count('ads'),
        )
        .order_by('account__name')
    )

    return render(request, 'admin_panel/db_tools.html', {
        'counts':        _counts(),
        'msg':           msg,
        'msg_type':      msg_type,
        'lmrb_quality':  lmrb_quality,
        'tc_quality':    tc_quality,
        'match_quality': match_quality,
        'freshness':     freshness,
        'competitor_accounts': competitor_account_summary,
        'all_accounts':  Account.objects.all().order_by('name'),
    })


# ═══════════════════════════════════════════════════════════════════════════════
# WhatsApp Notification Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _notify_missed_spots_whatsapp(account):
    """After verification runs, send WhatsApp to each channel officer that has
    missed spots in any schedule for this account. Attaches a missed-spots PDF."""
    from core.whatsapp import notify_missed_spots
    from core.models import MatchResult

    officers = ChannelOfficer.objects.filter(account=account).select_related('account')
    for officer in officers:
        wa = officer.effective_whatsapp
        if not wa:
            continue

        # Latest schedule for this channel
        schedule = Schedule.objects.filter(
            account=account, channel=officer.channel,
        ).order_by('-uploaded_at').first()

        schedule_pk     = schedule.pk             if schedule else 0
        month           = schedule.month          if schedule else ''
        schedule_number = schedule.schedule_number if schedule else ''
        start_date      = schedule.start_date      if schedule else None
        end_date        = schedule.end_date        if schedule else None

        missed_qs = MatchResult.objects.filter(
            account=account,
            channel=officer.channel,
            month=month,
            status__in=['not_aired', 'no_mapping'],
        )
        missed_count = missed_qs.count()
        if missed_count == 0:
            continue

        # Build PDF bytes for attachment
        pdf_bytes = b''
        try:
            all_rows = list(MatchResult.objects.filter(
                account=account,
                channel=officer.channel,
                month=month,
                status__in=['not_aired', 'no_mapping', 'programme_mismatch', 'late_telecast'],
            ).order_by('scheduled_date', 'brand'))
            pdf_bytes = _build_missed_ad_pdf(
                str(account), officer.channel, month, all_rows,
                schedule_number=schedule_number or '',
            )
        except Exception as e:
            print(f'[whatsapp notify] PDF generation failed (non-fatal): {e}')

        notify_missed_spots(
            officer_whatsapp=wa,
            account_name=str(account),
            channel=officer.channel,
            month=month,
            missed_count=missed_count,
            schedule_pk=schedule_pk,
            account_id=account.id,
            officer_name=officer.name,
            schedule_number=schedule_number or '',
            start_date=start_date,
            end_date=end_date,
            pdf_bytes=pdf_bytes,
        )


def _notify_reconciliation_done_whatsapp(account_id, channel, month, result):
    """After TC reconciliation, notify operations officers assigned to this channel."""
    from core.whatsapp import notify_reconciliation_done

    account = Account.objects.filter(id=account_id).first()
    if not account:
        return

    # Find operations users with WhatsApp numbers who have access to this account
    ops_users = User.objects.filter(
        role='operations',
        whatsapp_number__gt='',
        accounts=account,
    )
    for ops_user in ops_users:
        notify_reconciliation_done(
            ops_whatsapp=ops_user.whatsapp_number,
            account_name=str(account),
            channel=channel,
            month=month,
            matched=result.get('matched', 0),
            extra=result.get('extra', 0),
            lmrb_confirmed=result.get('lmrb_confirmed', 0),
            account_id=account_id,
        )


# ═══════════════════════════════════════════════════════════════════════════════
# Channel Officer — Self-service dashboard
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@role_required(['channel_officer'])
def channel_officer_dashboard(request):
    """
    Landing page for channel officers after login.
    Supports multiple accounts AND channels per officer (via multiple ChannelOfficer records).
    """
    user = request.user

    # All profiles for this officer (each = one account+channel combo)
    profiles = list(
        ChannelOfficer.objects.filter(user=user).select_related('account').order_by('account__name', 'channel')
    )

    # Sync User.accounts from profiles for legacy compatibility
    user_accounts = user.accounts.all().order_by('name')
    if not user_accounts.exists() and profiles:
        for p in profiles:
            user.accounts.add(p.account)
        user_accounts = user.accounts.all().order_by('name')

    # Resolve selected account
    filter_account_id = request.GET.get('account_id', '').strip()
    selected_account = None
    if filter_account_id:
        selected_account = user_accounts.filter(pk=filter_account_id).first()
    if selected_account is None and user_accounts.exists():
        selected_account = user_accounts.first()

    # Channels available for the selected account
    account_channels = []
    if selected_account:
        account_channels = [p.channel for p in profiles if p.account_id == selected_account.id]

    # Resolve selected channel
    filter_channel = request.GET.get('channel', '').strip()
    selected_channel = filter_channel if filter_channel in account_channels else (account_channels[0] if account_channels else '')

    # Active profile for selected account+channel
    profile = next((p for p in profiles if p.account_id == (selected_account.id if selected_account else None) and p.channel == selected_channel), None)
    if profile is None and profiles:
        profile = profiles[0]

    schedules  = []
    tc_reports = []
    summary    = {}

    if selected_account and selected_channel:
        from django.db.models import Sum as _Sum
        schedules = (
            Schedule.objects
            .filter(account=selected_account, channel=selected_channel)
            .order_by('-start_date', '-uploaded_at')
        )
        tc_reports = (
            TransmissionReport.objects
            .filter(account=selected_account, channel=selected_channel)
            .order_by('-uploaded_at')
        )
        total_spots     = schedules.aggregate(t=_Sum('row_count'))['t'] or 0
        latest_schedule = schedules.first()
        summary = {
            'schedule_count': schedules.count(),
            'total_spots':    total_spots,
            'tc_count':       tc_reports.count(),
            'latest_month':   latest_schedule.month    if latest_schedule else '',
            'latest_end':     latest_schedule.end_date if latest_schedule else None,
        }

    return render(request, 'channel_officers/dashboard.html', {
        'profile':           profile,
        'profiles':          profiles,
        'user_accounts':     user_accounts,
        'selected_account':  selected_account,
        'account_channels':  account_channels,
        'selected_channel':  selected_channel,
        'schedules':         schedules,
        'tc_reports':        tc_reports,
        'summary':           summary,
    })


@login_required
def tc_report_download(request, pk):
    """Serve the original TC file as a download (channel officers + all staff)."""
    user   = request.user
    report = get_object_or_404(TransmissionReport, pk=pk)

    # Channel officers may only download TCs for their own channel/accounts
    if user.role == 'channel_officer':
        allowed_accounts = list(user.accounts.values_list('id', flat=True))
        allowed_channels = list(ChannelOfficer.objects.filter(user=user).values_list('channel', flat=True))
        if not allowed_channels or report.account_id not in allowed_accounts or report.channel not in allowed_channels:
            return HttpResponse('Access denied', status=403)
    elif not _account_access(user, report.account_id):
        return HttpResponse('Access denied', status=403)

    try:
        file_path = report.file.path
        if not os.path.exists(file_path):
            return HttpResponse('File not found on server.', status=404)
        return FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename=report.original_filename or os.path.basename(file_path),
        )
    except Exception as e:
        return HttpResponse(f'Download failed: {e}', status=500)


# ═══════════════════════════════════════════════════════════════════════════════
# Channel Officer Management
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@role_required(['super_admin', 'admin'])
def channel_officer_list(request):
    """List all channel officers with their account/channel assignments."""
    officers = ChannelOfficer.objects.select_related('account', 'user').order_by('account__name', 'channel')
    accounts = Account.objects.all().order_by('name')
    return render(request, 'channel_officers/list.html', {
        'officers': officers,
        'accounts': accounts,
    })


@login_required
@role_required(['super_admin', 'admin'])
def channel_officer_create(request):
    """Create a new channel officer + their system login account."""
    import secrets as _secrets
    import string as _string

    accounts = Account.objects.all().order_by('name')
    channels = Channel.objects.all().order_by('name')

    if request.method == 'POST':
        name            = request.POST.get('name', '').strip()
        whatsapp        = request.POST.get('whatsapp_number', '').strip()
        account_ids     = request.POST.getlist('account_id[]')
        channels_list   = request.POST.getlist('channel[]')
        create_login    = request.POST.get('create_login') == '1'
        email           = request.POST.get('email', '').strip().lower()

        # Pair up account_ids and channels, skip blank rows
        assignments = [
            (aid.strip(), ch.strip())
            for aid, ch in zip(account_ids, channels_list)
            if aid.strip() and ch.strip()
        ]

        errors = []
        if not name:
            errors.append('Name is required.')
        if not assignments:
            errors.append('At least one Account / Channel assignment is required.')

        # Validate each assignment
        account_objs = {}
        for aid, ch in assignments:
            acc = Account.objects.filter(id=aid).first()
            if not acc:
                errors.append(f'Invalid account ID: {aid}')
            else:
                account_objs[aid] = acc

        login_user = None
        temp_password = None
        existing_user = None
        if create_login:
            if not email:
                errors.append('Email is required when creating a login.')
            else:
                existing_user = User.objects.filter(email=email).first()
                if existing_user:
                    # Allow reuse only if orphaned (no remaining officer profiles)
                    if ChannelOfficer.objects.filter(user=existing_user).exists():
                        errors.append(
                            f'{email} is already linked to another active officer. '
                            f'Edit that officer instead.'
                        )
                    # else: orphaned user — we'll reuse/reactivate below
                else:
                    chars = _string.ascii_letters + _string.digits + '!@#'
                    temp_password = ''.join(_secrets.choice(chars) for _ in range(12))

        if not errors:
            if create_login:
                if existing_user:
                    # Reactivate orphaned user with fresh password
                    chars = _string.ascii_letters + _string.digits + '!@#'
                    temp_password = ''.join(_secrets.choice(chars) for _ in range(12))
                    existing_user.name = name
                    existing_user.whatsapp_number = whatsapp
                    existing_user.role = 'channel_officer'
                    existing_user.is_active = True
                    existing_user.must_change_password = True
                    existing_user.set_password(temp_password)
                    existing_user.save()
                    login_user = existing_user
                else:
                    login_user = User.objects.create_user(
                        email=email,
                        name=name,
                        password=temp_password,
                        role='channel_officer',
                        created_by=request.user,
                        must_change_password=True,
                    )
                    login_user.whatsapp_number = whatsapp
                    login_user.save()

            # Create one ChannelOfficer per assignment
            for aid, ch in assignments:
                acc = account_objs.get(aid)
                if acc:
                    ChannelOfficer.objects.create(
                        account=acc,
                        channel=ch,
                        name=name,
                        whatsapp_number=whatsapp,
                        user=login_user,
                    )

            assignment_labels = ', '.join(
                f'{account_objs[aid].name}/{ch}' for aid, ch in assignments if aid in account_objs
            )
            msg = f'Officer "{name}" created for {assignment_labels}.'
            if temp_password:
                msg += f' Login: {email} / Temp password: <code class="font-mono font-bold">{temp_password}</code>'
            messages.success(request, msg)

            # Send WhatsApp welcome — use first assignment for account/channel display
            if whatsapp:
                try:
                    from core.whatsapp import notify_new_officer_created
                    base_url = get_setting('whatsapp_app_base_url', '').rstrip('/')
                    login_url = f'{base_url}/auth/login/' if base_url else '/auth/login/'
                    first_aid, first_ch = assignments[0]
                    first_acc = account_objs.get(first_aid)
                    wa_ok = notify_new_officer_created(
                        whatsapp=whatsapp,
                        name=name,
                        account_name=str(first_acc) if first_acc else '',
                        channel=first_ch,
                        email=email if create_login else '',
                        password=temp_password or '',
                        login_url=login_url,
                    )
                    if not wa_ok:
                        messages.warning(request,
                            f'Officer created but WhatsApp failed to send to {whatsapp}. '
                            f'If your Meta app is in Development mode, you must add {whatsapp} '
                            f'as a test recipient on developers.facebook.com → WhatsApp → API Setup.')
                except Exception as _e:
                    logger.warning('[WhatsApp] officer welcome failed (non-fatal): %s', _e)
                    messages.warning(request,
                        f'Officer created but WhatsApp notification failed: {_e}')

            return redirect('/dashboard/channel-officers/')

        import json as _json
        return render(request, 'channel_officers/form.html', {
            'accounts': accounts,
            'channels': channels,
            'errors':   errors,
            'fd': {
                'name':              request.POST.get('name', ''),
                'whatsapp_number':   request.POST.get('whatsapp_number', ''),
                'email':             request.POST.get('email', ''),
                'account_ids_json':  _json.dumps(account_ids),
                'channels_list_json': _json.dumps(channels_list),
            },
        })

    return render(request, 'channel_officers/form.html', {
        'accounts': accounts,
        'channels': channels,
        'fd': {'name': '', 'whatsapp_number': '', 'email': '', 'account_ids_json': '[]', 'channels_list_json': '[]'},
    })


@login_required
@role_required(['super_admin', 'admin'])
def channel_officer_edit(request, pk):
    """Edit or delete a channel officer."""
    officer  = get_object_or_404(ChannelOfficer, pk=pk)
    accounts = Account.objects.all().order_by('name')
    channels = Channel.objects.all().order_by('name')

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'delete':
            name = officer.name
            linked_user = officer.user
            officer.delete()
            # Delete the linked User only if they have no other officer profiles
            if linked_user:
                remaining = ChannelOfficer.objects.filter(user=linked_user).count()
                if remaining == 0:
                    linked_user.delete()
            messages.success(request, f'Officer "{name}" removed.')
            return redirect('/dashboard/channel-officers/')

        name       = request.POST.get('name', '').strip()
        whatsapp   = request.POST.get('whatsapp_number', '').strip()
        account_id = request.POST.get('account_id', '').strip()
        channel    = request.POST.get('channel', '').strip()

        if name and account_id and channel:
            account_obj = Account.objects.filter(id=account_id).first()
            if account_obj:
                officer.name            = name
                officer.whatsapp_number = whatsapp
                officer.account         = account_obj
                officer.channel         = channel
                officer.save()
                # Sync WhatsApp to linked user too
                if officer.user:
                    officer.user.whatsapp_number = whatsapp
                    officer.user.save()
                messages.success(request, f'Officer "{name}" updated.')
                return redirect('/dashboard/channel-officers/')

    return render(request, 'channel_officers/form.html', {
        'officer':  officer,
        'accounts': accounts,
        'channels': channels,
        'edit':     True,
        'fd': {'name': '', 'whatsapp_number': '', 'account_id': '', 'channel': '', 'email': ''},
    })


@login_required
@role_required(['super_admin', 'admin'])
def whatsapp_test(request):
    """Send a test WhatsApp message and show the raw API response for debugging."""
    import requests as _req
    from core.models import get_setting, get_setting_int
    from core import whatsapp as wa

    token    = get_setting('whatsapp_access_token', '')
    phone_id = get_setting('whatsapp_phone_number_id', '')
    # Accept ?to= param from test UI; fall back to saved test_number setting
    to = (request.GET.get('to', '') or get_setting('whatsapp_test_number', '')).strip().replace(' ', '')
    enabled  = get_setting_int('whatsapp_enabled', 0)
    base_url = get_setting('whatsapp_app_base_url', 'https://ad-monitoring-production.up.railway.app')

    if not to:
        messages.error(request, 'Enter a WhatsApp number in the test field first.')
        return redirect('/dashboard/settings/')
    if not token:
        messages.error(request, 'No Access Token set in Settings.')
        return redirect('/dashboard/settings/')
    if not phone_id:
        messages.error(request, 'No Phone Number ID set in Settings.')
        return redirect('/dashboard/settings/')
    if not enabled:
        messages.error(request, 'WhatsApp is disabled in Settings. Toggle Enabled to ON first.')
        return redirect('/dashboard/settings/')

    tmpl = request.GET.get('tmpl', '')
    login_url = base_url.rstrip('/') + '/auth/login/'

    if tmpl == 'officer_welcome':
        ok = wa.notify_new_officer_created(
            whatsapp=to,
            name='Test Officer',
            account_name='Ogilvy Nova',
            channel='Test TV',
            email='officer@test.com',
            password='TempPass123',
            login_url=login_url,
        )
        label = 'officer_welcome'

    elif tmpl == 'user_welcome':
        ok = wa.notify_new_user_created(
            whatsapp=to,
            name='Test User',
            email='user@test.com',
            password='TempPass123',
            role_label='Operations',
            login_url=login_url,
        )
        label = 'user_welcome'

    elif tmpl == 'missed_spots':
        ok = wa.notify_missed_spots(
            officer_whatsapp=to,
            account_name='Ogilvy Nova',
            channel='Test TV',
            month='March 2025',
            missed_count=5,
            schedule_pk=1,
            account_id=1,
            officer_name='Test Officer',
            schedule_number='101',
        )
        label = 'missed_spots_alert'

    else:
        # Plain text connectivity test
        to_clean = to.lstrip('+')
        url = f'https://graph.facebook.com/v19.0/{phone_id}/messages'
        payload = {
            'messaging_product': 'whatsapp',
            'to': to_clean,
            'type': 'text',
            'text': {'body': 'Test from Ad-Monitor. WhatsApp integration is working!', 'preview_url': False},
        }
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
        }
        try:
            resp = _req.post(url, json=payload, headers=headers, timeout=10)
            body = resp.json()
            if resp.status_code == 200 and 'messages' in body:
                msg_id = body['messages'][0].get('id', '?')
                messages.success(request,
                    f'Message accepted by Meta. Message ID: {msg_id}. '
                    f'Sending to: {to_clean}. Check your WhatsApp.')
            else:
                err = body.get('error', {})
                messages.error(request,
                    f'Meta API error {resp.status_code}: '
                    f'[{err.get("code","?")}] {err.get("message", resp.text)}. '
                    f'Recipient: {to_clean}')
        except Exception as exc:
            messages.error(request, f'Request failed: {exc}')
        return redirect('/dashboard/settings/')

    if ok:
        messages.success(request,
            f'✓ Template "{label}" sent to {to}. Check your WhatsApp.')
    else:
        messages.error(request,
            f'✗ Template "{label}" failed — check that the template is approved in Meta and '
            f'that the test number ({to}) is added as a recipient in the Meta API Setup page.')
    return redirect('/dashboard/settings/')


@login_required
@role_required(['super_admin', 'admin'])
def whatsapp_send_tc_reminders(request):
    """Manually trigger TC upload reminders for all channels whose schedules have ended
    and have no TC uploaded yet."""
    from core.whatsapp import notify_tc_upload_reminder
    from django.utils import timezone as _tz
    today = _tz.now().date()

    sent = 0
    for officer in ChannelOfficer.objects.select_related('account').all():
        wa = officer.effective_whatsapp
        if not wa:
            continue
        # Find schedules for this officer's channel that have ended but have no TC
        ended_schedules = Schedule.objects.filter(
            account=officer.account,
            channel=officer.channel,
            end_date__lt=today,
        ).exclude(
            transmission_reports__isnull=False,
        ).order_by('-end_date')

        for sch in ended_schedules:
            notify_tc_upload_reminder(
                officer_whatsapp=wa,
                account_name=str(officer.account),
                channel=officer.channel,
                month=sch.month,
                schedule_pk=sch.pk,
                account_id=officer.account_id,
                end_date=sch.end_date,
            )
            sent += 1

    messages.success(request, f'TC upload reminders sent: {sent} message(s).')
    return redirect('/dashboard/channel-officers/')


# ── WhatsApp Webhook (incoming message handler) ────────────────────────────────

@csrf_exempt
def whatsapp_webhook(request):
    """
    GET  — Meta webhook verification challenge
    POST — Receive incoming WhatsApp messages (for officer auto-registration)

    Setup in Meta Developer Console:
      WhatsApp → Configuration → Webhook URL: https://yourapp.railway.app/dashboard/whatsapp/webhook/
      Verify Token: same value as whatsapp_webhook_verify_token in Settings
      Subscribe to: messages
    """
    if request.method == 'GET':
        # Meta sends hub.mode, hub.verify_token, hub.challenge
        mode      = request.GET.get('hub.mode', '')
        token     = request.GET.get('hub.verify_token', '')
        challenge = request.GET.get('hub.challenge', '')
        expected  = get_setting('whatsapp_webhook_verify_token', '')

        # Debug: plain GET with no hub.mode → show status (helps confirm endpoint is live)
        if not mode and not token and not challenge:
            configured = bool(expected)
            return JsonResponse({
                'status': 'webhook endpoint is live',
                'verify_token_configured': configured,
                'hint': ('Set whatsapp_webhook_verify_token in /dashboard/settings/ '
                         'then use the same value as the Verify Token in Meta.') if not configured else
                        'Token is set. Use the same value in Meta Webhook Verify Token field.',
            })

        if mode == 'subscribe' and token == expected and expected:
            logger.info('[WhatsApp webhook] verification OK')
            return HttpResponse(challenge, content_type='text/plain')
        logger.warning('[WhatsApp webhook] verification failed: token=%s expected=%s', token, expected)
        return HttpResponse('Forbidden', status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except Exception:
            return HttpResponse('Bad JSON', status=400)

        try:
            for entry in data.get('entry', []):
                for change in entry.get('changes', []):
                    value = change.get('value', {})
                    for msg in value.get('messages', []):
                        _handle_incoming_whatsapp(msg)
        except Exception as exc:
            logger.error('[WhatsApp webhook] processing error: %s', exc)

        # Meta requires a 200 OK regardless of processing outcome
        return JsonResponse({'status': 'ok'})

    return HttpResponse('Method not allowed', status=405)


def _handle_incoming_whatsapp(msg: dict):
    """Process one incoming WhatsApp message — handle registration for officers and staff users."""
    from core.whatsapp import send_text
    from core.models import get_setting

    from_number = msg.get('from', '').strip()
    if not from_number:
        return

    # Normalise: add + if missing
    from_number_lookup = ('+' + from_number) if not from_number.startswith('+') else from_number

    msg_type = msg.get('type', '')
    msg_text = ''
    if msg_type == 'text':
        msg_text = msg.get('text', {}).get('body', '').strip().upper()

    # ── 1. Check ChannelOfficer ────────────────────────────────────────────────
    officer = (
        ChannelOfficer.objects.filter(whatsapp_number=from_number_lookup).first()
        or ChannelOfficer.objects.filter(whatsapp_number=from_number).first()
    )
    if officer:
        if not officer.is_whatsapp_registered:
            # Welcome message was already sent at creation — this is their reply/confirm
            officer.is_whatsapp_registered = True
            officer.save(update_fields=['is_whatsapp_registered'])
            send_text(
                from_number_lookup,
                f'✅ *{officer.name}* — you are now registered!\n\n'
                f'You will receive Ad-Monitoring alerts for *{officer.channel}* on this number.\n\n'
                f'_(Reply *HELP* any time for support)_',
            )
            logger.info('[WhatsApp webhook] channel officer registered: %s (%s)', officer.name, from_number_lookup)
        else:
            if msg_text in ('CONFIRM', 'YES', 'OK', 'HELP'):
                send_text(
                    from_number_lookup,
                    f'✅ *{officer.name}* — your WhatsApp notifications are already active for *{officer.channel}*.',
                )
        return

    # ── 2. Check staff User ────────────────────────────────────────────────────
    staff_user = (
        User.objects.filter(whatsapp_number=from_number_lookup).first()
        or User.objects.filter(whatsapp_number=from_number).first()
    )
    if staff_user:
        if not staff_user.is_whatsapp_registered:
            staff_user.is_whatsapp_registered = True
            staff_user.save(update_fields=['is_whatsapp_registered'])
            company = get_setting('whatsapp_company_name', 'Ogilvy Nova')
            send_text(
                from_number_lookup,
                f'✅ *{staff_user.name}* — you are now registered with *{company}*!\n\n'
                f'Your WhatsApp notifications are active.\n\n'
                f'Log in at: {get_setting("whatsapp_app_base_url", "").rstrip("/")}/auth/login/',
            )
            logger.info('[WhatsApp webhook] staff user registered: %s (%s)', staff_user.name, from_number_lookup)
        else:
            if msg_text in ('CONFIRM', 'YES', 'OK', 'HELP'):
                send_text(
                    from_number_lookup,
                    f'✅ *{staff_user.name}* — your WhatsApp notifications are already active.',
                )
        return

    # ── 3. Unknown number — send a polite "not registered" reply ──────────────
    logger.info('[WhatsApp webhook] message from unknown number %s — no matching user/officer', from_number)
    company = get_setting('whatsapp_company_name', 'Ogilvy Nova')
    send_text(
        from_number_lookup,
        f'👋 Hi! This is the *{company}* monitoring system.\n\n'
        f'Your number is not registered in our system. '
        f'If you believe this is an error, please contact your administrator.\n\n'
        f'_(Automated message from {company})_',
    )
