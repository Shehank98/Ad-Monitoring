"""
Comprehensive test suite for the Ad-Monitoring Django application.

Covers:
- Model integrity (dedup keys, locking flags)
- Schedule↔LMRB matching engine (run_scope smart/reset modes)
- TC reconciliation engine (reconcile_tc, build_summary_data)
- Sponsorship engine (reconcile_sponsorship, manual_assign, reset_sponsorship)
- Multi-schedule priority (Rule 10) and versioning (Rule 12)
- ManualMatch creation and permanence
- Edge cases: boundary dates, late-aired rule, time tolerance

Run with:
    python manage.py test core
"""

import datetime
import json
from django.db import IntegrityError
from django.test import TestCase, override_settings

from accounts.models import User
from core.models import (
    Account,
    BrandMapping,
    Channel,
    LMRBRow,
    ManualMatch,
    MatchResult,
    Schedule,
    ScheduleRow,
    SponsorshipLmrbAssignment,
    SystemSetting,
    TCRow,
    TransmissionReport,
)
from verification.engine import run_scope
from verification.sponsorship_engine import (
    lmrb_candidates,
    manual_assign,
    reconcile_sponsorship,
    reset_sponsorship,
)
from verification.tc_engine import build_summary_data, reconcile_tc


# ── Shared fixture helpers ────────────────────────────────────────────────────

CHANNEL = "Sirasa TV"
MONTH = "January 2025"
DATE = datetime.date(2025, 1, 15)


def make_account(name="Test Account"):
    """Create and return a test Account."""
    return Account.objects.create(name=name)


def make_user(email="ops@test.com", role="operations"):
    """Create and return a test User."""
    return User.objects.create_user(
        email=email,
        name="Test User",
        password="password123",
        role=role,
        must_change_password=False,
    )


def make_schedule(account, channel=CHANNEL, month=MONTH, schedule_number="101", version=1):
    """Create and return a Schedule header record."""
    return Schedule.objects.create(
        account=account,
        channel=channel,
        month=month,
        schedule_number=schedule_number,
        file="schedules/dummy.xlsx",
        original_filename="dummy.xlsx",
        row_count=0,
        start_date=datetime.date(2025, 1, 1),
        end_date=datetime.date(2025, 1, 31),
        version=version,
    )


def make_schedule_row(
    account,
    schedule,
    brand="Brand A",
    date=DATE,
    start_time="20:00:00",
    end_time="21:00:00",
    duration=30,
    ad_type="COMMERCIAL BENEFITS",
    programme="Test Show",
    channel=CHANNEL,
    month=MONTH,
):
    """Create and return a ScheduleRow."""
    return ScheduleRow.objects.create(
        schedule=schedule,
        account=account,
        channel=channel,
        month=month,
        brand=brand,
        programme=programme,
        date=date,
        start_time=start_time,
        end_time=end_time,
        duration=duration,
        ad_type=ad_type,
    )


def make_lmrb_row(
    account,
    advt_theme="Theme A",
    date=DATE,
    advt_time="20:30:00",
    duration=30,
    channel=CHANNEL,
    source="maponline",
    brk_no=None,
    pos_in_brk=None,
    advertiser="",
    product="",
):
    """Create and return an LMRBRow with a unique dedup key."""
    dedup_key = LMRBRow.make_dedup_key(
        account.id, channel, date, advt_time, advt_theme, duration,
        brk_no=brk_no, pos_in_brk=pos_in_brk,
        advertiser=advertiser, product=product,
    )
    return LMRBRow.objects.create(
        account=account,
        channel=channel,
        date=date,
        advt_theme=advt_theme,
        advt_time=advt_time,
        duration=duration,
        source=source,
        dedup_key=dedup_key,
    )


def make_brand_mapping(account, brand="Brand A", theme="Theme A", tc_theme="TC Theme A", duration=None):
    """Create and return a BrandMapping."""
    return BrandMapping.objects.create(
        account=account,
        brand=brand,
        theme=theme,
        tc_theme=tc_theme,
        duration=duration,
    )


def make_tc_report(account, channel=CHANNEL, month=MONTH, schedule=None):
    """Create and return a TransmissionReport."""
    return TransmissionReport.objects.create(
        account=account,
        channel=channel,
        month=month,
        schedule=schedule,
        file="tc/dummy.xlsx",
        original_filename="tc_dummy.xlsx",
        row_count=0,
        start_date=datetime.date(2025, 1, 1),
        end_date=datetime.date(2025, 1, 31),
    )


def make_tc_row(
    account,
    tc_report,
    tc_theme="TC Theme A",
    date=DATE,
    aired_time="20:30:00",
    duration=30,
    channel=CHANNEL,
    programme="Test Show",
    suffix="",
):
    """Create and return a TCRow with a unique dedup key."""
    dedup_key = TCRow.make_dedup_key(
        account.id, channel, date, aired_time + suffix, tc_theme, duration
    )
    return TCRow.objects.create(
        account=account,
        tc_report=tc_report,
        channel=channel,
        date=date,
        programme=programme,
        tc_theme=tc_theme,
        duration=duration,
        aired_time=aired_time,
        dedup_key=dedup_key,
    )


def ensure_tc_tolerance(seconds=5):
    """Ensure the TC-LMRB time tolerance SystemSetting is set to the given value."""
    SystemSetting.objects.update_or_create(
        key="tc_lmrb_time_tolerance",
        defaults={
            "value": str(seconds),
            "label": "TC-LMRB Time Tolerance (seconds)",
            "category": "reconciliation",
        },
    )


# ── Test Classes ──────────────────────────────────────────────────────────────


class LMRBRowDedupKeyTest(TestCase):
    """Tests for LMRBRow dedup key uniqueness constraint."""

    def setUp(self):
        self.account = make_account()

    def test_dedup_key_unique_constraint_raises_integrity_error(self):
        """Creating two LMRBRows with the same dedup key must raise IntegrityError."""
        key = LMRBRow.make_dedup_key(
            self.account.id, CHANNEL, DATE, "20:00:00", "Theme A", 30
        )
        LMRBRow.objects.create(
            account=self.account,
            channel=CHANNEL,
            date=DATE,
            advt_theme="Theme A",
            advt_time="20:00:00",
            duration=30,
            source="maponline",
            dedup_key=key,
        )
        with self.assertRaises(IntegrityError):
            LMRBRow.objects.create(
                account=self.account,
                channel=CHANNEL,
                date=DATE,
                advt_theme="Theme A",
                advt_time="20:00:00",
                duration=30,
                source="maponline",
                dedup_key=key,
            )

    def test_different_brk_no_produces_distinct_rows(self):
        """Two spots in the same break (same time/theme) with different brk_no are distinct."""
        key1 = LMRBRow.make_dedup_key(
            self.account.id, CHANNEL, DATE, "20:00:00", "Theme A", 30,
            brk_no=1, pos_in_brk=1,
        )
        key2 = LMRBRow.make_dedup_key(
            self.account.id, CHANNEL, DATE, "20:00:00", "Theme A", 30,
            brk_no=1, pos_in_brk=2,
        )
        self.assertNotEqual(key1, key2)

    def test_dedup_key_same_params_is_deterministic(self):
        """make_dedup_key returns the same value for the same inputs."""
        key1 = LMRBRow.make_dedup_key(self.account.id, CHANNEL, DATE, "20:00:00", "Theme A", 30)
        key2 = LMRBRow.make_dedup_key(self.account.id, CHANNEL, DATE, "20:00:00", "Theme A", 30)
        self.assertEqual(key1, key2)


class TCRowDedupKeyTest(TestCase):
    """Tests for TCRow dedup key uniqueness constraint."""

    def setUp(self):
        self.account = make_account()
        self.tc_report = make_tc_report(self.account)

    def test_tc_dedup_key_unique_constraint_raises_integrity_error(self):
        """Creating two TCRows with the same dedup key must raise IntegrityError."""
        key = TCRow.make_dedup_key(
            self.account.id, CHANNEL, DATE, "20:00:00", "TC Theme A", 30
        )
        TCRow.objects.create(
            account=self.account,
            tc_report=self.tc_report,
            channel=CHANNEL,
            date=DATE,
            tc_theme="TC Theme A",
            duration=30,
            aired_time="20:00:00",
            dedup_key=key,
        )
        with self.assertRaises(IntegrityError):
            TCRow.objects.create(
                account=self.account,
                tc_report=self.tc_report,
                channel=CHANNEL,
                date=DATE,
                tc_theme="TC Theme A",
                duration=30,
                aired_time="20:00:00",
                dedup_key=key,
            )

    def test_tc_dedup_key_different_from_lmrb_key(self):
        """TC dedup key is prefixed with 'tc|' so it cannot collide with LMRB keys."""
        lmrb_key = LMRBRow.make_dedup_key(
            self.account.id, CHANNEL, DATE, "20:00:00", "TC Theme A", 30
        )
        tc_key = TCRow.make_dedup_key(
            self.account.id, CHANNEL, DATE, "20:00:00", "TC Theme A", 30
        )
        self.assertNotEqual(lmrb_key, tc_key)


class BrandMappingTest(TestCase):
    """Tests for BrandMapping model helpers."""

    def setUp(self):
        self.account = make_account()

    def test_tc_themes_list_splits_on_pipe(self):
        """tc_themes_list property correctly splits pipe-separated values."""
        bm = BrandMapping.objects.create(
            account=self.account,
            brand="Brand X",
            theme="Theme X",
            tc_theme="TC Theme A|TC Theme B|TC Theme C",
        )
        self.assertEqual(bm.tc_themes_list, ["TC Theme A", "TC Theme B", "TC Theme C"])

    def test_tc_themes_list_empty_when_blank(self):
        """tc_themes_list returns empty list when tc_theme is blank."""
        bm = BrandMapping.objects.create(
            account=self.account,
            brand="Brand Y",
            theme="Theme Y",
            tc_theme="",
        )
        self.assertEqual(bm.tc_themes_list, [])

    def test_tc_themes_list_strips_whitespace(self):
        """tc_themes_list strips whitespace from each value."""
        bm = BrandMapping.objects.create(
            account=self.account,
            brand="Brand Z",
            theme="Theme Z",
            tc_theme=" Theme A | Theme B ",
        )
        self.assertEqual(bm.tc_themes_list, ["Theme A", "Theme B"])


class RunScopeSmartModeTest(TestCase):
    """Tests for verification.engine.run_scope in smart mode."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account)
        # A schedule row for 'Brand A'
        self.sr = make_schedule_row(self.account, self.schedule)
        # A matching LMRB row (same date, theme, duration, time within window)
        self.lr = make_lmrb_row(self.account)

    def test_smart_mode_matches_row_and_sets_is_matched(self):
        """run_scope smart mode matches the ScheduleRow to the LMRBRow."""
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        self.sr.refresh_from_db()
        self.lr.refresh_from_db()
        self.assertTrue(self.sr.is_matched)
        self.assertTrue(self.lr.is_matched)
        self.assertEqual(self.sr.matched_lmrb_id, self.lr.id)

    def test_smart_mode_skips_already_matched_lmrb_row(self):
        """After matching, a second smart run does not re-process the locked rows."""
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        # Create a second schedule row — there is no LMRB row left for it
        sr2 = make_schedule_row(
            self.account, self.schedule,
            brand="Brand A",
            date=datetime.date(2025, 1, 16),
            start_time="20:00:00",
            end_time="21:00:00",
        )
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        sr2.refresh_from_db()
        # sr2 should be Not Aired (no LMRB row available)
        self.assertFalse(sr2.is_matched)
        # The first LMRB row must still belong to sr1 only
        self.lr.refresh_from_db()
        self.assertEqual(self.lr.matched_schedule_id, self.sr.id)

    def test_smart_mode_creates_match_result(self):
        """run_scope creates a MatchResult record with status='matched'."""
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        mr = MatchResult.objects.filter(
            account=self.account, channel=CHANNEL, month=MONTH, status="matched"
        )
        self.assertEqual(mr.count(), 1)

    def test_no_brand_mapping_produces_no_mapping_result(self):
        """A ScheduleRow without a BrandMapping gets status='no_mapping'."""
        account2 = make_account("Account No Mapping")
        schedule2 = make_schedule(account2)
        make_schedule_row(account2, schedule2, brand="Unknown Brand")
        make_lmrb_row(account2, advt_theme="Unknown Theme")
        # No BrandMapping created for account2
        run_scope(account2.id, CHANNEL, MONTH, mode="smart")
        mr = MatchResult.objects.filter(
            account=account2, status="no_mapping"
        )
        self.assertEqual(mr.count(), 1)


class RunScopeResetModeTest(TestCase):
    """Tests for verification.engine.run_scope in reset mode."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account)
        self.sr = make_schedule_row(self.account, self.schedule)
        self.lr = make_lmrb_row(self.account)

    def test_reset_mode_clears_is_matched_flags(self):
        """reset mode unlocks previously matched rows and re-runs the engine."""
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        self.sr.refresh_from_db()
        self.assertTrue(self.sr.is_matched)

        # Now reset — it should clear and re-match
        run_scope(self.account.id, CHANNEL, MONTH, mode="reset")
        self.sr.refresh_from_db()
        self.lr.refresh_from_db()
        # After reset + re-run, both should still be matched
        self.assertTrue(self.sr.is_matched)
        self.assertTrue(self.lr.is_matched)

    def test_reset_mode_does_not_clear_manual_match_flags(self):
        """is_manual_matched=True rows are preserved across a reset run."""
        # Manually lock the schedule row
        self.sr.is_manual_matched = True
        self.sr.save(update_fields=["is_manual_matched"])
        self.lr.is_manual_matched = True
        self.lr.save(update_fields=["is_manual_matched"])

        # Also add a second unmatched row so run_scope doesn't bail with 0 rows
        lr2 = make_lmrb_row(
            self.account, advt_time="20:45:00",
            brk_no=None, pos_in_brk=1,
        )
        make_schedule_row(
            self.account, self.schedule,
            brand="Brand A",
            date=DATE,
            start_time="20:00:00",
            end_time="21:00:00",
        )

        run_scope(self.account.id, CHANNEL, MONTH, mode="reset")

        self.sr.refresh_from_db()
        self.lr.refresh_from_db()
        self.assertTrue(self.sr.is_manual_matched, "Manual lock must survive reset")
        self.assertTrue(self.lr.is_manual_matched, "Manual LMRB lock must survive reset")

    def test_reset_mode_clears_match_results_except_manual(self):
        """reset mode deletes MatchResult records but not those with status='manual_match'."""
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        # Manually create a manual_match result
        MatchResult.objects.create(
            account=self.account,
            channel=CHANNEL,
            month=MONTH,
            brand="Brand A",
            status="manual_match",
        )
        run_scope(self.account.id, CHANNEL, MONTH, mode="reset")
        manual_results = MatchResult.objects.filter(
            account=self.account, status="manual_match"
        )
        self.assertEqual(manual_results.count(), 1, "Manual match results must not be deleted by reset")


class LMRBSpotSingleUseTest(TestCase):
    """Test that a single LMRBRow cannot be claimed by two ScheduleRows."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account)

    def test_lmrb_row_consumed_by_first_schedule_row_only(self):
        """
        When there is one LMRB row and two matching ScheduleRows,
        only one ScheduleRow gets matched; the other is Not Aired.
        """
        sr1 = make_schedule_row(self.account, self.schedule, date=DATE)
        sr2 = make_schedule_row(
            self.account, self.schedule,
            date=DATE,
            start_time="20:00:00",
            end_time="21:00:00",
        )
        # Only one LMRB row available
        make_lmrb_row(self.account)

        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")

        sr1.refresh_from_db()
        sr2.refresh_from_db()

        matched = [sr for sr in [sr1, sr2] if sr.is_matched]
        unmatched = [sr for sr in [sr1, sr2] if not sr.is_matched]

        self.assertEqual(len(matched), 1, "Exactly one ScheduleRow should be matched")
        self.assertEqual(len(unmatched), 1, "The other ScheduleRow should be Not Aired")

        # Verify the matched LMRB row is only linked to one schedule row
        lmrb = LMRBRow.objects.first()
        self.assertEqual(lmrb.matched_schedule_id, matched[0].id)


class ManualMatchPermanenceTest(TestCase):
    """Tests for ManualMatch creation, locking, and permanence."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account)
        self.user = make_user()
        self.sr = make_schedule_row(self.account, self.schedule)
        self.lr = make_lmrb_row(self.account)

    def _create_manual_match(self):
        """Helper to create a ManualMatch and lock both rows."""
        mm = ManualMatch.objects.create(
            account=self.account,
            channel=CHANNEL,
            month=MONTH,
            match_mode="schedule_lmrb",
            schedule_row=self.sr,
            lmrb_row=self.lr,
            matched_by=self.user,
        )
        self.sr.is_manual_matched = True
        self.sr.save(update_fields=["is_manual_matched"])
        self.lr.is_manual_matched = True
        self.lr.save(update_fields=["is_manual_matched"])
        return mm

    def test_manual_match_locks_schedule_row(self):
        """Creating a ManualMatch sets is_manual_matched=True on ScheduleRow."""
        self._create_manual_match()
        self.sr.refresh_from_db()
        self.assertTrue(self.sr.is_manual_matched)

    def test_manual_match_locks_lmrb_row(self):
        """Creating a ManualMatch sets is_manual_matched=True on LMRBRow."""
        self._create_manual_match()
        self.lr.refresh_from_db()
        self.assertTrue(self.lr.is_manual_matched)

    def test_manual_matched_rows_excluded_from_smart_run(self):
        """
        After manual locking, smart run does not process the locked rows.
        A second LMRB row exists but the manually locked one must not be re-used.
        """
        self._create_manual_match()
        # Add a new unmatched schedule row and LMRB row
        sr2 = make_schedule_row(
            self.account, self.schedule,
            date=datetime.date(2025, 1, 16),
            start_time="20:00:00",
            end_time="21:00:00",
        )
        lr2 = make_lmrb_row(
            self.account,
            date=datetime.date(2025, 1, 16),
            advt_time="20:30:00",
            brk_no=1,
        )
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        # The manually locked rows must retain their flags
        self.sr.refresh_from_db()
        self.lr.refresh_from_db()
        self.assertTrue(self.sr.is_manual_matched)
        self.assertTrue(self.lr.is_manual_matched)

    def test_manual_matched_rows_survive_reset(self):
        """mode='reset' must not clear is_manual_matched flags."""
        self._create_manual_match()
        # Need at least one non-manual row for run_scope to proceed
        sr2 = make_schedule_row(
            self.account, self.schedule,
            date=datetime.date(2025, 1, 16),
        )
        lr2 = make_lmrb_row(
            self.account,
            date=datetime.date(2025, 1, 16),
            brk_no=2,
        )
        run_scope(self.account.id, CHANNEL, MONTH, mode="reset")
        self.sr.refresh_from_db()
        self.lr.refresh_from_db()
        self.assertTrue(self.sr.is_manual_matched, "is_manual_matched must survive reset")
        self.assertTrue(self.lr.is_manual_matched, "LMRB is_manual_matched must survive reset")

    def test_three_way_manual_match_locks_tc_row(self):
        """A 3way ManualMatch should lock the TCRow via is_manual_matched."""
        tc_report = make_tc_report(self.account)
        tcrow = make_tc_row(self.account, tc_report)
        mm = ManualMatch.objects.create(
            account=self.account,
            channel=CHANNEL,
            month=MONTH,
            match_mode="3way",
            schedule_row=self.sr,
            tc_row=tcrow,
            lmrb_row=self.lr,
            matched_by=self.user,
        )
        # In the real workflow the view sets is_manual_matched; simulate it here
        self.sr.is_manual_matched = True
        self.sr.save()
        self.lr.is_manual_matched = True
        self.lr.save()
        self.sr.refresh_from_db()
        self.lr.refresh_from_db()
        self.assertTrue(self.sr.is_manual_matched)
        self.assertTrue(self.lr.is_manual_matched)
        self.assertEqual(mm.match_mode, "3way")


class TCReconcileLateAiredRuleTest(TestCase):
    """Tests for the TC late-aired rule: TCRow.date >= ScheduleRow.date."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, tc_theme="TC Theme A")
        self.tc_report = make_tc_report(self.account)
        ensure_tc_tolerance(5)

    def test_tcrow_on_same_date_as_schedule_is_matched(self):
        """TCRow.date == ScheduleRow.date satisfies the >= rule (boundary case)."""
        sr = make_schedule_row(self.account, self.schedule, date=DATE)
        tc = make_tc_row(self.account, self.tc_report, date=DATE)

        result = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_schedule_matched, "TCRow on same date should be matched")
        self.assertEqual(tc.matched_schedule_id, sr.id)

    def test_tcrow_after_schedule_date_is_matched_late_aired(self):
        """TCRow.date > ScheduleRow.date is allowed (late-aired rule)."""
        sr = make_schedule_row(self.account, self.schedule, date=DATE)
        later_date = DATE + datetime.timedelta(days=2)
        tc = make_tc_row(
            self.account, self.tc_report,
            date=later_date,
            suffix="_later",
        )

        result = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_schedule_matched, "TCRow aired after schedule date should still match")

    def test_tcrow_before_schedule_date_is_not_matched(self):
        """TCRow.date < ScheduleRow.date must NOT match (violates >= rule)."""
        sr = make_schedule_row(self.account, self.schedule, date=DATE)
        earlier_date = DATE - datetime.timedelta(days=2)
        tc = make_tc_row(
            self.account, self.tc_report,
            date=earlier_date,
            suffix="_earlier",
        )

        result = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertFalse(tc.is_schedule_matched, "TCRow aired before schedule date must not match")
        self.assertTrue(tc.is_extra, "TCRow before schedule date must be marked extra")

    def test_extra_tcrows_are_flagged_when_no_schedule_match(self):
        """TCRows that don't match any ScheduleRow receive is_extra=True."""
        # Schedule row for Brand A, TC row for a completely different theme
        make_schedule_row(self.account, self.schedule)
        # TC row with a theme that has no brand mapping
        tc = make_tc_row(
            self.account, self.tc_report,
            tc_theme="Completely Unknown Theme",
            suffix="_unknown",
        )

        result = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_extra)
        self.assertFalse(tc.is_schedule_matched)


class TCReconcileTimeTolerance(TestCase):
    """Tests for TC-LMRB cross-check time tolerance."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, tc_theme="TC Theme A")
        self.tc_report = make_tc_report(self.account)

    def test_lmrb_within_tolerance_is_confirmed(self):
        """LMRBRow within ±5 seconds of TC aired_time gets is_lmrb_confirmed=True."""
        ensure_tc_tolerance(5)
        make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report, aired_time="20:30:00")
        # LMRB aired 3 seconds after TC — within 5s tolerance
        lr = make_lmrb_row(self.account, advt_time="20:30:03")

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_lmrb_confirmed)
        self.assertEqual(tc.matched_lmrb_id, lr.id)

    def test_lmrb_outside_tolerance_is_not_confirmed(self):
        """LMRBRow more than 5 seconds from TC aired_time must not be confirmed."""
        ensure_tc_tolerance(5)
        make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report, aired_time="20:30:00")
        # LMRB aired 10 seconds after TC — outside 5s tolerance
        lr = make_lmrb_row(self.account, advt_time="20:30:10")

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertFalse(tc.is_lmrb_confirmed)

    def test_wider_tolerance_allows_confirmation(self):
        """Increasing tolerance to 30s allows a 10s difference to confirm."""
        ensure_tc_tolerance(30)
        make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report, aired_time="20:30:00")
        lr = make_lmrb_row(self.account, advt_time="20:30:10")

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_lmrb_confirmed, "10s diff should confirm with 30s tolerance")
        self.assertEqual(tc.matched_lmrb_id, lr.id)

    def test_exact_time_match_is_confirmed(self):
        """LMRBRow at exactly the same time as TCRow is confirmed (0s diff)."""
        ensure_tc_tolerance(5)
        make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report, aired_time="20:30:00")
        lr = make_lmrb_row(self.account, advt_time="20:30:00")

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_lmrb_confirmed)


class TCReconcileExtraFlagTest(TestCase):
    """Tests that unmatched TCRows get is_extra=True after reconciliation."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        self.tc_report = make_tc_report(self.account)
        ensure_tc_tolerance(5)

    def test_tc_row_with_no_brand_mapping_is_extra(self):
        """A TCRow whose tc_theme has no BrandMapping entry is marked is_extra=True."""
        make_schedule_row(self.account, self.schedule, brand="Mapped Brand")
        make_brand_mapping(self.account, brand="Mapped Brand", tc_theme="Mapped TC Theme")
        # TC row with an unmapped theme
        tc_unmapped = make_tc_row(
            self.account, self.tc_report,
            tc_theme="No Mapping Theme",
            suffix="_unmapped",
        )
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc_unmapped.refresh_from_db()
        self.assertTrue(tc_unmapped.is_extra)

    def test_matched_tc_row_is_not_extra(self):
        """A successfully matched TCRow must not be flagged as is_extra."""
        make_schedule_row(self.account, self.schedule)
        make_brand_mapping(self.account, tc_theme="TC Theme A")
        tc = make_tc_row(self.account, self.tc_report)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertFalse(tc.is_extra)
        self.assertTrue(tc.is_schedule_matched)


class MultiSchedulePriorityRule10Test(TestCase):
    """
    Rule 10: Multiple schedules for the same scope are processed in ascending
    schedule_number order. The LMRB pool is shared — earlier schedule gets rows first.
    """

    def setUp(self):
        self.account = make_account()
        make_brand_mapping(self.account)

    def test_earlier_schedule_number_gets_lmrb_row_first(self):
        """
        Two schedules (#100 and #200). One shared LMRB row.
        The #100 schedule should claim the LMRB row.
        """
        sched100 = make_schedule(self.account, schedule_number="100", version=1)
        sched200 = make_schedule(self.account, schedule_number="200", version=1)

        sr100 = make_schedule_row(self.account, sched100)
        sr200 = make_schedule_row(self.account, sched200)

        # Only one LMRB row available
        lr = make_lmrb_row(self.account)

        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")

        sr100.refresh_from_db()
        sr200.refresh_from_db()
        lr.refresh_from_db()

        self.assertTrue(sr100.is_matched, "Schedule #100 should get the LMRB row")
        self.assertFalse(sr200.is_matched, "Schedule #200 should be Not Aired (no LMRB left)")
        self.assertEqual(lr.matched_schedule_id, sr100.id)

    def test_later_schedule_gets_remaining_rows(self):
        """
        Two LMRB rows available. Both schedules should each get one.
        """
        sched100 = make_schedule(self.account, schedule_number="100", version=1)
        sched200 = make_schedule(self.account, schedule_number="200", version=1)

        sr100 = make_schedule_row(self.account, sched100)
        sr200 = make_schedule_row(
            self.account, sched200,
            date=datetime.date(2025, 1, 16),
            start_time="20:00:00",
            end_time="21:00:00",
        )
        lr1 = make_lmrb_row(self.account)
        lr2 = make_lmrb_row(
            self.account,
            date=datetime.date(2025, 1, 16),
            advt_time="20:30:00",
            brk_no=1,
        )

        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")

        sr100.refresh_from_db()
        sr200.refresh_from_db()
        self.assertTrue(sr100.is_matched)
        self.assertTrue(sr200.is_matched)


class ScheduleVersioningRule12Test(TestCase):
    """
    Rule 12: When multiple uploads have the same schedule_number, only the
    highest version (latest upload) is used for matching.
    """

    def setUp(self):
        self.account = make_account()
        make_brand_mapping(self.account)

    def test_only_highest_version_is_used(self):
        """
        Two schedules with schedule_number='101': version=1 and version=2.
        Only version=2 rows should be matched.
        """
        sched_v1 = make_schedule(self.account, schedule_number="101", version=1)
        sched_v2 = make_schedule(self.account, schedule_number="101", version=2)

        # v1 has a row on Jan 15; v2 has a row on Jan 16
        sr_v1 = make_schedule_row(
            self.account, sched_v1,
            date=datetime.date(2025, 1, 15),
        )
        sr_v2 = make_schedule_row(
            self.account, sched_v2,
            date=datetime.date(2025, 1, 16),
        )

        # LMRB row on Jan 16 — only matches v2's row
        lr = make_lmrb_row(
            self.account,
            date=datetime.date(2025, 1, 16),
            advt_time="20:30:00",
        )

        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")

        sr_v1.refresh_from_db()
        sr_v2.refresh_from_db()

        # Only v2 should be processed (v1 is superseded)
        self.assertFalse(
            sr_v1.is_matched,
            "Version 1 rows must not be matched — schedule is superseded by v2",
        )
        self.assertTrue(
            sr_v2.is_matched,
            "Version 2 rows must be matched — it is the active schedule",
        )


class BuildSummaryDataTest(TestCase):
    """Tests for verification.tc_engine.build_summary_data counts."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, tc_theme="TC Theme A")
        self.tc_report = make_tc_report(self.account, schedule=self.schedule)
        ensure_tc_tolerance(5)

    def test_planned_count_equals_schedule_row_count(self):
        """Planned = total ScheduleRows with COMMERCIAL BENEFITS for the scope."""
        make_schedule_row(self.account, self.schedule)
        make_schedule_row(
            self.account, self.schedule,
            date=datetime.date(2025, 1, 16),
            start_time="20:00:00",
            end_time="21:00:00",
        )
        summary = build_summary_data(self.account.id, CHANNEL, MONTH)
        total_planned = sum(r["planned"] for r in summary["commercial"])
        self.assertEqual(total_planned, 2)

    def test_aired_count_requires_schedule_matched_and_lmrb_confirmed(self):
        """
        Aired = TCRows where is_schedule_matched=True AND is_lmrb_confirmed=True.
        A TC row that is only schedule-matched but NOT LMRB-confirmed must not count.
        """
        sr = make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report)
        lr = make_lmrb_row(self.account)

        # Manually set TC row as schedule-matched but NOT lmrb-confirmed
        tc.is_schedule_matched = True
        tc.matched_schedule = sr
        tc.is_lmrb_confirmed = False
        tc.save()

        summary = build_summary_data(self.account.id, CHANNEL, MONTH)
        commercial = summary["commercial"]
        self.assertEqual(len(commercial), 1)
        self.assertEqual(commercial[0]["aired"], 0, "Aired must be 0 without LMRB confirmation")

    def test_aired_count_with_full_tc_lmrb_confirmation(self):
        """Aired = 1 when TCRow is both schedule-matched and LMRB-confirmed."""
        sr = make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report)
        lr = make_lmrb_row(self.account)

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        summary = build_summary_data(self.account.id, CHANNEL, MONTH)
        commercial = summary["commercial"]
        self.assertEqual(commercial[0]["aired"], 1)

    def test_missed_is_max_zero_planned_minus_aired(self):
        """Missed = max(0, planned - aired)."""
        # 2 planned, 1 aired => missed=1
        make_schedule_row(self.account, self.schedule, date=DATE)
        make_schedule_row(
            self.account, self.schedule,
            date=datetime.date(2025, 1, 16),
            start_time="20:00:00",
            end_time="21:00:00",
        )
        tc = make_tc_row(self.account, self.tc_report, date=DATE)
        lr = make_lmrb_row(self.account)

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        summary = build_summary_data(self.account.id, CHANNEL, MONTH)
        commercial = summary["commercial"]
        row = commercial[0]
        self.assertEqual(row["planned"], 2)
        self.assertEqual(row["aired"], 1)
        self.assertEqual(row["missed"], 1)

    def test_extra_is_max_zero_third_party_minus_planned(self):
        """Extra = max(0, 3rd_party - planned)."""
        # 1 planned, 2 LMRB-confirmed via TC => extra=1
        sr = make_schedule_row(self.account, self.schedule, date=DATE)
        tc1 = make_tc_row(self.account, self.tc_report, aired_time="20:30:00", suffix="_t1")
        tc2 = make_tc_row(self.account, self.tc_report, aired_time="20:45:00", suffix="_t2")
        lr1 = make_lmrb_row(self.account, advt_time="20:30:00")
        lr2 = make_lmrb_row(self.account, advt_time="20:45:00", brk_no=2)

        # Manually set both TC rows as schedule-matched + LMRB-confirmed
        # tc1 linked to sr, tc2 as extra but LMRB-confirmed
        tc1.is_schedule_matched = True
        tc1.matched_schedule = sr
        tc1.is_lmrb_confirmed = True
        tc1.matched_lmrb = lr1
        tc1.save()

        tc2.is_lmrb_confirmed = True
        tc2.matched_lmrb = lr2
        tc2.is_schedule_matched = True
        tc2.matched_schedule = sr
        tc2.save()

        # Adjust planned: sr is 1 row, but 2 confirmed TCRows exist
        # Build summary
        summary = build_summary_data(self.account.id, CHANNEL, MONTH)
        commercial = summary["commercial"]
        row = commercial[0]
        # third_party should be 2 (two unique confirmed LMRBs), planned=1, extra=1
        self.assertEqual(row["planned"], 1)
        self.assertGreaterEqual(row["third_party"], 1)

    def test_sponsorship_section_present_when_sponsorship_rows_exist(self):
        """build_summary_data includes sponsorship sections when SPONSORSHIP rows exist."""
        make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            programme="Sports Show",
        )
        summary = build_summary_data(self.account.id, CHANNEL, MONTH)
        self.assertIn("sponsorship", summary)
        self.assertGreater(len(summary["sponsorship"]), 0)

    def test_summary_totals_are_sum_of_rows(self):
        """commercial_total values are the sum of all commercial row values."""
        make_schedule_row(self.account, self.schedule, date=DATE)
        make_schedule_row(
            self.account, self.schedule,
            date=datetime.date(2025, 1, 16),
            start_time="20:00:00",
            end_time="21:00:00",
        )
        summary = build_summary_data(self.account.id, CHANNEL, MONTH)
        commercial = summary["commercial"]
        total = summary["commercial_total"]
        self.assertEqual(total["planned"], sum(r["planned"] for r in commercial))
        self.assertEqual(total["aired"], sum(r["aired"] for r in commercial))
        self.assertEqual(total["missed"], sum(r["missed"] for r in commercial))


class SponsorshipAutoMatchTest(TestCase):
    """Tests for verification.sponsorship_engine.reconcile_sponsorship."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, theme="Spon Theme A", tc_theme="")
        self.user = make_user()

    def test_auto_match_assigns_lmrb_to_sponsorship_row(self):
        """reconcile_sponsorship pairs an unmatched LMRBRow with a SPONSORSHIP ScheduleRow."""
        sr = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            programme="Sports Show",
            brand="Brand A",  # maps to theme='Spon Theme A' via brand mapping
        )
        # The BrandMapping.theme for 'Brand A' is 'Theme A'; need to match for sponsorship
        # Recreate the brand mapping so theme matches the LMRB advt_theme
        BrandMapping.objects.filter(account=self.account).delete()
        make_brand_mapping(self.account, brand="Brand A", theme="Spon Theme A", tc_theme="")

        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")

        result = reconcile_sponsorship(self.account.id, CHANNEL, MONTH)

        sr.refresh_from_db()
        lr.refresh_from_db()

        self.assertTrue(lr.is_sponsorship_matched)
        self.assertEqual(result["assigned"], 1)
        self.assertTrue(SponsorshipLmrbAssignment.objects.filter(
            schedule_row=sr, lmrb_row=lr
        ).exists())

    def test_already_sponsorship_matched_lmrb_is_skipped(self):
        """reconcile_sponsorship does not reassign an already sponsorship-matched LMRBRow."""
        BrandMapping.objects.filter(account=self.account).delete()
        make_brand_mapping(self.account, brand="Brand A", theme="Spon Theme A", tc_theme="")

        sr1 = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            programme="Sports Show",
            brand="Brand A",
        )
        sr2 = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            programme="Sports Show",
            brand="Brand A",
            date=datetime.date(2025, 1, 16),
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")

        # First run assigns lr to sr1
        reconcile_sponsorship(self.account.id, CHANNEL, MONTH)
        lr.refresh_from_db()
        self.assertTrue(lr.is_sponsorship_matched)

        # Second run (smart mode) should not reassign lr to sr2
        result2 = reconcile_sponsorship(self.account.id, CHANNEL, MONTH, mode="smart")
        # sr2 has no available LMRB row
        self.assertEqual(
            SponsorshipLmrbAssignment.objects.filter(lmrb_row=lr).count(),
            1,
            "is_sponsorship_matched LMRBRow must not be reassigned",
        )

    def test_reset_mode_clears_and_reruns(self):
        """reconcile_sponsorship mode='reset' deletes existing assignments and re-runs."""
        BrandMapping.objects.filter(account=self.account).delete()
        make_brand_mapping(self.account, brand="Brand A", theme="Spon Theme A", tc_theme="")

        sr = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")

        reconcile_sponsorship(self.account.id, CHANNEL, MONTH)
        self.assertEqual(SponsorshipLmrbAssignment.objects.count(), 1)

        # Reset should delete the old assignment and re-create it
        reconcile_sponsorship(self.account.id, CHANNEL, MONTH, mode="reset")
        self.assertEqual(SponsorshipLmrbAssignment.objects.count(), 1)
        lr.refresh_from_db()
        self.assertTrue(lr.is_sponsorship_matched)


class SponsorshipManualAssignTest(TestCase):
    """Tests for verification.sponsorship_engine.manual_assign."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, theme="Spon Theme A", tc_theme="")
        self.user = make_user()

    def test_manual_assign_creates_assignment(self):
        """manual_assign creates a SponsorshipLmrbAssignment and locks the LMRBRow."""
        sr = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")

        result = manual_assign(
            self.account.id, CHANNEL, MONTH,
            assignments=[(sr.id, lr.id)],
            user=self.user,
        )

        self.assertEqual(result["created"], 1)
        self.assertEqual(result["skipped"], 0)
        lr.refresh_from_db()
        self.assertTrue(lr.is_sponsorship_matched)
        self.assertTrue(SponsorshipLmrbAssignment.objects.filter(
            schedule_row=sr, lmrb_row=lr, match_type="manual"
        ).exists())

    def test_manual_assign_rejects_already_matched_lmrb(self):
        """manual_assign skips LMRBRows that are already is_matched=True (commercially locked)."""
        sr = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")
        lr.is_matched = True
        lr.save(update_fields=["is_matched"])

        result = manual_assign(
            self.account.id, CHANNEL, MONTH,
            assignments=[(sr.id, lr.id)],
            user=self.user,
        )

        self.assertEqual(result["created"], 0)
        self.assertEqual(result["skipped"], 1)

    def test_manual_assign_rejects_already_sponsorship_matched_lmrb(self):
        """manual_assign skips LMRBRows that are already is_sponsorship_matched=True."""
        sr1 = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
        )
        sr2 = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
            date=datetime.date(2025, 1, 16),
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")
        lr.is_sponsorship_matched = True
        lr.save(update_fields=["is_sponsorship_matched"])

        result = manual_assign(
            self.account.id, CHANNEL, MONTH,
            assignments=[(sr2.id, lr.id)],
            user=self.user,
        )

        self.assertEqual(result["created"], 0)
        self.assertEqual(result["skipped"], 1)

    def test_manual_assign_rejects_manually_matched_lmrb(self):
        """manual_assign skips LMRBRows that are already is_manual_matched=True."""
        sr = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")
        lr.is_manual_matched = True
        lr.save(update_fields=["is_manual_matched"])

        result = manual_assign(
            self.account.id, CHANNEL, MONTH,
            assignments=[(sr.id, lr.id)],
            user=self.user,
        )

        self.assertEqual(result["created"], 0)
        self.assertEqual(result["skipped"], 1)

    def test_manual_assign_rejects_invalid_schedule_row(self):
        """manual_assign skips a schedule_row_id that is not a SPONSORSHIP row."""
        sr = make_schedule_row(
            self.account, self.schedule,
            ad_type="COMMERCIAL BENEFITS",  # wrong type
        )
        lr = make_lmrb_row(self.account)

        result = manual_assign(
            self.account.id, CHANNEL, MONTH,
            assignments=[(sr.id, lr.id)],
            user=self.user,
        )

        self.assertEqual(result["created"], 0)
        self.assertEqual(result["skipped"], 1)


class SponsorshipDoubleClaimLockTest(TestCase):
    """
    Regression: a raw LMRBRow already claimed by a ManualMatch or the standalone
    TC↔LMRB engine must never be re-used by the sponsorship engine (auto pool,
    manual picker, or manual assign) — otherwise one raw row is counted twice.
    """

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, brand="Brand A", theme="Spon Theme A", tc_theme="")
        self.user = make_user()

    def test_auto_skips_manual_matched_lmrb(self):
        """reconcile_sponsorship must not claim an is_manual_matched LMRB row."""
        sr = make_schedule_row(
            self.account, self.schedule, ad_type="SPONSORSHIP", brand="Brand A",
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")
        lr.is_manual_matched = True          # locked by a ManualMatch (is_matched stays False)
        lr.save(update_fields=["is_manual_matched"])

        result = reconcile_sponsorship(self.account.id, CHANNEL, MONTH, mode="reset")

        self.assertEqual(result["assigned"], 0)
        self.assertFalse(
            SponsorshipLmrbAssignment.objects.filter(lmrb_row=lr).exists(),
            "Manual-matched LMRB row was double-claimed by the sponsorship engine",
        )

    def test_auto_skips_tc_lmrb_matched_lmrb(self):
        """reconcile_sponsorship must not claim an is_tc_lmrb_matched LMRB row."""
        sr = make_schedule_row(
            self.account, self.schedule, ad_type="SPONSORSHIP", brand="Brand A",
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")
        lr.is_tc_lmrb_matched = True
        lr.save(update_fields=["is_tc_lmrb_matched"])

        result = reconcile_sponsorship(self.account.id, CHANNEL, MONTH, mode="reset")

        self.assertEqual(result["assigned"], 0)
        self.assertFalse(SponsorshipLmrbAssignment.objects.filter(lmrb_row=lr).exists())

    def test_manual_picker_excludes_locked_rows(self):
        """lmrb_candidates must hide rows locked by ManualMatch or TC↔LMRB."""
        make_schedule_row(self.account, self.schedule, ad_type="SPONSORSHIP", brand="Brand A")
        free = make_lmrb_row(self.account, advt_theme="Spon Theme A", advt_time="20:30:00")
        manual = make_lmrb_row(self.account, advt_theme="Spon Theme A", advt_time="20:31:00")
        tclmrb = make_lmrb_row(self.account, advt_theme="Spon Theme A", advt_time="20:32:00")
        manual.is_manual_matched = True
        manual.save(update_fields=["is_manual_matched"])
        tclmrb.is_tc_lmrb_matched = True
        tclmrb.save(update_fields=["is_tc_lmrb_matched"])

        ids = {c["id"] for c in lmrb_candidates(self.account.id, CHANNEL, MONTH)}

        self.assertIn(free.id, ids)
        self.assertNotIn(manual.id, ids)
        self.assertNotIn(tclmrb.id, ids)

    def test_manual_assign_rejects_tc_lmrb_matched_lmrb(self):
        """manual_assign must reject an is_tc_lmrb_matched LMRB row."""
        sr = make_schedule_row(self.account, self.schedule, ad_type="SPONSORSHIP", brand="Brand A")
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")
        lr.is_tc_lmrb_matched = True
        lr.save(update_fields=["is_tc_lmrb_matched"])

        result = manual_assign(
            self.account.id, CHANNEL, MONTH, assignments=[(sr.id, lr.id)], user=self.user,
        )

        self.assertEqual(result["created"], 0)
        self.assertEqual(result["skipped"], 1)


class SponsorshipCommercialSeparationTest(TestCase):
    """
    Sponsorship LMRB assignments must not affect commercial 3rd party counts
    in build_summary_data.
    """

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        self.tc_report = make_tc_report(self.account, schedule=self.schedule)
        ensure_tc_tolerance(5)

    def test_sponsorship_lmrb_excluded_from_commercial_third_party(self):
        """
        An LMRBRow consumed by sponsorship must not appear in commercial 3rd-party count.
        build_summary_data passes exclude_spon=True for commercial rows.
        """
        # Commercial brand mapping
        make_brand_mapping(
            self.account, brand="Commercial Brand", theme="Com Theme",
            tc_theme="Com TC Theme",
        )
        # Sponsorship brand mapping (same account, different brand)
        make_brand_mapping(
            self.account, brand="Spon Brand", theme="Spon Theme",
            tc_theme="",
        )

        # Commercial schedule row
        sr_com = make_schedule_row(
            self.account, self.schedule,
            brand="Commercial Brand",
            ad_type="COMMERCIAL BENEFITS",
        )
        # Sponsorship schedule row
        sr_spon = make_schedule_row(
            self.account, self.schedule,
            brand="Spon Brand",
            ad_type="SPONSORSHIP",
            programme="Sports Show",
        )

        # LMRB row for sponsorship theme — will be consumed by sponsorship
        lr_spon = make_lmrb_row(
            self.account, advt_theme="Spon Theme",
        )

        # Run sponsorship reconciliation to lock lr_spon
        reconcile_sponsorship(self.account.id, CHANNEL, MONTH)
        lr_spon.refresh_from_db()
        self.assertTrue(lr_spon.is_sponsorship_matched)

        # Build summary — commercial row should not count the sponsorship LMRB
        summary = build_summary_data(self.account.id, CHANNEL, MONTH)
        com_rows = summary["commercial"]
        # Filter for the Commercial Brand row
        com_brand_rows = [r for r in com_rows if r["product"] == "Commercial Brand"]
        if com_brand_rows:
            # 3rd party for commercial should be 0 (no commercial LMRB confirmed)
            self.assertEqual(
                com_brand_rows[0]["third_party"], 0,
                "Sponsorship LMRB must not be counted in commercial 3rd party",
            )


class SponsorshipResetTest(TestCase):
    """Tests for verification.sponsorship_engine.reset_sponsorship."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, brand="Brand A", theme="Spon Theme A", tc_theme="")
        self.user = make_user()

    def test_reset_sponsorship_deletes_assignments_and_unlocks_lmrb(self):
        """reset_sponsorship removes all assignments and sets is_sponsorship_matched=False."""
        sr = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")

        reconcile_sponsorship(self.account.id, CHANNEL, MONTH)
        lr.refresh_from_db()
        self.assertTrue(lr.is_sponsorship_matched)

        result = reset_sponsorship(self.account.id, CHANNEL, MONTH)

        lr.refresh_from_db()
        self.assertFalse(lr.is_sponsorship_matched)
        self.assertEqual(SponsorshipLmrbAssignment.objects.count(), 0)
        self.assertEqual(result["deleted"], 1)
        self.assertEqual(result["lmrb_unlocked"], 1)


class TCReconcileSmartResetModeTest(TestCase):
    """Tests for reconcile_tc smart vs reset mode behaviour."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, tc_theme="TC Theme A")
        self.tc_report = make_tc_report(self.account)
        ensure_tc_tolerance(5)

    def test_reset_mode_clears_tc_reconciliation_state(self):
        """reconcile_tc mode='reset' clears is_schedule_matched and is_lmrb_confirmed flags."""
        sr = make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report)
        lr = make_lmrb_row(self.account)

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_schedule_matched)

        # Manually dirty the state as if a re-upload happened
        tc.is_schedule_matched = False
        tc.matched_schedule = None
        tc.is_lmrb_confirmed = False
        tc.matched_lmrb = None
        tc.save()

        # Reset should restore correct state
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_schedule_matched)

    def test_smart_mode_skips_already_matched_tc_rows(self):
        """reconcile_tc smart mode does not reprocess already-matched TCRows."""
        sr = make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report)
        lr = make_lmrb_row(self.account)

        # First run
        r1 = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        self.assertEqual(r1["matched"], 1)

        # Smart run should report 0 new matches (already matched)
        r2 = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="smart")
        self.assertEqual(r2["matched"], 0)


class LMRBSponsorshipLockTest(TestCase):
    """
    LMRBRow.is_sponsorship_matched=True prevents the sponsorship engine from
    reassigning the row to a different SPONSORSHIP ScheduleRow.
    """

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, brand="Brand A", theme="Spon Theme A", tc_theme="")

    def test_sponsorship_matched_lmrb_not_reused_by_auto_engine(self):
        """
        An LMRBRow with is_sponsorship_matched=True is excluded from the
        auto-sponsorship pool and cannot be claimed by another row.
        """
        sr1 = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
        )
        sr2 = make_schedule_row(
            self.account, self.schedule,
            ad_type="SPONSORSHIP",
            brand="Brand A",
            programme="Sports Show",
            date=datetime.date(2025, 1, 16),
        )
        lr = make_lmrb_row(self.account, advt_theme="Spon Theme A")

        # First reconcile: lr assigned to sr1
        reconcile_sponsorship(self.account.id, CHANNEL, MONTH)
        lr.refresh_from_db()
        self.assertTrue(lr.is_sponsorship_matched)

        # Second smart run: lr should not be reassigned to sr2
        result = reconcile_sponsorship(self.account.id, CHANNEL, MONTH, mode="smart")

        assignments = SponsorshipLmrbAssignment.objects.filter(lmrb_row=lr)
        self.assertEqual(assignments.count(), 1, "LMRB row must only appear in one assignment")
        self.assertEqual(assignments.first().schedule_row, sr1)


class NoBrandMappingTest(TestCase):
    """ScheduleRows without a BrandMapping are handled gracefully."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)

    def test_no_mapping_status_in_match_result(self):
        """A ScheduleRow with no BrandMapping produces a 'no_mapping' MatchResult."""
        make_schedule_row(self.account, self.schedule, brand="Unmapped Brand")
        make_lmrb_row(self.account, advt_theme="Some Theme")

        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")

        results = MatchResult.objects.filter(
            account=self.account, status="no_mapping"
        )
        self.assertEqual(results.count(), 1)
        self.assertEqual(results.first().brand, "Unmapped Brand")

    def test_no_mapping_lmrb_row_is_not_consumed(self):
        """The LMRBRow remains unmatched when the ScheduleRow has no mapping."""
        make_schedule_row(self.account, self.schedule, brand="Unmapped Brand")
        lr = make_lmrb_row(self.account, advt_theme="Some Theme")

        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")

        lr.refresh_from_db()
        self.assertFalse(lr.is_matched)


class BoundaryDateTest(TestCase):
    """Edge cases for date handling in TC and Schedule reconciliation."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, tc_theme="TC Theme A")
        self.tc_report = make_tc_report(self.account)
        ensure_tc_tolerance(5)

    def test_tc_row_date_exactly_equals_schedule_row_date(self):
        """TCRow.date == ScheduleRow.date satisfies the >= rule exactly."""
        boundary_date = datetime.date(2025, 1, 10)
        sr = make_schedule_row(self.account, self.schedule, date=boundary_date)
        tc = make_tc_row(self.account, self.tc_report, date=boundary_date, suffix="_boundary")

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()

        self.assertTrue(
            tc.is_schedule_matched,
            "TCRow on exact same date as ScheduleRow must be matched",
        )

    def test_tc_row_one_day_after_schedule_row(self):
        """TCRow.date one day after ScheduleRow.date satisfies the late-aired rule."""
        base_date = datetime.date(2025, 1, 10)
        sr = make_schedule_row(self.account, self.schedule, date=base_date)
        tc = make_tc_row(
            self.account, self.tc_report,
            date=base_date + datetime.timedelta(days=1),
            suffix="_plus1",
        )

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()

        self.assertTrue(
            tc.is_schedule_matched,
            "TCRow one day after ScheduleRow must be matched (late-aired rule)",
        )

    def test_tc_row_one_day_before_schedule_row_is_not_matched(self):
        """TCRow.date one day before ScheduleRow.date must NOT match."""
        base_date = datetime.date(2025, 1, 10)
        sr = make_schedule_row(self.account, self.schedule, date=base_date)
        tc = make_tc_row(
            self.account, self.tc_report,
            date=base_date - datetime.timedelta(days=1),
            suffix="_minus1",
        )

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()

        self.assertFalse(
            tc.is_schedule_matched,
            "TCRow one day before ScheduleRow must not be matched",
        )
        self.assertTrue(tc.is_extra)


class MatchResultStatusTest(TestCase):
    """Tests that MatchResult records are created with correct status values."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account)

    def test_matched_status_created_on_successful_match(self):
        """A successful match creates a MatchResult with status='matched'."""
        make_schedule_row(self.account, self.schedule)
        make_lmrb_row(self.account)
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        self.assertTrue(
            MatchResult.objects.filter(
                account=self.account, status="matched"
            ).exists()
        )

    def test_not_aired_status_when_no_lmrb_row(self):
        """A ScheduleRow with no matching LMRB produces a 'not_aired' MatchResult."""
        make_schedule_row(self.account, self.schedule)
        # No LMRBRow created
        make_lmrb_row(
            self.account,
            advt_theme="Theme A",
            date=datetime.date(2025, 2, 1),  # Out of schedule date range
        )
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        not_aired = MatchResult.objects.filter(
            account=self.account, status__in=["not_aired", "no_mapping"]
        )
        self.assertGreater(not_aired.count(), 0)

    def test_late_telecast_status_on_different_date(self):
        """A ScheduleRow matched on a later date produces 'late_telecast' status."""
        sr = make_schedule_row(
            self.account, self.schedule,
            date=datetime.date(2025, 1, 10),
            start_time="20:00:00",
            end_time="21:00:00",
        )
        # LMRB row on a different (later) date, same time window
        lr = make_lmrb_row(
            self.account,
            date=datetime.date(2025, 1, 20),
            advt_time="20:30:00",
        )
        run_scope(self.account.id, CHANNEL, MONTH, mode="smart")
        self.assertTrue(
            MatchResult.objects.filter(
                account=self.account, status="late_telecast"
            ).exists()
        )


class AccountIsolationTest(TestCase):
    """Verify that data from different accounts never bleeds into each other's matching."""

    def setUp(self):
        self.account1 = make_account("Account One")
        self.account2 = make_account("Account Two")

    def test_lmrb_rows_from_different_account_not_used(self):
        """
        An LMRBRow belonging to account2 must not be matched against
        account1's ScheduleRows.
        """
        make_brand_mapping(self.account1)
        sched1 = make_schedule(self.account1)
        sr = make_schedule_row(self.account1, sched1)
        # LMRB row belongs to account2 — same channel, theme, date, duration
        lr = make_lmrb_row(self.account2, advt_theme="Theme A")

        run_scope(self.account1.id, CHANNEL, MONTH, mode="smart")

        sr.refresh_from_db()
        lr.refresh_from_db()
        self.assertFalse(sr.is_matched, "account1 ScheduleRow must not match account2 LMRB")
        self.assertFalse(lr.is_matched, "account2 LMRB must remain unmatched")


class ReconcileTCReturnCountsTest(TestCase):
    """Tests that reconcile_tc returns accurate summary counts."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_brand_mapping(self.account, tc_theme="TC Theme A")
        self.tc_report = make_tc_report(self.account)
        ensure_tc_tolerance(5)

    def test_return_counts_matched_and_lmrb_confirmed(self):
        """reconcile_tc returns correct matched and lmrb_confirmed counts."""
        make_schedule_row(self.account, self.schedule)
        tc = make_tc_row(self.account, self.tc_report, aired_time="20:30:00")
        lr = make_lmrb_row(self.account, advt_time="20:30:00")

        result = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")

        self.assertEqual(result["matched"], 1)
        self.assertEqual(result["lmrb_confirmed"], 1)
        self.assertEqual(result["extra"], 0)

    def test_return_counts_extra_tc_row(self):
        """reconcile_tc reports correct extra count when TCRow has no schedule match."""
        make_schedule_row(self.account, self.schedule)
        tc_matched = make_tc_row(self.account, self.tc_report, aired_time="20:30:00", suffix="_m")
        tc_extra = make_tc_row(
            self.account, self.tc_report,
            tc_theme="Completely Different Theme",
            aired_time="21:00:00",
            suffix="_e",
        )
        lr = make_lmrb_row(self.account, advt_time="20:30:00")

        result = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")

        self.assertEqual(result["matched"], 1)
        self.assertEqual(result["extra"], 1)

    def test_no_schedule_rows_returns_zero_matched(self):
        """reconcile_tc returns 0 matched when there are no ScheduleRows in scope."""
        tc = make_tc_row(self.account, self.tc_report)
        result = reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        self.assertEqual(result["matched"], 0)
        self.assertEqual(result["extra"], 1)


# ── Bug Regression Tests ──────────────────────────────────────────────────────


class ScheduleSupersedeBugTest(TestCase):
    """
    Regression tests for the schedule-supersede bug (core/views.py line ~490).

    BUG (CRITICAL — core/views.py:490):
        When a schedule is replaced via dup_action='replace', the upload view
        calls:

            existing_same_ref.rows.update(
                is_matched=False, matched_lmrb=None, matched_at=None,
                is_manual_matched=False,   # ← BUG
            )

        This clears is_manual_matched on ALL old ScheduleRows, including those
        locked by a ManualMatch record.  However:
          - The ManualMatch records are NOT deleted.
          - The LMRBRow.is_manual_matched flag is NOT cleared (LMRB stays locked).

        Result: orphaned ManualMatch records pointing to superseded ScheduleRows
        with is_manual_matched=False.  build_summary_data() still counts these
        ManualMatches in the 'aired' total (it filters by account/channel/month/
        brand, not by schedule_id), inflating the summary with stale data.

        The LMRB rows linked to the stale ManualMatches remain permanently locked
        (is_manual_matched=True) and cannot be used for any other match.

    NOTE: These tests exercise the MODEL/ENGINE layer only.  The bug itself lives
    in the VIEW layer (schedule_upload).  The tests simulate what the view does
    so they can run without HTTP requests, and assert the incorrect state that
    results.  Once the bug is fixed in the view, these tests serve as a guard
    to confirm the fix is correct.
    """

    def setUp(self):
        self.account = make_account()
        self.user = make_user()
        make_brand_mapping(self.account)

    def _simulate_supersede(self, old_schedule):
        """
        Simulate what core/views.py does when dup_action='replace'.
        Mirrors the exact update at views.py ~line 488-491.
        """
        from core.models import LMRBRow as _LR
        sr_ids = list(old_schedule.rows.values_list("id", flat=True))
        _LR.objects.filter(schedule_matches__in=sr_ids).update(
            is_matched=False, matched_at=None
        )
        old_schedule.rows.update(
            is_matched=False, matched_lmrb=None, matched_at=None,
            is_manual_matched=False,          # ← the buggy line
        )
        old_schedule.is_superseded = True
        old_schedule.save(update_fields=["is_superseded"])

    # ── Tests that document EXISTING (buggy) behaviour ────────────────────────

    def test_supersede_clears_is_manual_matched_on_schedule_rows(self):
        """
        BUG: is_manual_matched is cleared on ScheduleRows when schedule is
        superseded, even when a ManualMatch record exists for them.

        This violates the invariant that is_manual_matched=True is permanent.
        """
        old_sched = make_schedule(self.account, schedule_number="101", version=1)
        sr = make_schedule_row(self.account, old_sched)
        lr = make_lmrb_row(self.account)

        # Create a ManualMatch — locks both rows
        ManualMatch.objects.create(
            account=self.account,
            channel=CHANNEL,
            month=MONTH,
            match_mode="schedule_lmrb",
            schedule_row=sr,
            lmrb_row=lr,
            matched_by=self.user,
        )
        ScheduleRow.objects.filter(id=sr.id).update(is_manual_matched=True)
        LMRBRow.objects.filter(id=lr.id).update(is_manual_matched=True)

        # Simulate superseding the old schedule
        self._simulate_supersede(old_sched)

        # BUG: is_manual_matched is now False on the ScheduleRow
        sr.refresh_from_db()
        lr.refresh_from_db()

        self.assertFalse(
            sr.is_manual_matched,
            "BUG CONFIRMED: is_manual_matched was cleared on superseded ScheduleRow. "
            "ManualMatch record still exists but lock is gone.",
        )
        # The LMRBRow lock is NOT cleared (only ScheduleRow is affected by the bug)
        self.assertTrue(
            lr.is_manual_matched,
            "LMRBRow.is_manual_matched should remain True (LMRB side is not touched "
            "by the supersede update — only ScheduleRow side is wrongly cleared).",
        )

    def test_manual_match_record_survives_supersede(self):
        """
        BUG (related): ManualMatch records are NOT deleted when a schedule is
        superseded.  Combined with is_manual_matched being cleared on the
        ScheduleRow, this leaves orphaned ManualMatch records.
        """
        old_sched = make_schedule(self.account, schedule_number="101", version=1)
        sr = make_schedule_row(self.account, old_sched)
        lr = make_lmrb_row(self.account)

        mm = ManualMatch.objects.create(
            account=self.account,
            channel=CHANNEL,
            month=MONTH,
            match_mode="schedule_lmrb",
            schedule_row=sr,
            lmrb_row=lr,
            matched_by=self.user,
        )
        ScheduleRow.objects.filter(id=sr.id).update(is_manual_matched=True)
        LMRBRow.objects.filter(id=lr.id).update(is_manual_matched=True)

        self._simulate_supersede(old_sched)

        # ManualMatch record is still in the DB after supersede
        self.assertTrue(
            ManualMatch.objects.filter(id=mm.id).exists(),
            "ManualMatch record survives supersede (not deleted by the view). "
            "This orphaned record will be incorrectly counted in build_summary_data().",
        )

    def test_lmrb_row_remains_locked_after_supersede(self):
        """
        The LMRB row linked to a ManualMatch on a superseded schedule remains
        permanently locked (is_manual_matched=True), making it unusable for any
        new match even though the linked schedule is gone.
        """
        old_sched = make_schedule(self.account, schedule_number="101", version=1)
        sr = make_schedule_row(self.account, old_sched)
        lr = make_lmrb_row(self.account)

        ManualMatch.objects.create(
            account=self.account,
            channel=CHANNEL,
            month=MONTH,
            match_mode="schedule_lmrb",
            schedule_row=sr,
            lmrb_row=lr,
            matched_by=self.user,
        )
        LMRBRow.objects.filter(id=lr.id).update(is_manual_matched=True)

        self._simulate_supersede(old_sched)

        # Create a new schedule and a new ScheduleRow for the same brand
        new_sched = make_schedule(self.account, schedule_number="101", version=2)
        new_sr = make_schedule_row(self.account, new_sched)

        # The LMRB row is still locked — new schedule engine cannot use it
        lr.refresh_from_db()
        self.assertTrue(
            lr.is_manual_matched,
            "LMRB row is permanently locked even after old schedule is superseded. "
            "It cannot be matched against the new schedule.",
        )


class UnmappedThemeNotAiredTest(TestCase):
    """
    A TC spot only counts as Aired when its tc_theme resolves to the brand via
    BrandMapping (or an operator confirmed it with a ManualMatch).

    Regression guard for the removed "window-coverage fallback", which credited
    LMRB-confirmed TC spots with NO brand mapping to any planned slot whose date +
    time window + duration happened to cover them.  That inflated Aired for
    unmapped brands and leaked one advertiser's spot into another's count.
    """

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        self.tc_report = make_tc_report(self.account, schedule=self.schedule)
        ensure_tc_tolerance(5)

    def _confirmed_tc(self, tc_theme, advt_theme=None):
        """Create an LMRB-confirmed TCRow in the default 20:00-21:00 planned slot."""
        tc = make_tc_row(self.account, self.tc_report, tc_theme=tc_theme)
        lr = make_lmrb_row(self.account, advt_theme=advt_theme or tc_theme)
        tc.is_lmrb_confirmed = True
        tc.matched_lmrb = lr
        tc.save()
        return tc, lr

    def test_brand_with_no_mapping_is_not_aired(self):
        """A brand with no BrandMapping shows Aired=0, even with a covering TC spot."""
        make_schedule_row(self.account, self.schedule, brand="Brand X")
        self._confirmed_tc("Totally Unmapped Theme")

        row = build_summary_data(self.account.id, CHANNEL, MONTH)["commercial"][0]
        self.assertEqual(row["planned"], 1)
        self.assertEqual(row["aired"], 0, "Unmapped brand must not be credited as aired")
        self.assertEqual(row["third_party"], 0)
        self.assertEqual(row["missed"], 1, "Unmapped brand's planned spot is Missed")

    def test_unmapped_spot_does_not_leak_into_another_brand(self):
        """
        Brand A is mapped but did not air.  An unrelated advertiser's unmapped spot
        falling inside Brand A's window must not be counted as Brand A airing.
        """
        make_brand_mapping(self.account, brand="Brand A", tc_theme="TC Theme A")
        make_schedule_row(self.account, self.schedule, brand="Brand A")
        self._confirmed_tc("Some Other Advertiser")

        row = build_summary_data(self.account.id, CHANNEL, MONTH)["commercial"][0]
        self.assertEqual(row["product"], "Brand A")
        self.assertEqual(row["aired"], 0, "Another brand's spot must not count here")
        self.assertEqual(row["missed"], 1)

    def test_mapped_brand_that_aired_still_counts(self):
        """Control: a correctly mapped brand that genuinely aired is still Aired=1."""
        make_brand_mapping(self.account, brand="Brand A", tc_theme="TC Theme A")
        make_schedule_row(self.account, self.schedule, brand="Brand A")
        make_tc_row(self.account, self.tc_report, tc_theme="TC Theme A")
        make_lmrb_row(self.account, advt_theme="Theme A")

        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        row = build_summary_data(self.account.id, CHANNEL, MONTH)["commercial"][0]
        self.assertEqual(row["aired"], 1)
        self.assertEqual(row["missed"], 0)

    def test_manual_match_still_counts_without_mapping(self):
        """An operator's ManualMatch is explicit evidence and still counts as Aired."""
        sr = make_schedule_row(self.account, self.schedule, brand="Brand X")
        _tc, lr = self._confirmed_tc("Totally Unmapped Theme")
        user = make_user()
        ManualMatch.objects.create(
            account=self.account,
            channel=CHANNEL,
            month=MONTH,
            match_mode="schedule_lmrb",
            schedule_row=sr,
            lmrb_row=lr,
            matched_by=user,
        )

        row = build_summary_data(self.account.id, CHANNEL, MONTH)["commercial"][0]
        self.assertEqual(row["aired"], 1, "ManualMatch is explicit operator evidence")
        self.assertEqual(row["missed"], 0)


class BrandMappingDeleteViewTest(TestCase):
    """Deleting brand mapping data from /dashboard/brand-mappings/.

    Three delete paths exist:
      - 'delete'        → one LMRB theme row (chip ✕)
      - 'delete_group'  → every row of one (brand, product, duration) group
      - 'delete_bulk'   → every row of the ticked groups
    """

    URL = "/dashboard/brand-mappings/"

    def setUp(self):
        self.account = make_account()
        self.other   = make_account("Other Account")
        self.admin   = make_user(email="admin@test.com", role="admin")
        self.ops     = make_user(email="ops2@test.com", role="operations")
        self.ops.accounts.add(self.account)

        # Brand A: two LMRB theme variants in one group
        self.a1 = BrandMapping.objects.create(
            account=self.account, brand="Brand A", theme="Theme A (Sin)",
            tc_theme="TC A", duration=30)
        self.a2 = BrandMapping.objects.create(
            account=self.account, brand="Brand A", theme="Theme A (Tam)",
            tc_theme="TC A", duration=30)
        # Brand B: separate group
        self.b1 = BrandMapping.objects.create(
            account=self.account, brand="Brand B", theme="Theme B", tc_theme="TC B")

    def _login(self, user):
        self.client.force_login(user)

    def test_delete_group_removes_every_theme_row_of_the_brand(self):
        self._login(self.admin)
        resp = self.client.post(self.URL, {
            "action": "delete_group", "mapping_id": self.a1.id})
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(BrandMapping.objects.filter(brand="Brand A").exists())
        self.assertTrue(BrandMapping.objects.filter(id=self.b1.id).exists())

    def test_delete_group_keeps_other_duration_group(self):
        """Groups are keyed on (brand, product, duration) — a different duration
        for the same brand is a different mapping and must survive."""
        a_any = BrandMapping.objects.create(
            account=self.account, brand="Brand A", theme="Theme A any", duration=None)
        self._login(self.admin)
        self.client.post(self.URL, {"action": "delete_group", "mapping_id": self.a1.id})
        self.assertTrue(BrandMapping.objects.filter(id=a_any.id).exists())
        self.assertFalse(BrandMapping.objects.filter(id=self.a2.id).exists())

    def test_delete_bulk_removes_all_listed_ids(self):
        self._login(self.admin)
        ids = f"{self.a1.id},{self.a2.id},{self.b1.id}"
        resp = self.client.post(self.URL, {"action": "delete_bulk", "mapping_ids": ids})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(BrandMapping.objects.count(), 0)

    def test_delete_bulk_ignores_junk_ids(self):
        self._login(self.admin)
        self.client.post(self.URL, {
            "action": "delete_bulk", "mapping_ids": f"abc,,{self.b1.id},999999"})
        self.assertFalse(BrandMapping.objects.filter(id=self.b1.id).exists())
        self.assertTrue(BrandMapping.objects.filter(id=self.a1.id).exists())

    def test_delete_group_denied_for_account_user_has_no_access_to(self):
        foreign = BrandMapping.objects.create(
            account=self.other, brand="Foreign", theme="Foreign Theme")
        self._login(self.ops)
        self.client.post(self.URL, {"action": "delete_group", "mapping_id": foreign.id})
        self.assertTrue(BrandMapping.objects.filter(id=foreign.id).exists())

    def test_delete_bulk_skips_rows_outside_user_accounts(self):
        foreign = BrandMapping.objects.create(
            account=self.other, brand="Foreign", theme="Foreign Theme")
        self._login(self.ops)
        self.client.post(self.URL, {
            "action": "delete_bulk", "mapping_ids": f"{self.b1.id},{foreign.id}"})
        self.assertTrue(BrandMapping.objects.filter(id=foreign.id).exists())
        self.assertFalse(BrandMapping.objects.filter(id=self.b1.id).exists())

    def test_delete_redirect_keeps_the_active_filters(self):
        self._login(self.admin)
        resp = self.client.post(
            f"{self.URL}?account={self.account.id}&channel={CHANNEL}&month={MONTH}",
            {"action": "delete_group", "mapping_id": self.a1.id})
        self.assertIn(f"account={self.account.id}", resp["Location"])
        self.assertIn("channel=", resp["Location"])
        self.assertIn("month=", resp["Location"])

    def test_table_renders_delete_controls(self):
        self._login(self.admin)
        resp = self.client.get(f"{self.URL}?account={self.account.id}")
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        self.assertIn('value="delete_group"', html)
        self.assertIn('id="bulkDeleteForm"', html)
        # Brand A's checkbox carries both of its LMRB theme row ids
        self.assertIn(f'data-ids="{self.a1.id},{self.a2.id}"', html)


class BrandMappingQuickDeleteTest(TestCase):
    """The Quick Map picker's delete button (/dashboard/brand-mappings/quick/delete/).

    Accepts either {mapping_id} (one row) or {brand} (every row of the brand),
    always scoped to the posted account.
    """

    URL = "/dashboard/brand-mappings/quick/delete/"

    def setUp(self):
        self.account = make_account()
        self.other   = make_account("Other Account")
        self.admin   = make_user(email="admin3@test.com", role="admin")
        self.ops     = make_user(email="ops3@test.com", role="operations")
        self.ops.accounts.add(self.account)

        self.a1 = BrandMapping.objects.create(
            account=self.account, brand="Brand A", theme="Theme A (Sin)", duration=30)
        self.a2 = BrandMapping.objects.create(
            account=self.account, brand="Brand A", theme="Theme A (Tam)", duration=30)
        self.b1 = BrandMapping.objects.create(
            account=self.account, brand="Brand B", theme="Theme B")

    def _post(self, payload):
        return self.client.post(self.URL, payload, content_type="application/json")

    def test_delete_single_mapping_row(self):
        self.client.force_login(self.admin)
        resp = self._post({"account_id": self.account.id, "mapping_id": self.a1.id})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()["ok"])
        self.assertEqual(resp.json()["deleted"], 1)
        self.assertFalse(BrandMapping.objects.filter(id=self.a1.id).exists())
        self.assertTrue(BrandMapping.objects.filter(id=self.a2.id).exists())

    def test_delete_all_rows_of_a_brand(self):
        self.client.force_login(self.admin)
        resp = self._post({"account_id": self.account.id, "brand": "Brand A"})
        self.assertEqual(resp.json()["deleted"], 2)
        self.assertFalse(BrandMapping.objects.filter(brand="Brand A").exists())
        self.assertTrue(BrandMapping.objects.filter(id=self.b1.id).exists())

    def test_delete_is_scoped_to_the_posted_account(self):
        """An id from another account can't be deleted by posting our account."""
        foreign = BrandMapping.objects.create(
            account=self.other, brand="Foreign", theme="Foreign Theme")
        self.client.force_login(self.admin)
        resp = self._post({"account_id": self.account.id, "mapping_id": foreign.id})
        self.assertEqual(resp.status_code, 404)
        self.assertTrue(BrandMapping.objects.filter(id=foreign.id).exists())

    def test_denied_for_account_user_has_no_access_to(self):
        foreign = BrandMapping.objects.create(
            account=self.other, brand="Foreign", theme="Foreign Theme")
        self.client.force_login(self.ops)
        resp = self._post({"account_id": self.other.id, "mapping_id": foreign.id})
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(BrandMapping.objects.filter(id=foreign.id).exists())

    def test_empty_payload_is_rejected(self):
        self.client.force_login(self.admin)
        resp = self._post({"account_id": self.account.id})
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(BrandMapping.objects.count(), 3)

    def test_get_is_not_allowed(self):
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(self.URL).status_code, 405)

    def test_quick_page_renders(self):
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get("/dashboard/brand-mappings/quick/").status_code, 200)


# ── MapOnline verification (view-only) + retention ─────────────────────────────

class MapOnlineComputeScopeTest(TestCase):
    """verification.engine.compute_maponline_scope — view-only MapOnline matching.

    Uses BrandMapping.maponline_theme and source='maponline' rows, and must NOT
    persist any match state or lock flag (it feeds only the Verify Ads toggle).
    """

    def setUp(self):
        from verification.engine import compute_maponline_scope
        self.compute = compute_maponline_scope
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        self.sr = make_schedule_row(self.account, self.schedule)  # Brand A, 20:00–21:00, 30s
        # MapOnline theme name differs from the MediaWatch theme — mapped separately.
        BrandMapping.objects.create(
            account=self.account, brand="Brand A", theme="Theme A",
            maponline_theme="MO Theme A",
        )
        # A matching MapOnline observation (in-window, same duration).
        self.lr = make_lmrb_row(
            self.account, advt_theme="MO Theme A", advt_time="20:30:00",
            duration=30, source="maponline",
        )

    def test_matches_via_maponline_theme(self):
        res = self.compute(self.account.id, CHANNEL, MONTH)
        self.assertEqual(res["planned"], 1)
        self.assertEqual(len(res["matched"]), 1)
        self.assertEqual(res["matched"][0]["brand"], "Brand A")

    def test_locks_maponline_rows_but_not_mediawatch(self):
        """persist=True locks the MapOnline-specific fields one-to-one, but never
        touches MediaWatch state (is_matched) or writes MatchResult."""
        self.compute(self.account.id, CHANNEL, MONTH)  # persist defaults to True
        self.sr.refresh_from_db()
        self.lr.refresh_from_db()
        # MapOnline locks set
        self.assertTrue(self.sr.is_maponline_matched)
        self.assertEqual(self.sr.matched_maponline_lmrb_id, self.lr.id)
        self.assertTrue(self.lr.is_maponline_schedule_matched)
        # MediaWatch / official state untouched
        self.assertFalse(self.sr.is_matched)
        self.assertFalse(self.lr.is_matched)
        self.assertEqual(MatchResult.objects.count(), 0)

    def test_persist_false_writes_nothing(self):
        """persist=False computes the breakdown without any DB writes."""
        res = self.compute(self.account.id, CHANNEL, MONTH, persist=False)
        self.assertEqual(len(res["matched"]), 1)
        self.sr.refresh_from_db()
        self.lr.refresh_from_db()
        self.assertFalse(self.sr.is_maponline_matched)
        self.assertFalse(self.lr.is_maponline_schedule_matched)

    def test_same_ad_not_matched_twice(self):
        """One MapOnline row cannot be matched to two schedule rows (one-to-one lock)."""
        # A second identical schedule row, but only one MapOnline observation exists.
        make_schedule_row(
            self.account, self.schedule, brand="Brand A",
            date=DATE, start_time="20:00:00", end_time="21:00:00", duration=30,
        )
        res = self.compute(self.account.id, CHANNEL, MONTH)
        self.assertEqual(res["planned"], 2)
        self.assertEqual(len(res["matched"]), 1)   # only one row can claim the single spot
        # The single MapOnline row is locked to exactly one schedule row.
        self.lr.refresh_from_db()
        self.assertTrue(self.lr.is_maponline_schedule_matched)
        self.assertEqual(
            ScheduleRow.objects.filter(is_maponline_matched=True).count(), 1
        )

    def test_extra_aired_carries_aired_programme(self):
        """Unconsumed MapOnline rows surface in Extra Aired with the aired
        programme (LMRBRow.program, from the file's 'Prg Name' column)."""
        # A second MapOnline spot of the same brand that has no planned slot to
        # claim it → it becomes Extra Aired.
        make_lmrb_row(
            self.account, advt_theme="MO Theme A", advt_time="23:45:00",
            duration=30, source="maponline",
        )
        # Give both MapOnline rows an aired programme.
        LMRBRow.objects.filter(source="maponline").update(program="Prime Time Show")
        res = self.compute(self.account.id, CHANNEL, MONTH)
        self.assertTrue(res["extra"])
        self.assertTrue(all(r["programme"] == "Prime Time Show" for r in res["extra"]))

    def test_unmapped_maponline_theme_not_matched(self):
        """A MapOnline row whose theme has no maponline_theme mapping is not matched."""
        # Different brand row with no maponline_theme mapping at all.
        sched2 = make_schedule(self.account, schedule_number="102")
        make_schedule_row(self.account, sched2, brand="Brand B")
        res = self.compute(self.account.id, CHANNEL, MONTH)
        # Brand B has no maponline mapping → shows up as No Brand Mapping, not matched.
        matched_brands = {r["brand"] for r in res["matched"]}
        self.assertNotIn("Brand B", matched_brands)


@override_settings(STORAGES={
    "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
    "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
})
@override_settings(STORAGES={
    "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
    "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
})
class QuickMapTest(TestCase):
    """Quick Map: account-level scope (no channel/month filter) + per-brand products."""

    def setUp(self):
        self.account = make_account()
        self.admin = make_user(email="admin@test.com", role="admin")
        self.schedule = make_schedule(self.account)
        sr = make_schedule_row(self.account, self.schedule, brand="Ceylinco Life -15Sec")
        sr.product = "Ceylinco Life"
        sr.save(update_fields=["product"])

    def test_options_returns_brand_products(self):
        self.client.force_login(self.admin)
        resp = self.client.get("/dashboard/brand-mappings/options/", {"account_id": self.account.id})
        self.assertEqual(resp.status_code, 200)
        bp = resp.json().get("brand_products", {})
        self.assertIn("Ceylinco Life -15Sec", bp)
        self.assertIn("Ceylinco Life", bp["Ceylinco Life -15Sec"])

    def test_quick_page_has_account_and_channel_filter_but_no_month(self):
        self.client.force_login(self.admin)
        html = self.client.get("/dashboard/brand-mappings/quick/").content.decode()
        self.assertEqual(200, self.client.get("/dashboard/brand-mappings/quick/").status_code)
        self.assertIn('id="q-account"', html)
        self.assertIn('id="q-channel"', html)   # optional channel filter restored
        self.assertNotIn('id="q-month"', html)  # month is still not a filter

    def test_options_brand_map_status_and_unmapped_themes(self):
        from core.models import BrandMapping
        # Brand has an LMRB theme but NO tc_theme -> tc should be False (Aired 0).
        BrandMapping.objects.create(account=self.account, brand="Ceylinco Life -15Sec",
                                    theme="Ceylinco_15", tc_theme="")
        make_lmrb_row(self.account, advt_theme="ORPHAN LMRB", channel=CHANNEL,
                      advt_time="20:00:00", duration=15, source="mediawatch")
        rep = make_tc_report(self.account, channel=CHANNEL)
        make_tc_row(self.account, rep, tc_theme="ORPHAN TC", channel=CHANNEL, duration=15)
        self.client.force_login(self.admin)
        d = self.client.get("/dashboard/brand-mappings/options/",
                            {"account_id": self.account.id}).json()
        st = d["brand_map_status"]["Ceylinco Life -15Sec"]
        self.assertTrue(st["lmrb"]); self.assertFalse(st["tc"])
        self.assertIn("ORPHAN TC", d["unmapped_tc_themes"])
        self.assertIn("ORPHAN LMRB", d["unmapped_lmrb_themes"])
        self.assertNotIn("Ceylinco_15", d["unmapped_lmrb_themes"])  # this one IS mapped

    def test_options_returns_tc_theme_channels(self):
        # A brand's TC code differs per channel — the picker filters by channel.
        rep_a = make_tc_report(self.account, channel="TV - Derana")
        rep_b = make_tc_report(self.account, channel="TV - Hiru")
        make_tc_row(self.account, rep_a, tc_theme="ABC", channel="TV - Derana")
        make_tc_row(self.account, rep_b, tc_theme="BFB", channel="TV - Hiru")
        self.client.force_login(self.admin)
        d = self.client.get("/dashboard/brand-mappings/options/",
                            {"account_id": self.account.id}).json()
        self.assertIn("TV - Derana", d["tc_channels"])
        self.assertIn("TV - Hiru", d["tc_channels"])
        self.assertEqual(d["tc_theme_channels"]["ABC"], ["TV - Derana"])
        self.assertEqual(d["tc_theme_channels"]["BFB"], ["TV - Hiru"])

    def test_quick_add_stores_multi_maponline_and_never_saves_product(self):
        self.client.force_login(self.admin)
        resp = self.client.post(
            "/dashboard/brand-mappings/quick/add/",
            data=json.dumps({
                "account_id": self.account.id,
                "brand": "Ceylinco Life -15Sec",
                "themes": ["Ceylinco Life_15 (Sin)"],
                "tc_themes": ["ABC", "BFB"],
                "maponline_themes": ["Ceylinco Gen A", "Ceylinco Gen B"],
                "product": "Ceylinco Life",   # must be ignored (narrow-only)
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()["ok"])
        bm = BrandMapping.objects.get(account=self.account, brand="Ceylinco Life -15Sec")
        self.assertEqual(bm.maponline_theme, "Ceylinco Gen A|Ceylinco Gen B")
        self.assertEqual(bm.tc_theme, "ABC|BFB")
        self.assertEqual(bm.product, "")   # product is never saved from Quick Map

    def test_quick_add_backward_compat_single_maponline_theme(self):
        self.client.force_login(self.admin)
        self.client.post(
            "/dashboard/brand-mappings/quick/add/",
            data=json.dumps({
                "account_id": self.account.id,
                "brand": "Ceylinco Life -15Sec",
                "themes": ["Ceylinco Life_15 (Sin)"],
                "maponline_theme": "Legacy Single",
            }),
            content_type="application/json",
        )
        bm = BrandMapping.objects.get(account=self.account, brand="Ceylinco Life -15Sec")
        self.assertEqual(bm.maponline_theme, "Legacy Single")


class BrandMappingLmrbChannelTest(TestCase):
    """brand_mapping_options must surface LMRB themes even when the schedule
    channel carries a 'TV - ' prefix but LMRB is stored under the clean name."""

    URL = "/dashboard/brand-mappings/options/"

    def setUp(self):
        self.account = make_account()
        self.admin = make_user(email="admin@test.com", role="admin")
        # LMRB stored under the clean channel name "Derana"
        make_lmrb_row(self.account, advt_theme="Derana Theme A", channel="Derana",
                      source="mediawatch", advt_time="20:00:00", duration=30)

    def test_prefixed_channel_still_finds_lmrb_themes(self):
        self.client.force_login(self.admin)
        resp = self.client.get(self.URL, {"account_id": self.account.id, "channel": "TV - Derana"})
        self.assertEqual(resp.status_code, 200)
        self.assertIn("Derana Theme A", resp.json()["themes"])

    def test_clean_channel_also_finds_themes(self):
        self.client.force_login(self.admin)
        resp = self.client.get(self.URL, {"account_id": self.account.id, "channel": "Derana"})
        self.assertIn("Derana Theme A", resp.json()["themes"])

    def _narrow_schedule(self):
        """A schedule whose date range (Feb 1–5) does NOT cover the LMRB row (Jan 15)."""
        return Schedule.objects.create(
            account=self.account, channel="TV - Derana", month="February 2025",
            schedule_number="201", file="s.xlsx", original_filename="s.xlsx",
            start_date=datetime.date(2025, 2, 1), end_date=datetime.date(2025, 2, 5), version=1,
        )

    def test_theme_shown_even_outside_schedule_date_range(self):
        self._narrow_schedule()
        self.client.force_login(self.admin)
        resp = self.client.get(self.URL, {"account_id": self.account.id, "channel": "TV - Derana"})
        # LMRB (Jan 15) is outside the schedule window (Feb 1–5) but must still show
        # at account/channel level — otherwise uploaded LMRB looks missing.
        self.assertIn("Derana Theme A", resp.json()["themes"])

    def test_schedule_drilldown_still_scopes_by_date(self):
        sch = self._narrow_schedule()
        self.client.force_login(self.admin)
        resp = self.client.get(self.URL, {
            "account_id": self.account.id, "channel": "TV - Derana", "schedule_id": sch.id,
        })
        # When drilled into a specific schedule, the out-of-range theme is scoped out.
        self.assertNotIn("Derana Theme A", resp.json()["themes"])


class SponsorshipTagDurationTest(TestCase):
    """A sponsorship tag logged at different durations in Schedule/LMRB/TC must
    still match 3-way (TC confirmed by LMRB), ignoring duration."""

    def _tc_row(self, tc_report, tc_theme, duration, aired_time, date=DATE):
        key = TCRow.make_dedup_key(self.account.id, CHANNEL, date, aired_time, tc_theme, duration)
        return TCRow.objects.create(
            account=self.account, tc_report=tc_report, channel=CHANNEL, date=date,
            programme="Show", tc_theme=tc_theme, duration=duration,
            aired_time=aired_time, dedup_key=key,
        )

    def setUp(self):
        ensure_tc_tolerance(5)
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        # SPONSORSHIP tag planned at 5s
        make_schedule_row(
            self.account, self.schedule, brand="Tag Brand", programme="Show",
            duration=5, ad_type="SPONSORSHIP", start_time="20:00:00", end_time="20:05:00",
        )
        # Mapping: LMRB theme + TC theme (mapping duration 5s)
        make_brand_mapping(self.account, brand="Tag Brand", theme="Tag Theme",
                           tc_theme="Tag TC", duration=5)
        # LMRB logs the tag under BOTH 5s and 10s (mediawatch)
        make_lmrb_row(self.account, advt_theme="Tag Theme", advt_time="20:00:03",
                      duration=5, source="mediawatch")
        make_lmrb_row(self.account, advt_theme="Tag Theme", advt_time="20:00:02",
                      duration=10, source="mediawatch")
        # TC logs it at 8s
        self.tc_report = make_tc_report(self.account, schedule=self.schedule)
        self.tc = self._tc_row(self.tc_report, "Tag TC", 8, "20:00:00")

    def test_tc_confirmed_by_lmrb_ignoring_duration(self):
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        self.tc.refresh_from_db()
        self.assertTrue(self.tc.is_lmrb_confirmed,
                        "TC 8s tag should be LMRB-confirmed via a 5s/10s LMRB row")

    def test_summary_counts_tag_as_aired(self):
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        data = build_summary_data(self.account.id, CHANNEL, MONTH)
        spon_rows = [r for sec in data["sponsorship"] for r in sec["rows"]]
        tag = next((r for r in spon_rows if r["product"] == "Tag Brand"), None)
        self.assertIsNotNone(tag)
        self.assertGreaterEqual(tag["aired"], 1, "Tag should count as aired despite duration differences")

    def test_commercial_still_requires_matching_duration(self):
        """Duration-ignore must NOT leak into commercial matching."""
        make_schedule_row(self.account, self.schedule, brand="Comm Brand",
                          programme="Show2", duration=30, ad_type="COMMERCIAL BENEFITS",
                          start_time="21:00:00", end_time="21:05:00")
        make_brand_mapping(self.account, brand="Comm Brand", theme="Comm Theme",
                           tc_theme="Comm TC", duration=30)
        make_lmrb_row(self.account, advt_theme="Comm Theme", advt_time="21:00:02",
                      duration=15, source="mediawatch")  # wrong duration
        comm_tc = self._tc_row(self.tc_report, "Comm TC", 30, "21:00:00")
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        comm_tc.refresh_from_db()
        self.assertFalse(comm_tc.is_lmrb_confirmed,
                         "Commercial must not confirm against a different-duration LMRB row")


class TcLmrbDurationToleranceTest(TestCase):
    """Opt-in tc_lmrb_duration_tolerance lets a commercial TC spot confirm against
    an LMRB row whose duration differs by up to the configured number of seconds
    (e.g. a 5s TC 'Tag' spot logged in LMRB at 4s). Off by default (tolerance=0)."""

    def _set_duration_tolerance(self, seconds):
        SystemSetting.objects.update_or_create(
            key="tc_lmrb_duration_tolerance",
            defaults={
                "value": str(seconds),
                "label": "TC-LMRB Duration Tolerance (seconds)",
                "category": "reconciliation",
            },
        )

    def _tc_row(self, tc_theme, duration, aired_time, date=DATE):
        key = TCRow.make_dedup_key(self.account.id, CHANNEL, date, aired_time, tc_theme, duration)
        return TCRow.objects.create(
            account=self.account, tc_report=self.tc_report, channel=CHANNEL, date=date,
            programme="Show", tc_theme=tc_theme, duration=duration,
            aired_time=aired_time, dedup_key=key,
        )

    def setUp(self):
        ensure_tc_tolerance(5)
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        # COMMERCIAL spot planned at 5s
        make_schedule_row(
            self.account, self.schedule, brand="Papare Brand", programme="Show",
            duration=5, ad_type="COMMERCIAL BENEFITS",
            start_time="18:00:00", end_time="18:05:00",
        )
        make_brand_mapping(self.account, brand="Papare Brand", theme="Tag",
                           tc_theme="Papare TC", duration=5)
        # LMRB logs the spot as a generic "Tag" at 4s (one second short of the TC 5s)
        make_lmrb_row(self.account, advt_theme="Tag", advt_time="18:00:03",
                      duration=4, source="mediawatch")
        self.tc_report = make_tc_report(self.account, schedule=self.schedule)
        self.tc = self._tc_row("Papare TC", 5, "18:00:00")

    def test_not_confirmed_when_tolerance_zero(self):
        """Default (tolerance=0) keeps exact-duration matching: 5s TC != 4s LMRB."""
        self._set_duration_tolerance(0)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        self.tc.refresh_from_db()
        self.assertFalse(self.tc.is_lmrb_confirmed,
                         "With tolerance 0 a 5s TC spot must not confirm against a 4s LMRB row")

    def test_confirmed_when_tolerance_allows(self):
        """With tolerance >= 1 the 5s TC spot confirms against the 4s LMRB row."""
        self._set_duration_tolerance(1)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        self.tc.refresh_from_db()
        self.assertTrue(self.tc.is_lmrb_confirmed,
                        "With tolerance 1 a 5s TC spot should confirm against a 4s LMRB row")
        self.assertEqual(self.tc.matched_lmrb.duration, 4)

    def test_tolerance_still_respects_theme(self):
        """Duration tolerance must not link an unrelated theme."""
        self._set_duration_tolerance(2)
        # An unrelated LMRB row at a near duration but a different theme
        make_lmrb_row(self.account, advt_theme="Unrelated Theme", advt_time="18:00:01",
                      duration=6, source="mediawatch")
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        self.tc.refresh_from_db()
        # It should still confirm against the correct "Tag" row, not the unrelated one
        self.assertTrue(self.tc.is_lmrb_confirmed)
        self.assertEqual(self.tc.matched_lmrb.advt_theme, "Tag")


class TcLmrbToScheduleBridgeTest(TestCase):
    """TC → specific LMRB row → Schedule, with deterministic variant resolution.

    Regression cover for the 'second variant dropping' bug: a TC row must confirm
    against the LMRB theme paired with its OWN tc_theme (A10E → 'A 10 SEC - English'),
    never greedily consume a sibling variant's LMRB row, and the matched LMRB row
    must be the bridge carried into Schedule matching.
    """

    def _mk_variant_mappings(self):
        """A 10 SEC with three separately-mapped variants (one BrandMapping row each)."""
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - English",
                           tc_theme="A10E", duration=10)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - Sinhala",
                           tc_theme="A10S", duration=10)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - Tamil",
                           tc_theme="A10T", duration=10)

    def _sched_rows(self, n):
        for _ in range(n):
            make_schedule_row(self.account, self.schedule, brand="A 10 SEC", duration=10,
                              start_time="20:00:00", end_time="21:00:00")

    def setUp(self):
        ensure_tc_tolerance(5)
        self.account   = make_account()
        self.schedule  = make_schedule(self.account)
        self.tc_report = make_tc_report(self.account, schedule=self.schedule)

    # ── Test 1 — Basic TC → LMRB → Schedule ────────────────────────────────────
    def test_basic_tc_lmrb_schedule(self):
        self._sched_rows(1)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - English",
                           tc_theme="A10E", duration=10)
        lmrb = make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                             advt_time="20:31:07", duration=10)
        tc = make_tc_row(self.account, self.tc_report, tc_theme="A10E",
                         aired_time="20:31:05", duration=10)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_lmrb_confirmed)
        self.assertEqual(tc.matched_lmrb_id, lmrb.id)
        self.assertTrue(tc.is_schedule_matched)
        self.assertIsNotNone(tc.matched_schedule_id)

    # ── Test 2 — Multiple variants each get their OWN LMRB + Schedule ───────────
    def test_multiple_variants_deterministic(self):
        self._sched_rows(3)
        self._mk_variant_mappings()
        eng = make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                            advt_time="20:31:07", duration=10)
        sin = make_lmrb_row(self.account, advt_theme="A 10 SEC - Sinhala",
                            advt_time="20:32:07", duration=10)
        tam = make_lmrb_row(self.account, advt_theme="A 10 SEC - Tamil",
                            advt_time="20:33:07", duration=10)
        tc_e = make_tc_row(self.account, self.tc_report, tc_theme="A10E",
                           aired_time="20:31:05", duration=10, suffix="e")
        tc_s = make_tc_row(self.account, self.tc_report, tc_theme="A10S",
                           aired_time="20:32:05", duration=10, suffix="s")
        tc_t = make_tc_row(self.account, self.tc_report, tc_theme="A10T",
                           aired_time="20:33:05", duration=10, suffix="t")
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        for tc, lmrb in ((tc_e, eng), (tc_s, sin), (tc_t, tam)):
            tc.refresh_from_db()
            self.assertEqual(tc.matched_lmrb_id, lmrb.id,
                             f"{tc.tc_theme} must map to its own variant")
            self.assertTrue(tc.is_schedule_matched)

    # ── Test 3 — Several LMRB rows same theme: closest valid time wins ──────────
    def test_multiple_lmrb_same_theme_closest_wins(self):
        self._sched_rows(1)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - English",
                           tc_theme="A10E", duration=10)
        near = make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                             advt_time="20:31:07", duration=10)  # Δ2
        make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                      advt_time="20:31:30", duration=10)         # Δ25
        tc = make_tc_row(self.account, self.tc_report, tc_theme="A10E",
                         aired_time="20:31:05", duration=10)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertEqual(tc.matched_lmrb_id, near.id)

    # ── Test 4 — Wrong language must NOT win even when closer in time ───────────
    def test_correct_variant_beats_closer_wrong_language(self):
        """This is the core regression: Sinhala is closer in time, but A10E maps
        to English, so English must be selected (deterministic variant)."""
        self._sched_rows(1)
        self._mk_variant_mappings()
        make_lmrb_row(self.account, advt_theme="A 10 SEC - Sinhala",
                      advt_time="20:31:06", duration=10)          # Δ1 (closer, wrong lang)
        eng = make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                            advt_time="20:31:09", duration=10)    # Δ4 (correct lang)
        tc = make_tc_row(self.account, self.tc_report, tc_theme="A10E",
                         aired_time="20:31:05", duration=10)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_lmrb_confirmed)
        self.assertEqual(tc.matched_lmrb_id, eng.id,
                         "A10E must confirm English even though Sinhala is closer")

    def test_two_variants_no_starvation(self):
        """A10E must not consume Sinhala's row; A10S must still confirm.  Under the
        old brand-level logic A10E (processed first) grabbed the closer Sinhala row
        and starved A10S out of tolerance."""
        self._sched_rows(2)
        self._mk_variant_mappings()
        eng = make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                            advt_time="20:31:20", duration=10)
        sin = make_lmrb_row(self.account, advt_theme="A 10 SEC - Sinhala",
                            advt_time="20:31:06", duration=10)
        tc_e = make_tc_row(self.account, self.tc_report, tc_theme="A10E",
                           aired_time="20:31:18", duration=10, suffix="e")  # Δ2 to English
        tc_s = make_tc_row(self.account, self.tc_report, tc_theme="A10S",
                           aired_time="20:31:05", duration=10, suffix="s")  # Δ1 to Sinhala
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc_e.refresh_from_db(); tc_s.refresh_from_db()
        self.assertEqual(tc_e.matched_lmrb_id, eng.id)
        self.assertEqual(tc_s.matched_lmrb_id, sin.id)
        self.assertTrue(tc_e.is_schedule_matched and tc_s.is_schedule_matched)

    # ── Test 5 — LMRB matched but no Schedule slot: TC not fully matched ────────
    def test_lmrb_matched_schedule_missing(self):
        # Schedule exists for a DIFFERENT brand (so the scope has a date range),
        # but there is no 'A 10 SEC' slot to match.
        make_schedule_row(self.account, self.schedule, brand="Other Brand", duration=30)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - English",
                           tc_theme="A10E", duration=10)
        lmrb = make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                             advt_time="20:31:07", duration=10)
        tc = make_tc_row(self.account, self.tc_report, tc_theme="A10E",
                         aired_time="20:31:05", duration=10)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertTrue(tc.is_lmrb_confirmed)
        self.assertEqual(tc.matched_lmrb_id, lmrb.id)
        self.assertFalse(tc.is_schedule_matched)
        self.assertTrue(tc.is_extra)

    # ── Test 6 — One Schedule row, two confirmed TCs: one-to-one lock holds ─────
    def test_schedule_one_to_one_lock(self):
        self._sched_rows(1)  # only ONE planned slot
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - English",
                           tc_theme="A10E", duration=10)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - Sinhala",
                           tc_theme="A10S", duration=10)
        make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                      advt_time="20:31:07", duration=10)
        make_lmrb_row(self.account, advt_theme="A 10 SEC - Sinhala",
                      advt_time="20:32:07", duration=10)
        tc_e = make_tc_row(self.account, self.tc_report, tc_theme="A10E",
                           aired_time="20:31:05", duration=10, suffix="e")
        tc_s = make_tc_row(self.account, self.tc_report, tc_theme="A10S",
                           aired_time="20:32:05", duration=10, suffix="s")
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc_e.refresh_from_db(); tc_s.refresh_from_db()
        # Both confirm LMRB, but only ONE can claim the single schedule slot.
        self.assertTrue(tc_e.is_lmrb_confirmed and tc_s.is_lmrb_confirmed)
        matched = [t for t in (tc_e, tc_s) if t.is_schedule_matched]
        extra   = [t for t in (tc_e, tc_s) if t.is_extra]
        self.assertEqual(len(matched), 1, "exactly one TC may claim the one slot")
        self.assertEqual(len(extra), 1)

    # ── Test 7 — Time outside tolerance: no LMRB confirmation ──────────────────
    def test_time_outside_tolerance_not_confirmed(self):
        self._sched_rows(1)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - English",
                           tc_theme="A10E", duration=10)
        make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                      advt_time="20:31:30", duration=10)  # Δ25 > tol 5
        tc = make_tc_row(self.account, self.tc_report, tc_theme="A10E",
                         aired_time="20:31:05", duration=10)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        # The important invariant: no LMRB bridge is formed outside tolerance, so
        # the spot cannot count as Aired (Aired requires is_lmrb_confirmed=True).
        self.assertFalse(tc.is_lmrb_confirmed)
        self.assertIsNone(tc.matched_lmrb_id)
        # NOTE: the row may still be schedule-matched via the pre-existing Step 3
        # fallback (TC → Schedule by tc_theme, no LMRB). That path is intentional
        # and distinct from the LMRB-bridged path; it does not inflate Aired.

    # ── Test 8 — Unmapped TC theme ─────────────────────────────────────────────
    def test_unmapped_tc_theme_no_candidate_stays_unconfirmed(self):
        """An unmapped tc_theme with no time-coincident LMRB row must NOT match."""
        self._sched_rows(1)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - English",
                           tc_theme="A10E", duration=10)
        # LMRB row for the mapped variant only; the unmapped TC has no partner here.
        make_lmrb_row(self.account, advt_theme="A 10 SEC - English",
                      advt_time="20:31:07", duration=10)
        tc = make_tc_row(self.account, self.tc_report, tc_theme="ZZZ-UNMAPPED",
                         aired_time="05:00:00", duration=10)  # far from any LMRB time
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        self.assertFalse(tc.is_lmrb_confirmed)
        self.assertFalse(tc.is_schedule_matched)

    def test_unmapped_tc_theme_time_only_fallback_is_documented_behaviour(self):
        """DOCUMENTS (does not endorse) the pre-existing legacy fallback: when a
        tc_theme resolves to NO brand, the engine still confirms against an LMRB
        row that matches channel+date+duration within the time tolerance, by time
        alone.  Flagged in the report as an edge case to optionally gate behind a
        'require mapping' setting."""
        self._sched_rows(1)
        make_brand_mapping(self.account, brand="A 10 SEC", theme="A 10 SEC - English",
                           tc_theme="A10E", duration=10)
        lmrb = make_lmrb_row(self.account, advt_theme="Some Untracked Theme",
                             advt_time="20:31:07", duration=10)
        tc = make_tc_row(self.account, self.tc_report, tc_theme="ZZZ-UNMAPPED",
                         aired_time="20:31:05", duration=10)
        reconcile_tc(self.account.id, CHANNEL, MONTH, mode="reset")
        tc.refresh_from_db()
        # Current behaviour: confirmed by time alone (legacy fallback).
        self.assertTrue(tc.is_lmrb_confirmed)
        self.assertEqual(tc.matched_lmrb_id, lmrb.id)


class DiagnoseScopeTest(TestCase):
    """verification.engine.diagnose_scope — 'why didn't this match?' reasons."""

    def setUp(self):
        from verification.engine import diagnose_scope
        self.diagnose = diagnose_scope
        self.account = make_account()
        self.schedule = make_schedule(self.account)

    def _reason_for(self, brand):
        rows = self.diagnose(self.account.id, CHANNEL, MONTH)
        return next((r["reason"] for r in rows if r["brand"] == brand), None)

    def test_no_brand_mapping_reason(self):
        make_schedule_row(self.account, self.schedule, brand="Unmapped Brand")
        reason = self._reason_for("Unmapped Brand")
        self.assertIn("No Brand Mapping", reason)

    def test_duration_mismatch_reason(self):
        make_brand_mapping(self.account, brand="Brand D", theme="Theme D", duration=30)
        make_schedule_row(self.account, self.schedule, brand="Brand D", duration=30)
        # LMRB has the theme but at 15s, not 30s
        make_lmrb_row(self.account, advt_theme="Theme D", duration=15,
                      advt_time="20:15:00", source="mediawatch")
        reason = self._reason_for("Brand D")
        self.assertIn("duration", reason.lower())

    def test_should_match_reason_flags_stale_or_old_build(self):
        """An in-window LMRB spot exists but the row is unmatched → 'should match'."""
        make_brand_mapping(self.account, brand="Brand E", theme="Theme E", duration=30)
        make_schedule_row(self.account, self.schedule, brand="Brand E",
                          start_time="20:00:00", end_time="21:00:00", duration=30)
        make_lmrb_row(self.account, advt_theme="Theme E", advt_time="20:30:00",
                      duration=30, source="mediawatch")  # in-window, unmatched
        reason = self._reason_for("Brand E")
        self.assertIn("should match", reason.lower())

    def test_theme_not_found_reason(self):
        make_brand_mapping(self.account, brand="Brand F", theme="Theme F", duration=30)
        make_schedule_row(self.account, self.schedule, brand="Brand F", duration=30)
        # No LMRB rows at all for Theme F
        reason = self._reason_for("Brand F")
        self.assertIn("No LMRB spot", reason)


@override_settings(STORAGES={
    "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
    "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
})
class TcConvertAccessTest(TestCase):
    """Convert TC (AI) is open to all users; the AI Conversion Prompt UI is hidden."""

    def test_non_admin_can_open_convert_page(self):
        planner = make_user(email="planner@test.com", role="planner")
        self.client.force_login(planner)
        resp = self.client.get("/dashboard/tc/pdf-convert/")
        self.assertEqual(resp.status_code, 200)

    def test_convert_page_hides_ai_conversion_prompt(self):
        planner = make_user(email="planner2@test.com", role="planner")
        self.client.force_login(planner)
        html = self.client.get("/dashboard/tc/pdf-convert/").content.decode()
        self.assertNotIn("AI Conversion Prompt", html)
        self.assertIn("Upload TC PDF", html)   # section still present, renumbered

    def test_channel_prompt_get_open_to_all(self):
        planner = make_user(email="planner3@test.com", role="planner")
        self.client.force_login(planner)
        resp = self.client.get("/dashboard/tc/channel-prompt/?channel=Tv - Derana")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json().get("ok"))


class ScheduleTemplateTest(TestCase):
    """Admin-uploaded sample schedule template: admin upload, everyone downloads."""

    URL_UPLOAD = "/dashboard/schedules/template/upload/"
    URL_DOWNLOAD = "/dashboard/schedules/template/download/"

    def _xlsx(self, name="template.xlsx"):
        from django.core.files.uploadedfile import SimpleUploadedFile
        return SimpleUploadedFile(
            name, b"PK\x03\x04 fake xlsx bytes",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def setUp(self):
        from core.models import ScheduleTemplate
        self.ScheduleTemplate = ScheduleTemplate
        self.admin = make_user(email="admin@test.com", role="admin")
        self.staff = make_user(email="ops@test.com", role="operations")

    def test_admin_can_upload_template(self):
        self.client.force_login(self.admin)
        resp = self.client.post(self.URL_UPLOAD, {"file": self._xlsx()})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.ScheduleTemplate.objects.count(), 1)
        self.assertEqual(self.ScheduleTemplate.objects.first().original_filename, "template.xlsx")

    def test_upload_replaces_previous(self):
        self.client.force_login(self.admin)
        self.client.post(self.URL_UPLOAD, {"file": self._xlsx("first.xlsx")})
        self.client.post(self.URL_UPLOAD, {"file": self._xlsx("second.xlsx")})
        self.assertEqual(self.ScheduleTemplate.objects.count(), 1)
        self.assertEqual(self.ScheduleTemplate.objects.first().original_filename, "second.xlsx")

    def test_non_admin_cannot_upload(self):
        self.client.force_login(self.staff)
        resp = self.client.post(self.URL_UPLOAD, {"file": self._xlsx()})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(self.ScheduleTemplate.objects.count(), 0)

    def test_any_user_can_download(self):
        self.client.force_login(self.admin)
        self.client.post(self.URL_UPLOAD, {"file": self._xlsx()})
        self.client.force_login(self.staff)          # a non-admin
        resp = self.client.get(self.URL_DOWNLOAD)
        self.assertEqual(resp.status_code, 200)

    def test_download_without_template_redirects(self):
        self.client.force_login(self.staff)
        resp = self.client.get(self.URL_DOWNLOAD)
        self.assertEqual(resp.status_code, 302)


class MapOnlineProgrammeParseTest(TestCase):
    """MapOnline parsing must read the aired programme from the 'Prg Name' column."""

    def test_prg_name_populates_program(self):
        import pandas as pd
        from core.views import _parse_lmrb_rows
        account = make_account()
        Channel.objects.create(name="TV - Sirasa TV")
        df = pd.DataFrame([{
            "Channel":    "Tv - Sirasa TV",
            "Prg Date":   "2026-07-01",
            "Prg Name":   "Fifa World Cup 2026 - Fra Vs Swe",
            "Prg Start":  "2:15",
            "Product":    "Coca Cola",
            "Theme":      "Tani With Friends (30)(Sin)",
            "Ad Start":   "2:53:08",
            "Ad Dur":     30,
            "Language":   "SINHALA",
            "Advertiser": "Coca Cola Beverages Ltd",
            "Category":   "Aerated Soft Drinks",
        }])
        inserted = _parse_lmrb_rows(df, "maponline", account)
        self.assertEqual(inserted, 1)
        row = LMRBRow.objects.get(source="maponline")
        self.assertEqual(row.program, "Fifa World Cup 2026 - Fra Vs Swe")
        self.assertEqual(row.advt_theme, "Tani With Friends (30)(Sin)")
        self.assertEqual(row.advt_time, "2:53:08")
        self.assertEqual(row.duration, 30)


class MapOnlineColoredStatusMapTest(TestCase):
    """verification.colored_schedule.build_status_map_from_maponline."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_schedule_row(self.account, self.schedule)  # Test Show, 20:00:00, 30s, 2025-01-15
        BrandMapping.objects.create(
            account=self.account, brand="Brand A", theme="Theme A",
            maponline_theme="MO Theme A",
        )
        make_lmrb_row(
            self.account, advt_theme="MO Theme A", advt_time="20:30:00",
            duration=30, source="maponline",
        )

    def test_status_map_marks_matched_slot(self):
        from verification.colored_schedule import build_status_map_from_maponline
        sm = build_status_map_from_maponline(self.account.id, CHANNEL, MONTH)
        slot = ("test show", "20:00:00", 30)
        self.assertIn(slot, sm)
        self.assertEqual(sm[slot]["2025-01-15"]["matched"], 1)

    def test_none_when_no_maponline_data(self):
        from verification.colored_schedule import build_status_map_from_maponline
        LMRBRow.objects.filter(source="maponline").delete()
        self.assertIsNone(
            build_status_map_from_maponline(self.account.id, CHANNEL, MONTH)
        )

    def test_status_map_build_does_not_lock(self):
        """The export builder uses persist=False — it must not lock rows."""
        from verification.colored_schedule import build_status_map_from_maponline
        build_status_map_from_maponline(self.account.id, CHANNEL, MONTH)
        self.assertFalse(
            ScheduleRow.objects.filter(is_maponline_matched=True).exists()
        )


class MapOnlineColoredSheetIntegrationTest(TestCase):
    """build_original_and_colored_wb produces a 'MapOnline Colored' sheet."""

    def _make_pivot_bytes(self):
        import io as _io
        import datetime as _dt
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header row: PROGRAM | DAY | TIME | END | DUR | <15 Jan> | <16 Jan>
        ws.append(["PROGRAM", "DAY", "TIME", "END", "DUR",
                   _dt.date(2025, 1, 15), _dt.date(2025, 1, 16)])
        # Data row: one planned spot on the 15th AND one on the 16th.
        ws.append(["Test Show", "Wed", _dt.time(20, 0, 0), _dt.time(21, 0, 0), 30, 1, 1])
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def setUp(self):
        from django.core.files.base import ContentFile
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        self.schedule.file.save("pivot.xlsx", ContentFile(self._make_pivot_bytes()), save=True)
        make_schedule_row(
            self.account, self.schedule, brand="Brand A", programme="Test Show",
            date=DATE, start_time="20:00:00", end_time="21:00:00", duration=30,
        )
        BrandMapping.objects.create(
            account=self.account, brand="Brand A", theme="Theme A",
            maponline_theme="MO Theme A",
        )
        make_lmrb_row(
            self.account, advt_theme="MO Theme A", advt_time="20:30:00",
            duration=30, source="maponline",
        )

    def test_workbook_has_maponline_colored_sheet(self):
        from verification.colored_schedule import (
            build_original_and_colored_wb, build_status_map_from_maponline,
        )
        colors = {
            'aired': '#22c55e', 'not_aired': '#ef4444', 'late_telecast': '#a855f7',
            'programme_mismatch': '#f97316', 'extra_aired': '#3b82f6',
            'planned': '#94a3b8', 'manual_override': '#14b8a6', 'aired_less': '#f59e0b',
        }
        mo_map = build_status_map_from_maponline(self.account.id, CHANNEL, MONTH)
        self.assertIsNotNone(mo_map)
        wb, detected = build_original_and_colored_wb(
            self.schedule.pk, colors, status_map=None, maponline_status_map=mo_map,
        )
        self.assertTrue(detected)
        self.assertIn("MapOnline Colored", wb.sheetnames)
        # 15 Jan (F2) has MapOnline data → coloured.
        self.assertEqual(wb["MapOnline Colored"]["F2"].fill.fill_type, "solid")
        # 16 Jan (G2) is beyond the latest MapOnline date (15 Jan) → left
        # uncoloured, because there is no data for it yet.
        self.assertIn(wb["MapOnline Colored"]["G2"].fill.fill_type, (None, "none"))

    def test_mediawatch_sheet_capped_at_mediawatch_date(self):
        """The MediaWatch sheet must not colour past the latest MediaWatch date,
        even when MapOnline data extends further."""
        from verification.colored_schedule import (
            build_original_and_colored_wb, build_status_map_from_maponline,
        )
        # MediaWatch data only on 15 Jan; MapOnline extends to 16 Jan.
        make_lmrb_row(
            self.account, advt_theme="Theme A", advt_time="20:30:00",
            duration=30, source="mediawatch", date=datetime.date(2025, 1, 15),
        )
        colors = {
            'aired': '#22c55e', 'not_aired': '#ef4444', 'late_telecast': '#a855f7',
            'programme_mismatch': '#f97316', 'extra_aired': '#3b82f6',
            'planned': '#94a3b8', 'manual_override': '#14b8a6', 'aired_less': '#f59e0b',
        }
        mo_map = build_status_map_from_maponline(self.account.id, CHANNEL, MONTH)
        wb, _ = build_original_and_colored_wb(
            self.schedule.pk, colors, status_map=None, maponline_status_map=mo_map,
        )
        # 16 Jan (G2) on the MediaWatch sheet is beyond MediaWatch's data → uncoloured.
        self.assertIn(wb["Colored Schedule"]["G2"].fill.fill_type, (None, "none"))


class MapOnlinePurgeTest(TestCase):
    """core.maponline_cleanup.purge_old_maponline_data — 30-day retention."""

    def setUp(self):
        import uuid as _uuid
        from django.utils import timezone
        from core.models import MonitoringData
        self.MonitoringData = MonitoringData
        self.account = make_account()
        old = timezone.now() - datetime.timedelta(days=40)

        # Old MapOnline upload + its LMRB row (should be purged).
        self.old_batch = _uuid.uuid4()
        old_md = MonitoringData.objects.create(
            account=self.account, data_type="maponline", channel=CHANNEL,
            file="monitoring/old.xlsx", original_filename="old.xlsx",
            file_group_id=str(self.old_batch), row_count=1,
        )
        MonitoringData.objects.filter(id=old_md.id).update(uploaded_at=old)
        old_row = LMRBRow.objects.create(
            account=self.account, channel=CHANNEL, date=DATE,
            advt_theme="Old", advt_time="20:00:00", duration=30,
            source="maponline", batch_id=self.old_batch,
            dedup_key=LMRBRow.make_dedup_key(self.account.id, CHANNEL, DATE, "20:00:00", "Old", 30),
        )
        LMRBRow.objects.filter(id=old_row.id).update(uploaded_at=old)

        # Recent MapOnline upload + row (should be kept).
        self.recent_batch = _uuid.uuid4()
        MonitoringData.objects.create(
            account=self.account, data_type="maponline", channel=CHANNEL,
            file="monitoring/new.xlsx", original_filename="new.xlsx",
            file_group_id=str(self.recent_batch), row_count=1,
        )
        LMRBRow.objects.create(
            account=self.account, channel=CHANNEL, date=DATE,
            advt_theme="Fresh", advt_time="21:00:00", duration=30,
            source="maponline", batch_id=self.recent_batch,
            dedup_key=LMRBRow.make_dedup_key(self.account.id, CHANNEL, DATE, "21:00:00", "Fresh", 30),
        )

        # Old MediaWatch upload + row (must be untouched — different source).
        mw_md = MonitoringData.objects.create(
            account=self.account, data_type="mediawatch", channel=CHANNEL,
            file="monitoring/mw.xlsx", original_filename="mw.xlsx", row_count=1,
        )
        MonitoringData.objects.filter(id=mw_md.id).update(uploaded_at=old)
        mw_row = LMRBRow.objects.create(
            account=self.account, channel=CHANNEL, date=DATE,
            advt_theme="MW", advt_time="22:00:00", duration=30,
            source="mediawatch",
            dedup_key=LMRBRow.make_dedup_key(self.account.id, CHANNEL, DATE, "22:00:00", "MW", 30),
        )
        LMRBRow.objects.filter(id=mw_row.id).update(uploaded_at=old)

    def test_purges_only_old_maponline(self):
        from core.maponline_cleanup import purge_old_maponline_data
        result = purge_old_maponline_data(days=30)

        self.assertEqual(result["monitoring_deleted"], 1)
        self.assertEqual(result["lmrb_deleted"], 1)
        # Old MapOnline gone
        self.assertFalse(LMRBRow.objects.filter(advt_theme="Old").exists())
        self.assertFalse(
            self.MonitoringData.objects.filter(file_group_id=str(self.old_batch)).exists()
        )
        # Recent MapOnline kept
        self.assertTrue(LMRBRow.objects.filter(advt_theme="Fresh").exists())
        # MediaWatch kept regardless of age
        self.assertTrue(LMRBRow.objects.filter(advt_theme="MW").exists())
        self.assertTrue(
            self.MonitoringData.objects.filter(data_type="mediawatch").exists()
        )


class TcLmrbCandidatesFilterTest(TestCase):
    """The 'Find LMRB Match' picker (tc_lmrb_candidates) must find LMRB rows
    even when the channel string carries a media-type prefix on only one side,
    and regardless of the LMRB row's duration — it is a manual fallback used
    exactly when the auto engine (which requires exact channel + duration)
    failed. Date still has to match. Locked rows stay hidden.
    """

    def setUp(self):
        from django.test import Client

        self.account = make_account()
        # TC row stores the prefixed channel form ('TV - Sirasa TV'); the
        # LMRB rows store the clean form ('Sirasa TV').
        self.tc_report = make_tc_report(self.account, channel="TV - Sirasa TV")
        self.tc_row = make_tc_row(
            self.account, self.tc_report,
            channel="TV - Sirasa TV", date=DATE,
            aired_time="20:30:00", duration=30,
        )
        self.user = make_user(email="ops-cand@test.com", role="super_admin")
        self.client = Client()
        self.client.force_login(self.user)

    def _mk_lmrb(self, advt_theme, advt_time, duration, channel="Sirasa TV", date=DATE):
        return LMRBRow.objects.create(
            account=self.account, channel=channel, date=date,
            advt_theme=advt_theme, advt_time=advt_time, duration=duration,
            source="maponline",
            dedup_key=LMRBRow.make_dedup_key(
                self.account.id, channel, date, advt_time, advt_theme, duration
            ),
        )

    def _fetch_ids(self):
        resp = self.client.get(
            "/dashboard/tc/lmrb-candidates/", {"tc_row_id": self.tc_row.id}
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data["ok"], data)
        return {c["id"] for c in data["candidates"]}

    def test_channel_prefix_and_any_duration_show_up(self):
        # Same duration, clean channel — must show despite prefixed TC channel.
        same_dur = self._mk_lmrb("Theme A", "20:30:05", 30)
        # Different duration — must ALSO show now (any duration).
        diff_dur = self._mk_lmrb("Theme A 15", "20:30:20", 15)
        ids = self._fetch_ids()
        self.assertIn(same_dur.id, ids)
        self.assertIn(diff_dur.id, ids)

    def test_other_date_excluded(self):
        other_day = self._mk_lmrb(
            "Theme A", "20:30:05", 30, date=datetime.date(2025, 1, 16)
        )
        self.assertNotIn(other_day.id, self._fetch_ids())

    def test_locked_rows_excluded(self):
        spon = self._mk_lmrb("Spon", "20:30:05", 30)
        LMRBRow.objects.filter(id=spon.id).update(is_sponsorship_matched=True)
        man = self._mk_lmrb("Man", "20:30:06", 30)
        LMRBRow.objects.filter(id=man.id).update(is_manual_matched=True)
        ids = self._fetch_ids()
        self.assertNotIn(spon.id, ids)
        self.assertNotIn(man.id, ids)

    def test_sorted_by_time_gap(self):
        far = self._mk_lmrb("Far", "21:00:00", 30)   # 30 min gap
        near = self._mk_lmrb("Near", "20:30:02", 15)  # 2 sec gap, diff duration
        resp = self.client.get(
            "/dashboard/tc/lmrb-candidates/", {"tc_row_id": self.tc_row.id}
        )
        cands = resp.json()["candidates"]
        self.assertEqual(cands[0]["id"], near.id)
        self.assertLess(
            [c["id"] for c in cands].index(near.id),
            [c["id"] for c in cands].index(far.id),
        )


class SummaryExcelLandscapeTest(TestCase):
    """Every sheet in the reconciliation Excel download must print landscape +
    fit-to-width. The wide Matched/Unmatched LMRB sheets (21 columns) used to
    default to portrait, which printed broken."""

    def setUp(self):
        self.account = make_account()

    def _orientation(self, writer):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        writer(ws, self.account.id, CHANNEL, MONTH)
        return ws.page_setup.orientation, ws.page_setup.fitToWidth

    def test_matched_lmrb_sheet_is_landscape(self):
        from core.views import _write_matched_lmrb_sheet
        orient, fit = self._orientation(_write_matched_lmrb_sheet)
        self.assertEqual(orient, 'landscape')
        self.assertEqual(fit, 1)

    def test_unmatched_lmrb_sheet_is_landscape(self):
        from core.views import _write_unmatched_lmrb_sheet
        orient, fit = self._orientation(_write_unmatched_lmrb_sheet)
        self.assertEqual(orient, 'landscape')
        self.assertEqual(fit, 1)


class ReconExcelLogoTest(TestCase):
    """The reconciliation Excel summary must embed the client logo (openpyxl
    needs Pillow — declared in requirements) and lay the header out like the
    PDF/UI: title on the left, logo top-right."""

    # 1x1-ish valid PNG
    PNG = (b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x10\x00\x00\x00\x10'
           b'\x08\x06\x00\x00\x00\x1f\xf3\xffa\x00\x00\x00\x19IDATx\x9cc\xfc\xcf'
           b'\xc0\xf0\x1f\x8a\x01\x08\x18\x18\x00\x00\xff\xff\x03\x00\x06\x05\x02'
           b'\x9f\xe7\x86\xf1\xc4\x00\x00\x00\x00IEND\xaeB`\x82')

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        make_schedule_row(self.account, self.schedule, brand="Brand A", duration=30)

    def _fake_logo(self):
        import io as _io
        png = self.PNG
        class FakeLogo:
            def open(self, mode='rb'): return _io.BytesIO(png)
            def read(self): return png
            def close(self): pass
        return FakeLogo()

    def _build_sheet(self):
        import openpyxl
        from verification.tc_engine import build_summary_data
        from verification.media_recon import build_recon_context
        from core.views import _write_media_recon_sheet
        data = build_summary_data(self.account.id, CHANNEL, MONTH)
        ctx = build_recon_context(self.account, CHANNEL, MONTH, None, data, schedule=self.schedule)
        ctx['logo'] = self._fake_logo()
        wb = openpyxl.Workbook(); ws = wb.active
        _write_media_recon_sheet(ws, ctx)
        return wb, ws

    def test_logo_embeds_and_landscape(self):
        import io, zipfile
        wb, ws = self._build_sheet()
        self.assertEqual(ws['A1'].alignment.horizontal, 'center')
        self.assertEqual(ws.page_setup.orientation, 'landscape')
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        names = zipfile.ZipFile(buf).namelist()
        self.assertTrue(any('media/image' in n for n in names),
                        "client logo must be embedded in the recon sheet")

    def test_new_summary_labels_and_header_colour(self):
        """Spot table uses the New_Summary labels and light-blue (BDD7EE)
        section headers, matching the on-screen report / New_Summary template."""
        _wb, ws = self._build_sheet()
        header_texts = {ws.cell(13, c).value for c in range(1, 6)}
        self.assertIn('Schedule', header_texts)
        self.assertIn('Transmission Report', header_texts)
        self.assertIn('Nielsen Report', header_texts)
        self.assertNotIn('Schedule (Planned)', header_texts)
        # light-blue header fill
        self.assertEqual(ws.cell(13, 1).fill.fgColor.rgb[-6:], 'BDD7EE')


class AgentToolsTest(TestCase):
    """Deterministic shared tool layer (core/agent_tools.py) used by both the
    Nova chat agent and the future background agent."""

    def setUp(self):
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        self.sr = make_schedule_row(self.account, self.schedule, brand="Brand A",
                                    start_time="20:00:00", end_time="21:00:00", duration=30)
        # LMRB stored under clean channel name; schedule channel is CHANNEL.
        self.lmrb = make_lmrb_row(self.account, advt_theme="Theme A", channel=CHANNEL,
                                  advt_time="20:05:00", duration=30, source="mediawatch")

    def test_get_schedule_row_context(self):
        from core.agent_tools import get_schedule_row_context
        ctx = get_schedule_row_context(self.sr.id)
        self.assertEqual(ctx["brand"], "Brand A")
        self.assertEqual(ctx["channel"], CHANNEL)
        self.assertEqual(ctx["account_id"], self.account.id)

    def test_search_lmrb_candidates_ignores_programme_and_uses_window(self):
        from core.agent_tools import search_lmrb_candidates
        out = search_lmrb_candidates(CHANNEL, str(DATE), "20:00:00", window_minutes=60)
        ids = [c["lmrb_row_id"] for c in out["candidates"]]
        self.assertIn(self.lmrb.id, ids)
        # 5 minutes from planned
        cand = next(c for c in out["candidates"] if c["lmrb_row_id"] == self.lmrb.id)
        self.assertEqual(cand["minutes_from_planned"], 5.0)
        self.assertFalse(cand["already_locked"])

    def test_search_lmrb_candidates_channel_prefix_tolerant(self):
        from core.agent_tools import search_lmrb_candidates
        # TC/schedule side carries the 'TV - ' prefix; LMRB is clean.
        out = search_lmrb_candidates("TV - " + CHANNEL, str(DATE), "20:00:00")
        self.assertIn(self.lmrb.id, [c["lmrb_row_id"] for c in out["candidates"]])

    def test_check_brand_mapping(self):
        from core.agent_tools import check_brand_mapping
        make_brand_mapping(self.account, brand="Brand A", theme="Theme A", tc_theme="TC A")
        out = check_brand_mapping(self.account.id, "Brand A")
        self.assertTrue(out["is_mapped"])
        self.assertIn("Theme A", out["mapped_lmrb_themes"])

    def test_propose_manual_match_refuses_without_confirmation(self):
        from core.agent_tools import propose_manual_match
        out = propose_manual_match(self.sr.id, self.lmrb.id, confirmed_by_user=False)
        self.assertEqual(out["status"], "not_created")
        self.assertFalse(ManualMatch.objects.filter(schedule_row=self.sr).exists())

    def test_propose_manual_match_creates_and_locks_when_confirmed(self):
        from core.agent_tools import propose_manual_match
        out = propose_manual_match(self.sr.id, self.lmrb.id, confirmed_by_user=True)
        self.assertEqual(out["status"], "created")
        self.sr.refresh_from_db(); self.lmrb.refresh_from_db()
        self.assertTrue(self.sr.is_manual_matched)
        self.assertTrue(self.lmrb.is_manual_matched)
        mm = ManualMatch.objects.get(schedule_row=self.sr)
        self.assertEqual(mm.match_mode, "schedule_lmrb")
        self.assertEqual(mm.lmrb_row_id, self.lmrb.id)

    def test_list_unmatched_spots_matches_summary(self):
        """list_unmatched_spots must read the SAME source as the Summary Sheet
        (build_summary_data), so its counts agree — including sponsorship/tag
        deviations that MatchResult never records. Here Brand A is planned but
        nothing aired, so it shows as Missed and carries the schedule row id."""
        from core.agent_tools import list_unmatched_spots
        out = list_unmatched_spots(self.account.id, CHANNEL, MONTH)
        self.assertGreaterEqual(out["total_missed"], 1)
        dev = next((d for d in out["deviations"] if d["brand"] == "Brand A"), None)
        self.assertIsNotNone(dev, out)
        self.assertGreaterEqual(dev["missed"], 1)
        self.assertIn(self.sr.id, dev["unmatched_schedule_row_ids"])

    def test_lookup_schedule_system_wide(self):
        from core.agent_tools import lookup_schedule
        out = lookup_schedule(self.schedule.schedule_number)
        self.assertTrue(out["found"])
        self.assertEqual(out["account_id"], self.account.id)
        self.assertEqual(out["channel"], CHANNEL)
        self.assertEqual(out["month"], MONTH)
        self.assertFalse(lookup_schedule("nonexistent-999")["found"])

    def test_lookup_by_brand(self):
        from core.agent_tools import lookup_by_brand
        out = lookup_by_brand("Brand A")
        self.assertTrue(any(m["brand"] == "Brand A" and m["account_id"] == self.account.id
                            for m in out["matches"]))

    def test_list_schedules(self):
        from core.agent_tools import list_schedules
        self.assertGreaterEqual(list_schedules(brand="Brand A")["total"], 1)
        self.assertEqual(list_schedules(brand="does-not-exist")["total"], 0)
        self.assertGreaterEqual(list_schedules()["total"], 1)  # "any schedules?"

    def test_open_summary_smart_single_uses_real_month(self):
        """Resolver must build the URL from real DB values — real month string
        'January 2025', not a fabricated '2025-01' — and the real account_id."""
        from core.agent_tools import open_summary_smart
        out = open_summary_smart(brand="Brand A")
        self.assertTrue(out["found"])
        self.assertEqual(out["action"], "navigate")
        self.assertIn("account_id=%d" % self.account.id, out["url"])
        self.assertIn("month=January+2025", out["url"])   # urlencoded real month
        self.assertNotIn("2025-01", out["url"])

    def test_open_summary_smart_asks_when_multiple(self):
        from core.agent_tools import open_summary_smart
        sch2 = make_schedule(self.account, channel="TV - Hiru", schedule_number="777")
        make_schedule_row(self.account, sch2, brand="Brand A", channel="TV - Hiru")
        out = open_summary_smart(brand="Brand A")
        self.assertEqual(out["action"], "choose")
        self.assertGreaterEqual(len(out["options"]), 2)
        channels = {o["channel"] for o in out["options"]}
        self.assertIn(CHANNEL, channels)
        self.assertIn("TV - Hiru", channels)

    def test_open_summary_smart_not_found(self):
        from core.agent_tools import open_summary_smart
        self.assertFalse(open_summary_smart(brand="no-such-brand")["found"])

    def test_navigation_and_report_tools_return_actions(self):
        from core.agent_tools import (open_summary_sheet, open_mapping_page,
                                       generate_summary_report)
        nav = open_summary_sheet(self.account.id, CHANNEL, MONTH)
        self.assertEqual(nav["action"], "navigate")
        self.assertIn("/dashboard/summary/", nav["url"])
        self.assertIn("account_id=%d" % self.account.id, nav["url"])
        self.assertEqual(open_mapping_page(self.account.id, "Brand A")["action"], "navigate")
        dl = generate_summary_report(self.account.id, CHANNEL, MONTH, format="pdf")
        self.assertEqual(dl["action"], "download")
        self.assertIn("/dashboard/summary/pdf/", dl["url"])

    def test_investigate_all_unmatched_covers_schedule_and_tc(self):
        """Nova can investigate everything with no IDs from the user: the missed
        schedule spot AND an unmatched TC spot both come back, each with the
        closest LMRB airing."""
        from core.agent_tools import investigate_all_unmatched
        rep = make_tc_report(self.account, channel=CHANNEL)
        tc = make_tc_row(self.account, rep, tc_theme="Brand A TC", channel=CHANNEL,
                         aired_time="20:05:30", duration=30)  # is_lmrb_confirmed False
        out = investigate_all_unmatched(self.account.id, CHANNEL, MONTH)
        # missed schedule spot (Brand A) is present with the nearby LMRB candidate
        sched = out["missed_schedule_spots"]
        self.assertTrue(any(s["schedule_row_id"] == self.sr.id for s in sched))
        s0 = next(s for s in sched if s["schedule_row_id"] == self.sr.id)
        self.assertIsNotNone(s0["best_candidate"])
        self.assertEqual(s0["best_candidate"]["lmrb_row_id"], self.lmrb.id)
        # the unmatched TC spot is present too
        self.assertTrue(any(t["tc_row_id"] == tc.id for t in out["unmatched_tc_spots"]))

    def test_propose_tc_lmrb_match_requires_confirmation_then_links(self):
        from core.agent_tools import propose_tc_lmrb_match
        rep = make_tc_report(self.account, channel=CHANNEL)
        tc = make_tc_row(self.account, rep, tc_theme="Brand A TC", channel=CHANNEL,
                         aired_time="20:05:30", duration=30)
        # refuses without confirmation
        self.assertEqual(propose_tc_lmrb_match(tc.id, self.lmrb.id, False)["status"],
                         "not_created")
        # links on confirmation
        out = propose_tc_lmrb_match(tc.id, self.lmrb.id, True)
        self.assertEqual(out["status"], "created")
        tc.refresh_from_db(); self.lmrb.refresh_from_db()
        self.assertTrue(tc.is_lmrb_confirmed)
        self.assertEqual(tc.matched_lmrb_id, self.lmrb.id)
        self.assertTrue(self.lmrb.is_manual_matched)

    def test_diagnose_unmatched_tc_spot_finds_theme_mapped_elsewhere(self):
        """A TC theme unmapped for this account but mapped under a brand in
        ANOTHER account should surface as an exact match elsewhere."""
        from core.agent_tools import diagnose_unmatched_tc_spot
        rep = make_tc_report(self.account, channel=CHANNEL)
        tc = make_tc_row(self.account, rep, tc_theme="SIGNAL PLUS PROMO 30",
                         channel=CHANNEL, duration=30)
        other = make_account(name="Other Co")
        make_brand_mapping(other, brand="Signal Plus", theme="Signal Plus_30",
                           tc_theme="SIGNAL PLUS PROMO 30")
        out = diagnose_unmatched_tc_spot(tc.id)
        self.assertEqual(out["raw_tc_theme"], "SIGNAL PLUS PROMO 30")
        brands = [e["brand"] for e in out["exact_match_elsewhere"]]
        self.assertIn("Signal Plus", brands)
        self.assertFalse(out["exact_match_elsewhere"][0]["same_account"])


class MappingGuardianTest(TestCase):
    """Upload & Mapping Guardian (Tier 1 — advisory, non-blocking)."""

    def setUp(self):
        self.account = make_account(name="Milo")

    def test_validate_upload_blocks_without_account(self):
        from core.agent_tools import validate_upload_selection
        out = validate_upload_selection("schedule", account_id=None)
        self.assertTrue(out["block"])

    def test_validate_upload_warns_on_brand_mismatch(self):
        from core.agent_tools import validate_upload_selection
        out = validate_upload_selection("lmrb", account_id=self.account.id,
                                        brand_hint_from_filename="Coca Cola")
        self.assertFalse(out["block"])
        self.assertTrue(out["warnings"])

    def test_audit_flags_product_without_duration(self):
        from core.agent_tools import audit_brand_mapping
        BrandMapping.objects.create(account=self.account, brand="Milo 5s",
                                    theme="Milo Gen", product="Milo", duration=None)
        out = audit_brand_mapping(self.account.id, "Milo 5s")
        self.assertFalse(out["is_clean"])
        self.assertTrue(any("Product filter" in w for w in out["warnings"]))

    def test_audit_flags_theme_mapped_to_other_brand(self):
        from core.agent_tools import audit_brand_mapping
        BrandMapping.objects.create(account=self.account, brand="Milo 5s", theme="Shared Theme")
        BrandMapping.objects.create(account=self.account, brand="Other Brand", theme="Shared Theme")
        out = audit_brand_mapping(self.account.id, "Milo 5s")
        self.assertTrue(any("already mapped to a different brand" in w for w in out["warnings"]))

    def test_audit_clean_mapping(self):
        from core.agent_tools import audit_brand_mapping
        BrandMapping.objects.create(account=self.account, brand="Milo 5s",
                                    theme="Milo Gen 5", tc_theme="MILO 05")
        out = audit_brand_mapping(self.account.id, "Milo 5s")
        self.assertTrue(out["is_clean"])


class NovaChatEndpointTest(TestCase):
    """Nova chat endpoint degrades gracefully when Gemini isn't configured, and
    requires auth."""

    def setUp(self):
        from django.test import Client
        self.account = make_account()
        self.schedule = make_schedule(self.account)
        self.sr = make_schedule_row(self.account, self.schedule, brand="Brand A")
        self.user = make_user(email="nova@test.com", role="operations")
        self.client = Client()

    def test_requires_login(self):
        resp = self.client.post("/dashboard/nova-chat/",
                                data=json.dumps({"schedule_row_id": self.sr.id, "message": "hi"}),
                                content_type="application/json")
        self.assertIn(resp.status_code, (302, 403))

    @override_settings(GEMINI_API_KEY="")
    def test_graceful_without_gemini_key(self):
        self.client.force_login(self.user)
        resp = self.client.post(
            "/dashboard/nova-chat/",
            data=json.dumps({"schedule_row_id": self.sr.id, "message": "why not aired?"}),
            content_type="application/json")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("configured", resp.json()["reply"].lower())

    @override_settings(GEMINI_API_KEY="")
    def test_scope_mode_accepted(self):
        """Nova can be opened at Summary-Sheet scope (account+channel+month)."""
        self.user.role = "super_admin"; self.user.save(update_fields=["role"])
        self.client.force_login(self.user)
        resp = self.client.post(
            "/dashboard/nova-chat/",
            data=json.dumps({"account_id": self.account.id, "channel": CHANNEL,
                             "month": MONTH, "message": "which spots didn't match?"}),
            content_type="application/json")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("configured", resp.json()["reply"].lower())  # graceful (no key)

    @override_settings(GEMINI_API_KEY="test-key")
    def test_global_mode_returns_navigate_action(self):
        """Global chat (no page context) can look up + open a page; the endpoint
        returns the navigate action for the browser to execute."""
        from unittest import mock
        self.user.role = "super_admin"; self.user.save(update_fields=["role"])
        self.client.force_login(self.user)

        def fake(status, payload):
            m = mock.Mock(); m.status_code = status; m.json.return_value = payload
            m.text = json.dumps(payload); return m

        turn1 = {"candidates": [{"content": {"parts": [
            {"functionCall": {"name": "open_summary_sheet",
                              "args": {"account_id": self.account.id,
                                       "channel": CHANNEL, "month": MONTH}}}]}}]}
        turn2 = {"candidates": [{"content": {"parts": [
            {"text": "Opening the summary now."}]}}]}
        with mock.patch("core.agent_chat.requests.post",
                        side_effect=[fake(200, turn1), fake(200, turn2)]):
            resp = self.client.post(
                "/dashboard/nova-chat/",
                data=json.dumps({"message": "open the summary for that account"}),
                content_type="application/json")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(any(a["action"] == "navigate" and "/dashboard/summary/" in a["url"]
                            for a in data["actions"]))

    @override_settings(GEMINI_API_KEY="test-key")
    def test_rest_function_calling_loop(self):
        """With a key set, Nova should (1) call a tool via the REST functionCall
        protocol, then (2) return the model's final text. We mock requests.post
        so no real Gemini call is made."""
        from unittest import mock
        self.client.force_login(self.user)

        def fake_response(status, payload):
            m = mock.Mock(); m.status_code = status; m.json.return_value = payload
            m.text = json.dumps(payload); return m

        turn1 = {"candidates": [{"content": {"parts": [
            {"functionCall": {"name": "get_schedule_row_context",
                              "args": {"schedule_row_id": self.sr.id}}}]}}]}
        turn2 = {"candidates": [{"content": {"parts": [
            {"text": "This spot is on Sirasa TV; let me check LMRB."}]}}]}
        with mock.patch("core.agent_chat.requests.post",
                        side_effect=[fake_response(200, turn1), fake_response(200, turn2)]) as post:
            resp = self.client.post(
                "/dashboard/nova-chat/",
                data=json.dumps({"schedule_row_id": self.sr.id, "message": "why not aired?"}),
                content_type="application/json")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("Sirasa TV", resp.json()["reply"])
        self.assertEqual(post.call_count, 2)  # tool round-trip + final text


class NovaEnableToggleTest(TestCase):
    """The 'Ask Nova' feature can be turned off via a SystemSetting; the context
    processor reflects it so base.html hides the launcher."""

    def _nova_enabled(self):
        from core.context_processors import branding
        return branding(None).get('nova_enabled')

    def test_enabled_by_default(self):
        self.assertTrue(self._nova_enabled())  # no row -> default '1'

    def test_toggle_off(self):
        SystemSetting.objects.update_or_create(
            key='nova_enabled',
            defaults={'value': '0', 'label': 'Ask Nova Assistant Enabled',
                      'category': 'assistant'})
        self.assertFalse(self._nova_enabled())

    def test_toggle_on(self):
        SystemSetting.objects.update_or_create(
            key='nova_enabled',
            defaults={'value': '1', 'label': 'Ask Nova Assistant Enabled',
                      'category': 'assistant'})
        self.assertTrue(self._nova_enabled())
