"""Colored schedule pivot export.

Reads the original uploaded pivot-format Excel file and produces a new workbook
in the same layout, with date-column cells coloured by reconciliation status:

    matched / manual_match  → green  (schedule_color_aired)
    not_aired / no_mapping  → red    (schedule_color_not_aired)
    late_telecast           → purple (schedule_color_late_telecast)
    programme_mismatch      → orange (schedule_color_programme_mismatch)
    extra_aired             → blue   (schedule_color_extra_aired)
    planned (no data yet)   → grey   (schedule_color_planned)

If status_map is None, all date cells are written in the "planned" colour.

When a date cell has mixed statuses the main row shows the first status and
additional sub-rows (with blank left columns) carry the remaining statuses.

Late-telecast spots also appear in the ACTUAL aired date column (purple), even
if that column is beyond the schedule's original date range (new columns are
appended to the right).
"""

import io
import logging
from collections import defaultdict
from datetime import date as _date

logger = logging.getLogger(__name__)

# ── Sentinel row-type constants ───────────────────────────────────────────────
_ROW_DATA    = 'data'
_ROW_OTHER   = 'other'   # section header, brand header, total, blank

# ── Skip keywords (same as schedule_converter) ────────────────────────────────
_SKIP_KW = ('TOTAL', 'VAT', 'SSCL', 'SUMMARY', 'ROI', 'CPRP',
            'COMMERCIAL ONLY', 'ALL EXPO', 'NOTE:', 'GRAND TOTAL')

# Statuses that map to "aired as planned"
_AIRED_STATUSES = {'matched', 'manual_match'}
# Statuses that map to "not aired"
_NOT_AIRED_STATUSES = {'not_aired', 'no_mapping'}


def _hex_to_fill(hex_color: str):
    """Return an openpyxl PatternFill from a CSS hex string like '#22c55e'."""
    from openpyxl.styles import PatternFill
    h = hex_color.lstrip('#').upper()
    return PatternFill('solid', fgColor=h)


def _copy_cell_style(src_cell, dst_cell):
    """Copy basic style (font bold/size/color, alignment) from src to dst."""
    from copy import copy
    try:
        if src_cell.font:
            dst_cell.font = copy(src_cell.font)
        if src_cell.alignment:
            dst_cell.alignment = copy(src_cell.alignment)
        if src_cell.fill and src_cell.fill.fill_type not in (None, 'none'):
            dst_cell.fill = copy(src_cell.fill)
        if src_cell.border:
            dst_cell.border = copy(src_cell.border)
    except Exception:
        pass  # best-effort only


def _normalize(s) -> str:
    return str(s).strip().lower() if s else ''


def _try_int(val):
    """Return int(val) or None if val is not a valid integer."""
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _normalize_time(raw) -> str:
    """
    Convert an openpyxl time cell value to a normalised 'HH:MM:SS' string.

    Excel stores times as a fraction of a day. openpyxl may return:
      - datetime.time(9, 0, 0)           → "09:00:00"
      - datetime.datetime(1899,12,30,9,0) → we want "09:00:00", not the full repr
      - str "09:00:00"                    → keep as-is
    """
    from datetime import time as _time, datetime as _datetime
    if raw is None:
        return ''
    if isinstance(raw, _time):
        return raw.strftime('%H:%M:%S')
    if isinstance(raw, _datetime):
        return raw.strftime('%H:%M:%S')
    return _normalize(str(raw))


# ── Status map builder ────────────────────────────────────────────────────────

def build_status_map(account_id, channel, month) -> dict:
    """
    Build a per-slot, per-date status count dict from MatchResult records.

    Late-telecast handling (key design decision):
      The PLANNED date of a late spot shows as NOT_AIRED (red) because the ad
      was genuinely missing on its original date.  The ACTUAL aired date receives
      a 'late_telecast' count (purple) via the 'late_actual' side-channel so that
      the cell on the actual date shows total airings including the delayed spot.

    Returns
    -------
    dict keyed by (norm_programme, norm_start_time, duration_int) →
        dict keyed by date_str ('YYYY-MM-DD') →
            {
              'matched': int,
              'not_aired': int,          # true missed + late-planned-date spots
              'late_telecast': int,      # (only on actual aired date via late_actual)
              'programme_mismatch': int,
              'manual_match': int,
              'extra_aired': int,
              'late_actual': {actual_date_str: int},  # actual aired date → count
              'late_to':     {actual_date_str: int},  # for comment: where it moved to
            }
    """
    from core.models import MatchResult

    qs = (
        MatchResult.objects
        .filter(account_id=account_id, channel=channel, month=month)
        .exclude(status='extra_aired')
        .values(
            'brand', 'programme', 'scheduled_date', 'planned_start',
            'duration', 'status', 'aired_date',
        )
    )

    status_map: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    late_actual_map: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for row in qs:
        prog   = _normalize(row['programme'])
        start  = _normalize(row['planned_start'])
        dur    = row['duration'] or 0
        s_date = row['scheduled_date']
        status = row['status']

        if not s_date:
            continue

        date_str = s_date.strftime('%Y-%m-%d')
        slot_key = (prog, start, dur)

        if status == 'late_telecast':
            # Planned date shows RED (not aired on its original date)
            status_map[slot_key][date_str]['not_aired'] += 1
            if row['aired_date']:
                actual_str = row['aired_date'].strftime('%Y-%m-%d')
                # late_actual → used to add purple count on the actual aired date
                late_actual_map[slot_key][date_str][actual_str] += 1
                # late_to stored separately (can't use defaultdict(int) for nested dict)
        elif status in _NOT_AIRED_STATUSES:
            status_map[slot_key][date_str]['not_aired'] += 1
        else:
            status_map[slot_key][date_str][status] += 1

    # Embed late_actual and late_to into status_map (keyed off the PLANNED date).
    # late_to == late_actual from the planned-date perspective; stored as a plain
    # dict so the cell-comment code can iterate it without touching defaultdict(int).
    for slot_key, date_dict in late_actual_map.items():
        for date_str, actual_dict in date_dict.items():
            status_map[slot_key][date_str]['late_actual'] = dict(actual_dict)
            status_map[slot_key][date_str]['late_to'] = dict(actual_dict)

    return dict(status_map)


def build_status_map_from_tc(account_id, channel, month) -> dict:
    """
    Build the per-slot, per-date status map from TC ↔ LMRB reconciliation data,
    reconciled so the colours AGREE with the Summary Sheet's aired/missed numbers.

    Per planned ScheduleRow:
        matched (green)   – aired (see below)
        late_telecast     – aired on a LATER date (planned cell red, actual purple)
        not_aired (red)   – genuinely missed

    A row is treated as aired when ANY of these hold, in order:
      1. A confirmed TCRow is one-to-one linked to it (is_schedule_matched +
         is_lmrb_confirmed).
      2. An unclaimed confirmed TCRow falls within its date + time-window + duration
         (window coverage — same rule as the TC three-way "Not Aired" list).
      3. Brand-level alignment: the Summary counts this brand/programme as aired
         (by tc_theme / assignment) more than the rows greened by 1–2, so the
         remaining red rows are promoted to green up to the Summary's aired count.
    This guarantees the coloured sheet's green/red totals match the Summary card
    (e.g. "48 aired / 0 missed" → no red cells).
    """
    from collections import defaultdict as _dd
    from core.models import ScheduleRow, TCRow
    from verification.tc_engine import _time_to_secs, build_summary_data

    # ── 1. Confirmed TCRow linked one-to-one to each ScheduleRow ──────────────
    tc_by_sched: dict = {}
    for tc in (
        TCRow.objects.filter(
            account_id=account_id, channel__iexact=channel, tc_report__month=month,
            is_schedule_matched=True, is_lmrb_confirmed=True,
            matched_schedule__isnull=False,
        ).select_related('matched_lmrb')
    ):
        tc_by_sched[tc.matched_schedule_id] = tc

    rows = [
        sr for sr in ScheduleRow.objects.filter(
            account_id=account_id, channel=channel, month=month,
        ) if sr.date
    ]
    commercial_rows  = [sr for sr in rows if sr.ad_type == 'COMMERCIAL BENEFITS']
    sponsorship_rows = [sr for sr in rows if sr.ad_type == 'SPONSORSHIP']

    status_map: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    late_actual_map: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    # ══ COMMERCIAL: per-row matched / late / not-aired, reconciled to Summary ══
    per_row: list = []
    for sr in commercial_rows:
        tc = tc_by_sched.get(sr.id)
        if tc is None:
            per_row.append({'sr': sr, 'status': 'not_aired', 'actual': None})
        elif tc.date and tc.date > sr.date:
            per_row.append({'sr': sr, 'status': 'late', 'actual': tc.date})
        else:
            per_row.append({'sr': sr, 'status': 'matched', 'actual': None})

    # Window coverage: unclaimed confirmed TCRow inside the row's date+window+dur.
    avail = [
        {'date': t.date, 'secs': _time_to_secs(t.aired_time), 'dur': t.duration, 'used': False}
        for t in TCRow.objects.filter(
            account_id=account_id, channel__iexact=channel, tc_report__month=month,
            is_lmrb_confirmed=True, matched_schedule__isnull=True,
        )
    ]

    def _window_covers(sr):
        ss = _time_to_secs(sr.start_time)
        ee = _time_to_secs(sr.end_time)
        if ss is None or ee is None:
            return False
        if ee < ss and ss > 43200:
            ee += 86400
        for c in avail:
            if c['used'] or c['secs'] is None or c['date'] != sr.date:
                continue
            if sr.duration is not None and c['dur'] is not None and c['dur'] != sr.duration:
                continue
            cs = c['secs'] + 86400 if ee > 86400 and c['secs'] < ss else c['secs']
            if ss <= cs <= ee + 600:
                c['used'] = True
                return True
        return False

    for e in sorted([x for x in per_row if x['status'] == 'not_aired'],
                    key=lambda x: x['sr'].date):
        if _window_covers(e['sr']):
            e['status'] = 'matched'

    # Align commercial green count to the Summary aired number.
    try:
        sd = build_summary_data(account_id, channel, month)
    except Exception:
        logger.exception('build_status_map_from_tc: summary alignment failed')
        sd = None
    if sd:
        comm_aired = {
            (_normalize(r['product']), r.get('dur')): r.get('aired', 0)
            for r in sd.get('commercial', [])
        }
        groups: dict = _dd(list)
        for e in per_row:
            sr = e['sr']
            groups[(_normalize(sr.brand), sr.duration if sr.duration is not None else None)].append(e)
        for key, entries in groups.items():
            target = comm_aired.get(key)
            if target is None:
                continue
            target = min(target, len(entries))
            greened = sum(1 for e in entries if e['status'] in ('matched', 'late'))
            need = target - greened
            for e in sorted([x for x in entries if x['status'] == 'not_aired'],
                            key=lambda x: x['sr'].date):
                if need <= 0:
                    break
                e['status'] = 'matched'
                need -= 1

    for e in per_row:
        sr = e['sr']
        slot_key = (_normalize(sr.programme), _normalize_time(sr.start_time), sr.duration or 0)
        date_str = sr.date.strftime('%Y-%m-%d')
        if e['status'] == 'late':
            status_map[slot_key][date_str]['not_aired'] += 1
            if e['actual']:
                late_actual_map[slot_key][date_str][e['actual'].strftime('%Y-%m-%d')] += 1
        elif e['status'] == 'matched':
            status_map[slot_key][date_str]['matched'] += 1
        else:
            status_map[slot_key][date_str]['not_aired'] += 1

    # ══ SPONSORSHIP: per-day tag count (planned vs LMRB-observed that day) ══════
    # For each sponsorship slot/date compare the planned tag count with the number
    # of LMRB tag observations on that day for the brand:
    #   aired >= planned          → green   (note "N extra aired" if aired > planned)
    #   0 < aired < planned        → under   (note "N missed")
    #   aired == 0                 → red     (note "N missed")
    if sponsorship_rows:
        from core.models import Schedule as _Sch, LMRBRow as _LR
        from django.db.models import Min as _Min, Max as _Max
        from verification.tc_engine import (
            _build_lmrb_theme_map, _lmrb_themes_for_brand, _lmrb_channel_q,
        )
        lmrb_theme_map = _build_lmrb_theme_map(account_id)
        _dd2 = _Sch.objects.filter(
            account_id=account_id, channel=channel, month=month,
        ).aggregate(s=_Min('start_date'), e=_Max('end_date'))
        dmin, dmax = _dd2.get('s'), _dd2.get('e')

        brand_themes = {}
        for sr in sponsorship_rows:
            nb = _normalize(sr.brand)
            if nb not in brand_themes:
                brand_themes[nb] = _lmrb_themes_for_brand(sr.brand, sr.duration, lmrb_theme_map)

        # Observed LMRB tag count per (brand, date).
        aired_bd: dict = _dd(lambda: _dd(int))
        lq = _LR.objects.filter(_lmrb_channel_q(channel), account_id=account_id).exclude(advt_theme='')
        if dmin and dmax:
            lq = lq.filter(date__range=(dmin, dmax))
        for lr in lq.values('advt_theme', 'product', 'date'):
            if not lr['date']:
                continue
            ds = lr['date'].strftime('%Y-%m-%d')
            lt = _normalize(lr['advt_theme'])
            lp = _normalize(lr['product']) if lr['product'] else ''
            for nb, themes in brand_themes.items():
                hit = False
                for entry in themes:
                    t = entry[0]
                    p = entry[1] if len(entry) > 1 else ''
                    tmatch = lt.startswith(t[:-1]) if t.endswith('*') else (lt == t)
                    if tmatch and (not p or lp == p):
                        hit = True
                        break
                if hit:
                    aired_bd[nb][ds] += 1

        spon_planned: dict = _dd(lambda: _dd(int))
        slot_brands: dict = _dd(set)
        for sr in sponsorship_rows:
            sk = (_normalize(sr.programme), _normalize_time(sr.start_time), sr.duration or 0)
            ds = sr.date.strftime('%Y-%m-%d')
            spon_planned[sk][ds] += 1
            slot_brands[sk].add(_normalize(sr.brand))

        for sk, by_date in spon_planned.items():
            brands = slot_brands[sk]
            for ds, planned in by_date.items():
                aired = sum(aired_bd[nb].get(ds, 0) for nb in brands)
                cell = status_map[sk][ds]
                if planned > 0 and aired >= planned:
                    cell['matched'] += planned
                    if aired > planned:
                        cell['extra_note'] = f'{aired - planned} extra aired (observed {aired}, planned {planned}).'
                elif aired > 0:
                    cell['aired_less'] += planned
                    cell['miss_note'] = f'{planned - aired} missed (aired {aired} of {planned}).'
                else:
                    cell['not_aired'] += planned
                    cell['miss_note'] = f'{planned} missed (none aired).'

    # Embed late-telecast metadata (commercial different-date arrivals).
    for slot_key, date_dict in late_actual_map.items():
        for date_str, actual_dict in date_dict.items():
            status_map[slot_key][date_str]['late_actual'] = dict(actual_dict)
            status_map[slot_key][date_str]['late_to'] = dict(actual_dict)

    return dict(status_map)


def build_status_map_auto(account_id, channel, month):
    """
    Pick the colour source automatically:

      1. If TC reconciliation data exists for the scope (any LMRB-confirmed TCRow),
         colour from TC ↔ LMRB matched data (the channel-certificate ground truth).
      2. Else if Schedule ↔ LMRB MatchResult records exist, colour from those.
      3. Else return None (all date cells fall back to the 'planned' colour).
    """
    from core.models import MatchResult, TCRow

    tc_exists = TCRow.objects.filter(
        account_id=account_id, channel__iexact=channel,
        tc_report__month=month, is_lmrb_confirmed=True,
    ).exists()
    if tc_exists:
        return build_status_map_from_tc(account_id, channel, month)

    if MatchResult.objects.filter(
        account_id=account_id, channel=channel, month=month,
    ).exists():
        return build_status_map(account_id, channel, month)

    return None


def build_status_map_from_maponline(account_id, channel, month) -> dict | None:
    """Build the per-slot, per-date status map from MapOnline matching.

    Uses compute_maponline_scope(persist=False) (read-only — no locks written for
    an export) and maps its result rows onto the same {slot_key: {date: {...}}}
    shape the colouring code expects:
        matched            → green
        programme_mismatch → orange
        late_telecast      → planned date RED + actual date PURPLE
        not_aired/no_mapping → red

    Returns None when there is no MapOnline data for the scope (so the export
    simply omits the extra sheet).
    """
    from core.models import LMRBRow
    from verification.engine import compute_maponline_scope, _lmrb_channel_q

    has_maponline = LMRBRow.objects.filter(
        _lmrb_channel_q(channel), account_id=account_id, source='maponline',
    ).exists()
    if not has_maponline:
        return None

    res = compute_maponline_scope(account_id, channel, month, persist=False)

    status_map: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    late_actual_map: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    def _slot(r):
        return (_normalize(r.get('programme')),
                _normalize(r.get('planned_start')),
                r.get('duration') or 0)

    for r in res.get('matched', []):
        if r.get('planned_date'):
            status_map[_slot(r)][r['planned_date']]['matched'] += 1
    for r in res.get('prog_mismatch', []):
        if r.get('planned_date'):
            status_map[_slot(r)][r['planned_date']]['programme_mismatch'] += 1
    for r in res.get('late_telecast', []):
        if not r.get('planned_date'):
            continue
        status_map[_slot(r)][r['planned_date']]['not_aired'] += 1
        if r.get('aired_date'):
            late_actual_map[_slot(r)][r['planned_date']][r['aired_date']] += 1
    for r in res.get('not_aired', []) + res.get('no_mapping', []):
        if r.get('planned_date'):
            status_map[_slot(r)][r['planned_date']]['not_aired'] += 1

    for slot_key, date_dict in late_actual_map.items():
        for date_str, actual_dict in date_dict.items():
            status_map[slot_key][date_str]['late_actual'] = dict(actual_dict)
            status_map[slot_key][date_str]['late_to'] = dict(actual_dict)

    return dict(status_map)


# ── Pivot structure detection (openpyxl-based) ───────────────────────────────

def _detect_pivot_openpyxl(ws):
    """
    Detect the header row and date columns in an openpyxl worksheet.

    Returns (header_row_idx, col_map, date_col_map) where:
      header_row_idx  – 1-based row index of the header row (or None)
      col_map         – {'programme': int, 'day': int, 'start_time': int,
                         'end_time': int, 'duration': int} (1-based col indices)
      date_col_map    – {col_idx: 'YYYY-MM-DD'} (1-based)

    Accepts a wide variety of column naming conventions so that different
    client schedule file formats are handled without code changes.
    """
    from datetime import datetime as _dt

    _PROGRAMME_NAMES = {'PROGRAM', 'PROGRAMME', 'PROG', 'PROG NAME',
                        'PROGRAMME NAME', 'PROGRAM NAME', 'PRG', 'PRG NAME'}
    _DAY_NAMES       = {'DAY', 'DAY OF WEEK', 'WEEKDAY', 'WK DAY', 'WEEK DAY'}
    _TIME_NAMES      = {'TIME', 'START TIME', 'START', 'FROM', 'TIME (START)',
                        'START_TIME', 'TIME START', 'STARTTIME', 'AD TIME'}
    _DUR_NAMES       = {'DUR', 'DURATION', 'SPOT DUR', 'DUR (SEC)',
                        'DURATION (SEC)', 'SECONDS', 'SECS', 'SEC', 'AD DUR'}

    header_row_idx = None
    col_map = {}

    # Scan first 25 rows; check up to 15 columns for the programme header
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(25, ws.max_row)), start=1):
        for c_idx, cell in enumerate(row[:15], start=1):
            val = str(cell.value or '').strip().upper()
            if val in _PROGRAMME_NAMES:
                header_row_idx = row_idx
                col_map['programme'] = c_idx
                break
        if header_row_idx:
            break

    if not header_row_idx:
        return None, {}, {}

    # Walk the header row for other semantic columns
    header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
    for cell in header_row:
        val = str(cell.value or '').strip().upper()
        c   = cell.column
        if val in _DAY_NAMES and 'day' not in col_map:
            col_map['day'] = c
        elif val in _TIME_NAMES and 'start_time' not in col_map:
            col_map['start_time'] = c
            col_map['end_time']   = c + 1
        elif val in _DUR_NAMES and 'duration' not in col_map:
            col_map['duration'] = c
        elif val == 'BRAND' and 'brand' not in col_map:
            col_map['brand'] = c

    if not col_map.get('day') or not col_map.get('start_time') or not col_map.get('duration'):
        return None, {}, {}

    min_date_col = max(col_map.values()) + 1
    date_col_map = {}
    for cell in header_row:
        if cell.column < min_date_col:
            continue
        val = cell.value
        if isinstance(val, (_dt, _date)):
            try:
                if isinstance(val, _dt):
                    date_str = val.strftime('%Y-%m-%d')
                else:
                    date_str = val.strftime('%Y-%m-%d')
                date_col_map[cell.column] = date_str
            except Exception:
                pass

    return header_row_idx, col_map, date_col_map


def _detect_pivot_via_converter(raw_bytes):
    """
    Detect the pivot structure using the SAME logic as the upload parser
    (verification/schedule_converter), so any file that imported as a pivot at
    upload time is detected here identically.  This is the user's chosen source of
    truth for the colored export.

    Returns (header_row_idx, col_map, date_col_map) with 1-based indices to match
    the openpyxl convention used by build_colored_schedule_wb, or (None, {}, {}).
    """
    import pandas as pd
    from verification.schedule_converter import (
        _build_col_map, _detect_header_row, _find_date_columns,
    )

    try:
        raw = pd.read_excel(io.BytesIO(raw_bytes), header=None)
    except Exception:
        return None, {}, {}

    header_row, prog_col = _detect_header_row(raw)
    if header_row is None:
        return None, {}, {}

    col_map0   = _build_col_map(raw, header_row, prog_col)        # 0-based
    date_cols0 = _find_date_columns(raw, header_row, col_map0)    # {0-based col: 'YYYY-MM-DD'}

    if any(k not in col_map0 for k in ('day', 'start_time', 'duration')) or not date_cols0:
        return None, {}, {}

    # openpyxl uses 1-based row/column indices.
    col_map      = {k: v + 1 for k, v in col_map0.items()}
    date_col_map = {c + 1: d for c, d in date_cols0.items()}
    return header_row + 1, col_map, date_col_map


def _is_skip_row(val_str: str) -> bool:
    v = val_str.strip().upper()
    if not v or v in ('NONE', 'NAN'):
        return True
    return any(kw in v for kw in _SKIP_KW)


def _is_data_row(ws_row, col_map) -> bool:
    """True when the row has non-empty DAY and TIME values."""
    from datetime import time as _time, datetime as _datetime
    day  = ws_row[col_map.get('day', 1) - 1].value
    time_val = ws_row[col_map.get('start_time', 1) - 1].value
    day_ok  = bool(day and str(day).strip())
    # time_val can be a datetime/time object (always truthy even at midnight)
    time_ok = isinstance(time_val, (_time, _datetime)) or bool(
        time_val and str(time_val).strip()
    )
    return day_ok and time_ok


# ── DB-based fallback workbook builder ───────────────────────────────────────

def _build_wb_from_db(schedule, colors: dict, status_map=None):
    """
    Build a colored pivot workbook purely from ScheduleRow DB records.
    Used when the original uploaded file is no longer on disk (e.g. after Railway redeploy).
    """
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from core.models import ScheduleRow
    from collections import defaultdict

    rows_qs = (
        ScheduleRow.objects
        .filter(schedule=schedule)
        .order_by('brand', 'programme', 'start_time', 'date')
    )

    fills = {
        'aired':               _hex_to_fill(colors.get('aired',               '#22c55e')),
        'not_aired':           _hex_to_fill(colors.get('not_aired',           '#ef4444')),
        'late_telecast':       _hex_to_fill(colors.get('late_telecast',       '#a855f7')),
        'programme_mismatch':  _hex_to_fill(colors.get('programme_mismatch',  '#f97316')),
        'extra_aired':         _hex_to_fill(colors.get('extra_aired',         '#3b82f6')),
        'planned':             _hex_to_fill(colors.get('planned',             '#94a3b8')),
        'manual_override':     _hex_to_fill(colors.get('manual_override',     '#14b8a6')),
    }
    status_to_fill = {
        'matched': fills['aired'], 'manual_match': fills['manual_override'],
        'not_aired': fills['not_aired'], 'late_telecast': fills['late_telecast'],
        'programme_mismatch': fills['programme_mismatch'],
        'extra_aired': fills['extra_aired'], 'planned': fills['planned'],
    }

    # Collect all unique dates sorted
    all_dates = sorted({r.date for r in rows_qs if r.date})
    if not all_dates:
        wb = Workbook()
        wb.active.cell(1, 1, 'No schedule rows found in database.')
        return wb

    # Collect any extra late-telecast actual dates
    extra_dates: set = set()
    if status_map:
        for slot_dict in status_map.values():
            for d_dict in slot_dict.values():
                for actual_date in (d_dict.get('late_actual') or {}).keys():
                    actual_d = _date.fromisoformat(actual_date) if isinstance(actual_date, str) else actual_date
                    if actual_d not in all_dates:
                        extra_dates.add(actual_date)
    sorted_extra_dates = sorted(extra_dates)

    # Group rows by (brand, programme, start_time, duration) → {date: count}
    slot_groups: dict = defaultdict(lambda: defaultdict(int))
    for r in rows_qs:
        key = (r.brand or '', r.programme or '', r.start_time or '', r.duration or 0)
        if r.date:
            slot_groups[key][r.date.strftime('%Y-%m-%d')] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = 'Colored Schedule'

    # Fixed left columns: Brand, Programme, Start Time, Duration
    LEFT_COLS = ['Brand', 'Programme', 'Start Time', 'Duration (s)']
    n_left = len(LEFT_COLS)

    # Header row
    hdr_font = Font(bold=True, size=9)
    for ci, label in enumerate(LEFT_COLS, start=1):
        c = ws.cell(1, ci, label)
        c.font = hdr_font
    for di, d in enumerate(all_dates, start=n_left + 1):
        c = ws.cell(1, di, d.strftime('%d/%m'))
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center')
    for ei, d_str in enumerate(sorted_extra_dates, start=n_left + len(all_dates) + 1):
        c = ws.cell(1, ei, d_str)
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center')

    # Build date → output column mapping
    date_str_to_col = {d.strftime('%Y-%m-%d'): n_left + i for i, d in enumerate(all_dates, start=1)}
    for ei, d_str in enumerate(sorted_extra_dates, start=1):
        date_str_to_col[d_str] = n_left + len(all_dates) + ei

    # Data rows
    current_brand = None
    out_row = 2
    for (brand, programme, start_time, duration), date_counts in sorted(slot_groups.items()):
        if brand != current_brand:
            # Brand header row
            c = ws.cell(out_row, 1, brand)
            c.font = Font(bold=True, size=9)
            out_row += 1
            current_brand = brand

        # Write left columns
        ws.cell(out_row, 1, brand).font = Font(size=9)
        ws.cell(out_row, 2, programme).font = Font(size=9)
        ws.cell(out_row, 3, start_time).font = Font(size=9)
        ws.cell(out_row, 4, duration).font = Font(size=9)

        # Write date columns
        slot_key = (_normalize(programme), _normalize(start_time), duration)
        slot_status = (status_map or {}).get(slot_key, {})

        for d_str, planned_count in sorted(date_counts.items()):
            col = date_str_to_col.get(d_str)
            if not col:
                continue
            if not status_map:
                status = 'planned'
                count = planned_count
            else:
                d_data = slot_status.get(d_str, {})
                statuses = _ordered_statuses({k: v for k, v in d_data.items() if k != 'late_actual'})
                if statuses:
                    status, count = statuses[0]
                else:
                    status, count = 'planned', planned_count
            cell = ws.cell(out_row, col, count)
            cell.fill = status_to_fill.get(status, fills['planned'])
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center')

        # Extra late-telecast columns
        if status_map:
            for d_str, extra_col_idx in [(d, date_str_to_col[d]) for d in sorted_extra_dates]:
                count = sum(
                    d_dict.get('late_actual', {}).get(d_str, 0)
                    for d_dict in slot_status.values()
                )
                if count:
                    cell = ws.cell(out_row, extra_col_idx, count)
                    cell.fill = fills['late_telecast']
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal='center')

        out_row += 1

    # Legend
    _write_legend(ws, out_row + 1, fills)
    ws.append(['', '', '', '', 'NOTE: Rebuilt from DB — original file unavailable.'])
    return wb


# ── Workbook builder ──────────────────────────────────────────────────────────

def build_colored_schedule_wb(schedule_pk, colors: dict, status_map=None):
    """
    Generate a colored pivot-format Excel workbook.

    Parameters
    ----------
    schedule_pk : int
        Primary key of the Schedule object whose original file is the template.
    colors : dict
        {status_key: '#rrggbb'} — keys:
            'aired', 'not_aired', 'late_telecast', 'programme_mismatch',
            'extra_aired', 'planned'
    status_map : dict or None
        Output of build_status_map(). None → all date cells in 'planned' colour.

    Returns
    -------
    openpyxl.Workbook
    """
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from core.models import Schedule

    schedule = Schedule.objects.get(pk=schedule_pk)

    # Load the original file bytes from storage (works for local + Firebase).
    # Never fall back to DB reconstruction — it produces wrong programme names and order.
    if not schedule.file or not schedule.file.name:
        raise FileNotFoundError(
            f'No file is linked to schedule "{schedule.original_filename}". '
            f'Please re-upload the schedule to generate a colored export.'
        )
    try:
        raw_bytes = schedule.file.read()
    except Exception as exc:
        logger.error(
            'build_colored_schedule_wb: failed to read schedule pk=%s '
            'file="%s" from storage: %s',
            schedule_pk, schedule.file.name, exc,
        )
        raise FileNotFoundError(
            f'The original schedule file "{schedule.original_filename}" '
            f'(stored as "{schedule.file.name}") could not be read from storage. '
            f'Please re-upload the schedule.'
        ) from exc

    if not raw_bytes:
        raise FileNotFoundError(
            f'The original schedule file "{schedule.original_filename}" is empty. '
            f'Please re-upload the schedule.'
        )

    wb_src = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    ws_src = wb_src.active
    # Detect the pivot structure. Prefer the upload parser's proven (pandas-based)
    # detection so any file that imported as a pivot at upload is detected here
    # too; fall back to the openpyxl-based detector for extra leniency.
    header_row_idx, col_map, date_col_map = _detect_pivot_via_converter(raw_bytes)
    if not header_row_idx or not date_col_map:
        header_row_idx, col_map, date_col_map = _detect_pivot_openpyxl(ws_src)
    if not header_row_idx or not date_col_map:
        # Pivot structure genuinely not detectable. NEVER silently rebuild from the
        # database — that produces a system-generated layout, not the user's file.
        # Return the original uploaded file unchanged, with a clear warning row.
        logger.warning(
            'build_colored_schedule_wb: pivot structure not detected for schedule pk=%s; '
            'returning the original file unchanged (no DB rebuild).', schedule_pk
        )
        wb_orig = openpyxl.load_workbook(io.BytesIO(raw_bytes))
        ws_orig = wb_orig.active
        ws_orig.insert_rows(1)
        warn = ws_orig.cell(
            1, 1,
            'NOTE: Could not auto-detect the schedule layout, so status colours were '
            'not applied. This is your original uploaded file, unchanged.',
        )
        warn.font = Font(bold=True, color='B91C1C')
        return wb_orig

    # Build colour fills
    fills = {
        'aired':               _hex_to_fill(colors.get('aired',               '#22c55e')),
        'not_aired':           _hex_to_fill(colors.get('not_aired',           '#ef4444')),
        'late_telecast':       _hex_to_fill(colors.get('late_telecast',       '#a855f7')),
        'programme_mismatch':  _hex_to_fill(colors.get('programme_mismatch',  '#f97316')),
        'extra_aired':         _hex_to_fill(colors.get('extra_aired',         '#3b82f6')),
        'planned':             _hex_to_fill(colors.get('planned',             '#94a3b8')),
        'manual_override':     _hex_to_fill(colors.get('manual_override',     '#14b8a6')),
    }
    # matched/manual_match both use 'aired' fill
    status_to_fill = {
        'matched':             fills['aired'],
        'manual_match':        fills['manual_override'],
        'not_aired':           fills['not_aired'],
        'late_telecast':       fills['late_telecast'],
        'programme_mismatch':  fills['programme_mismatch'],
        'extra_aired':         fills['extra_aired'],
        'planned':             fills['planned'],
    }

    # Collect any EXTRA date columns needed for late-telecast actual dates
    extra_dates: set = set()
    if status_map:
        for slot_dict in status_map.values():
            for d_dict in slot_dict.values():
                for actual_date in (d_dict.get('late_actual') or {}).keys():
                    if actual_date not in date_col_map.values():
                        extra_dates.add(actual_date)
    sorted_extra_dates = sorted(extra_dates)

    # Assign column indices for extra dates (after last existing date column)
    max_existing_col  = max(date_col_map.keys()) if date_col_map else ws_src.max_column
    extra_date_col_map: dict = {}   # date_str → 1-based col index in output
    for i, d in enumerate(sorted_extra_dates, start=1):
        extra_date_col_map[d] = max_existing_col + i

    # Full date col map for output (existing + extra)
    all_date_cols: dict = dict(date_col_map)          # src_col → date_str
    # existing date cols keep same column index in output

    # ── Build a list of output rows (each is a list of (col, value, fill)) ──
    # We'll collect rows in a list first so we can append sub-rows as needed.
    output_rows = []   # list of list-of-dicts: [{col, value, fill, font, alignment}]
    src_row_meta = []  # parallel list: {'type': 'data'/'other', 'src_row': list_of_cells}

    # Track current brand context for slot key matching
    current_brand = ''

    for row_idx, src_row in enumerate(
        ws_src.iter_rows(min_row=1, max_row=ws_src.max_row), start=1
    ):
        row_cells = list(src_row)

        # Build base output cells (copy everything from source)
        out_cells = []
        for cell in row_cells:
            out_cells.append({
                'col':       cell.column,
                'value':     cell.value,
                'fill':      None,    # will be overwritten for date cols on data rows
                'src_cell':  cell,
            })
        # Also prepare empty cells for extra date columns
        for d, col in extra_date_col_map.items():
            out_cells.append({
                'col':      col,
                'value':    None,
                'fill':     None,
                'src_cell': None,
            })

        if row_idx < header_row_idx:
            # Pre-header rows: copy as-is
            output_rows.append(out_cells)
            src_row_meta.append({'type': _ROW_OTHER})
            continue

        if row_idx == header_row_idx:
            # Header row: strip time from date column values, add extra date headers
            for oc in out_cells:
                if oc['col'] in date_col_map:
                    date_str = date_col_map[oc['col']]  # 'YYYY-MM-DD'
                    try:
                        d = _date.fromisoformat(date_str)
                        oc['value'] = d.strftime('%d %b')  # e.g. "20 Mar"
                    except Exception:
                        oc['value'] = date_str
            for d, col in extra_date_col_map.items():
                for oc in out_cells:
                    if oc['col'] == col:
                        try:
                            oc['value'] = _date.fromisoformat(d).strftime('%d %b')
                        except Exception:
                            oc['value'] = d
                        break
            output_rows.append(out_cells)
            src_row_meta.append({'type': _ROW_OTHER})
            continue

        # Data / brand / section rows
        prog_cell_val = str(row_cells[col_map['programme'] - 1].value or '').strip()

        if _is_skip_row(prog_cell_val):
            output_rows.append(out_cells)
            src_row_meta.append({'type': _ROW_OTHER})
            continue

        if not _is_data_row(row_cells, col_map):
            # Brand / section header row
            current_brand = prog_cell_val
            output_rows.append(out_cells)
            src_row_meta.append({'type': _ROW_OTHER})
            continue

        # ── Data row ──────────────────────────────────────────────────────────
        prog       = _normalize(prog_cell_val)
        start_raw  = row_cells[col_map['start_time'] - 1].value
        start_time = _normalize_time(start_raw)
        try:
            dur = int(float(row_cells[col_map['duration'] - 1].value or 0))
        except (ValueError, TypeError):
            dur = 0

        slot_key = (prog, start_time, dur)

        # We may need to write sub-rows for mixed statuses.
        # For each date column, determine per-status counts.
        # Structure: {status: {date_col: count}}
        # Then order: main row gets 1st status, sub-rows get the rest.
        # All sub-rows share the same left-column values (blank).

        # Collect per-date status data
        # per_date[date_str] keys: status counts + special metadata keys:
        #   'late_actual'  {actual_date: n}  – where late spots from THIS date ended up
        #   'late_to'      {actual_date: n}  – same, for comment on planned-date cell
        #   'late_from'    {orig_date: n}    – late spots FROM other dates arriving HERE
        per_date: dict = {}
        for src_col, date_str in all_date_cols.items():
            src_cell_val = row_cells[src_col - 1].value if src_col <= len(row_cells) else None
            try:
                planned_count = int(float(src_cell_val)) if src_cell_val is not None else 0
            except (ValueError, TypeError):
                planned_count = 0

            if not planned_count:
                per_date[date_str] = {}
                continue

            if status_map is None:
                per_date[date_str] = {'planned': planned_count}
            else:
                slot_data = status_map.get(slot_key, {})
                date_data = slot_data.get(date_str, {})
                if not date_data:
                    per_date[date_str] = {'planned': planned_count}
                else:
                    # Copy all non-zero numeric statuses and metadata dicts
                    per_date[date_str] = {
                        k: v for k, v in date_data.items()
                        if k not in ('late_actual', 'late_to') and isinstance(v, int) and v > 0
                    }
                    if 'late_actual' in date_data:
                        per_date[date_str]['late_actual'] = date_data['late_actual']
                    if 'late_to' in date_data:
                        per_date[date_str]['late_to'] = dict(date_data['late_to'])
                    if 'prog_note' in date_data:
                        per_date[date_str]['prog_note'] = date_data['prog_note']

        # Skip rows where every date column is zero (hidden / unused template rows).
        total_planned = sum(
            _try_int(row_cells[c - 1].value) or 0
            for c in all_date_cols
        )
        if total_planned == 0:
            continue

        # For EVERY date column (scheduled + extra), add late arrivals from other dates.
        # A spot planned for date A but aired on date B: date A → RED, date B → PURPLE +1.
        all_candidate_dates = set(all_date_cols.values()) | set(extra_date_col_map.keys())
        if status_map:
            slot_data = status_map.get(slot_key, {})
            for orig_date_str, orig_dict in slot_data.items():
                la = orig_dict.get('late_actual', {})
                for actual_date_str, late_count in la.items():
                    if late_count <= 0:
                        continue
                    if actual_date_str not in per_date:
                        per_date[actual_date_str] = {}
                    entry = per_date[actual_date_str]
                    entry['late_telecast'] = entry.get('late_telecast', 0) + late_count
                    # Track which original dates contributed (for comment)
                    if 'late_from' not in entry:
                        entry['late_from'] = {}
                    entry['late_from'][orig_date_str] = (
                        entry['late_from'].get(orig_date_str, 0) + late_count
                    )

        # Determine maximum number of sub-rows needed across all dates
        # (e.g. 3 statuses in one date → 2 sub-rows in addition to main)
        _META_KEYS = {'late_actual', 'late_to', 'late_from', 'prog_note'}
        max_sub_rows = 0
        for date_str, counts in per_date.items():
            statuses = [s for s in counts if s not in _META_KEYS and isinstance(counts[s], int) and counts[s] > 0]
            max_sub_rows = max(max_sub_rows, len(statuses) - 1)

        # Build main row + sub-rows
        # Main row: first status per date
        # Sub-rows[i]: (i+1)-th status per date (blank left columns)

        sub_rows = [[] for _ in range(max_sub_rows)]

        # Update date cells in main out_cells
        def _build_comment(date_str: str, counts: dict) -> str | None:
            """Build a comment string for a date cell, or None if not needed."""
            late_to   = counts.get('late_to',   {})
            late_from = counts.get('late_from', {})
            prog_note = counts.get('prog_note')
            parts = []
            if late_to:
                for actual, n in sorted(late_to.items()):
                    d = _fmt_date(actual)
                    parts.append(
                        f'{n} spot(s) NOT aired on this date — aired on {d} instead.'
                    )
            if late_from:
                for orig, n in sorted(late_from.items()):
                    d = _fmt_date(orig)
                    parts.append(
                        f'Includes {n} delayed spot(s) originally planned for {d}.'
                    )
            if prog_note:
                parts.append(f'⚠ {prog_note}')
            return '\n'.join(parts) if parts else None

        for src_col, date_str in all_date_cols.items():
            counts = per_date.get(date_str, {})
            statuses = _ordered_statuses(counts)
            comment_text = _build_comment(date_str, counts)
            main_entry = next((oc for oc in out_cells if oc['col'] == src_col), None)
            if not statuses:
                if main_entry:
                    main_entry['value'] = None
                continue
            if main_entry:
                s0, c0 = statuses[0]
                main_entry['value']   = c0
                main_entry['fill']    = status_to_fill.get(s0, fills['planned'])
                main_entry['comment'] = comment_text
            for i, (si, ci) in enumerate(statuses[1:]):
                sub_rows[i].append({'col': src_col, 'status': si, 'count': ci,
                                    'comment': comment_text if i == 0 else None})

        # Extra date columns (late telecast actual arrivals from dates outside schedule)
        for date_str, extra_col in extra_date_col_map.items():
            counts = per_date.get(date_str, {})
            statuses = _ordered_statuses(counts)
            comment_text = _build_comment(date_str, counts)
            main_entry = next((oc for oc in out_cells if oc['col'] == extra_col), None)
            if not statuses:
                if main_entry:
                    main_entry['value'] = None
                continue
            if main_entry:
                s0, c0 = statuses[0]
                main_entry['value']   = c0
                main_entry['fill']    = status_to_fill.get(s0, fills['planned'])
                main_entry['comment'] = comment_text
            for i, (si, ci) in enumerate(statuses[1:]):
                if i < len(sub_rows):
                    sub_rows[i].append({'col': extra_col, 'status': si, 'count': ci,
                                        'comment': comment_text if i == 0 else None})

        output_rows.append(out_cells)
        src_row_meta.append({'type': _ROW_DATA, 'n_sub': max_sub_rows})

        # Append sub-rows (blank left columns, coloured date cells only)
        for sub_list in sub_rows:
            sub_out = [{
                'col': oc['col'], 'value': None, 'fill': None, 'src_cell': None
            } for oc in out_cells if oc['col'] not in {x['col'] for x in sub_list}]
            for entry in sub_list:
                sub_out.append({
                    'col':   entry['col'],
                    'value': entry['count'],
                    'fill':  status_to_fill.get(entry['status'], fills['planned']),
                    'src_cell': None,
                })
            output_rows.append(sub_out)
            src_row_meta.append({'type': _ROW_OTHER})

    # ── Write output workbook ─────────────────────────────────────────────────
    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = ws_src.title or 'Colored Schedule'

    # Copy column widths
    for col_letter, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = dim.width
    for extra_col in extra_date_col_map.values():
        ws_dst.column_dimensions[get_column_letter(extra_col)].width = 10

    # Find last row that actually has a value, to avoid writing the legend
    # hundreds of rows below when the source file has trailing empty rows.
    last_data_row = 1
    src_rows_list = list(ws_src.iter_rows())
    for r_idx, out_cells in enumerate(output_rows, start=1):
        src_r = r_idx if r_idx <= len(src_rows_list) else None
        if src_r and src_r in ws_src.row_dimensions:
            ws_dst.row_dimensions[r_idx].height = ws_src.row_dimensions[src_r].height

        has_content = False
        for oc in sorted(out_cells, key=lambda x: x['col']):
            dst_cell = ws_dst.cell(row=r_idx, column=oc['col'])
            dst_cell.value = oc['value']
            src_c = oc.get('src_cell')
            if src_c:
                _copy_cell_style(src_c, dst_cell)
            if oc.get('fill'):
                dst_cell.fill = oc['fill']
                dst_cell.font = Font(bold=True, size=9)
                dst_cell.alignment = Alignment(horizontal='center', vertical='center')
            if oc.get('comment'):
                from openpyxl.comments import Comment
                dst_cell.comment = Comment(oc['comment'], 'Ad Monitor')
            if oc['value'] is not None:
                has_content = True
        if has_content:
            last_data_row = r_idx

    # ── Legend (vertical colour key) ──────────────────────────────────────────
    _write_legend(ws_dst, last_data_row + 2, fills)

    return wb_dst


# ── Legend helper ─────────────────────────────────────────────────────────────

def _write_legend(ws, start_row: int, fills: dict):
    """
    Write a vertical colour-key table starting at start_row.

    Layout (each row):
      Col A : coloured cell with short label
      Col B : full description
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    legend_items = [
        ('aired',              'Aired as Planned',   'Ad was detected on the scheduled date and within the planned time window'),
        ('not_aired',          'Not Aired / Missed', 'Ad was not detected on its planned date (either missed entirely or aired on a different date — hover cell for details)'),
        ('late_telecast',      'Late Arrival',       'This date received spots that were originally planned for an earlier date (hover cell for details)'),
        ('programme_mismatch', 'Programme Mismatch', 'Ad aired after the planned time window (same date)'),
        ('extra_aired',        'Extra Aired',        'Ad detected by monitoring but not present in the schedule'),
        ('aired_less',         'Aired Under Planned','Sponsorship: fewer aired than planned that day (hover cell for the missed count)'),
        ('manual_override',    'Manual Override',    'Match confirmed manually by an operator'),
        ('planned',            'Planned Only',       'Scheduled but no reconciliation data available yet'),
    ]
    # Only render legend rows for colours that exist in the fills map.
    legend_items = [it for it in legend_items if it[0] in fills]

    # Header
    hdr = ws.cell(start_row, 1, 'COLOUR LEGEND')
    hdr.font = Font(bold=True, size=10, color='FFFFFF')
    hdr.fill = PatternFill('solid', fgColor='334155')
    hdr.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    ws.row_dimensions[start_row].height = 18

    for i, (status_key, short_label, description) in enumerate(legend_items, start=1):
        r = start_row + i
        ws.row_dimensions[r].height = 16

        # Col A: coloured badge with short label
        a = ws.cell(r, 1, f'  {short_label}  ')
        a.fill = fills[status_key]
        a.font = Font(bold=True, size=9, color='FFFFFF')
        a.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        a.border = border

        # Col B: description
        b = ws.cell(r, 2, description)
        b.font = Font(size=9, color='334155')
        b.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
        b.border = border

    # Widen col B to fit descriptions
    ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width or 0, 60)
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 0, 22)


# ── Helpers ───────────────────────────────────────────────────────────────────

_STATUS_ORDER = [
    'matched', 'manual_match', 'aired_less', 'not_aired', 'late_telecast',
    'programme_mismatch', 'extra_aired', 'planned',
]


_META_KEYS_GLOBAL = {'late_actual', 'late_to', 'late_from', 'prog_note',
                     'miss_note', 'extra_note'}


def _ordered_statuses(counts: dict) -> list:
    """Return [(status, count)] in display order, skipping metadata and zero counts."""
    result = []
    seen = set()
    for s in _STATUS_ORDER:
        if s in counts and isinstance(counts[s], int) and counts[s] > 0:
            result.append((s, counts[s]))
            seen.add(s)
    for s, c in counts.items():
        if s not in seen and s not in _META_KEYS_GLOBAL and isinstance(c, int) and c > 0:
            result.append((s, c))
    return result


def _fmt_date(date_str: str) -> str:
    """'2026-03-20' → '20 Mar 2026'."""
    try:
        return _date.fromisoformat(date_str).strftime('%d %b %Y')
    except Exception:
        return date_str


# ── Combined export: Original + Colored sheets from the uploaded file ──────────

def build_original_and_colored_wb(schedule_pk, colors: dict, status_map=None,
                                  maponline_status_map=None):
    """
    Build a workbook whose FIRST sheet is the user's original uploaded schedule
    (exact copy — all formatting, merged cells, column widths, row heights and
    formulas preserved) and SECOND sheet is the same layout with reconciliation
    status colours applied IN-PLACE.

    When maponline_status_map is provided, a THIRD sheet ("MapOnline Colored") is
    added — the same pivot layout coloured from MapOnline matching results, so the
    schedule can be eyeballed against MapOnline independently of MediaWatch.

    The original Firebase file is the single source of truth for visual
    presentation; sheets are never rebuilt from database rows.  openpyxl's
    copy_worksheet() preserves cell styles, merged cells and dimensions — only
    fills + cell comments are added to the colored copy.

    Returns (workbook, detected) where `detected` is True when the pivot layout
    was recognised and colours were applied; False means the Colored sheet is an
    uncoloured (but still exact) copy of the original.
    """
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.comments import Comment
    from datetime import time as _t, datetime as _dtt
    from core.models import Schedule

    schedule = Schedule.objects.get(pk=schedule_pk)
    if not schedule.file or not schedule.file.name:
        raise FileNotFoundError(
            f'No file is linked to schedule "{schedule.original_filename}". '
            f'Please re-upload the schedule to generate the export.'
        )
    try:
        raw_bytes = schedule.file.read()
    except Exception as exc:
        raise FileNotFoundError(
            f'The original schedule file "{schedule.original_filename}" could not be '
            f'read from storage. Please re-upload the schedule.'
        ) from exc
    if not raw_bytes:
        raise FileNotFoundError(
            f'The original schedule file "{schedule.original_filename}" is empty. '
            f'Please re-upload the schedule.'
        )

    # Base workbook preserves formulas + all formatting.
    wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes))
    # The schedule lives on the first sheet (matches the upload parser, sheet 0).
    ws_orig = wb.worksheets[0]
    for extra in wb.worksheets[1:]:
        wb.remove(extra)              # keep the clean 4-sheet output structure
    ws_orig.title = 'Original Schedule'

    # A data_only copy is used purely to read numeric cell values for detection
    # and per-cell planned counts (formula cells would otherwise be strings).
    wb_vals = openpyxl.load_workbook(_io.BytesIO(raw_bytes), data_only=True)
    ws_vals = wb_vals.worksheets[0]

    # Exact-formatting copy that we colour.
    ws_color = wb.copy_worksheet(ws_orig)
    ws_color.title = 'Colored Schedule'

    fills = {
        'aired':              _hex_to_fill(colors.get('aired',              '#22c55e')),
        'not_aired':          _hex_to_fill(colors.get('not_aired',          '#ef4444')),
        'late_telecast':      _hex_to_fill(colors.get('late_telecast',      '#a855f7')),
        'programme_mismatch': _hex_to_fill(colors.get('programme_mismatch', '#f97316')),
        'extra_aired':        _hex_to_fill(colors.get('extra_aired',        '#3b82f6')),
        'planned':            _hex_to_fill(colors.get('planned',            '#94a3b8')),
        'manual_override':    _hex_to_fill(colors.get('manual_override',    '#14b8a6')),
        'aired_less':         _hex_to_fill(colors.get('aired_less',         '#f59e0b')),
    }
    status_to_fill = {
        'matched': fills['aired'], 'manual_match': fills['manual_override'],
        'not_aired': fills['not_aired'], 'late_telecast': fills['late_telecast'],
        'programme_mismatch': fills['programme_mismatch'],
        'extra_aired': fills['extra_aired'], 'planned': fills['planned'],
        'aired_less': fills['aired_less'],
    }

    header_row_idx, col_map, date_col_map = _detect_pivot_via_converter(raw_bytes)
    if not header_row_idx or not date_col_map:
        header_row_idx, col_map, date_col_map = _detect_pivot_openpyxl(ws_vals)

    if not header_row_idx or not date_col_map:
        # Layout not recognised — return the exact copy uncoloured rather than
        # rebuilding anything from the database.
        logger.warning(
            'build_original_and_colored_wb: pivot not detected for schedule pk=%s; '
            'Colored sheet is an uncoloured copy.', schedule_pk
        )
        return wb, False

    _apply_status_colors_inplace(
        ws_color, ws_vals, header_row_idx, col_map, date_col_map,
        status_map, fills, status_to_fill,
    )

    # Optional extra sheet coloured from MapOnline matching (independent source).
    if maponline_status_map is not None:
        ws_mo = wb.copy_worksheet(ws_orig)
        ws_mo.title = 'MapOnline Colored'
        _apply_status_colors_inplace(
            ws_mo, ws_vals, header_row_idx, col_map, date_col_map,
            maponline_status_map, fills, status_to_fill,
        )

    return wb, True


def _apply_status_colors_inplace(ws_color, ws_vals, header_row_idx, col_map,
                                 date_col_map, status_map, fills, status_to_fill):
    """Paint reconciliation status colours onto an exact copy of the pivot sheet.

    Shared by the MediaWatch "Colored Schedule" sheet and the "MapOnline Colored"
    sheet — both use the same {slot_key: {date: {status counts}}} status_map shape,
    so the colouring is identical; only the source of the map differs.
    """
    from datetime import time as _t, datetime as _dtt
    from openpyxl.styles import Font, Alignment
    from openpyxl.comments import Comment

    # Per-slot late arrivals INTO each date (planned date is RED, actual date PURPLE).
    late_in: dict = {}
    if status_map:
        for slot_key, date_dict in status_map.items():
            for planned_date, d in date_dict.items():
                for actual, n in (d.get('late_actual') or {}).items():
                    if n <= 0:
                        continue
                    entry = late_in.setdefault(slot_key, {}).setdefault(
                        actual, {'n': 0, 'from': {}})
                    entry['n'] += n
                    entry['from'][planned_date] = entry['from'].get(planned_date, 0) + n

    for r in range(header_row_idx + 1, ws_vals.max_row + 1):
        prog_str = str(ws_vals.cell(r, col_map['programme']).value or '').strip()
        if _is_skip_row(prog_str):
            continue
        day_val  = ws_vals.cell(r, col_map['day']).value
        time_val = ws_vals.cell(r, col_map['start_time']).value
        day_ok  = bool(day_val and str(day_val).strip())
        time_ok = isinstance(time_val, (_t, _dtt)) or bool(time_val and str(time_val).strip())
        if not (day_ok and time_ok):
            continue   # brand / section header row — leave untouched

        prog  = _normalize(prog_str)
        start = _normalize_time(time_val)
        try:
            dur = int(float(ws_vals.cell(r, col_map['duration']).value or 0))
        except (ValueError, TypeError):
            dur = 0
        slot_key      = (prog, start, dur)
        slot_data     = (status_map or {}).get(slot_key, {})
        slot_late_in  = late_in.get(slot_key, {})

        for src_col, date_str in date_col_map.items():
            raw_v = ws_vals.cell(r, src_col).value
            try:
                planned = int(float(raw_v)) if raw_v not in (None, '') else 0
            except (ValueError, TypeError):
                planned = 0

            date_data = slot_data.get(date_str, {})
            statuses  = _ordered_statuses({
                k: v for k, v in date_data.items() if k not in _META_KEYS_GLOBAL
            })
            late_here = slot_late_in.get(date_str)
            cell      = ws_color.cell(r, src_col)
            parts     = []

            # Paint a status colour ONLY on cells that actually hold a planned
            # count.  Statuses are keyed per (programme, start, dur) slot, so a
            # slot that appears on several rows would otherwise paint empty
            # cells on the duplicate rows — green boxes with no spot number.
            if statuses and planned > 0:
                primary = statuses[0][0]
                cell.fill = status_to_fill.get(primary, fills['planned'])
                for actual, n in sorted((date_data.get('late_to') or {}).items()):
                    parts.append(f'{n} spot(s) not aired on this date — aired on {_fmt_date(actual)} instead.')
                if date_data.get('miss_note'):
                    parts.append(date_data['miss_note'])
                if date_data.get('extra_note'):
                    parts.append(date_data['extra_note'])
            elif planned > 0:
                cell.fill = fills['planned']

            if late_here:
                if not statuses:
                    cell.fill = fills['late_telecast']
                    if raw_v in (None, ''):
                        cell.value = late_here['n']
                        cell.font = Font(bold=True, size=9)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                for orig, n in sorted(late_here['from'].items()):
                    parts.append(f'Includes {n} delayed spot(s) planned for {_fmt_date(orig)}.')

            if parts:
                cell.comment = Comment('\n'.join(parts), 'Ad Monitor')

    _write_legend(ws_color, ws_vals.max_row + 3, fills)
