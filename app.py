import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
from fuzzywuzzy import fuzz
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Set page configuration
st.set_page_config(page_title="Ad Airing Verification", layout="wide")

# Core helper functions
def read_excel_file(file):
 try: 
     return pd.read_excel(file)
 except Exception as e: 
     st.error(f"Error reading file: {str(e)}")
     return None

def normalize_text(text):
 return str(text).lower().strip() if not pd.isna(text) else ""

def prepare_dataframes(map_df, schedule_df, map_date_col, sch_date_col):
 # Handle potential errors with column names
 try:
     # Standardize date columns
     map_df[map_date_col] = pd.to_datetime(map_df[map_date_col])
     schedule_df[sch_date_col] = pd.to_datetime(schedule_df[sch_date_col])
     
     # Create Date column for consistency
     map_df['Date'] = map_df[map_date_col]
     schedule_df['Date'] = schedule_df[sch_date_col]
     
     # Add normalized theme columns if Advt_Theme exists
     if 'Advt_Theme' in map_df.columns:
         map_df['Normalized_Theme'] = map_df['Advt_Theme'].apply(normalize_text)
     if 'Advt_Theme' in schedule_df.columns:
         schedule_df['Normalized_Theme'] = schedule_df['Advt_Theme'].apply(normalize_text)
     
     return map_df, schedule_df
 except KeyError as e:
     st.error(f"Column not found: {e}. Please check your data and column selections.")
     st.write("Available columns in Map Online data:", ', '.join(map_df.columns))
     st.write("Available columns in Schedule data:", ', '.join(schedule_df.columns))
     return None, None

def match_ads(schedule_df, map_df, theme_mapping, date_tolerance=0):
 """Match schedule with map online data including shifted ads detection"""
 if schedule_df is None or map_df is None or schedule_df.empty or map_df.empty:
     return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
 
 # Create mapping dictionary
 sch_to_map = {}
 for mapping in theme_mapping:
     sch_theme = normalize_text(mapping['schedule_theme'])
     map_theme = normalize_text(mapping['map_online_theme'])
     sch_dur = mapping.get('schedule_duration')
     map_dur = mapping.get('map_duration')
     sch_to_map[(sch_theme, sch_dur)] = (map_theme, map_dur)
 
 # Track all matched map online rows to avoid double counting
 matched_map_indices = set()
 
 # Results containers
 matched_results = []     # Ads aired as scheduled (within date tolerance)
 shifted_ads = []         # Ads aired but on different dates than scheduled
 unmatched_schedule = []  # Ads not aired at all
 extra_aired = []         # Extra ads aired beyond schedule
 
 # Process each schedule row individually for more granular matching
 progress_bar = st.progress(0)
 total_rows = len(schedule_df)
 
 for idx, sch_row in enumerate(schedule_df.itertuples()):
     progress_bar.progress((idx + 1) / total_rows)
     
     sch_date = getattr(sch_row, 'Date')
     # Get theme - prefer normalized version if it exists
     if hasattr(sch_row, 'Normalized_Theme'):
         sch_theme = getattr(sch_row, 'Normalized_Theme')
     else:
         sch_theme = normalize_text(getattr(sch_row, 'Advt_Theme', ''))
         
     sch_dur = getattr(sch_row, 'Dur', None)
     sch_program = getattr(sch_row, 'Program', '')
     
     # Find matching map theme
     key = (sch_theme, sch_dur)
     if key not in sch_to_map:
         # Try without duration
         key = (sch_theme, None)
         if key not in sch_to_map:
             # No mapping found for this theme
             unmatched_schedule.append({
                 'Scheduled_Date': sch_date,
                 'Theme': getattr(sch_row, 'Advt_Theme', ''),
                 'Program': sch_program,
                 'Duration': sch_dur,
                 'Map_Theme': None,
                 'Reason': 'No Theme Mapping'
             })
             continue
     
     map_theme, map_dur = sch_to_map[key]
     
     # Find date range with tolerance
     date_min = sch_date - timedelta(days=date_tolerance)
     date_max = sch_date + timedelta(days=date_tolerance)
     
     # Filter map data by theme 
     # Use normalized theme if it exists, otherwise filter by the original theme
     if 'Normalized_Theme' in map_df.columns:
         map_theme_matches = map_df[map_df['Normalized_Theme'] == map_theme]
     else:
         map_theme_matches = map_df[map_df['Advt_Theme'].apply(normalize_text) == map_theme]
         
     # Further filter by duration if specified
     if map_dur is not None and 'Dur' in map_theme_matches.columns:
         map_theme_matches = map_theme_matches[map_theme_matches['Dur'] == map_dur]
     
     # Find matches within date tolerance
     date_matches = map_theme_matches[
         (map_theme_matches['Date'] >= date_min) & 
         (map_theme_matches['Date'] <= date_max)
     ]
     
     # Only consider rows not already matched
     date_matches = date_matches[~date_matches.index.isin(matched_map_indices)]
     
     if not date_matches.empty:
         # Found a match within date tolerance
         best_match_idx = date_matches.index[0]
         matched_map_indices.add(best_match_idx)
         
         best_match = map_df.loc[best_match_idx]
         
         matched_results.append({
             'Scheduled_Date': sch_date,
             'Aired_Date': best_match['Date'],
             'Program': best_match.get('Program', ''),
             'Map_Theme': best_match['Advt_Theme'],  # Use Map Online theme
             'Duration': best_match.get('Dur', sch_dur),
             'Schedule_Theme': getattr(sch_row, 'Advt_Theme', ''),
             'Status': 'Aired as Scheduled'
         })
     else:
         # Check if ad aired on a different date (shifted)
         # Exclude already matched rows and date tolerance window
         other_date_matches = map_theme_matches[
             ~map_theme_matches.index.isin(matched_map_indices) & 
             ~((map_theme_matches['Date'] >= date_min) & (map_theme_matches['Date'] <= date_max))
         ]
         
         if not other_date_matches.empty:
             # Ad aired but on a different date
             best_match_idx = other_date_matches.index[0]
             matched_map_indices.add(best_match_idx)
             
             best_match = map_df.loc[best_match_idx]
             
             shifted_ads.append({
                 'Scheduled_Date': sch_date,
                 'Aired_Date': best_match['Date'],
                 'Program': best_match.get('Program', ''),
                 'Map_Theme': best_match['Advt_Theme'],  # Use Map Online theme
                 'Duration': best_match.get('Dur', sch_dur),
                 'Schedule_Theme': getattr(sch_row, 'Advt_Theme', ''),
                 'Days_Difference': (best_match['Date'] - sch_date).days
             })
         else:
             # Ad not aired at all
             unmatched_schedule.append({
                 'Scheduled_Date': sch_date,
                 'Theme': getattr(sch_row, 'Advt_Theme', ''),
                 'Program': sch_program,
                 'Duration': sch_dur,
                 'Map_Theme': map_theme,  # Show mapped Map Online theme
                 'Reason': 'Not Aired'
             })
 
 # Find extra ads (aired but not in schedule)
 for idx, map_row in map_df.iterrows():
     if idx not in matched_map_indices:
         extra_aired.append({
             'Date': map_row['Date'],
             'Program': map_row.get('Program', ''),
             'Map_Theme': map_row['Advt_Theme'],
             'Duration': map_row.get('Dur', ''),
             'Status': 'Extra Aired'
         })
 
 # Reset progress bar
 progress_bar.empty()
 
 # Convert to dataframes
 matched_df = pd.DataFrame(matched_results) if matched_results else pd.DataFrame()
 shifted_df = pd.DataFrame(shifted_ads) if shifted_ads else pd.DataFrame()
 unmatched_df = pd.DataFrame(unmatched_schedule) if unmatched_schedule else pd.DataFrame()
 extra_df = pd.DataFrame(extra_aired) if extra_aired else pd.DataFrame()
 
 return matched_df, shifted_df, unmatched_df, extra_df

def create_pdf_report(matched_df, shifted_df, unmatched_df, extra_df, channel):
 """Generate PDF report with tables for all ad categories"""
 pdf = FPDF()
 pdf.add_page()
 
 # Title and header
 pdf.set_font("Arial", "B", 16)
 pdf.cell(0, 10, f"Ad Verification Report - {channel}", 0, 1, "C")
 pdf.set_font("Arial", "", 10)
 pdf.cell(0, 6, f"Report Date: {datetime.now().strftime('%Y-%m-%d')}", 0, 1)
 
 # Summary statistics
 pdf.ln(5)
 pdf.set_font("Arial", "B", 12)
 pdf.cell(0, 8, "Summary", 0, 1)
 pdf.set_font("Arial", "", 10)
 
 total_scheduled = len(matched_df) + len(shifted_df) + len(unmatched_df)
 total_aired = len(matched_df)
 total_shifted = len(shifted_df)
 total_missed = len(unmatched_df)
 
 pdf.cell(0, 6, f"Total Scheduled Ads: {total_scheduled}", 0, 1)
 pdf.cell(0, 6, f"Ads Aired as Scheduled: {total_aired}", 0, 1)
 pdf.cell(0, 6, f"Ads Aired on Different Date: {total_shifted}", 0, 1)
 pdf.cell(0, 6, f"Ads Not Aired: {total_missed}", 0, 1)
 pdf.cell(0, 6, f"Compliance Rate: {round((total_aired/total_scheduled)*100, 1)}%" if total_scheduled > 0 else "N/A", 0, 1)
 
 # SECTION 1: Aired Ads Table
 pdf.ln(10)
 pdf.set_font("Arial", "B", 12)
 pdf.cell(0, 8, "Ads Aired as Scheduled", 0, 1)
 
 if not matched_df.empty:
     cols = ['Scheduled_Date', 'Program', 'Map_Theme', 'Duration']
     col_widths = [30, 60, 60, 20]
     
     # Table header
     pdf.set_font("Arial", "B", 10)
     for i, col in enumerate(cols):
         pdf.cell(col_widths[i], 7, col.replace('_', ' '), 1, 0, 'C')
     pdf.ln()
     
     # Table data
     pdf.set_font("Arial", "", 9)
     for _, row in matched_df.iterrows():
         if pdf.get_y() > 270:  # Check for page break
             pdf.add_page()
             pdf.set_font("Arial", "B", 10)
             for i, col in enumerate(cols):
                 pdf.cell(col_widths[i], 7, col.replace('_', ' '), 1, 0, 'C')
             pdf.ln()
             pdf.set_font("Arial", "", 9)
         
         # Format date as string
         date_str = row['Scheduled_Date'].strftime('%Y-%m-%d') if isinstance(row['Scheduled_Date'], pd.Timestamp) else str(row['Scheduled_Date'])
         
         pdf.cell(col_widths[0], 7, date_str, 1)
         pdf.cell(col_widths[1], 7, str(row['Program'])[:30], 1)
         pdf.cell(col_widths[2], 7, str(row['Map_Theme'])[:30], 1)
         pdf.cell(col_widths[3], 7, str(row['Duration']), 1)
         pdf.ln()
 else:
     pdf.cell(0, 8, "No ads aired as scheduled.", 0, 1)
 
 # SECTION 2: Shifted Ads Table
 pdf.add_page()
 pdf.set_font("Arial", "B", 12)
 pdf.cell(0, 8, "Ads Aired on Different Date", 0, 1)
 
 if not shifted_df.empty:
     cols = ['Scheduled_Date', 'Aired_Date', 'Program', 'Map_Theme', 'Duration', 'Days_Difference']
     col_widths = [25, 25, 40, 50, 15, 20]
     
     # Table header
     pdf.set_font("Arial", "B", 10)
     for i, col in enumerate(cols):
         pdf.cell(col_widths[i], 7, col.replace('_', ' '), 1, 0, 'C')
     pdf.ln()
     
     # Table data
     pdf.set_font("Arial", "", 9)
     for _, row in shifted_df.iterrows():
         if pdf.get_y() > 270:
             pdf.add_page()
             pdf.set_font("Arial", "B", 10)
             for i, col in enumerate(cols):
                 pdf.cell(col_widths[i], 7, col.replace('_', ' '), 1, 0, 'C')
             pdf.ln()
             pdf.set_font("Arial", "", 9)
         
         # Format dates as strings
         sch_date = row['Scheduled_Date'].strftime('%Y-%m-%d') if isinstance(row['Scheduled_Date'], pd.Timestamp) else str(row['Scheduled_Date'])
         air_date = row['Aired_Date'].strftime('%Y-%m-%d') if isinstance(row['Aired_Date'], pd.Timestamp) else str(row['Aired_Date'])
         
         pdf.cell(col_widths[0], 7, sch_date, 1)
         pdf.cell(col_widths[1], 7, air_date, 1)
         pdf.cell(col_widths[2], 7, str(row['Program'])[:20], 1)
         pdf.cell(col_widths[3], 7, str(row['Map_Theme'])[:25], 1)
         pdf.cell(col_widths[4], 7, str(row['Duration']), 1)
         pdf.cell(col_widths[5], 7, str(row['Days_Difference']), 1)
         pdf.ln()
 else:
     pdf.cell(0, 8, "No ads aired on different dates.", 0, 1)
 
 # SECTION 3: Missed Ads Table
 pdf.add_page()
 pdf.set_font("Arial", "B", 12)
 pdf.cell(0, 8, "Ads Not Aired", 0, 1)
 
 if not unmatched_df.empty:
     cols = ['Scheduled_Date', 'Map_Theme', 'Program', 'Duration']
     col_widths = [30, 70, 50, 20]
     
     # Table header
     pdf.set_font("Arial", "B", 10)
     for i, col in enumerate(cols):
         pdf.cell(col_widths[i], 7, col.replace('_', ' '), 1, 0, 'C')
     pdf.ln()
     
     # Table data
     pdf.set_font("Arial", "", 9)
     for _, row in unmatched_df.iterrows():
         if pdf.get_y() > 270:
             pdf.add_page()
             pdf.set_font("Arial", "B", 10)
             for i, col in enumerate(cols):
                 pdf.cell(col_widths[i], 7, col.replace('_', ' '), 1, 0, 'C')
             pdf.ln()
             pdf.set_font("Arial", "", 9)
         
         # Format date as string
         date_str = row['Scheduled_Date'].strftime('%Y-%m-%d') if isinstance(row['Scheduled_Date'], pd.Timestamp) else str(row['Scheduled_Date'])
         
         pdf.cell(col_widths[0], 7, date_str, 1)
         pdf.cell(col_widths[1], 7, str(row['Map_Theme'])[:35], 1)  # Show mapped Map Online theme
         pdf.cell(col_widths[2], 7, str(row['Program'])[:25], 1)
         pdf.cell(col_widths[3], 7, str(row['Duration']), 1)
         pdf.ln()
 else:
     pdf.cell(0, 8, "All scheduled ads were aired.", 0, 1)
     
 return pdf

# Main application
def main():
 st.title("Ad Airing Verification Tool")
 
 # File uploads
 col1, col2 = st.columns(2)
 with col1:
     map_file = st.file_uploader("Upload Map Online Data", type=["xlsx", "xls"])
 with col2:
     schedule_file = st.file_uploader("Upload Schedule Data", type=["xlsx", "xls"])
 
 # Initialize session state
 if 'theme_mapping' not in st.session_state:
     st.session_state.theme_mapping = []
 
 # Process data if files are uploaded
 if map_file and schedule_file:
     map_df = read_excel_file(map_file)
     schedule_df = read_excel_file(schedule_file)
     
     if map_df is not None and schedule_df is not None:
         # Fix columns with spaces and standardize column names
         map_df.columns = map_df.columns.str.strip()
         schedule_df.columns = schedule_df.columns.str.strip()
         
         # Show available columns for debugging
         with st.expander("Show Available Columns"):
             st.write("Map Online Data Columns:", map_df.columns.tolist())
             st.write("Schedule Data Columns:", schedule_df.columns.tolist())
             
             # Display first few rows of data to better understand the structure
             st.write("Map Online Data Preview:")
             st.dataframe(map_df.head(3))
             st.write("Schedule Data Preview:")
             st.dataframe(schedule_df.head(3))
             
         # Detect or select columns
         st.subheader("1. Select Data Columns")
         
         # Date columns
         map_date_options = [col for col in map_df.columns if any(date_term in str(col).lower() for date_term in ['date', 'day', 'dt'])]
         sch_date_options = [col for col in schedule_df.columns if any(date_term in str(col).lower() for date_term in ['date', 'day', 'dt'])]
         
         col1, col2 = st.columns(2)
         with col1:
             map_date_idx = 0
             if map_date_options:
                 try:
                     map_date_idx = map_df.columns.get_loc(map_date_options[0])
                 except:
                     pass
             map_date_col = st.selectbox("Map Online Date Column", 
                                       options=map_df.columns, 
                                       index=map_date_idx)
         with col2:
             sch_date_idx = 0
             if sch_date_options:
                 try:
                     sch_date_idx = schedule_df.columns.get_loc(sch_date_options[0])
                 except:
                     pass
             sch_date_col = st.selectbox("Schedule Date Column", 
                                       options=schedule_df.columns,
                                       index=sch_date_idx)
         
         # Channel name handling - special case for your data
         # First see if "Channel" exists exactly
         if "Channel" in map_df.columns:
             map_channel_col = "Channel"
         else:
             # Create a new Channel column with a constant value from user input
             st.write("No 'Channel' column found in data. Please enter the channel name:")
             channel_name = st.text_input("Channel Name", "Unknown Channel")
             map_df["Channel"] = channel_name
             map_channel_col = "Channel"
         
         # Theme columns
         map_theme_options = [col for col in map_df.columns if any(theme_term in str(col).lower() for theme_term in ['theme', 'product', 'brand', 'advertiser'])]
         sch_theme_options = [col for col in schedule_df.columns if any(theme_term in str(col).lower() for theme_term in ['theme', 'product', 'brand', 'advertiser'])]
         
         col1, col2 = st.columns(2)
         with col1:
             map_theme_idx = 0
             if map_theme_options:
                 try:
                     map_theme_idx = map_df.columns.get_loc(map_theme_options[0])
                 except:
                     pass
             map_theme_col = st.selectbox("Map Online Theme/Product Column", 
                                        options=map_df.columns,
                                        index=map_theme_idx)
         with col2:
             sch_theme_idx = 0
             if sch_theme_options:
                 try:
                     sch_theme_idx = schedule_df.columns.get_loc(sch_theme_options[0])
                 except:
                     pass
             sch_theme_col = st.selectbox("Schedule Theme Column", 
                                        options=schedule_df.columns,
                                        index=sch_theme_idx)
         
         # Program columns
         map_program_options = [col for col in map_df.columns if any(prog_term in str(col).lower() for prog_term in ['prog', 'program', 'show', 'name']) and col != map_date_col]
         sch_program_options = [col for col in schedule_df.columns if any(prog_term in str(col).lower() for prog_term in ['prog', 'program', 'show'])]
         
         col1, col2 = st.columns(2)
         with col1:
             map_prog_idx = 0
             if map_program_options:
                 try:
                     map_prog_idx = map_df.columns.get_loc(map_program_options[0])
                 except:
                     pass
             map_program_col = st.selectbox("Map Online Program Column", 
                                          options=map_df.columns,
                                          index=map_prog_idx)
         with col2:
             sch_prog_idx = 0
             if sch_program_options:
                 try:
                     sch_prog_idx = schedule_df.columns.get_loc(sch_program_options[0])
                 except:
                     pass
             sch_program_col = st.selectbox("Schedule Program Column", 
                                          options=schedule_df.columns,
                                          index=sch_prog_idx)
         
         # Duration columns
         map_dur_options = [col for col in map_df.columns if any(dur_term in str(col).lower() for dur_term in ['dur', 'length', 'time'])]
         sch_dur_options = [col for col in schedule_df.columns if any(dur_term in str(col).lower() for dur_term in ['dur', 'length', 'time'])]
         
         col1, col2 = st.columns(2)
         with col1:
             map_dur_idx = 0
             if map_dur_options:
                 try:
                     map_dur_idx = map_df.columns.get_loc(map_dur_options[0])
                 except:
                     pass
             map_dur_col = st.selectbox("Map Online Duration Column", 
                                      options=map_df.columns,
                                      index=map_dur_idx)
         with col2:
             sch_dur_idx = 0
             if sch_dur_options:
                 try:
                     sch_dur_idx = schedule_df.columns.get_loc(sch_dur_options[0])
                 except:
                     pass
             sch_dur_col = st.selectbox("Schedule Duration Column", 
                                      options=schedule_df.columns,
                                      index=sch_dur_idx)
         
         # Rename key columns for consistency
         map_df = map_df.rename(columns={
             map_theme_col: 'Advt_Theme',
             map_program_col: 'Program',
             map_dur_col: 'Dur',
         })
         
         schedule_df = schedule_df.rename(columns={
             sch_theme_col: 'Advt_Theme',
             sch_program_col: 'Program',
             sch_dur_col: 'Dur',
         })
         
         # Prepare dataframes
         map_df, schedule_df = prepare_dataframes(map_df, schedule_df, map_date_col, sch_date_col)
         
         if map_df is not None and schedule_df is not None:
             # UI for filtering
             st.subheader("2. Select Channel and Date Range")
             
             # Channel selection (from Map Online data)
             # Check if the column exists and has values
             if map_channel_col in map_df.columns:
                 channels = sorted(map_df[map_channel_col].dropna().unique())
                 if len(channels) > 0:
                     selected_channel = st.selectbox("Select Channel", channels)
                 else:
                     selected_channel = st.text_input("Enter Channel Name", "Unknown Channel")
             else:
                 selected_channel = st.text_input("Enter Channel Name", "Unknown Channel")
                 # Add channel column to map_df
                 map_df[map_channel_col] = selected_channel
             
             # Date range selection
             min_date = min(map_df[map_date_col].min(), schedule_df[sch_date_col].min())
             max_date = max(map_df[map_date_col].max(), schedule_df[sch_date_col].max())
             
             col1, col2 = st.columns(2)
             with col1:
                 start_date = st.date_input("Start Date", min_date)
             with col2:
                 end_date = st.date_input("End Date", max_date)
             
             # Filter data by channel (Map Online only) and date
             if map_channel_col in map_df.columns:
                 filtered_map = map_df[
                     (map_df[map_channel_col] == selected_channel) & 
                     (map_df[map_date_col] >= pd.Timestamp(start_date)) & 
                     (map_df[map_date_col] <= pd.Timestamp(end_date))
                 ]
             else:
                 filtered_map = map_df[
                     (map_df[map_date_col] >= pd.Timestamp(start_date)) & 
                     (map_df[map_date_col] <= pd.Timestamp(end_date))
                 ]
             
             filtered_schedule = schedule_df[
                 (schedule_df[sch_date_col] >= pd.Timestamp(start_date)) & 
                 (schedule_df[sch_date_col] <= pd.Timestamp(end_date))
             ]
             
             # Theme mapping interface
             st.subheader("3. Create Theme Mappings")
             
             col1, col2 = st.columns(2)
             with col1:
                 st.write("Schedule Theme and Duration")
                 sch_themes = sorted(filtered_schedule['Advt_Theme'].dropna().unique())
                 sel_sch_theme = st.selectbox("Schedule Theme", [""] + list(sch_themes))
                 
                 # Get durations for selected theme
                 if sel_sch_theme:
                     sch_durations = sorted(filtered_schedule[filtered_schedule['Advt_Theme'] == sel_sch_theme]['Dur'].dropna().unique())
                     sel_sch_dur = st.selectbox("Schedule Duration", [""] + [str(int(d)) for d in sch_durations if pd.notna(d)])
                     if sel_sch_dur:
                         sel_sch_dur = int(sel_sch_dur)
                 else:
                     sel_sch_dur = None
             
             with col2:
                 st.write("Map Online Theme and Duration")
                 map_themes = sorted(filtered_map['Advt_Theme'].dropna().unique())
                 sel_map_theme = st.selectbox("Map Online Theme", [""] + list(map_themes))
                 
                 # Get durations for selected theme
                 if sel_map_theme:
                     map_durations = sorted(filtered_map[filtered_map['Advt_Theme'] == sel_map_theme]['Dur'].dropna().unique())
                     sel_map_dur = st.selectbox("Map Online Duration", [""] + [str(int(d)) for d in map_durations if pd.notna(d)])
                     if sel_map_dur:
                         sel_map_dur = int(sel_map_dur)
                 else:
                     sel_map_dur = None
             
             # Add mapping button
             if st.button("Add Mapping"):
                 if sel_sch_theme and sel_map_theme:
                     # Check for duplicates
                     is_duplicate = any(m['schedule_theme'] == sel_sch_theme and 
                                    m.get('schedule_duration') == sel_sch_dur and
                                    m['map_online_theme'] == sel_map_theme and
                                    m.get('map_duration') == sel_map_dur 
                                    for m in st.session_state.theme_mapping)
                     
                     if not is_duplicate:
                         mapping = {
                             'schedule_theme': sel_sch_theme,
                             'map_online_theme': sel_map_theme
                         }
                         if sel_sch_dur:
                             mapping['schedule_duration'] = sel_sch_dur
                         if sel_map_dur:
                             mapping['map_duration'] = sel_map_dur
                             
                         st.session_state.theme_mapping.append(mapping)
                         st.success("Mapping added!")
                     else:
                         st.warning("This mapping already exists.")
                 else:
                     st.warning("Please select both themes.")
             
             # Display current mappings
             if st.session_state.theme_mapping:
                 st.write("Current Mappings:")
                 mapping_data = []
                 for m in st.session_state.theme_mapping:
                     mapping_data.append({
                         'Schedule Theme': m['schedule_theme'],
                         'Schedule Duration': m.get('schedule_duration', 'Any'),
                         'Map Online Theme': m['map_online_theme'],
                         'Map Online Duration': m.get('map_duration', 'Any')
                     })
                 st.dataframe(pd.DataFrame(mapping_data))
                 
                 # Option to clear mappings
                 if st.button("Clear All Mappings"):
                     st.session_state.theme_mapping = []
                     st.success("All mappings cleared.")
             
             # Generate report
             st.subheader("4. Generate Report")
             
             date_tolerance = st.slider("Date matching tolerance (days)", 0, 7, 0)
             
             if st.button("Generate Report"):
                 if not st.session_state.theme_mapping:
                     st.warning("Please create at least one theme mapping first.")
                 else:
                     with st.spinner("Processing data..."):
                         # Match ads using the mapping
                         matched_df, shifted_df, unmatched_df, extra_df = match_ads(
                             filtered_schedule, filtered_map, 
                             st.session_state.theme_mapping,
                             date_tolerance
                         )
                         
                         # Display summary statistics
                         total_scheduled = len(matched_df) + len(shifted_df) + len(unmatched_df)
                         total_aired = len(matched_df)
                         total_shifted = len(shifted_df)
                         
                         col1, col2, col3, col4 = st.columns(4)
                         with col1:
                             st.metric("Scheduled Ads", total_scheduled)
                         with col2:
                             st.metric("Aired as Scheduled", total_aired)
                         with col3:
                             st.metric("Aired Different Date", total_shifted)
                         with col4:
                             compliance = round((total_aired/total_scheduled)*100, 1) if total_scheduled > 0 else 0
                             st.metric("Compliance", f"{compliance}%")
                         
                         # Display results in tabs
                         tab1, tab2, tab3, tab4 = st.tabs([
                             "Aired as Scheduled", 
                             "Aired on Different Date", 
                             "Not Aired", 
                             "Extra Aired"
                         ])
                         
                         with tab1:
                             if not matched_df.empty:
                                 display_cols = ['Scheduled_Date', 'Program', 'Map_Theme', 'Duration']
                                 cols_to_display = [col for col in display_cols if col in matched_df.columns]
                                 st.dataframe(matched_df[cols_to_display])
                             else:
                                 st.info("No ads aired as scheduled.")
                                 
                         with tab2:
                             if not shifted_df.empty:
                                 display_cols = ['Scheduled_Date', 'Aired_Date', 'Program', 'Map_Theme', 'Duration', 'Days_Difference']
                                 cols_to_display = [col for col in display_cols if col in shifted_df.columns]
                                 st.dataframe(shifted_df[cols_to_display])
                             else:
                                 st.info("No ads found aired on different dates.")
                                 
                         with tab3:
                             if not unmatched_df.empty:
                                 display_cols = ['Scheduled_Date', 'Map_Theme', 'Program', 'Duration']
                                 cols_to_display = [col for col in display_cols if col in unmatched_df.columns]
                                 st.dataframe(unmatched_df[cols_to_display])
                             else:
                                 st.info("All scheduled ads were aired.")
                                 
                         with tab4:
                             if not extra_df.empty:
                                 display_cols = ['Date', 'Program', 'Map_Theme', 'Duration']
                                 cols_to_display = [col for col in display_cols if col in extra_df.columns]
                                 st.dataframe(extra_df[cols_to_display])
                             else:
                                 st.info("No extra aired ads found.")
                         
                         # Generate PDF report
                         pdf = create_pdf_report(matched_df, shifted_df, unmatched_df, extra_df, selected_channel)
                         
                         # Save to buffer
                         pdf_buffer = io.BytesIO()
                         pdf_buffer.write(pdf.output(dest="S").encode("latin1"))
                         pdf_buffer.seek(0)
                         
                         # Download button
                         st.download_button(
                             "Download PDF Report",
                             data=pdf_buffer,
                             file_name=f"{selected_channel}_ad_verification_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                             mime="application/pdf"
                         )
 else:
     st.info("Please upload both Map Online and Schedule files to begin.")

if __name__ == "__main__":
 main()