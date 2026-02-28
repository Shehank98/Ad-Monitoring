"""
Ad Verification view – matches schedule against MapOnline or MediaWatch data.
Supports both file-upload mode (quick use) and stored-data mode (pull from Firestore).
"""
import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import datetime, timedelta
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from fpdf import FPDF

from components.auth import get_session_user
from components.db import get_schedules, get_monitoring_data, get_schedule_file, get_monitoring_file
from components.styles import page_header


# ── Core processing (ported from original app.py) ─────────────────────────────

def normalize_text(text):
    return str(text).lower().strip() if not pd.isna(text) else ""


def match_ads(schedule_df, map_df, theme_mapping, date_tolerance=0):
    if schedule_df is None or map_df is None or schedule_df.empty or map_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    sch_to_map = {}
    for mapping in theme_mapping:
        sch_theme = normalize_text(mapping["schedule_theme"])
        map_theme = normalize_text(mapping["map_online_theme"])
        sch_dur   = mapping.get("schedule_duration")
        map_dur   = mapping.get("map_duration")
        sch_to_map[(sch_theme, sch_dur)] = (map_theme, map_dur)

    matched_map_indices = set()
    matched_results = []
    shifted_ads     = []
    unmatched_schedule = []
    extra_aired     = []

    progress = st.progress(0)
    total    = len(schedule_df)

    for idx, sch_row in enumerate(schedule_df.itertuples()):
        progress.progress((idx + 1) / total)
        sch_date    = getattr(sch_row, "Date")
        sch_theme   = normalize_text(getattr(sch_row, "Advt_Theme", ""))
        sch_dur     = getattr(sch_row, "Dur", None)
        sch_program = getattr(sch_row, "Program", "")

        key = (sch_theme, sch_dur)
        if key not in sch_to_map:
            key = (sch_theme, None)
        if key not in sch_to_map:
            unmatched_schedule.append({
                "Scheduled_Date": sch_date,
                "Theme":          getattr(sch_row, "Advt_Theme", ""),
                "Program":        sch_program,
                "Duration":       sch_dur,
                "Map_Theme":      None,
                "Reason":         "No Theme Mapping",
            })
            continue

        map_theme, map_dur = sch_to_map[key]
        date_min = sch_date - timedelta(days=date_tolerance)
        date_max = sch_date + timedelta(days=date_tolerance)

        if "Normalized_Theme" in map_df.columns:
            theme_matches = map_df[map_df["Normalized_Theme"] == map_theme]
        else:
            theme_matches = map_df[map_df["Advt_Theme"].apply(normalize_text) == map_theme]

        if map_dur is not None and "Dur" in theme_matches.columns:
            theme_matches = theme_matches[theme_matches["Dur"] == map_dur]

        date_matches = theme_matches[
            (theme_matches["Date"] >= date_min) & (theme_matches["Date"] <= date_max)
        ]
        date_matches = date_matches[~date_matches.index.isin(matched_map_indices)]

        if not date_matches.empty:
            best_idx = date_matches.index[0]
            matched_map_indices.add(best_idx)
            best = map_df.loc[best_idx]
            matched_results.append({
                "Scheduled_Date": sch_date,
                "Aired_Date":     best["Date"],
                "Program":        best.get("Program", ""),
                "Map_Theme":      best["Advt_Theme"],
                "Duration":       best.get("Dur", sch_dur),
                "Schedule_Theme": getattr(sch_row, "Advt_Theme", ""),
                "Status":         "Aired as Scheduled",
            })
        else:
            other = theme_matches[
                ~theme_matches.index.isin(matched_map_indices) &
                ~((theme_matches["Date"] >= date_min) & (theme_matches["Date"] <= date_max))
            ]
            if not other.empty:
                best_idx = other.index[0]
                matched_map_indices.add(best_idx)
                best = map_df.loc[best_idx]
                shifted_ads.append({
                    "Scheduled_Date":  sch_date,
                    "Aired_Date":      best["Date"],
                    "Program":         best.get("Program", ""),
                    "Map_Theme":       best["Advt_Theme"],
                    "Duration":        best.get("Dur", sch_dur),
                    "Schedule_Theme":  getattr(sch_row, "Advt_Theme", ""),
                    "Days_Difference": (best["Date"] - sch_date).days,
                })
            else:
                unmatched_schedule.append({
                    "Scheduled_Date": sch_date,
                    "Theme":          getattr(sch_row, "Advt_Theme", ""),
                    "Program":        sch_program,
                    "Duration":       sch_dur,
                    "Map_Theme":      map_theme,
                    "Reason":         "Not Aired",
                })

    for idx, row in map_df.iterrows():
        if idx not in matched_map_indices:
            extra_aired.append({
                "Date":     row["Date"],
                "Program":  row.get("Program", ""),
                "Map_Theme": row["Advt_Theme"],
                "Duration": row.get("Dur", ""),
                "Status":   "Extra Aired",
            })

    progress.empty()

    return (
        pd.DataFrame(matched_results)    if matched_results    else pd.DataFrame(),
        pd.DataFrame(shifted_ads)        if shifted_ads        else pd.DataFrame(),
        pd.DataFrame(unmatched_schedule) if unmatched_schedule else pd.DataFrame(),
        pd.DataFrame(extra_aired)        if extra_aired        else pd.DataFrame(),
    )


def prepare_dataframes(map_df, schedule_df, map_date_col, sch_date_col):
    try:
        map_df[map_date_col]      = pd.to_datetime(map_df[map_date_col])
        schedule_df[sch_date_col] = pd.to_datetime(schedule_df[sch_date_col])
        map_df["Date"]            = map_df[map_date_col]
        schedule_df["Date"]       = schedule_df[sch_date_col]
        if "Advt_Theme" in map_df.columns:
            map_df["Normalized_Theme"]      = map_df["Advt_Theme"].apply(normalize_text)
        if "Advt_Theme" in schedule_df.columns:
            schedule_df["Normalized_Theme"] = schedule_df["Advt_Theme"].apply(normalize_text)
        return map_df, schedule_df
    except KeyError as e:
        st.error(f"Column not found: {e}")
        return None, None


def create_pdf_report(matched_df, shifted_df, unmatched_df, extra_df, channel):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, f"Ad Verification Report – {channel}", 0, 1, "C")
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 6, f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}", 0, 1)
    pdf.ln(4)

    total_sch = len(matched_df) + len(shifted_df) + len(unmatched_df)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "Summary", 0, 1)
    pdf.set_font("Arial", "", 10)
    for line in [
        f"Total Scheduled: {total_sch}",
        f"Aired as Scheduled: {len(matched_df)}",
        f"Aired on Different Date: {len(shifted_df)}",
        f"Not Aired: {len(unmatched_df)}",
        f"Extra Aired: {len(extra_df)}",
        f"Compliance Rate: {round(len(matched_df)/total_sch*100,1)}%" if total_sch else "N/A",
    ]:
        pdf.cell(0, 6, line, 0, 1)

    def _table(title, df, cols, widths):
        pdf.add_page()
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, title, 0, 1)
        if df.empty:
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 8, "None.", 0, 1)
            return
        pdf.set_font("Arial", "B", 9)
        for c, w in zip(cols, widths):
            pdf.cell(w, 7, c.replace("_", " "), 1, 0, "C")
        pdf.ln()
        pdf.set_font("Arial", "", 8)
        for _, row in df.iterrows():
            if pdf.get_y() > 270:
                pdf.add_page()
                pdf.set_font("Arial", "B", 9)
                for c, w in zip(cols, widths):
                    pdf.cell(w, 7, c.replace("_", " "), 1, 0, "C")
                pdf.ln()
                pdf.set_font("Arial", "", 8)
            for c, w in zip(cols, widths):
                val = row.get(c, "")
                if isinstance(val, pd.Timestamp):
                    val = val.strftime("%Y-%m-%d")
                pdf.cell(w, 7, str(val)[:int(w/2)], 1)
            pdf.ln()

    _table("Aired as Scheduled",       matched_df,   ["Scheduled_Date","Program","Map_Theme","Duration"], [35,55,65,20])
    _table("Aired on Different Date",  shifted_df,   ["Scheduled_Date","Aired_Date","Program","Map_Theme","Days_Difference"], [30,30,45,55,20])
    _table("Not Aired",                unmatched_df, ["Scheduled_Date","Map_Theme","Program","Duration"], [35,65,55,20])
    _table("Extra Aired",              extra_df,     ["Date","Program","Map_Theme","Duration"],           [35,55,65,20])
    return pdf


# ── MapOnline-specific column handling ────────────────────────────────────────

def _prepare_maponline(df: pd.DataFrame) -> pd.DataFrame:
    """
    MapOnline uses 'Theme' as the ad theme column and 'Prg Date' as the date.
    Rename to the system's internal column names.
    """
    df = df.copy()
    rename = {}
    if "Theme"     in df.columns and "Advt_Theme" not in df.columns:
        rename["Theme"]    = "Advt_Theme"
    if "Prg Date"  in df.columns and "Date"       not in df.columns:
        rename["Prg Date"] = "_PrgDate"
    if "Prg Name"  in df.columns and "Program"    not in df.columns:
        rename["Prg Name"] = "Program"
    if "Ad Dur"    in df.columns and "Dur"         not in df.columns:
        rename["Ad Dur"]   = "Dur"
    df.rename(columns=rename, inplace=True)
    return df


def _prepare_mediawatch(df: pd.DataFrame) -> pd.DataFrame:
    """
    MediaWatch stores date across Dd/Mn/Yr columns and theme in Advt_Theme.
    Build a Date column and rename Program/Duration.
    """
    df = df.copy()
    if {"Dd", "Mn", "Yr"}.issubset(df.columns) and "Date" not in df.columns:
        df["Date"] = pd.to_datetime(
            df["Yr"].astype(str) + "-" + df["Mn"].astype(str).str.zfill(2) + "-" + df["Dd"].astype(str).str.zfill(2),
            errors="coerce",
        )
    if "Program" not in df.columns and "Program" in df.columns:
        pass  # already correct
    if "Dur" not in df.columns and "Dur" in df.columns:
        pass
    return df


# ── UI: file-upload mode ──────────────────────────────────────────────────────

def _file_upload_mode():
    data_source = st.radio(
        "Monitoring data source",
        ["MapOnline", "MediaWatch (LMRB)"],
        horizontal=True,
    )
    col1, col2 = st.columns(2)
    with col1:
        mon_file  = st.file_uploader(
            f"Upload {data_source} Data", type=["xlsx", "xls"], key="mon_upload")
    with col2:
        sch_file  = st.file_uploader(
            "Upload Schedule Data",        type=["xlsx", "xls"], key="sch_upload")
    return mon_file, sch_file, data_source


# ── UI: stored-data mode ──────────────────────────────────────────────────────

def _stored_data_mode():
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**Select Monitoring Data**")
        data_source = st.radio("Source", ["MapOnline", "MediaWatch (LMRB)"],
                               horizontal=True, key="stored_src")
        dtype_key   = "maponline" if "MapOnline" in data_source else "mediawatch"
        mon_list    = get_monitoring_data(data_type=dtype_key)
        if not mon_list:
            st.warning("No monitoring data uploaded yet.")
            return None, None, data_source
        mon_options = {
            f"{d['channel']} | {d['start_date']} → {d['end_date']} | {d['filename']}": d["doc_id"]
            for d in mon_list
        }
        sel_mon = st.selectbox("Monitoring dataset", list(mon_options.keys()))
        mon_doc_id = mon_options[sel_mon]

    with col2:
        st.markdown("**Select Schedule**")
        sch_list = get_schedules()
        if not sch_list:
            st.warning("No schedules uploaded yet.")
            return None, None, data_source
        sch_options = {
            f"{s['account']} | {s['channel']} | {s['month']} | #{s['schedule_number']} | {s['filename']}": s["doc_id"]
            for s in sch_list
        }
        sel_sch    = st.selectbox("Schedule", list(sch_options.keys()))
        sch_doc_id = sch_options[sel_sch]

    mon_bytes = get_monitoring_file(mon_doc_id)
    sch_bytes = get_schedule_file(sch_doc_id)

    if mon_bytes is None or sch_bytes is None:
        st.error("Could not load the selected data from the system.")
        return None, None, data_source

    mon_io = io.BytesIO(mon_bytes)
    sch_io = io.BytesIO(sch_bytes)
    return mon_io, sch_io, data_source


# ── Main verification workflow ────────────────────────────────────────────────

def show():
    page_header("Verify Ads", "Match scheduled ads against monitoring data")

    if "theme_mapping" not in st.session_state:
        st.session_state.theme_mapping = []

    # Data source selection
    mode = st.radio("Data input mode", ["Upload Files", "Use Stored Data"],
                    horizontal=True, key="verify_mode")

    if mode == "Upload Files":
        mon_file, sch_file, data_source = _file_upload_mode()
    else:
        mon_file, sch_file, data_source = _stored_data_mode()

    if mon_file is None or sch_file is None:
        st.info("Select monitoring data and schedule to continue.")
        return

    try:
        mon_df = pd.read_excel(mon_file)
        sch_df = pd.read_excel(sch_file)
    except Exception as e:
        st.error(f"Error reading files: {e}")
        return

    mon_df.columns = mon_df.columns.str.strip()
    sch_df.columns = sch_df.columns.str.strip()

    # Pre-process source-specific columns
    if "MapOnline" in data_source:
        mon_df = _prepare_maponline(mon_df)
    else:
        mon_df = _prepare_mediawatch(mon_df)

    with st.expander("Preview loaded data"):
        c1, c2 = st.columns(2)
        c1.write(f"**Monitoring data** – {len(mon_df):,} rows")
        c1.dataframe(mon_df.head(3))
        c2.write(f"**Schedule** – {len(sch_df):,} rows")
        c2.dataframe(sch_df.head(3))

    # ── Step 1: column mapping ──────────────────────────────────────────────
    st.subheader("1. Column Mapping")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**Monitoring Data columns**")
        # Date column
        date_guess = next((c for c in mon_df.columns if "date" in c.lower() or c in ("_PrgDate", "Date")), mon_df.columns[0])
        mon_date_col = st.selectbox("Date column (monitoring)", mon_df.columns,
                                    index=list(mon_df.columns).index(date_guess) if date_guess in mon_df.columns else 0)
        if "_PrgDate" in mon_df.columns and mon_date_col == "_PrgDate":
            mon_df["Date"] = pd.to_datetime(mon_df["_PrgDate"], errors="coerce")

        theme_guess = "Advt_Theme" if "Advt_Theme" in mon_df.columns else mon_df.columns[0]
        mon_theme_col = st.selectbox("Theme / Product column (monitoring)", mon_df.columns,
                                     index=list(mon_df.columns).index(theme_guess) if theme_guess in mon_df.columns else 0)
        prog_guess = "Program" if "Program" in mon_df.columns else mon_df.columns[0]
        mon_prog_col  = st.selectbox("Programme column (monitoring)", mon_df.columns,
                                     index=list(mon_df.columns).index(prog_guess) if prog_guess in mon_df.columns else 0)
        dur_guess = "Dur" if "Dur" in mon_df.columns else ("Ad Dur" if "Ad Dur" in mon_df.columns else mon_df.columns[0])
        mon_dur_col   = st.selectbox("Duration column (monitoring)", mon_df.columns,
                                     index=list(mon_df.columns).index(dur_guess) if dur_guess in mon_df.columns else 0)

    with col2:
        st.markdown("**Schedule columns**")
        sch_date_guess = next((c for c in sch_df.columns if "date" in c.lower()), sch_df.columns[0])
        sch_date_col   = st.selectbox("Date column (schedule)", sch_df.columns,
                                      index=list(sch_df.columns).index(sch_date_guess) if sch_date_guess in sch_df.columns else 0)
        brand_guess = "Brand" if "Brand" in sch_df.columns else sch_df.columns[0]
        sch_theme_col  = st.selectbox("Brand / Theme column (schedule)", sch_df.columns,
                                      index=list(sch_df.columns).index(brand_guess) if brand_guess in sch_df.columns else 0)
        prog_sch_guess = "Programme" if "Programme" in sch_df.columns else ("Program" if "Program" in sch_df.columns else sch_df.columns[0])
        sch_prog_col   = st.selectbox("Programme column (schedule)", sch_df.columns,
                                      index=list(sch_df.columns).index(prog_sch_guess) if prog_sch_guess in sch_df.columns else 0)
        dur_sch_guess  = "Duration" if "Duration" in sch_df.columns else sch_df.columns[0]
        sch_dur_col    = st.selectbox("Duration column (schedule)", sch_df.columns,
                                      index=list(sch_df.columns).index(dur_sch_guess) if dur_sch_guess in sch_df.columns else 0)

    # Rename to internal names
    mon_df = mon_df.rename(columns={mon_theme_col: "Advt_Theme",
                                     mon_prog_col:   "Program",
                                     mon_dur_col:    "Dur"})
    sch_df = sch_df.rename(columns={sch_theme_col:  "Advt_Theme",
                                     sch_prog_col:   "Program",
                                     sch_dur_col:    "Dur"})

    mon_df, sch_df = prepare_dataframes(mon_df, sch_df, mon_date_col, sch_date_col)
    if mon_df is None:
        return

    # ── Step 2: channel & date filter ──────────────────────────────────────
    st.subheader("2. Channel & Date Range")

    if "Channel" in mon_df.columns:
        channels = sorted(mon_df["Channel"].dropna().unique())
        selected_channel = st.selectbox("Channel", channels) if channels else st.text_input("Channel", "Unknown")
    else:
        selected_channel = st.text_input("Channel name", "Unknown Channel")
        mon_df["Channel"] = selected_channel

    min_date = min(mon_df[mon_date_col].min(), sch_df[sch_date_col].min())
    max_date = max(mon_df[mon_date_col].max(), sch_df[sch_date_col].max())
    c1, c2  = st.columns(2)
    start_d = c1.date_input("Start Date", min_date)
    end_d   = c2.date_input("End Date",   max_date)

    if "Channel" in mon_df.columns:
        filt_mon = mon_df[
            (mon_df["Channel"] == selected_channel) &
            (mon_df[mon_date_col] >= pd.Timestamp(start_d)) &
            (mon_df[mon_date_col] <= pd.Timestamp(end_d))
        ]
    else:
        filt_mon = mon_df[
            (mon_df[mon_date_col] >= pd.Timestamp(start_d)) &
            (mon_df[mon_date_col] <= pd.Timestamp(end_d))
        ]
    filt_sch = sch_df[
        (sch_df[sch_date_col] >= pd.Timestamp(start_d)) &
        (sch_df[sch_date_col] <= pd.Timestamp(end_d))
    ]
    st.caption(f"Filtered: {len(filt_mon):,} monitoring rows | {len(filt_sch):,} schedule rows")

    # ── Step 3: theme mapping ───────────────────────────────────────────────
    st.subheader("3. Theme Mapping")
    st.info(
        "Map each **Schedule theme/brand** to the corresponding theme name "
        "in the monitoring data. You can add multiple mappings."
    )

    col1, col2 = st.columns(2)
    with col1:
        st.write("**Schedule Side**")
        sch_themes = sorted(filt_sch["Advt_Theme"].dropna().unique())
        sel_sch_th = st.selectbox("Schedule Theme", [""] + list(sch_themes), key="th_sch")
        if sel_sch_th:
            sch_durs = sorted(filt_sch[filt_sch["Advt_Theme"] == sel_sch_th]["Dur"].dropna().unique())
            sel_sch_dur = st.selectbox("Duration (optional)", ["Any"] + [str(int(d)) for d in sch_durs], key="dur_sch")
            sel_sch_dur = int(sel_sch_dur) if sel_sch_dur != "Any" else None
        else:
            sel_sch_dur = None

    with col2:
        st.write("**Monitoring Data Side**")
        mon_themes = sorted(filt_mon["Advt_Theme"].dropna().unique())
        sel_mon_th = st.selectbox("Monitoring Theme", [""] + list(mon_themes), key="th_mon")
        if sel_mon_th:
            mon_durs = sorted(filt_mon[filt_mon["Advt_Theme"] == sel_mon_th]["Dur"].dropna().unique())
            sel_mon_dur = st.selectbox("Duration (optional)", ["Any"] + [str(int(d)) for d in mon_durs], key="dur_mon")
            sel_mon_dur = int(sel_mon_dur) if sel_mon_dur != "Any" else None
        else:
            sel_mon_dur = None

    if st.button("➕ Add Mapping"):
        if sel_sch_th and sel_mon_th:
            dup = any(
                m["schedule_theme"] == sel_sch_th and
                m.get("schedule_duration") == sel_sch_dur and
                m["map_online_theme"] == sel_mon_th and
                m.get("map_duration") == sel_mon_dur
                for m in st.session_state.theme_mapping
            )
            if not dup:
                mapping = {"schedule_theme": sel_sch_th, "map_online_theme": sel_mon_th}
                if sel_sch_dur: mapping["schedule_duration"] = sel_sch_dur
                if sel_mon_dur: mapping["map_duration"]      = sel_mon_dur
                st.session_state.theme_mapping.append(mapping)
                st.success("Mapping added!")
            else:
                st.warning("Duplicate mapping – already exists.")
        else:
            st.warning("Select both themes before adding a mapping.")

    if st.session_state.theme_mapping:
        rows = []
        for m in st.session_state.theme_mapping:
            rows.append({
                "Schedule Theme":    m["schedule_theme"],
                "Sch Duration":      m.get("schedule_duration", "Any"),
                "Monitoring Theme":  m["map_online_theme"],
                "Mon Duration":      m.get("map_duration", "Any"),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        if st.button("🗑 Clear All Mappings"):
            st.session_state.theme_mapping = []
            st.rerun()

    # ── Step 4: run ─────────────────────────────────────────────────────────
    st.subheader("4. Run Verification")
    date_tol = st.slider("Date matching tolerance (days)", 0, 7, 0)

    if st.button("▶ Run Verification", type="primary"):
        if not st.session_state.theme_mapping:
            st.warning("Add at least one theme mapping first.")
            return

        with st.spinner("Processing…"):
            matched, shifted, unmatched, extra = match_ads(
                filt_sch, filt_mon, st.session_state.theme_mapping, date_tol
            )

        total_sch = len(matched) + len(shifted) + len(unmatched)
        compliance = round(len(matched) / total_sch * 100, 1) if total_sch else 0

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Scheduled",          total_sch)
        c2.metric("Aired as Scheduled", len(matched))
        c3.metric("Aired (Shifted)",    len(shifted))
        c4.metric("Compliance",         f"{compliance}%")

        tab1, tab2, tab3, tab4 = st.tabs([
            f"✅ Aired ({len(matched)})",
            f"⏰ Shifted ({len(shifted)})",
            f"❌ Not Aired ({len(unmatched)})",
            f"➕ Extra ({len(extra)})",
        ])
        with tab1:
            st.dataframe(matched,   use_container_width=True, hide_index=True) if not matched.empty   else st.info("None.")
        with tab2:
            st.dataframe(shifted,   use_container_width=True, hide_index=True) if not shifted.empty   else st.info("None.")
        with tab3:
            st.dataframe(unmatched, use_container_width=True, hide_index=True) if not unmatched.empty else st.info("None.")
        with tab4:
            st.dataframe(extra,     use_container_width=True, hide_index=True) if not extra.empty     else st.info("None.")

        # PDF export
        pdf     = create_pdf_report(matched, shifted, unmatched, extra, selected_channel)
        buf     = io.BytesIO()
        buf.write(pdf.output(dest="S").encode("latin1"))
        buf.seek(0)
        st.download_button(
            "⬇ Download PDF Report",
            data=buf,
            file_name=f"{selected_channel}_verification_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
        )
